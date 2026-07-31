#!/usr/bin/env python3
"""
Sample MNTN deliverable — exercises every sheet type in mntn_xlsx so you can open it and judge the
LOOK. All numbers are SYNTHETIC (clearly labelled), entities are placeholders. Regenerate with:

    python3 lib/mntn_xlsx_demo.py

Writes to the scratchpad and to My Drive/Tickets/_FORMAT_SAMPLE/ so it opens from anywhere.
"""

import os
import sys

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from lib.mntn_xlsx import FMT, MntnWorkbook, rag_threshold

GEN = "2026-07-21"

# --- synthetic headline table: MM vs 3P by vertical -------------------------
rows = [
    # vertical,            mm_ivr, tp_ivr, mm_cpv, tp_cpv, mm_roas, tp_roas, mm_adv, tp_adv
    ("Retail & Ecommerce", 0.0061, 0.0011, 3.85, 11.90, 3.4, 1.7, 142, 88),
    ("Auto", 0.0048, 0.0009, 4.60, 13.10, 2.9, 1.5, 63, 40),
    ("Financial Services", 0.0044, 0.0012, 5.10, 12.40, 2.6, 1.4, 55, 33),
    ("Health & Wellness", 0.0058, 0.0013, 3.95, 10.80, 3.1, 1.6, 71, 49),
    ("Home & Garden", 0.0052, 0.0010, 4.20, 12.90, 3.0, 1.5, 60, 37),
    ("Travel", 0.0046, 0.0011, 4.75, 11.60, 2.7, 1.5, 38, 24),
    ("Professional Svcs", 0.0039, 0.0009, 5.40, 13.80, 2.4, 1.3, 44, 27),
    ("Media & Content", 0.0055, 0.0012, 4.05, 11.20, 3.2, 1.6, 33, 21),
]
head = pd.DataFrame(
    [
        {
            "Sales vertical": v,
            "MM IVR": mi,
            "3P IVR": ti,
            "IVR advantage": mi / ti,
            "MM CPV": mc,
            "3P CPV": tc,
            "MM ROAS": mr,
            "3P ROAS": tr,
            "IVR advantage.": None,
            "MM advertisers": ma,
            "3P advertisers": ta,
        }
        for (v, mi, ti, mc, tc, mr, tr, ma, ta) in rows
    ]
)
head = head.drop(columns=["IVR advantage."]).sort_values("IVR advantage", ascending=False)

# --- synthetic full scorecard: every group x vertical -----------------------
groups = ["MM (gated)", "MM (no gate)", "MM restricted", "3P"]
mult = {"MM (gated)": 1.0, "MM (no gate)": 0.55, "MM restricted": 0.7, "3P": 0.2}
full_rows = []
for v, mi, ti, mc, tc, mr, tr, ma, ta in rows:
    for g in groups:
        k = mult[g]
        full_rows.append(
            {
                "Sales vertical": v,
                "Group": g,
                "Advertisers": int(ma * (0.9 if g.startswith("MM") else 1.0) if g != "3P" else ta),
                "Campaigns": int((ma if g != "3P" else ta) * 1.8),
                "Spend": round((ma if g != "3P" else ta) * 21000 * k),
                "Impressions": int((ma if g != "3P" else ta) * 5_100_000 * k),
                "IVR (median)": round(mi * (k if g != "3P" else 1.0) if g != "3P" else ti, 5),
                "CPV (median)": round(mc / (k if g != "3P" else 1.0) if g != "3P" else tc, 2),
                "ROAS (median)": round(mr * (k if g != "3P" else 1.0) if g != "3P" else tr, 2),
            }
        )
full = pd.DataFrame(full_rows)

# --- synthetic detail rows --------------------------------------------------
detail = pd.DataFrame(
    [
        {
            "Advertiser": f"Advertiser {chr(65 + i)}",
            "Campaign ID": 400000 + i,
            "Sales vertical": rows[i % len(rows)][0],
            "Group": groups[i % 4],
            "Impressions": 5_200_000 - i * 130_000,
            "Visits": int((5_200_000 - i * 130_000) * (0.0060 - i * 0.0003)),
            "Spend": 92000 - i * 4200,
            "IVR": 0.0060 - i * 0.0003,
            "CPV": 3.9 + i * 0.7,
            "ROAS": 3.4 - i * 0.18,
        }
        for i in range(11)
    ]
)

# ---------------------------------------------------------------------------
wb = MntnWorkbook(
    title="Segment Performance Scorecard",
    ticket="SAMPLE-000",
    subtitle="FORMAT SAMPLE — synthetic data. Shows the standard MNTN .xlsx look.",
    period="Jan-Jun 2026 (illustrative)",
    generated=GEN,
    status="Sample",
)

wb.table(
    "MM vs 3P by vertical",
    head,
    finding="MNTN Matched leads visit rate in every vertical",
    method="Advertiser-weighted medians. Visit rate higher is better; cost per visit lower is better. Synthetic sample data.",
    formats={
        "MM IVR": FMT.PCT2,
        "3P IVR": FMT.PCT2,
        "IVR advantage": FMT.MULT,
        "MM CPV": FMT.USD,
        "3P CPV": FMT.USD,
        "MM ROAS": FMT.ROAS,
        "3P ROAS": FMT.ROAS,
        "MM advertisers": FMT.INT,
        "3P advertisers": FMT.INT,
    },
    heat={
        "MM IVR": "high",
        "3P IVR": "high",
    },  # heat and rag never on the same column (they collide in Excel)
    rag={"IVR advantage": rag_threshold(good_above=4.0, bad_below=2.0)},
    kind="headline",
    widths={"Sales vertical": 22},
    toc="The headline — MM vs 3P visit rate, cost per visit and ROAS by vertical",
    query="sample_cohort.sql",  # names + deep-links this sheet's source query on the Queries tab
)

wb.table(
    "Full scorecard",
    full,
    finding="Every audience group by vertical",
    method="Gated MM leads; un-gated and restricted MM fall between gated MM and 3P. Synthetic sample data.",
    formats={
        "Advertisers": FMT.INT,
        "Campaigns": FMT.INT,
        "Spend": FMT.USD0,
        "Impressions": FMT.INT,
        "IVR (median)": FMT.PCT2,
        "CPV (median)": FMT.USD,
        "ROAS (median)": FMT.ROAS,
    },
    heat={"IVR (median)": "high", "CPV (median)": "low", "ROAS (median)": "high"},
    kind="data",
    widths={"Sales vertical": 22, "Group": 15, "Spend": 12, "Impressions": 13},
    freeze="C",
    toc="Full breakout — all four groups across every vertical",
)

wb.table(
    "Segment detail",
    detail,
    finding="Campaign-level rows for your own pivots",
    method="Raw sample rows. Filter and pivot freely. Synthetic sample data.",
    formats={
        "Impressions": FMT.INT,
        "Visits": FMT.INT,
        "Spend": FMT.USD0,
        "IVR": FMT.PCT2,
        "CPV": FMT.USD,
        "ROAS": FMT.ROAS,
    },
    band=False,
    kind="detail",
    widths={"Advertiser": 16},
    toc="Row-level detail for auditing and custom pivots",
)

wb.glossary(
    "Read me",
    intro="One workbook, one look. Here's how to read it and what each metric means.",
    rows=[
        (
            "How to use this",
            "Start on Overview for the takeaways and a clickable index. Each data tab states its finding at the top, with the method just under it in grey. Everything is filterable; the SQL is on the Queries tab.",
        ),
        ("", ""),
        ("Metrics", "All rates are over impressions."),
        ("IVR", "Visit rate. Visits divided by impressions (visits = views + clicks)."),
        ("CPV", "Cost per visit. Spend divided by visits. Lower is better."),
        ("ROAS", "Revenue divided by spend. Directional for prospecting."),
        (
            "IVR advantage",
            "MM visit rate divided by 3P visit rate, shown as a multiple (e.g. 4.1x).",
        ),
        ("", ""),
        (
            "Weighting",
            "Headline numbers are advertiser-weighted medians so one large account can't set the result.",
        ),
        (
            "Colour",
            "Darker green = larger on that metric (per-column magnitude). On lift columns a semantic scheme applies: red = significant negative, amber = not significant, green deepens with lift. Navy headers, banded rows.",
        ),
        ("", ""),
        (
            "This file",
            "FORMAT SAMPLE with synthetic data — it exists to show the standard MNTN .xlsx look, not to report real numbers.",
        ),
    ],
)

wb.sql(
    "Queries",
    "-- sample_cohort.sql - SAMPLE cohort SQL (illustrative)\n"
    "SELECT\n"
    "  sales_vertical,\n"
    "  audience_group,\n"
    "  APPROX_QUANTILES(visits / impressions, 100)[OFFSET(50)] AS ivr_median,\n"
    "  SUM(spend) / NULLIF(SUM(visits), 0)                      AS cpv_pooled,\n"
    "  COUNT(DISTINCT advertiser_id)                            AS advertisers\n"
    "FROM `dw-main-silver.summarydata.sum_by_campaign_by_day`\n"
    "WHERE DATE(day) BETWEEN '2026-01-01' AND '2026-06-30'\n"
    "  AND objective_id IN (1, 5, 6)   -- prospecting\n"
    "GROUP BY sales_vertical, audience_group\n"
    "ORDER BY sales_vertical, audience_group;",
    note="BigQuery SQL used to produce these numbers. Copy-paste to re-run and audit. (Sample.)",
)

wb.notes(
    "Method & caveats",
    intro="Read before quoting any number.",
    blocks=[
        (
            "Advertiser-weighted medians",
            "Each advertiser counts once. This keeps a single large account from setting the headline. Columns labelled 'pooled' are impression- or spend-weighted instead.",
        ),
        (
            "ROAS is directional",
            "Prospecting, last-touch. Revenue mostly lands in retargeting (excluded here), and some pixels are unreliable. Visit rate and cost per visit are the solid metrics.",
        ),
        (
            "Minimum volume",
            "Advertisers under 20,000 impressions in a group are dropped from the median so a tiny campaign can't swing it.",
        ),
        (
            "This is a sample",
            "All figures are synthetic. The purpose is to demonstrate the shared format, not to report findings.",
        ),
    ],
)

wb.cover(
    takeaways=[
        "MNTN Matched delivers roughly 4x the visit rate of 3P segments, for the typical advertiser.",
        "The advantage holds in all eight verticals — it is not driven by one account or category.",
        "Configured well (intent gate on, audience broad), the gap widens to about 6x on visit rate.",
    ]
)

# --- save + verify ----------------------------------------------------------
scratch = os.environ.get(
    "CLAUDE_SCRATCH",
    "/private/tmp/claude-501/-Users-malachi-Developer-work-mntn-workspace/3b55570d-c509-4bdc-a8b6-68fa3f480871/scratchpad",
)
local = wb.save_local(os.path.join(scratch, "mntn_format_sample.xlsx"))
print("saved local :", local)
try:
    drive = wb.save_drive("_FORMAT_SAMPLE", "MNTN xlsx Format Sample")
    print("saved drive :", drive)
except Exception as e:  # Drive not mounted in some contexts
    print("drive save skipped:", e)

# reopen to prove it's valid
chk = load_workbook(local)
print("tabs        :", chk.sheetnames)
