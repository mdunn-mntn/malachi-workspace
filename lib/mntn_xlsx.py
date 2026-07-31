#!/usr/bin/env python3
"""
mntn_xlsx — the MNTN shared .xlsx deliverable builder.

One import, one look. Every ticket's shareable workbook is built with this module so they all read
as one polished system: a branded cover, finding-led table sheets, a glossary, the SQL behind the
numbers, and method notes — color-coded tabs, frozen headers, autofilters, decimal-stored %/$ formats.

Design goals (from the deliverable standard, documentation/docs/xlsx_deliverable_standard.md):
  * The audience should feel real care went into it — coloring, spacing, typography, borders.
  * The numbers stay auditable: percents are stored as DECIMALS with true % number formats (survive
    a re-type in Excel/Sheets), the SQL rides along on its own tab, and nothing is pre-scaled.
  * Swapping in official MNTN brand hexes / logo is a ONE-LINE change (edit BRAND / pass logo_path).

Quick start
-----------
    import sys; sys.path.insert(0, "<workspace-root>")   # or install editable
    from lib.mntn_xlsx import MntnWorkbook, FMT

    wb = MntnWorkbook(
        title="MM vs 3P Segment Scorecard",
        ticket="AUDI-1141",
        subtitle="Prospecting performance by vertical — trailing 6 months",
        period="Jan-Jun 2026",
    )
    wb.table("MM vs 3P by vertical", df,
             finding="MNTN Matched leads visit rate in every vertical",
             method="Advertiser-weighted medians; prospecting only; visits = views + clicks.",
             formats={"MM IVR": FMT.PCT2, "3P IVR": FMT.PCT2, "IVR advantage": FMT.MULT},
             heat={"MM IVR": "high", "3P IVR": "high"},
             kind="headline", toc="The headline: MM vs 3P visit rate, CPV and ROAS by vertical")
    wb.glossary("Read me", intro="How to read this workbook.", rows=[("IVR", "Visit rate ..."), ...])
    wb.sql("Queries", open("....sql").read(), note="BigQuery cohort SQL used to produce these numbers.")
    wb.cover(takeaways=["...", "...", "..."])            # call LAST — builds the clickable contents
    wb.save_drive("AUDI-1141", "MM vs 3P Scorecard")     # -> My Drive/Tickets/AUDI-1141/AUDI-1141 MM vs 3P Scorecard.xlsx

Notes
-----
  * Store rates as DECIMALS (0.0046) and pass a % format — never pre-scale to "0.46".
  * Use None (not "") for truly-empty cells so an adjacent long label can overflow.
  * Keep headers <= 2 words; the sheet title says which group. Column widths fit the WHOLE header.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.properties import PageSetupProperties

if TYPE_CHECKING:
    from collections.abc import Callable, Iterable

    from openpyxl.cell.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# BRAND — the single source of truth for the look. Swap these for official MNTN
# hexes when we have them; everything downstream re-colors automatically.
# ---------------------------------------------------------------------------
BRAND = {
    # Official MNTN brand palette (brand.mountain.com, 2025 guidelines).
    "INK": "191E28",  # Slate Grey (deepest) — cover band bg, footer, brand copy on light
    "PRIMARY": "262E3C",  # Slate Grey — table header fills, finding titles, row labels
    "ACCENT": "1AC9AA",  # Mountain Green (core brand color) — cover rule, key numbers, takeaway ticks
    "LINK": "0AABC5",  # Mountain Blue — hyperlinks (better small-text contrast than the green)
    "BAND": "E4F7F2",  # zebra band on data rows — light Mountain Green tint (brand, not grey)
    "PAPER": "F6F6F6",  # Glacier White — off-white fill / neutral heat start
    "GREY": "5C6675",  # subtitles, methodology lines (mid slate)
    "MUTE": "98A2B3",  # footnotes, appendix tab color
    "LINE": "DCE3EA",  # thin cell borders / rules
    "POS": "1AC9AA",  # Mountain Green — good delta / RAG green
    "NEG": "D1495B",  # bad delta / RAG red (brand has no red; reserved — reads "bad" to execs)
    "WARN": "E9A23B",  # caution / RAG amber
    "WHITE": "FFFFFF",
}

# Assets + brand override. Drop the official MNTN logo at assets/mntn_logo.png and it is used on every
# cover automatically. Drop official hexes in assets/brand.json ({"PRIMARY": "...", "ACCENT": "..."}) and
# they override the defaults above — no code edit needed. See lib/assets/README.md.
_ASSET_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
_DEFAULT_LOGO = os.path.join(_ASSET_DIR, "mntn_logo.png")
try:
    import json

    _bpath = os.path.join(_ASSET_DIR, "brand.json")
    if os.path.exists(_bpath):
        with open(_bpath) as _bf:
            BRAND.update({k: str(v).lstrip("#").upper() for k, v in json.load(_bf).items()})
except Exception:
    pass  # malformed override never breaks a build; fall back to defaults

# Typography. Inter is the official MNTN body/UI font (open-license) and renders natively in Google
# Sheets — the actual delivery surface. Set FONT_BODY = "Arial" if recipients open in desktop Excel
# without Inter installed and you need guaranteed-identical metrics. Consolas/Menlo for SQL.
FONT_BODY = "Inter"
FONT_MONO = "Consolas"


@dataclass(frozen=True)
class _Fmt:
    """Number formats. Percents assume the cell holds a DECIMAL (0.0046 -> '0.46%')."""

    INT: str = "#,##0"
    NUM1: str = "0.0"
    NUM2: str = "0.00"
    PCT1: str = "0.0%"
    PCT2: str = "0.00%"
    PCT3: str = "0.000%"
    USD: str = '"$"#,##0.00'
    USD0: str = '"$"#,##0'
    USD2: str = '"$"#,##0.00'
    MULT: str = '0.0"x"'  # advantage ratios: 4.1x
    ROAS: str = '0.00"x"'
    DATE: str = "yyyy-mm-dd"


FMT = _Fmt()

# Tab color by role, so the tab strip itself is a color-coded legend. Distinct MNTN brand hues:
# a dark anchor for the cover, bright greens/blues for content, muted greys for the appendix.
# NOTE: apply with a leading "FF" (opaque) — a bare 6-hex string makes openpyxl store alpha 00
# (transparent), which renders as NO tab color in Google Sheets. See _new_sheet().
# HEAT — the SEQUENTIAL magnitude ramp, Mountain Green light->saturated (dataviz rule: one hue,
# light->dark; never a rainbow). Use `heat=` for pure-magnitude columns (counts, rates with no sign).
# For EFFECT/LIFT columns (signed, with a significance flag) use `signal=` instead — it paints
# red = significant negative, amber = not significant, and this same green ramp scaled by magnitude
# for significant positives. The green portion of both modes shares LIGHT->DARK below.
HEAT = {
    "LIGHT": "E4F7F2",  # near-white Mountain Green tint (smallest values)
    "MID": "8CE0CE",  # mid Mountain Green
    "DARK": "1AC9AA",  # Mountain Green (largest / "best" values)
    # The pale END of a heat RAMP is floored here (not LIGHT) so EVERY cell in a heat column reads as a
    # visible tint — never near-white. Otherwise a 2-row heat (e.g. Cost per incremental) puts one row dark
    # and the other near-white, which looks like only one row is highlighted. "If we highlight one, highlight
    # all." Also keeps the palest heat cell distinct from the zebra BAND (E4F7F2).
    "FLOOR": "A7E9DC",  # visible light Mountain Green — the palest a ramp cell ever gets
}

TAB = {
    "cover": "191E28",  # Slate INK — dark anchor / "start here"
    "headline": "1AC9AA",  # Mountain Green — the hero content tab
    "data": "0AABC5",  # Mountain Blue — content
    "detail": "26D1EA",  # Mountain Blue (light) — supporting detail
    "glossary": "22E5BE",  # Mountain Green (light) — the Read me / reference
    "sql": "667085",  # Slate grey — appendix
    "notes": "98A2B3",  # Light slate grey — appendix
}

# Reusable style objects -----------------------------------------------------
_THIN = Side(style="thin", color=BRAND["LINE"])
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NOBORDER = Border()
# header cells get a thick Mountain Green underline — ties the slate header to the brand
_HEADER_BORDER = Border(
    left=_THIN, right=_THIN, top=_THIN, bottom=Side(style="thick", color=BRAND["ACCENT"])
)
# ALIGNMENT STANDARD (apply consistently — never leave a cell on Excel's implicit general/bottom default):
#   - numbers / short codes / Yes-No flags  -> center + vcenter        (_CEN_FLAT)
#   - single-line text / labels / links / headers -> left + vcenter    (_LEFT_MID_FLAT)
#   - multi-line wrapped prose (glossary defs, notes, footnotes, wrapped table cells) -> left + TOP (_LEFT / _LEFT_MID)
#     (top so a tall wrapped row starts its text at the top, not floating in the middle)
# A header cell matches its column's body horizontally and is ALWAYS vertically centered.
_CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CEN_FLAT = Alignment(horizontal="center", vertical="center")
_CEN_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_LEFT_MID = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_MID_FLAT = Alignment(
    horizontal="left", vertical="center"
)  # single-line left cells (headers, meta, links)
_RIGHT = Alignment(horizontal="right", vertical="center")


def _font(
    size: int = 10,
    bold: bool = False,
    italic: bool = False,
    color: str = "000000",
    name: str = FONT_BODY,
) -> Font:
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)


def _fill(hex_: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_)


def _lerp_hex(a: str, b: str, t: float) -> str:
    """Linear-interpolate two 6-hex colors (t in [0,1]) -> a 6-hex string. Used to scale the
    green magnitude ramp per cell in signal() coloring."""
    a, b = a.lstrip("#"), b.lstrip("#")
    ar, ag, ab = int(a[0:2], 16), int(a[2:4], 16), int(a[4:6], 16)
    br, bg, bb = int(b[0:2], 16), int(b[2:4], 16), int(b[4:6], 16)
    r, g, bl = (round(x + (y - x) * t) for x, y in ((ar, br), (ag, bg), (ab, bb)))
    return f"{r:02X}{g:02X}{bl:02X}"


_DASH_RE = re.compile(r"\s*[—–]\s*")


def _demdash(s: str) -> str:
    """Replace em/en dashes with a spaced hyphen. People read '—' as AI-written, so no MNTN
    deliverable should ship one. ASCII hyphens ('sub-vertical', '2026-07-21') are untouched."""
    return _DASH_RE.sub(" - ", s) if isinstance(s, str) else s


def _to_native(v: Any) -> Any:
    """numpy/pandas -> json-native, NaN/NaT -> None so Excel shows a truly empty cell.
    String cells are em-dash-sanitized on the way in."""
    if v is None:
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating, float)):
        return None if pd.isna(v) else float(v)
    if isinstance(v, (np.bool_,)):
        return bool(v)
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return _demdash(v) if isinstance(v, str) else v


class MntnWorkbook:
    """A branded, multi-sheet MNTN deliverable. Build content sheets, then call cover() last."""

    def __init__(
        self,
        title: str,
        ticket: str,
        subtitle: str = "",
        period: str = "",
        owner: str = "Malachi Dunn · Audience Intelligence",
        logo_path: str | None = None,
        generated: str | None = None,
        status: str = "Final",
    ) -> None:
        self.title = _demdash(title)
        self.ticket = ticket.upper().strip()
        self.subtitle = _demdash(subtitle)
        self.period = period
        self.owner = owner
        # explicit logo_path wins; else use the canonical asset if it's been dropped in
        cand = logo_path or _DEFAULT_LOGO
        self.logo_path = cand if (cand and os.path.exists(cand)) else None
        self.generated = generated  # 'YYYY-MM-DD' string; pass one for reproducible files
        self.status = status
        self._toc: list[tuple[str, str, str]] = []  # (sheet_name, one-line description, role)
        self._issues: list[
            str
        ] = []  # build-time violations; save_*() prints them and RAISES so a broken workbook can't ship
        self._query_tabs: list[
            str
        ] = []  # titles of sql() tabs, so a table's query= can deep-link into one
        self._pending_query_links: list[
            tuple[str, int, str]
        ] = []  # (data_sheet_title, footnote_row, query_filename); resolved at save

        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # start empty; cover() is added last and moved to front
        p = self.wb.properties
        p.title = f"{self.ticket} — {title}"
        p.subject = subtitle
        p.creator = owner
        p.keywords = f"MNTN, {self.ticket}, Audience Intelligence"
        p.category = "Analysis deliverable"

    # -- internal sheet scaffolding -----------------------------------------
    def _new_sheet(self, name: str, role: str) -> Worksheet:
        ws = self.wb.create_sheet(name[:31])  # Excel 31-char tab limit
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "FF" + TAB.get(
            role, BRAND["PRIMARY"]
        )  # FF = opaque (bare hex -> alpha 00 = invisible)
        ws.sheet_view.zoomScale = 100
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        for side in ("left", "right", "top", "bottom"):
            setattr(ws.page_margins, side, 0.4)
        return ws

    def _titleblock(self, ws: Worksheet, finding: str, method: str, ncols: int = 1) -> None:
        """Finding-led title (states the finding) + grey italic methodology line.

        The subtitle is merged across the table columns and wrapped, so it never runs off to the
        right past the table below it. Its row height is fitted later in table() once column widths
        are known (Excel/Sheets won't auto-fit a merged cell). Applies to every table sheet."""
        self._sheet_title(ws, finding, ncols)
        if method:
            m = ws.cell(row=2, column=1, value=_demdash(method))
            m.font = _font(10, italic=True, color=BRAND["GREY"])
            m.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if ncols > 1:
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        # thin Mountain Green accent rule (row 3) closing the header block off from the content below.
        # Keep the subtitle to ONE short method line; definitions/caveats live on the Read me tab.
        self._accent_rule(ws, ncols)

    def _accent_rule(self, ws: Worksheet, ncols: int, row: int = 3, height: int = 5) -> None:
        """A thin Mountain Green rule spanning the content width — the shared header-to-body separator
        used on every sheet (data tables AND the Read me / Queries / Method reference tabs), so they
        all read as one system."""
        for cc in range(1, max(ncols, 1) + 1):
            ws.cell(row=row, column=cc).fill = _fill(BRAND["ACCENT"])
        ws.row_dimensions[row].height = height

    def _sheet_title(self, ws: Worksheet, text: str, ncols: int = 1) -> Cell:
        """Write the row-1 sheet title with top breathing room: a taller row with the title
        BOTTOM-aligned, so there is whitespace above it and the title no longer jams into the top
        edge (the cover keeps the brand band; content/reference tabs just get clean top air).
        Merged + wrapped across the table columns (ncols>1) so a long finding wraps within the table
        width instead of overflowing off the right edge; height is fitted later in table() once the
        column widths are known. Used on every non-cover sheet so the top spacing is uniform."""
        c = ws.cell(row=1, column=1, value=_demdash(text))
        c.font = _font(15, bold=True, color=BRAND["PRIMARY"])
        # wrap ONLY when merged across a table (ncols>1). On a single-column tab (glossary/notes/sql,
        # ncols=1) the cell is unmerged, so wrap would squeeze the title into column A — leave it
        # unwrapped there so it spills across to the right like a normal full-width title.
        c.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=(ncols > 1))
        if ncols > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.row_dimensions[
            1
        ].height = 34  # ~2.2x default -> the air sits above the bottom-aligned title
        return c

    def _fit_title_height(
        self,
        ws: Worksheet,
        text: str,
        total_width_chars: float,
        per_line: float = 19.0,
        pad: float = 9.0,
        base: float = 34.0,
        maxlines: int = 3,
    ) -> None:
        """Size the merged row-1 title height to its wrapped text at the table width (Excel won't
        auto-fit a merged cell). 15pt bold fits ~0.6 chars per width-unit; keeps the base top air."""
        if not text:
            return
        cpl = max(int(total_width_chars * 0.60), 18)  # chars per line for 15pt bold across the span
        lines = min(max(1, -(-len(_demdash(text)) // cpl)), maxlines)
        ws.row_dimensions[1].height = max(base, lines * per_line + pad)

    def _fit_subtitle_height(
        self,
        ws: Worksheet,
        row: int,
        text: str,
        total_width_chars: float,
        per_line: float = 13.5,
        pad: float = 4.0,
        maxlines: int = 5,
    ) -> None:
        """Set a merged, wrapped subtitle row's height to fit its text at the table width.
        total_width_chars = summed width of the merged columns; Excel won't auto-fit merged cells."""
        if not text:
            return
        cpl = max(int(total_width_chars * 0.92), 24)  # ~chars per line across the merged span
        lines = max(1, -(-len(_demdash(text)) // cpl))
        ws.row_dimensions[row].height = max(16.0, min(lines, maxlines) * per_line + pad)

    @staticmethod
    def _wrap_lines(text: str, width_chars: int) -> int:
        """Word-aware line count for `text` wrapped to `width_chars`. Returns a large number if any single
        word is wider than the column (unbreakable overflow -> guaranteed clip)."""
        words, lines, cur = str(text).split(), 0, 0
        for w in words:
            if len(w) > width_chars:
                return 99
            if cur == 0:
                cur = len(w)
            elif cur + 1 + len(w) <= width_chars:
                cur += 1 + len(w)
            else:
                lines += 1
                cur = len(w)
        return lines + (1 if cur else 0)

    def _fit_header_height(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        start_row: int,
        colw: dict[str, int],
        per_line: float = 15.0,
        pad: float = 8.0,
        base: float = 30.0,
    ) -> None:
        """Size the HEADER row to the TALLEST wrapped header at its column width, so a column title never
        clips (the old fixed 30pt row was the root cause of clipped headers). Bold ~11pt -> ~0.85 chars/unit."""
        maxlines = 1
        for col in df.columns:
            w = max(int(colw.get(col, 12) * 0.85), 4)
            maxlines = max(maxlines, self._wrap_lines(col, w))
        ws.row_dimensions[start_row].height = max(base, min(maxlines, 4) * per_line + pad)

    def _issue(self, sheet: str, msg: str) -> None:
        """Record a build-time violation. save_*() prints all of them and RAISES, so a workbook that breaks
        a hard rule (char cap, etc.) cannot be produced -- the mistake fails the build instead of shipping."""
        self._issues.append(f"[{sheet}] {msg}")

    def _raise_if_issues(self) -> None:
        if self._issues:
            import sys

            report = "\n  - ".join(self._issues)
            print(
                f"[mntn_xlsx] BUILD BLOCKED — {len(self._issues)} rule violation(s):\n  - {report}",
                file=sys.stderr,
            )
            raise ValueError(
                f"mntn_xlsx: {len(self._issues)} workbook rule violation(s) — see stderr. "
                f"Fix them (trim text / widen columns), the file was NOT written."
            )

    def _resolve_query_links(self) -> None:
        """Deep-link each table's Source footnote to its query= block on the Query tab. Runs at save (after
        every sheet exists). A query= naming a file that isn't on the Query tab fails the build."""
        for sheet_title, foot_row, fname in self._pending_query_links:
            target = None
            for qtab in self._query_tabs:
                wq = self.wb[qtab]
                for cell in wq["A"]:
                    v = cell.value
                    if isinstance(v, str) and v.lstrip().startswith("--") and fname in v:
                        target = (qtab, cell.row)
                        break
                if target:
                    break
            if not target:
                self._issue(
                    sheet_title,
                    f"query '{fname}' referenced (query=) but no header naming it is on "
                    f"the Query tab -- add the filename to that query's header comment",
                )
                continue
            qtab, qrow = target
            fc = self.wb[sheet_title].cell(row=foot_row, column=1)  # merged footnote anchor cell
            # display MUST be the full footnote text: Google Sheets renders a hyperlink's display over the
            # cell value, so display=fname alone would hide Source/Period/Generated. Whole line stays clickable.
            fc.hyperlink = Hyperlink(
                ref=f"A{foot_row}", location=f"'{qtab}'!A{qrow}", display=str(fc.value)
            )

    def _footnote(self, ws: Worksheet, text: str, row: int, ncols: int) -> None:
        if not text:
            return
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
        c = ws.cell(row=row, column=1, value=text)
        c.font = _font(9, italic=True, color=BRAND["MUTE"])
        c.alignment = _LEFT

    def _autosize(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        widths: dict[str, int] | None = None,
        first_col: int | None = None,
        cap: int = 38,
        filter_pad: int = 7,
    ) -> dict[str, int]:
        """Size each column to the WIDER of (a) its longest header WORD + padding for bold text and the
        autofilter dropdown icon — so a header word never breaks mid-word ("Spend" -> "Spen/d") — and
        (b) its actual DATA, up to `cap` (longer data wraps on word boundaries).

        Returns {column_name: final_width_chars} so row heights can be sized to the wrapped text.
        Explicit `widths` are honored as-is (up to a sane ceiling) so a caller can make a prose column
        genuinely wide.
        """
        widths = widths or {}
        final: dict[str, int] = {}
        for j, col in enumerate(df.columns, 1):
            if col in widths:
                w = min(max(int(widths[col]), 8), 72)  # honor caller intent, sane ceiling
            else:
                header = str(col)
                longest_word = max([len(x) for x in header.split()] or [len(header)])
                cd = df[col].dropna()
                if df[col].dtype == object:
                    data_w = int(cd.astype(str).map(len).head(200).max()) if len(cd) else 0
                else:
                    # numbers are display-FORMATTED (%, $, commas) — never measure the raw float repr
                    # (str(0.00189)="0.00189372…"). Estimate from magnitude + room for separators/symbol.
                    mx = (
                        pd.to_numeric(cd, errors="coerce")
                        .replace([np.inf, -np.inf], np.nan)
                        .abs()
                        .max()
                    )
                    int_digits = len(f"{int(mx):,}") if (pd.notna(mx) and mx >= 1) else 3
                    data_w = int_digits + 5
                # header word must fit (never break mid-word); data fits up to cap (wraps beyond); floor 10
                need = max(longest_word + filter_pad, min(data_w, cap) + 2, 10)
                w = min(need, cap)
            ws.column_dimensions[get_column_letter(j)].width = w
            final[col] = w
        if first_col:
            ws.column_dimensions["A"].width = first_col
            final[df.columns[0]] = first_col
        return final

    def _fit_row_heights(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        start: int,
        colw: dict[str, int],
        per_line: float = 15.0,
        pad: float = 6.0,
        minh: float = 18.0,
        maxlines: int = 10,
    ) -> None:
        """Size each data row to its tallest wrapped cell so nothing clips (Excel won't auto-fit).

        colw = {column: width_chars} from _autosize. Slightly over-estimates lines (chars-per-line
        uses 0.85 * width) so text errs tall rather than clipped.
        """
        for i, (_, r) in enumerate(df.iterrows(), 1):
            rr = start + i
            lines = 1
            for col in df.columns:
                v = r[col]
                if v is None:
                    continue
                s = str(v)
                cpl = max(int(colw.get(col, 12) * 0.85), 4)
                need = sum(max(1, -(-len(seg) // cpl)) for seg in s.split("\n"))
                lines = max(lines, need)
            ws.row_dimensions[rr].height = max(minh, min(lines, maxlines) * per_line + pad)

    @staticmethod
    def _wrap_rows(
        ws: Worksheet,
        col_letter: str,
        rows: list[tuple[int, str | None]],
        width_chars: int,
        per_line: float = 15.0,
        pad: float = 6.0,
        minh: float = 16.0,
        maxlines: int = 40,
    ) -> None:
        """Size a single wrapped text column's rows (used by notes/glossary prose cells)."""
        cpl = max(int(width_chars * 0.9), 8)
        for rr, text in rows:
            if text is None:
                continue
            need = sum(max(1, -(-len(seg) // cpl)) for seg in str(text).split("\n"))
            ws.row_dimensions[rr].height = max(minh, min(need, maxlines) * per_line + pad)

    # -- public: table sheet -------------------------------------------------
    def table(
        self,
        name: str,
        df: pd.DataFrame,
        finding: str,
        method: str = "",
        formats: dict[str, str] | None = None,
        heat: dict[str, str] | None = None,
        rag: dict[str, Callable[[Any], str | None]] | None = None,
        signal: dict[str, dict[str, str]] | None = None,
        band: bool = True,
        kind: str = "data",
        toc: str = "",
        widths: dict[str, int] | None = None,
        first_col_width: int | None = None,
        freeze: str = "A",
        query: str = "",
    ) -> Worksheet:
        """Add a styled table sheet.

        finding  : sheet title that STATES THE FINDING (not the metric).
        method   : one grey line of methodology under the title.
        formats  : {column_name: number_format}. Store %/$ as decimals; pass FMT.PCT2 etc.
        heat     : {column_name: 'high'|'low'|'neutral'} -> sequential single-hue (green) magnitude
                   ramp. Use for PURE-MAGNITUDE columns (counts, rates, no sign). 'low' inverts (cost).
        signal   : {column_name: {'sig': <sig_column>}} -> SEMANTIC effect/lift coloring:
                   red = significant negative, amber = not significant, green (scaled by magnitude)
                   = significant positive. Use for signed EFFECT/LIFT columns. Omit 'sig' to skip the
                   amber rule (then just negative=red, positive=green-scaled). Don't also pass heat on
                   the same column.
        rag      : {column_name: fn(value)->'POS'|'NEG'|'WARN'|None} -> traffic-light cell fills.
        kind     : 'headline' | 'data' | 'detail' (drives tab color).
        query    : source .sql filename. Named inline in the bottom Source line and deep-linked to that
                   query's block on the Query tab (resolved at save; a query not on the tab fails the build).
        """
        formats = formats or {}
        heat = heat or {}
        rag = rag or {}
        signal = signal or {}
        ws = self._new_sheet(name, "headline" if kind == "headline" else kind)
        ncols = len(df.columns)
        self._titleblock(ws, finding, method, ncols)
        start = 4

        # header row (slate fill, white text, Mountain Green underline)
        for j, col in enumerate(df.columns, 1):
            c = ws.cell(row=start, column=j, value=str(col))
            c.font = _font(11, bold=True, color=BRAND["WHITE"])
            c.fill = _fill(BRAND["PRIMARY"])
            c.alignment = _CEN
            c.border = _HEADER_BORDER
        ws.row_dimensions[start].height = 30

        # body
        for i, (_, r) in enumerate(df.iterrows(), 1):
            rr = start + i
            for j, col in enumerate(df.columns, 1):
                v = _to_native(r[col])
                c = ws.cell(row=rr, column=j, value=v)
                c.border = _BORDER
                # numbers stay centered on one line; text wraps so long cells never clip
                c.alignment = (
                    _LEFT_MID
                    if (j == 1 and isinstance(v, str))
                    else (_CEN_WRAP if isinstance(v, str) else _CEN_FLAT)
                )
                if col in formats:
                    c.number_format = formats[col]
                if j == 1:
                    c.font = _font(10, bold=True, color=BRAND["PRIMARY"])
                if band and i % 2 == 0:
                    c.fill = _fill(BRAND["BAND"])
                if col in rag and v is not None:
                    key = rag[col](v)
                    if key in ("POS", "NEG", "WARN"):
                        c.fill = _fill(BRAND[key])  # solid semantic fill
                        c.font = _font(10, bold=True, color=BRAND["WHITE"])

        n = len(df)
        last_col = get_column_letter(ncols)
        # per-column heat scales
        for col, direction in heat.items():
            if col not in df.columns:
                continue
            jj = get_column_letter(list(df.columns).index(col) + 1)
            rng = f"{jj}{start + 1}:{jj}{start + n}"
            if direction == "neutral":
                rule = ColorScaleRule(
                    start_type="min",
                    start_color=BRAND["PAPER"],
                    end_type="max",
                    end_color=BRAND["PRIMARY"],
                )
            else:
                # SEQUENTIAL single-hue ramp (Mountain Blue), light->saturated. Magnitude gets one hue,
                # never a red-yellow-green rainbow (that painted the lowest positive value red). 'high':
                # bigger = darker; 'low' (cost): smaller = darker. So darker always reads "better".
                # pale end floored to HEAT["FLOOR"] (not near-white LIGHT) so every cell reads as tinted
                if direction == "high":
                    s, m, e = HEAT["FLOOR"], HEAT["MID"], HEAT["DARK"]
                else:
                    s, m, e = HEAT["DARK"], HEAT["MID"], HEAT["FLOOR"]
                rule = ColorScaleRule(
                    start_type="min",
                    start_color=s,
                    mid_type="percentile",
                    mid_value=50,
                    mid_color=m,
                    end_type="max",
                    end_color=e,
                )
            ws.conditional_formatting.add(rng, rule)

        # semantic effect/lift coloring: amber = not significant, red = significant negative,
        # green (deeper = more lift) = significant positive. Per-cell fills (needs the value + sign +
        # significance flag, which a single ColorScaleRule can't express).
        def _is_sig(x: Any) -> bool:
            return x.strip().lower() in ("yes", "true", "y", "1") if isinstance(x, str) else bool(x)

        for col, spec in signal.items():
            if col not in df.columns:
                continue
            jcol = list(df.columns).index(col) + 1
            vals = pd.to_numeric(df[col], errors="coerce")
            sigcol = spec.get("sig")
            has_sig = bool(sigcol) and sigcol in df.columns
            flags = [(_is_sig(df[sigcol].iloc[i]) if has_sig else True) for i in range(len(df))]
            pos = sorted(
                vals.iloc[i]
                for i in range(len(df))
                if pd.notna(vals.iloc[i]) and vals.iloc[i] > 0 and flags[i]
            )
            npos = len(pos)
            for i in range(len(df)):
                v = vals.iloc[i]
                if pd.isna(v):
                    continue
                c = ws.cell(row=start + 1 + i, column=jcol)
                if has_sig and not flags[i]:  # inconclusive
                    c.fill = _fill(BRAND["WARN"])
                    c.font = _font(10, bold=True, color=BRAND["INK"])
                elif v < 0:  # significant negative
                    c.fill = _fill(BRAND["NEG"])
                    c.font = _font(10, bold=True, color=BRAND["WHITE"])
                else:  # significant positive -> green by RANK
                    # rank-based (not linear): even gradient, so a skewed tail can't wash the rest pale
                    # or flatten the top. Bigger lift = deeper green, evenly stepped.
                    rank = (sum(1 for p in pos if p < v) / (npos - 1)) if npos > 1 else 1.0
                    # floor at 0.30 so the palest cell still reads clearly green — HEAT["LIGHT"] equals
                    # the zebra BAND, so a rank-0 cell at t=0 would vanish into a banded row.
                    t = 0.30 + 0.70 * rank
                    c.fill = _fill(_lerp_hex(HEAT["LIGHT"], HEAT["DARK"], t))
                    c.font = _font(10, bold=True, color=BRAND["INK"])

        ws.freeze_panes = f"{freeze}{start + 1}"
        ws.auto_filter.ref = f"A{start}:{last_col}{start + n}"
        colw = self._autosize(ws, df, widths=widths, first_col=first_col_width)
        self._fit_row_heights(ws, df, start, colw)
        self._fit_header_height(
            ws, df, start, colw
        )  # size the header row to its wrapped headers -> never clips
        # rows 1 (title) and 2 (subtitle) are merged across the table; size heights to the wrapped
        # text at table width so a long finding/method wraps in place instead of running off the edge.
        table_width = sum(colw.get(c, 12) for c in df.columns)
        self._fit_title_height(ws, finding, table_width)
        self._fit_subtitle_height(ws, 2, method, table_width)
        foot_row = start + n + 2
        self._footnote(
            ws,
            f"Source: {self.ticket}"
            + (f"  ·  Query: {query}" if query else "")
            + (f"  ·  Period: {self.period}" if self.period else "")
            + (f"  ·  Generated {self.generated}" if self.generated else ""),
            foot_row,
            ncols,
        )
        if query:
            self._pending_query_links.append((ws.title, foot_row, query))
        if toc:
            self._toc.append((ws.title, toc, kind))
        return ws

    # -- public: glossary / read-me -----------------------------------------
    def glossary(
        self,
        name: str,
        rows: list[tuple[str, str]],
        intro: str = "",
        toc: str = "How to read this workbook",
        body_width: int = 104,
        max_def_chars: int = 220,
        max_entries: int = 14,
    ) -> Worksheet:
        """Two-column term/definition sheet (term bold in A, definition wrapped in B).
        A row of ('', '') renders a blank spacer; a row of ('Header', '') renders a section band.

        Terseness guard: a glossary entry is a term + 1-2 tight sentences (<= max_def_chars, ~3 lines),
        not a paragraph, and a Read me stays <= max_entries. Overflows print a BUILD-time warning so a
        glossary can't silently sprawl into prose (move why/how reasoning to the Method/notes tab). Warn,
        don't truncate; raise the caps explicitly per-call if a deliverable genuinely needs it."""
        ents = [(k, v) for k, v in rows if k and v]
        over = [(k, len(v)) for k, v in ents if len(v) > max_def_chars]
        # a too-long definition is a HARD fail (exact rule) so it can't ship; too-many-entries stays a warn
        # because max_entries is a deliberate per-call knob (a workbook with more tabs legitimately needs more).
        for k, nch in over:
            self._issue(
                name,
                f"glossary def '{k}' is {nch} chars (cap {max_def_chars}) — trim to 1-2 sentences, move why/how to the Method tab",
            )
        if len(ents) > max_entries:
            import sys

            print(
                f"[mntn_xlsx] Read me '{name}': {len(ents)} entries > {max_entries} "
                "(raise max_entries only if the tab count genuinely warrants it).",
                file=sys.stderr,
            )
        ws = self._new_sheet(name, "glossary")
        self._sheet_title(ws, self.title)
        sub = f"{self.ticket}." + (f"  {_demdash(intro)}" if intro else "")
        subc = ws.cell(row=2, column=1, value=sub)
        subc.font = _font(10, italic=True, color=BRAND["GREY"])
        subc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells("A2:B2")
        self._accent_rule(ws, 2)
        r = 4
        def_rows: list[tuple[int, str | None]] = []
        for k, v in rows:
            is_section = bool(k) and not v  # (heading, '') -> a section band (visual grouper)
            kc = ws.cell(row=r, column=1, value=(_demdash(k) or None))
            vc = ws.cell(row=r, column=2, value=(_demdash(v) or None))
            if is_section:
                for cc in (1, 2):  # light Mountain-Green band across both columns
                    ws.cell(row=r, column=cc).fill = _fill(BRAND["BAND"])
                kc.font = _font(10, bold=True, color=BRAND["PRIMARY"])
                kc.alignment = _LEFT_MID_FLAT
            else:
                kc.font = _font(10, bold=True, color=BRAND["PRIMARY"] if v else BRAND["INK"])
                kc.alignment = _LEFT
                vc.font = _font(10)
                vc.alignment = _LEFT
            def_rows.append((r, v))
            r += 1
        a_width = max((len(str(k)) for k, _ in rows), default=24) + 3
        ws.column_dimensions["A"].width = a_width
        ws.column_dimensions["B"].width = body_width
        self._wrap_rows(ws, "B", def_rows, body_width)
        self._fit_subtitle_height(ws, 2, sub, a_width + body_width)
        if toc:
            self._toc.append((ws.title, toc, "glossary"))
        return ws

    @staticmethod
    def _cap_comment_runs(sql_text: str, cap: int = 3) -> tuple[str, int]:
        """Cap every SQL comment HEADER at `cap` lines so the Query tab never becomes a wall of grey.
        A run of consecutive `--` lines (blank lines between them are treated as interior to the same
        header and collapsed) is trimmed to its first `cap` lines. Code and code-separating blanks are
        untouched. Returns (text, dropped_comment_lines). Rule: a query header is a label, not prose."""
        out, run, dropped = [], 0, 0
        for ln in sql_text.split("\n"):
            s = ln.strip()
            if s.startswith("--"):
                run += 1
                if run <= cap:
                    out.append(ln)
                else:
                    dropped += 1
            elif s == "" and run > 0:
                pass  # collapse blanks inside a comment header (keeps blank-split blocks as one run)
            else:
                if run > 0:
                    out.append("")  # one clean separator between the capped header and the code
                run = 0
                out.append(ln)
        return "\n".join(out), dropped

    # -- public: SQL / queries ----------------------------------------------
    def sql(
        self,
        name: str,
        sql_text: str,
        note: str = "",
        toc: str = "The SQL behind the numbers",
        width: int = 120,
        max_comment_run: int = 3,
    ) -> Worksheet:
        """Add a Query tab: the SQL verbatim on a code panel, comment headers capped and receded."""
        ws = self._new_sheet(name, "sql")
        self._query_tabs.append(ws.title)  # a table()'s query= can deep-link into this tab
        self._sheet_title(ws, "Queries used (for validation)")
        if note:
            nc = ws.cell(row=2, column=1, value=_demdash(note))
            nc.font = _font(10, italic=True, color=BRAND["GREY"])
            nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        self._accent_rule(ws, 1)
        sql_text, dropped = self._cap_comment_runs(sql_text, max_comment_run)
        if dropped:
            import sys

            print(
                f"[mntn_xlsx] Query '{name}' trimmed {dropped} comment line(s) over the "
                f"{max_comment_run}-line header cap. Tighten the SQL comment headers.",
                file=sys.stderr,
            )
        r = 4
        for line in sql_text.split("\n"):  # SQL body left verbatim (never sanitized)
            c = ws.cell(row=r, column=1, value=line)
            # comments recede (muted italic), code stays dark; whole block on a light code-panel fill
            is_comment = line.lstrip().startswith("--")
            c.font = _font(
                9,
                name=FONT_MONO,
                italic=is_comment,
                color=BRAND["GREY"] if is_comment else BRAND["INK"],
            )
            c.fill = _fill(BRAND["PAPER"])
            c.alignment = _LEFT_MID_FLAT
            r += 1
        ws.column_dimensions["A"].width = width
        self._fit_subtitle_height(ws, 2, note, width)
        if toc:
            self._toc.append((ws.title, toc, "sql"))
        return ws

    def sql_dir(
        self,
        name: str,
        directory: str,
        order: list[str] | None = None,
        ignore: Iterable[str] | None = None,
        headers: dict[str, str] | None = None,
        note: str = "",
        collapse_aids: bool = True,
        aid_placeholder: str = "/* the AID list, same as the main query */",
    ) -> Worksheet:
        """Build the Query tab from EVERY .sql file in `directory` (minus `ignore`), so a newly-added query
        can never be forgotten — the default is 'included', not 'you remembered to add it'.

        order:   list of filenames to lead with (the rest follow, sorted). ignore: filenames to omit
        (superseded / one-off diagnostic queries — omitting requires a deliberate act). headers: {filename:
        one-line header}; otherwise the file's first `-- ...` line, else the filename. AID UNNEST lists are
        collapsed to a placeholder. The whole thing reuses sql() (comment-cap + styling)."""
        import glob as _glob
        import re as _re

        files = sorted(_glob.glob(os.path.join(directory, "*.sql")))
        ignore = set(ignore or [])
        chosen = [f for f in files if os.path.basename(f) not in ignore]
        if order:
            rank = {n: i for i, n in enumerate(order)}
            chosen.sort(
                key=lambda f: (rank.get(os.path.basename(f), len(order)), os.path.basename(f))
            )
        headers = headers or {}
        parts = []
        for f in chosen:
            base = os.path.basename(f)
            raw = Path(f).read_text().strip()
            first_comment = next(
                (ln.strip() for ln in raw.splitlines() if ln.strip().startswith("--")), None
            )
            hdr = headers.get(base) or first_comment or f"-- {base}"
            if not hdr.lstrip().startswith("--"):
                hdr = "-- " + hdr
            body = _re.sub(r"\A(\s*--[^\n]*\n)+", "", raw)  # strip the file's own comment block
            if collapse_aids:
                body = _re.sub(
                    r"UNNEST\(\[.*?\]\)", f"UNNEST([ {aid_placeholder} ])", body, flags=_re.S
                )
            parts.append(hdr.split("\n")[0].strip() + "\n\n" + body.strip())
        return self.sql(name, "\n\n\n".join(parts) + "\n", note=note)

    def check_queries_covered(
        self, query_text: str, directory: str, ignore: Iterable[str] | None = None
    ) -> None:
        """HARD-fail the build if a .sql file in `directory` is NOT present in the Query tab text — so a
        newly-added query can't be forgotten. Use with a hand-curated sql() Query tab (sql_dir() already
        guarantees coverage). `ignore` = filenames deliberately kept out (superseded / one-off diagnostics)."""
        import glob as _glob
        import re as _re

        def norm(s: str) -> str:
            s = _re.sub(r"--[^\n]*", "", s)  # strip comments
            s = _re.sub(r"UNNEST\(\[.*?\]\)", "UNNEST([])", s, flags=_re.S)  # neutralize AID lists
            return _re.sub(r"\s+", " ", s).strip().lower()

        tab = norm(query_text)
        ignore = set(ignore or [])
        for f in sorted(_glob.glob(os.path.join(directory, "*.sql"))):
            base = os.path.basename(f)
            if base in ignore:
                continue
            body = norm(Path(f).read_text())
            if body and body not in tab:
                self._issue(
                    "Query",
                    f"'{base}' is NOT in the Query tab — add it (or list it in ignore= if superseded/diagnostic)",
                )

    # -- public: long-form notes / method -----------------------------------
    def notes(
        self,
        name: str,
        blocks: list[tuple[str, str]],
        intro: str = "",
        toc: str = "Method & caveats",
        body_width: int = 110,
        max_block_chars: int = 320,
    ) -> Worksheet:
        """blocks = list of (heading, body). heading '' -> continuation paragraph.

        Each block body leads with its answer and stays <= max_block_chars (the narrative-explainer cap);
        a longer block is a HARD build failure (move detail to a linked tab or cut it). Same discipline as
        the glossary terseness guard, for the Method/caveats prose."""
        for head, body in blocks:
            if body and len(str(body)) > max_block_chars:
                self._issue(
                    name,
                    f"block '{head or body[:30]}' is {len(str(body))} chars (cap {max_block_chars}) — trim it or split",
                )
        ws = self._new_sheet(name, "notes")
        self._sheet_title(ws, name)
        if intro:
            ic = ws.cell(row=2, column=1, value=_demdash(intro))
            ic.font = _font(10, italic=True, color=BRAND["GREY"])
            ic.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        self._accent_rule(ws, 1)
        r = 4
        body_rows: list[tuple[int, str | None]] = []
        for head, body in blocks:
            if head:
                hc = ws.cell(row=r, column=1, value=_demdash(head))
                hc.font = _font(11, bold=True, color=BRAND["PRIMARY"])
                hc.fill = _fill(BRAND["BAND"])  # light green band = section header
                hc.alignment = _LEFT_MID_FLAT
                r += 1
            bc = ws.cell(row=r, column=1, value=_demdash(body))
            bc.font = _font(10)
            bc.alignment = _LEFT
            body_rows.append((r, body))
            r += 2
        ws.column_dimensions["A"].width = body_width
        self._wrap_rows(ws, "A", body_rows, body_width)
        self._fit_subtitle_height(ws, 2, intro, body_width)
        if toc:
            self._toc.append((ws.title, toc, "notes"))
        return ws

    # -- public: branded cover (call LAST) ----------------------------------
    def cover(self, takeaways: list[str] | None = None, name: str = "Overview") -> Worksheet:
        """Create the branded cover and move it to the front. Builds the clickable contents
        from every sheet added so far. takeaways = up to 3 headline bullets (Rule of Three)."""
        takeaways = (takeaways or [])[:3]
        ws = self._new_sheet(name, "cover")
        span = 8
        wide = get_column_letter(span)

        # brand band (rows 1-3): INK fill, wordmark / logo left
        for rr in (1, 2, 3):
            for cc in range(1, span + 1):
                ws.cell(row=rr, column=cc).fill = _fill(BRAND["INK"])
            ws.row_dimensions[rr].height = 22
        ws.merge_cells(f"A1:{wide}3")
        if self.logo_path:
            try:
                from PIL import Image as PILImage

                target_h = 44  # px; fits inside the 3-row band with breathing room
                lg = PILImage.open(self.logo_path).convert("RGBA")
                w = max(1, int(lg.width * target_h / lg.height))
                # Render the resized logo to a PERSISTENT temp file, not a BytesIO: openpyxl re-reads the
                # image ref on EVERY wb.save(), so save_local()+save_drive() would exhaust a one-shot
                # buffer ("I/O operation on closed file"). A file path is re-readable across saves.
                render_path = os.path.join(_ASSET_DIR, "_logo_render.png")
                lg.resize((w, target_h)).save(render_path, format="PNG")
                self._logo_render = render_path  # keep a ref; the file persists (gitignored)
                ws.add_image(XLImage(render_path), "A2")  # row 2 -> vertically centered in the band
            except Exception:
                self.logo_path = None
        if not self.logo_path:
            wm = ws.cell(row=1, column=1, value="MNTN")
            wm.font = _font(26, bold=True, color=BRAND["WHITE"])
            wm.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # accent rule (row 4)
        for cc in range(1, span + 1):
            ws.cell(row=4, column=cc).fill = _fill(BRAND["ACCENT"])
        ws.row_dimensions[4].height = 5

        # title + subtitle
        ws.merge_cells(f"A6:{wide}6")
        t = ws.cell(row=6, column=1, value=self.title)
        t.font = _font(24, bold=True, color=BRAND["INK"])
        t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[6].height = 34
        if self.subtitle:
            ws.merge_cells(f"A7:{wide}7")
            s = ws.cell(row=7, column=1, value=self.subtitle)
            s.font = _font(12, italic=True, color=BRAND["GREY"])
            s.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # meta strip
        meta = [
            ("Ticket", self.ticket),
            ("Period", self.period or "—"),
            ("Prepared by", self.owner),
            ("Status", self.status),
        ]
        if self.generated:
            meta.insert(2, ("Generated", self.generated))
        r = 9
        for label, val in meta:
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = _font(9, bold=True, color=BRAND["MUTE"])
            lc.alignment = _LEFT_MID_FLAT
            c = ws.cell(row=r, column=2, value=val)
            c.font = _font(11, bold=True, color=BRAND["PRIMARY"])
            c.alignment = _LEFT_MID_FLAT
            r += 1

        # key takeaways (Rule of Three)
        r += 1
        if takeaways:
            ws.cell(row=r, column=1, value="Key takeaways").font = _font(
                13, bold=True, color=BRAND["INK"]
            )
            r += 1
            for i, tk in enumerate(takeaways, 1):
                ws.cell(row=r, column=1, value=str(i)).font = _font(
                    12, bold=True, color=BRAND["ACCENT"]
                )
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
                c = ws.cell(row=r, column=2, value=_demdash(tk))
                c.font = _font(11, color=BRAND["INK"])
                c.alignment = _LEFT_MID
                ws.row_dimensions[r].height = 30
                r += 1

        # contents (clickable)
        r += 1
        ws.cell(row=r, column=1, value="Contents").font = _font(13, bold=True, color=BRAND["INK"])
        r += 1
        # both header cells LEFT + vertically CENTER, matching the left-aligned links/descriptions below
        # (alignment standard: header cells match their column's body horizontally, always vcenter).
        th = ws.cell(row=r, column=1, value="Tab")
        th.font = _font(9, bold=True, color=BRAND["WHITE"])
        th.fill = _fill(BRAND["PRIMARY"])
        th.alignment = _LEFT_MID_FLAT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        hc = ws.cell(row=r, column=2, value="What's on it")
        hc.font = _font(9, bold=True, color=BRAND["WHITE"])
        hc.fill = _fill(BRAND["PRIMARY"])
        hc.alignment = _LEFT_MID_FLAT
        r += 1
        for sheet_name, desc, _role in self._toc:
            link = ws.cell(row=r, column=1, value=sheet_name)
            link.hyperlink = Hyperlink(
                ref=f"A{r}", location=f"'{sheet_name}'!A1", display=sheet_name
            )
            link.font = _font(10, bold=True, color=BRAND["LINK"], name=FONT_BODY)
            link.alignment = _LEFT_MID
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
            d = ws.cell(row=r, column=2, value=_demdash(desc))
            d.font = _font(10, color=BRAND["GREY"])
            d.alignment = _LEFT_MID
            r += 1

        # footer
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        f = ws.cell(
            row=r,
            column=1,
            value=f"MNTN · Audience Intelligence · {self.ticket}"
            + (f" · generated {self.generated}" if self.generated else "")
            + " · Internal",
        )
        f.font = _font(9, italic=True, color=BRAND["MUTE"])

        # column widths for the cover — col A must fit the longest tab name in Contents
        longest_tab = max((len(n) for n, _, _ in self._toc), default=16)
        ws.column_dimensions["A"].width = min(max(longest_tab + 2, 16), 28)
        for cc in range(2, span + 1):
            ws.column_dimensions[get_column_letter(cc)].width = 16

        # move to front and select
        self.wb.move_sheet(ws, -(len(self.wb.worksheets) - 1))
        self.wb.active = 0
        return ws

    # -- save ---------------------------------------------------------------
    def save_local(self, path: str) -> str:
        """Resolve links, fail the build on any rule violation, then write the .xlsx to `path`."""
        self._resolve_query_links()  # deep-link each Source footnote to its query block (may add issues)
        self._raise_if_issues()  # a rule violation fails the build here -> no broken file is ever written
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        self.wb.save(path)
        return path

    def save_drive(self, ticket_key: str, filename_desc: str, drive_root: str | None = None) -> str:
        """Write straight into the mounted Google Drive: My Drive/Tickets/<KEY>/<KEY> <Desc>.xlsx"""
        self._resolve_query_links()  # deep-link each Source footnote to its query block (may add issues)
        self._raise_if_issues()  # broken workbook cannot reach Drive
        root = drive_root or os.path.expanduser(
            "~/Library/CloudStorage/GoogleDrive-malachi@mountain.com/My Drive/Tickets"
        )
        folder = os.path.join(root, ticket_key.upper().strip())
        os.makedirs(folder, exist_ok=True)
        fname = f"{ticket_key.upper().strip()} {filename_desc}.xlsx"
        path = os.path.join(folder, fname)
        self.wb.save(path)
        return path


# RAG helpers ----------------------------------------------------------------
def rag_threshold(
    good_above: float | None = None,
    bad_below: float | None = None,
    reverse: bool = False,
) -> Callable[[Any], str | None]:
    """Return a fn(value)->'POS'|'WARN'|'NEG' for use in table(rag=...)."""

    def f(v: Any) -> str | None:
        try:
            x = float(v)
        except (TypeError, ValueError):
            return None
        hi, lo = good_above, bad_below
        if hi is not None and x >= hi:
            return "NEG" if reverse else "POS"
        if lo is not None and x <= lo:
            return "POS" if reverse else "NEG"
        return "WARN"

    return f


if __name__ == "__main__":
    print("mntn_xlsx — import this module. See documentation/docs/xlsx_deliverable_standard.md")
