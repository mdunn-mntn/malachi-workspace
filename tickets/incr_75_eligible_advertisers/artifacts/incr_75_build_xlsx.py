"""INCR-75 — build the deliverable Excel workbook from the scorer CSVs.

Sheets:
  1. Funnel Waterfall      — start -> remaining per hard filter + tier/power split
  2. All Advertisers       — every advertiser, per-filter flags, failed_at_filter, tier (audit)
  3. Final Eligible (tiered)— eligible only, row-colored by tier, all user-required columns.
                             Columns tagged [CAN-DETECT] (future-test sensitivity/MDE) vs
                             [MEASURED NOW] (actual ghost-bid lift) vs [PRIOR] (past-test evidence).
  4. Method & Caveats      — definitions, targets, pitfalls, two-family legend
  5. Spend -> MDE curve     — achievable MDE vs monthly spend at INCR-75 eligible-cohort medians
  6. Column Glossary       — plain-English definition of every column, grouped by section (appendix)
  7. Current Lift (ghost-bid)— actual measured treat-vs-holdout lift (entry-cohort, exclude 06-22,
                             7d-from-first-bid window); confirmation-vs-tier + confirmed shortlist

Style cloned from ti_1053 build_deliverable.py (navy header, tier fills, freeze panes).
Reads ../outputs/incr_75_{all_flagged,final_tiered,funnel_counts}.csv.
"""
import csv
import math
import statistics
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "outputs"
BER = ROOT.parent / "ber_2250_incrementality_overhaul"  # reused TI-884 engine lives here
sys.path.insert(0, str(BER / "ti_884_power_sample_size_analysis" / "artifacts"))
from ti_884_mde_calculator import mde_binomial  # noqa: E402

TEST_MONTHS = 56 / 30.4

# ---- palette / styles ----
NAVY = PatternFill("solid", fgColor="1F3A5F")
TOP = PatternFill("solid", fgColor="E8F0E3")     # green
MID = PatternFill("solid", fgColor="FBF3E2")     # amber
LOW = PatternFill("solid", fgColor="F0F0F0")     # gray
GREEN = PatternFill("solid", fgColor="D6EAD0")
RED = PatternFill("solid", fgColor="F6D9D5")
AMBER = PatternFill("solid", fgColor="FBF0D0")
TIER_FILL = {"Top": TOP, "Mid": MID, "Low": LOW, "EXCLUDED": LOW}
WHITE = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="top")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

USD = '$#,##0'
PCT2 = '0.00%'
PP = '0.0" pp"'
NUM1 = '0.0'


def read_csv(name):
    return list(csv.DictReader(open(OUT / name)))


def fnum(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def header_row(ws, headers, r):
    for j, h in enumerate(headers, 1):
        c = ws.cell(r, j, h)
        c.fill = NAVY; c.font = WHITE; c.alignment = WRAP; c.border = THIN


def title_block(ws, ncol, title, subtitle):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(1, 1, title).font = Font(bold=True, size=13, color="1F3A5F")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    s = ws.cell(2, 1, subtitle)
    s.font = Font(italic=True, size=9, color="555555"); s.alignment = WRAP
    ws.row_dimensions[2].height = 46


# column spec for the metric sheets: (csv_key, header, fmt, width)
# fmt: int | usd | pct (raw fraction) | pctx (already *100 -> /100) | pp | num | text
FINAL_COLS = [
    ("final_tier", "Value tier", "text", 8),
    ("value_score", "Value score (0–100)", "num", 11),
    ("advertiser_id", "Advertiser ID", "int", 11),
    ("advertiser_name", "Advertiser", "text", 26),
    ("vertical_buckets", "Industry (vertical)", "text", 20),
    ("avg_monthly_spend", "Avg monthly spend", "usd", 15),
    ("ivr", "Visit rate (IVR)", "pct", 11),
    ("cvr", "Conversion rate (CVR)", "pct", 11),
    ("mde_ivr_at_normal_pct", "[CAN-DETECT] Smallest IVR lift a test detects", "pctx", 17),
    ("can_hit_ivr_5pct_8w", "[CAN-DETECT] Powered for a 5% IVR lift, ≤8w?", "text", 14),
    ("can_hit_ivr_10pct_8w", "[CAN-DETECT] Powered for a 10% IVR lift, ≤8w?", "text", 14),
    ("budget_for_mde_ivr_5pct", "[CAN-DETECT] Test $ to prove a 5% IVR lift", "usd", 16),
    ("budget_for_mde_ivr_10pct", "[CAN-DETECT] Test $ to prove a 10% IVR lift", "usd", 16),
    ("req_monthly_spend_ivr_10pct", "[CAN-DETECT] Monthly $ run-rate (10% IVR)", "usd", 16),
    ("extra_spend_ivr_10pct_abs", "Extra spend needed (10% IVR)", "usd", 14),
    ("extra_spend_ivr_10pct_pct", "Extra spend, % over current", "pctx", 12),
    ("ivr_ask_band", "Budget-ask feasibility", "text", 13),
    ("close_to_ivr_min", "IVR spend feasible (at/near min)?", "text", 12),
    ("mde_cvr_at_normal_pct", "[CAN-DETECT] Smallest CVR lift a test detects", "pctx", 17),
    ("can_hit_cvr_15pct_8w", "[CAN-DETECT] Powered for a loose 15% CVR lift?", "text", 14),
    ("budget_for_mde_cvr_15pct", "[CAN-DETECT] Test $ to prove a 15% CVR lift", "usd", 16),
    ("req_monthly_spend_cvr_15pct", "[CAN-DETECT] Monthly $ run-rate (15% CVR)", "usd", 16),
    ("close_to_cvr_min", "CVR spend feasible (at/near min)?", "text", 12),
    ("prior_lift_pp", "[PRIOR] Lift from a past MNTN test", "pp", 14),
    ("prior_lift_source", "[PRIOR] Source of the prior lift", "text", 14),
    ("mde_ivr_direct_56d_pct", "[CAN-DETECT] Smallest IVR lift: real reach", "pctx", 17),
    ("cpm", "CPM (cost/1k imps)", "usd", 11),
    ("imps_per_ip", "Impressions per IP", "num", 10),
    ("distinct_ips_30d", "Unique IPs reached (30d)", "int", 13),
    # ---- [MEASURED NOW] current ghost-bid lift (Beeswax leg; entry-cohort, excl 06-22, 7d window) ----
    # trimmed to 3 cols (full detail — pp/z/p/CI/signal — stays in the CSV + Sheet 7)
    ("current_lift_confirms", "[MEASURED NOW] Real-lift verdict (10d)", "text", 15),
    ("current_rel_lift", "[MEASURED NOW] Real lift, relative", "relx", 14),
    ("ghost_vis_clean", "[MEASURED NOW] Holdout visits (trust)", "int", 13),
]


def write_value_cell(ws, r, j, key, fmt, row):
    v = row.get(key, "")
    c = ws.cell(r, j); c.border = THIN; c.alignment = CTR if fmt in ("int", "num", "pp", "pctx") else WRAP
    if fmt == "text":
        c.value = v; c.alignment = WRAP if key in ("advertiser_name", "vertical_buckets", "prior_lift_source") else CTR
        return c
    num = fnum(v)
    if num is None:
        c.value = "—"; c.alignment = CTR; return c
    if fmt == "int":
        c.value = int(num)
    elif fmt == "usd":
        c.value = round(num); c.number_format = USD
    elif fmt == "pct":              # raw fraction
        c.value = num; c.number_format = PCT2
    elif fmt == "pctx":            # already *100 -> back to fraction for % format
        c.value = num / 100.0; c.number_format = PCT2
    elif fmt == "pp":
        c.value = num; c.number_format = PP
    elif fmt == "pp3":                 # small percentage-point value (bid-grain ITT)
        c.value = num; c.number_format = '0.000" pp"'
    elif fmt == "relx":               # already in percent-points (18.0 -> "18.0%")
        c.value = round(num, 1); c.number_format = '0.0"%"'
    elif fmt == "num":
        c.value = round(num, 1); c.number_format = NUM1
    return c


# ---------------- Sheet 1: Funnel waterfall ----------------
def sheet_funnel(wb, funnel, tiers, n_prior):
    ws = wb.create_sheet("1. Funnel Waterfall")
    title_block(ws, 6,
                "INCR-75 — Eligible-Advertiser Funnel for Incrementality Lift Tests",
                "Hard filters only (per request, spend/IVR-position/power are scored not cut). "
                "Starting universe = advertisers that delivered in the trailing 30 days. "
                f"{tiers['Top']+tiers['Mid']+tiers['Low']:,} eligible; {n_prior} have prior demonstrated lift.")
    r = 4
    header_row(ws, ["Step", "Filter", "Type", "Threshold", "Removed", "Remaining", "% of start"], r)
    for s in funnel:
        r += 1
        vals = [s["step"] if s["step"] != 99 else "→", s["filter"], s["type"], s["threshold"],
                int(s["removed"]), int(s["remaining"]), s["pct_of_start"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.border = THIN
            c.alignment = CTR if j in (1, 3, 5, 6, 7) else WRAP
            if j == 5 and int(s["removed"]) > 0:
                c.fill = RED
            if s["filter"].startswith("FINAL"):
                c.font = Font(bold=True, color="1F3A5F")
    # tier / power split block
    r += 2
    ws.cell(r, 1, "Eligible set — value tiers").font = Font(bold=True, size=11, color="1F3A5F"); r += 1
    header_row(ws, ["Tier", "Count", "Definition"], r)
    defns = {
        "Top": "Clears 5% IVR MDE + score≥60 + mid-spend + movable IVR + low saturation AND a 'confirmed +' measured lift (run first)",
        "Mid": "Clears 10% IVR MDE (or a-priori-Top demoted for lacking a confirmed measured lift, pending window maturity)",
        "Low": "Eligible but needs a large budget bump to power, or saturated / spend far from sweet spot",
    }
    for t in ("Top", "Mid", "Low"):
        r += 1
        for j, v in enumerate([t, tiers[t], defns[t]], 1):
            c = ws.cell(r, j, v); c.border = THIN; c.fill = TIER_FILL[t]
            c.alignment = CTR if j in (1, 2) else WRAP
    widths = [7, 42, 8, 40, 11, 12, 11]
    for j, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = wd
    ws.freeze_panes = "A5"


# ---------------- Sheet 2: All advertisers (audit) ----------------
def sheet_all(wb, rows):
    ws = wb.create_sheet("2. All Advertisers")
    cols = [
        ("advertiser_id", "Advertiser ID", "int", 11), ("advertiser_name", "Advertiser", "text", 26),
        ("vertical_buckets", "Industry (vertical)", "text", 20), ("active", "Active?", "text", 8),
        ("is_b2b", "B2B?", "text", 6), ("avg_monthly_spend", "Avg monthly spend", "usd", 15),
        ("ivr", "Visit rate (IVR)", "pct", 11), ("cvr", "Conversion rate (CVR)", "pct", 11),
        ("visiting_ips_30d", "Visiting IPs (30d)", "int", 12),
        ("distinct_ips_30d", "Unique IPs reached (30d)", "int", 12),
        ("pass_f1_clean_active", "Pass: clean & active", "bool", 10),
        ("pass_f2_not_b2b", "Pass: not B2B", "bool", 9),
        ("pass_f3_measurable_ivr", "Pass: measurable IVR", "bool", 10),
        ("failed_at_filter", "Disposition", "text", 16),
        ("final_tier", "Value tier", "text", 9), ("value_score", "Value score (0–100)", "num", 11),
    ]
    title_block(ws, len(cols),
                "INCR-75 — All Advertisers (audit trail)",
                "Every advertiser that delivered in the trailing 30d, with each hard-filter pass/fail and final "
                "disposition. PASSED rows are tiered; others show where they dropped out. Sorted PASSED-first by value score.")
    r = 4
    header_row(ws, [h for _, h, _, _ in cols], r)
    for row in rows:
        r += 1
        passed = row["failed_at_filter"] == "PASSED"
        base = TIER_FILL.get(row["final_tier"], LOW) if passed else LOW
        for j, (key, _h, fmt, _w) in enumerate(cols, 1):
            if fmt == "bool":
                val = row.get(key, "").upper() == "TRUE"
                c = ws.cell(r, j, "Y" if val else "N"); c.alignment = CTR; c.border = THIN
                c.fill = GREEN if val else RED
            else:
                c = write_value_cell(ws, r, j, key, fmt, row)
                c.fill = base
                if key == "failed_at_filter" and not passed:
                    c.fill = LOW
    for j, (_k, _h, _f, wd) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = wd
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{r}"


# ---------------- Sheet 3: Final eligible (tiered) ----------------
def sheet_final(wb, rows):
    ws = wb.create_sheet("3. Final Eligible (tiered)")
    title_block(ws, len(FINAL_COLS),
                "INCR-75 — Final Eligible Advertisers (tiered Top → Low)",
                "All must-pass advertisers (not-B2B, active, measurable IVR), ranked by value score. "
                "TWO column families (see the tag in each header): [CAN-DETECT] = the smallest lift a FUTURE 8-wk test "
                "COULD prove (test sensitivity — lower is better; NOT a result); [MEASURED NOW] = the ACTUAL lift already "
                "measured vs a live holdout (judged significant-vs-ZERO, not vs 5%; higher/positive is better). "
                "MEASURED-LIFT GATE (staged): advertisers with a significant NEGATIVE measured lift are excluded; "
                "TOP tier now REQUIRES a 'confirmed +' measured lift (a-priori Tops without it demote to Mid); "
                "'flat so far' / 'too early' advertisers stay eligible and re-gate as the 10-day window matures toward 30d. "
                "MDE is RELATIVE (5% on 2% IVR = detect 2.1%). 8-week test ≈ 1.84 months. "
                "Highlights: green=Top, amber=Mid, gray=Low.")
    r = 4
    header_row(ws, [h for _, h, _, _ in FINAL_COLS], r)
    ws.row_dimensions[4].height = 60  # tagged headers wrap to ~3 lines
    for row in rows:
        r += 1
        fill = TIER_FILL.get(row["final_tier"], LOW)
        for j, (key, _h, fmt, _w) in enumerate(FINAL_COLS, 1):
            c = write_value_cell(ws, r, j, key, fmt, row)
            c.fill = fill
            if key in ("can_hit_ivr_5pct_8w", "can_hit_ivr_10pct_8w", "can_hit_cvr_15pct_8w",
                       "close_to_ivr_min", "close_to_cvr_min"):
                v = (row.get(key) or "").lower()
                if v == "yes":
                    c.fill = GREEN
                elif v == "no":
                    c.fill = RED
                elif v == "no_data":
                    c.fill = LOW
            if key == "ivr_ask_band":
                v = (row.get(key) or "")
                c.fill = {"easy": GREEN, "none": GREEN, "stretch": AMBER, "unreasonable": RED}.get(v, fill)
            if key == "prior_lift_pp" and fnum(row.get(key)) is not None:
                c.font = Font(bold=True, color="1F6B2E")
            if key == "current_lift_confirms":
                v = (row.get(key) or "")
                c.fill = {"confirmed +": GREEN, "flat so far": LOW,
                          "too early": LOW, "no data yet": LOW}.get(v, fill)
    for j, (_k, _h, _f, wd) in enumerate(FINAL_COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = wd
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(FINAL_COLS))}{r}"


# ---------------- Sheet 4: Method & Caveats ----------------
def sheet_method(wb, medians):
    ws = wb.create_sheet("4. Method & Caveats")
    ws.column_dimensions["A"].width = 120
    lines = [
        ("INCR-75 — Method & Caveats", "h"),
        ("Goal", "h2"),
        ("Identify live MNTN advertisers that are the best candidates for incremental-lift studies: measurable-but-"
         "movable IVR, smaller/lesser-known brands, mostly-net-new audience, powerable within 4–8 weeks.", "p"),
        ("READ THIS FIRST — two column families (the tag in each header tells you which)", "h2"),
        ("The workbook has two kinds of “lift %” columns that mean OPPOSITE things — never compare them directly:\n"
         "• [CAN-DETECT] = test SENSITIVITY. The smallest lift a FUTURE ~8-week test COULD prove at this advertiser's "
         "spend/reach. LOWER is better. It is NOT a result — nothing has been measured; it just says how fine an "
         "instrument the budget buys. (This family drives the eligibility score.)\n"
         "• [MEASURED NOW] = actual RESULT. The real treatment-vs-holdout visit lift already measured from live "
         "ghost-bid data (~10 days so far). HIGHER/positive is better. Judged SIGNIFICANT-vs-ZERO (is there a real "
         "effect?), NOT against the 5% detectable floor — a significant +3% is good.\n"
         "Key trap: an advertiser can be well-powered under [CAN-DETECT] (small floor for a future 8-wk test) yet "
         "“unconfirmed” under [MEASURED NOW] because the current ~10 days of live data are still too thin — different "
         "time horizons, different questions.", "p"),
        ("Is MDE relative or absolute?", "h2"),
        ("RELATIVE. MDE_rel = MDE_abs / baseline. For a 0.5% IVR advertiser, a 5% MDE means detecting a 0.525% IVR "
         "(a 5% proportional lift, +0.025 percentage points) — NOT 5.5%. Matches how lift is reported at MNTN.", "p"),
        ("Reasonable MDE — IVR vs CVR", "h2"),
        ("IVR: both 5% (credible) and 10% (realistic) computed; eligibility tiers on 10%, Top tier requires 5% at "
         "normal spend. CVR is ~7–10x harder (baseline ~30x lower; MDE_rel ∝ √((1−p)/p) explodes as p→0) — a 5% CVR "
         "MDE needs ~$2–5M/mo, so CVR is INFORMATIONAL only, never a gate. The 15% CVR target is a FEASIBILITY CEILING "
         "(a tight CVR MDE is unaffordable for nearly all advertisers), NOT a claim that CVR lifts are ~15% — if "
         "anything CVR's true relative lift is comparable to or smaller than IVR's. Judge CVR on the "
         "'[CAN-DETECT] Smallest CVR lift a test detects' column, not the 15% yes/no.", "p"),
        ("What counts as 'enough' spend?", "h2"),
        ("Per-advertiser, not a flat number: enough = running 8 weeks at typical spend accumulates enough treated IPs "
         "to push IVR MDE ≤ target. Encoded in the power columns. Low-spend advertisers fail to power and lack the ROI "
         "to justify a test.", "p"),
        ("Reasonable extra-spend ask", "h2"),
        ("Banded label (never a cut): ≤25% over normal = easy, 25–50% = stretch, >50% = unreasonable.", "p"),
        ("Baseline definition (settled, TI-1019 §7b/§7e)", "h2"),
        ("IVR = distinct visiting∩served IPs / distinct served cost_impression_log.ip (per-IP Bernoulli rate). "
         "NOT graph.visits (event count, 3.36x inflated) or graph.usersreached (device-blended, ~2x). All-funnel grain, "
         "matching the team's MDE-prefill calculator + the gary-ql resolver.", "p"),
        ("Power math", "h2"),
        ("TI-884 Lewis-Rao two-proportion z-test, z=2.80 (α=0.05, power=0.80), 10% holdout, var_reduction=1.0 "
         "(no CUPED/ghost-ad/stratified reduction). Budget-for-MDE is the total test budget to reach the required "
         "distinct treated IPs; required monthly spend = budget / 1.84 (8wk ≈ 1.84 months).", "p"),
        (f"INCR-75 eligible-cohort medians: IVR {medians['ivr']*100:.2f}%, CVR {medians['cvr']*100:.3f}%, "
         f"CPM ${medians['cpm']:.2f}, imps/IP {medians['ipi']:.1f}.", "p"),
        ("Filter funnel", "h2"),
        ("HARD (membership): (1) clean & active; (2) not B2B — exclude the 'B2B Software & Services' vertical bucket; "
         "(3) measurable IVR — ≥100 visiting IPs and IVR>0; (4) measured lift not negative — exclude advertisers whose "
         "live ghost-bid lift is significantly NEGATIVE (staged gate, 2026-07-02). SCORED (tier, not cut): mid-spend sweet "
         "spot ($25k–$200k/mo), IVR band position (peak 3–6%, >12% = saturated/hard-to-move), powerability at 5%/10%, "
         "brand-size (spend rank + reach-to-spend), audience saturation (reach-to-spend), and a prior-demonstrated-lift bonus. "
         "TIER GATE: Top additionally requires a 'confirmed +' measured lift (a-priori Tops without it demote to Mid until "
         "the 10-day ghost window matures toward 30d ~late-July).", "p"),
        ("Pitfalls", "h2"),
        ("• spend_required uses 30d imps/IP and is an OPTIMISTIC floor for large budget gaps (imps/IP grows with window "
         "length); the '[CAN-DETECT] Smallest IVR lift: real reach' column is the no-extrapolation cross-check.\n"
         "• The [CAN-DETECT] columns are a POWER screen — eligibility ≠ guaranteed lift; the [MEASURED NOW] columns are "
         "the actual result. Matt Brorby's clean population read (entry-cohort, exclude 06-22) is a small POSITIVE "
         "(~+5%), and lift rises from top-intent (≈0) to mid-intent (highest) — movability lives in mid-IVR, which is "
         "what this screen rewards.\n"
         "• CVR MDE reported only when ≥50 converting IPs (else small-p noise).\n"
         "• Managed-service advertisers can log ~$0 spend with real delivery; ratios are SAFE-guarded.\n"
         "• Prior lift: TI-933 = Select clickpass visit-rate pp (significant only); TI-837 = guid total-traffic pp "
         "(all-funnel, includes retargeting — a permissive 'has shown lift' signal, not a prospecting-incrementality "
         "estimate).", "p"),
        ("Sources", "h2"),
        ("Metrics: queries/incr_75_advertiser_metrics.sql (forks TI-1019). Engine: TI-884 ti_884_mde_calculator.py. "
         "Prior lift: TI-933 ti_933_per_advertiser_lift.csv, TI-837 ti_837_lift_30adv_7day_v5. Window: trailing 30d "
         "(rates/CPM/imps-per-IP), trailing 56d (direct power), trailing 12mo (typical-active-month spend).", "p"),
    ]
    r = 1
    for text, kind in lines:
        c = ws.cell(r, 1, text); c.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "h":
            c.font = Font(bold=True, size=14, color="1F3A5F")
        elif kind == "h2":
            c.font = Font(bold=True, size=11, color="1F3A5F")
        else:
            c.font = Font(size=10, color="333333")
            ws.row_dimensions[r].height = 14 * (1 + text.count("\n") + len(text) // 110)
        r += 1
    ws.sheet_view.showGridLines = False


# ---------------- Sheet 5: Spend -> MDE curve ----------------
def sheet_curve(wb, medians):
    ws = wb.create_sheet("5. Spend → MDE curve")
    title_block(ws, 5,
                "Spend → Achievable MDE (at INCR-75 eligible-cohort medians)",
                f"At median IVR {medians['ivr']*100:.2f}%, CVR {medians['cvr']*100:.3f}%, CPM ${medians['cpm']:.2f}, "
                f"imps/IP {medians['ipi']:.1f}. var_reduction=1.0. Read: a $X/mo advertiser detects ≥ this relative lift "
                "in an 8-week test. Use to size how much budget unlocks a target MDE.")
    r = 4
    header_row(ws, ["Monthly spend", "8-wk test budget", "IVR MDE (rel)", "CVR MDE (rel)", "IVR verdict @5%"], r)
    grid = [10_000, 25_000, 50_000, 75_000, 100_000, 150_000, 200_000, 300_000, 500_000, 1_000_000]
    for ms in grid:
        r += 1
        test_budget = ms * TEST_MONTHS
        imps = test_budget / medians["cpm"] * 1000
        treated = imps / medians["ipi"]
        holdout = treated * (0.10 / 0.90)
        _, mv = mde_binomial(treated, holdout, medians["ivr"], var_reduction=1.0)
        _, mc = mde_binomial(treated, holdout, medians["cvr"], var_reduction=1.0)
        verdict = "well-powered" if mv <= 0.05 else "borderline" if mv <= 0.10 else "underpowered"
        cells = [ms, round(test_budget), mv, mc, verdict]
        for j, v in enumerate(cells, 1):
            c = ws.cell(r, j, v); c.border = THIN; c.alignment = CTR
            if j in (1, 2):
                c.number_format = USD
            if j in (3, 4):
                c.number_format = PCT2
            if j == 5:
                c.fill = {"well-powered": GREEN, "borderline": AMBER, "underpowered": RED}[verdict]
    for j, wd in enumerate([16, 16, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(j)].width = wd
    ws.freeze_panes = "A5"


# ---------------- Sheet 6: Column glossary (appendix) ----------------
# (section, column-as-shown, plain definition, how it's computed / notes)
GLOSSARY = [
    ("§", "IDENTITY & RANKING", "", ""),
    ("", "Value tier", "Priority bucket for running a test. Top = run first; Low = eligible but lowest priority.",
     "Top = [CAN-DETECT] can detect a 5% IVR lift at current spend + value-score≥60 + mid-spend + movable IVR + unsaturated, AND (staged gate) a 'confirmed +' [MEASURED NOW] lift. Mid = can detect 10% at current spend (or a-priori-Top demoted for lacking a confirmed lift). Low = needs a big budget bump or is saturated / off the spend sweet-spot. Advertisers with a significant-NEGATIVE measured lift are EXCLUDED (F4). 'apriori_tier' (in the CSV) shows the tier before the measured gate."),
    ("", "Value score (0–100)", "Composite ranking score within the eligible set; higher = better candidate.",
     "Power margin (30) + mid-spend fit (20) + smaller-brand/movability (20) + IVR-band position (15) + low audience saturation (15) + prior-lift bonus (+10)."),
    ("", "Advertiser ID", "MNTN advertiser_id.", "From core advertiser dimension."),
    ("", "Advertiser", "Advertiser company name.", "advertisers.company_name."),
    ("", "Industry (vertical)", "Advertiser's industry bucket(s).", "fpa_advertiser_verticals (type=0 industry bucket)."),

    ("§", "SPEND & UNIT ECONOMICS", "", ""),
    ("", "Avg monthly spend", "The advertiser's typical monthly budget — the baseline we test against.",
     "Median monthly media spend across active months (>$1k) over the last 12 months. Robust to on/off & seasonal advertisers."),
    ("", "CPM (cost/1k imps)", "Advertiser-paid cost per 1,000 impressions.",
     "(media + data + platform spend) ÷ impressions × 1000, trailing 30d."),
    ("", "Impressions per IP", "Average ad frequency — impressions delivered per unique IP.",
     "impressions ÷ distinct served IPs (30d). Higher = more frequency, fewer NEW IPs per dollar."),
    ("", "Unique IPs reached (30d)", "Reach: distinct IPs served at least one ad in the last 30 days.",
     "COUNT(DISTINCT ip) from cost_impression_log, 30d."),

    ("§", "RATES (the baseline we measure lift on)", "", ""),
    ("", "Visit rate (IVR)", "Share of served households (IPs) that then visited the site. The headline KPI.",
     "distinct visiting-AND-served IPs ÷ distinct served IPs (30d). A per-IP probability, NOT impressions-based."),
    ("", "Conversion rate (CVR)", "Share of served households (IPs) that converted. ~30× rarer than visits.",
     "distinct converting-AND-served IPs ÷ distinct served IPs (30d)."),

    ("§", "[CAN-DETECT] IVR POWER — smallest lift a FUTURE test could PROVE (test sensitivity, NOT a measured result)", "", ""),
    ("", "[CAN-DETECT] Smallest IVR lift a test detects",
     "The smallest true IVR (visit-rate) lift a FUTURE ~8-week test could statistically prove at this advertiser's current spend/reach. Sensitivity of the measuring instrument, not a result; lower is better.",
     "RELATIVE MDE: 5% = a 5% proportional lift (2.00%→2.10% IVR), NOT 5pp. TI-884 Lewis-Rao two-proportion z-test (z=2.80, α=.05 two-sided, power=.80), 10% holdout, var_reduction=1.0, from current 8-wk reach. Never a measured outcome."),
    ("", "[CAN-DETECT] Powered for a 5% IVR lift, ≤8w?", "Yes = a future 8-wk test at today's spend is sensitive enough to PROVE a 5% IVR lift IF one exists (the credible bar). A capability verdict, not a claim a 5% lift was found.",
     "Yes if current 8-wk spend reaches the distinct served IPs needed for a 5% relative IVR MDE. Green=Yes, red=No. Says nothing about any observed effect."),
    ("", "[CAN-DETECT] Powered for a 10% IVR lift, ≤8w?", "Same, for a 10% lift (the easier realistic bar). Still test sensitivity, not a measured result.",
     "Yes if current 8-wk spend reaches the IPs for a 10% relative IVR MDE. This 10% flag is the eligibility/tiering gate; 5% is the tighter stretch bar."),
    ("", "[CAN-DETECT] Test $ to prove a 5% / 10% IVR lift",
     "Total future-test dollars needed to make the instrument sensitive enough to prove that lift. A budget to BUY power — not spend already committed, not a lift earned.",
     "TI-884 Lewis-Rao: dollars to reach the required distinct served IPs at the advertiser's own CPM & impressions/IP, no variance reduction. 5% costs ~4× the 10% figure."),
    ("", "[CAN-DETECT] Monthly $ run-rate (10% IVR)", "The 10%-IVR total-test budget as a monthly run-rate, comparable to normal monthly spend. A prospective requirement, not a measured lift.",
     "Total test spend ÷ 1.84 (an 8-week test ≈ 1.84 months)."),
    ("", "Extra spend needed (10% IVR)", "Additional dollars beyond what they'd spend anyway, to reach the 10% bar. $0 if already powered.",
     "max(0, total test spend for 10% − current 8-wk spend)."),
    ("", "Extra spend, % over current", "That extra as a percentage of their current 8-week spend.",
     "extra ÷ (current 8-wk spend)."),
    ("", "Budget-ask feasibility", "How big the budget ask is, in plain terms.",
     "none = already powered · easy = ≤25% more · stretch = 25–50% more · unreasonable = >50% more."),
    ("", "IVR spend feasible (at/near min)?", "Is this advertiser spend-feasible for an IVR test? Yes = already AT/OVER the IVR spend minimum (no ask needed) OR a reasonable (≤50%) bump away. No = would need an unreasonable (>50%) spend increase. Already-spending-plenty advertisers are Yes (not an issue).",
     "Yes if current 8-wk spend ≥ (10% IVR budget) / 1.5. One-sided — spending well over the minimum is never flagged No."),
    ("", "[CAN-DETECT] Smallest IVR lift: real reach",
     "Same smallest-detectable IVR lift as the modeled column, but from the ACTUAL distinct IPs reached in the last 56 days instead of a spend projection. A no-extrapolation cross-check — still instrument sensitivity, still NOT a measured lift; lower is better.",
     "RENAMED to kill the old 'measured 8-wk reach' trap: 'measured' there meant measured REACH (distinct served IPs from impression logs), never a measured lift. Same Lewis-Rao MDE on real 56d reach; a large gap vs the modeled column flags a frequency/reach quirk."),

    ("§", "[CAN-DETECT] CVR POWER — informational only (conversions need ~$2M+/mo)", "", ""),
    ("", "[CAN-DETECT] Smallest CVR lift a test detects",
     "Smallest CVR (conversion-rate) lift a future 8-wk test could prove at today's spend. Usually large (10s–100s %) because CVR is ~30× rarer than IVR. Judge CVR on THIS, not the 15% yes/no; sensitivity, not a result.",
     "Same relative Lewis-Rao math on the CVR base rate; mde_rel grows as p→0. Lower = better powered. '—' if <50 converting IPs. Informational only; CVR is never a gate."),
    ("", "[CAN-DETECT] Powered for a loose 15% CVR lift?",
     "Yes = even a deliberately LOOSE 15% CVR bar is powerable ≤8 wks (usually No). 15% is a feasibility CEILING, NOT a claim CVR lifts are ~15%, and not a result.",
     "Yes if current 8-wk spend reaches the IPs for a 15% relative CVR MDE; 'no_data' if <50 converting IPs. A tight CVR floor is unaffordable for nearly everyone. Never an elimination criterion."),
    ("", "[CAN-DETECT] Test $ to prove a 15% CVR lift", "Total future-test dollars to prove even a loose 15% conversion lift.",
     "Same math as IVR, on the CVR baseline. Far higher because CVR is ~30× rarer."),
    ("", "[CAN-DETECT] Monthly $ run-rate (15% CVR)", "The 15%-CVR total-test figure as a monthly run-rate.", "Total ÷ 1.84."),
    ("", "CVR spend feasible (at/near min)?", "Same as the IVR flag, for the loose 15% CVR bar: Yes = already at/over OR a ≤50% bump away; No = needs >50% more; 'no_data' if <50 converting IPs.", "Yes if current 8-wk spend ≥ (15% CVR budget) / 1.5."),

    ("§", "[PRIOR] PRIOR EVIDENCE — a result from an EARLIER test (neither a future-test floor nor the current ghost-bid result)", "", ""),
    ("", "[PRIOR] Lift from a past MNTN test", "A positive lift this advertiser already showed in a PAST MNTN test (a bonus signal in the score). Historical — not the current [MEASURED NOW] ghost-bid result and not a [CAN-DETECT] sensitivity number.",
     "In percentage points (pp). Blank if none."),
    ("", "[PRIOR] Source of the prior lift", "Which past study the prior lift came from.",
     "TI-933 = Select clickpass visit-rate test (significant only). TI-837 = ghost-bid guid total-traffic (all-funnel; permissive 'has shown lift' signal)."),

    ("§", "[MEASURED NOW] CURRENT LIFT — actual result vs a live ghost-bid holdout (judged vs ZERO, not vs 5%)", "", ""),
    ("", "[MEASURED NOW] Real-lift verdict (10d)", "The measured verdict over the current ~10-day ghost-bid window, and the driver of the staged gate. Four values (see the Holdout-visits column to tell 'flat' from 'thin').",
     "confirmed + = ≥20 holdout visits, two-sided p<.05, POSITIVE (this is what a Top advertiser must have). flat so far = ≥100 holdout visits, not significant (enough data, ~0 effect so far). too early = <100 holdout visits, not significant (thin / window too short). no data yet = not in the ghost-bid table. (A significant-NEGATIVE verdict → advertiser EXCLUDED, so it never appears here.) Judged SIGNIFICANT-vs-ZERO, NOT vs the 5% MDE — a significant +3% is confirmed +. Green = confirmed +. Full stats (pp, z, p, CI) are in the CSV + this sheet."),
    ("", "[MEASURED NOW] Real lift, relative", "The actual visit lift measured vs a live never-served holdout, as a % of the holdout baseline — the number to LEAD with. +18% = treated visit rate is 18% above holdout. Higher/positive is better.",
     "(treat_vr − ghost_vr)/ghost_vr on the clean set. Entry-cohort, EXCLUDE 06-22, 7-day-from-first-bid window (Matt Brorby). Bar is ZERO + significance, NOT 5%. Rank by THIS magnitude+direction; the z (in the CSV) is N-inflated so it's a floor, not the headline. ~10 days so far; bid-grain ITT (diluted by win rate). Compare to '[MEASURED NOW] Real lift, pp' in the CSV: relative = pp ÷ holdout baseline (pp is the raw gap, tiny because the denominator is all bid-eligible IPs)."),
    ("", "[MEASURED NOW] Holdout visits (trust)", "How many holdout households actually visited — the sample size behind the verdict. Read it WITH the verdict: 'flat so far' + a big number = enough data, ~0 so far; 'too early' + a small number = not enough data yet.",
     "Distinct clean holdout (ghost) IPs with a visit. <20 → can't be significant; <100 → verdict caps at 'too early'. This is [MEASURED NOW]'s own clock (how mature the live window is), distinct from the [CAN-DETECT] future-test power."),
    ("", "[MEASURED NOW] Method, gate & caveats", "Why these numbers are trustworthy, how the gate uses them, AND the limits.",
     "STAGED GATE: significant-NEGATIVE advertisers are excluded (F4); TOP requires 'confirmed +'; 'flat so far'/'too early' stay eligible and re-gate as the window grows (chosen because today's 10-day window makes most advertisers inconclusive for lack of TIME, not lack of lift). Beeswax/JVM-bidder leg (source bid_price_log); the MNTN Rust-bidder leg isn't folded in yet. Entry-cohort + exclude-06-22 removes the left-edge stock that manufactured a spurious negative; on the clean set ghost_frac lands on 0.10 and pooled lift = +5% (z≈26). The gold ghost_bid_rollup is all-time / can't drop 06-22 → reads spuriously negative — don't use it yet. Tables now accumulate (no TTL) → a true ≥30-day window ~late-July. Directional per INCR-69."),

    ("§", "ALL-ADVERTISERS SHEET — extra columns", "", ""),
    ("", "Active?", "Whether the advertiser is currently active.", "advertisers.active."),
    ("", "B2B?", "Whether the advertiser is B2B (these are excluded).", "In the 'B2B Software & Services' vertical bucket."),
    ("", "Visiting IPs (30d)", "Distinct IPs that visited the site (the IVR numerator).", "COUNT(DISTINCT ip) from clickpass_log ∩ served, 30d."),
    ("", "Pass: clean & active / not B2B / measurable IVR", "The three hard filters (Y/N).",
     "clean & active = active, named, served. not B2B = not in the B2B bucket. measurable IVR = ≥100 visiting IPs & IVR>0."),
    ("", "Disposition", "Outcome: PASSED (eligible) or the first filter it failed.", "PASSED / F1_clean_active / F2_not_b2b / F3_measurable_ivr."),
]


def sheet_glossary(wb):
    ws = wb.create_sheet("6. Column Glossary")
    title_block(ws, 3,
                "Column Glossary (appendix)",
                "Plain-English definition of every column on the 'All Advertisers' and 'Final Eligible' sheets. "
                "Key reminder: every lift / MDE figure is RELATIVE — a 5% MDE on a 0.5% visit rate means detecting 0.525%, "
                "not 5.5%. All budgets assume NO variance reduction; an 8-week test ≈ 1.84 months of spend.")
    r = 4
    header_row(ws, ["Column", "What it means", "How it's computed / notes"], r)
    for tag, col, definition, computed in GLOSSARY:
        r += 1
        if tag == "§":
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            c = ws.cell(r, 1, col); c.font = Font(bold=True, color="FFFFFF", size=10.5)
            c.fill = PatternFill("solid", fgColor="33506E"); c.alignment = Alignment(vertical="center")
            ws.row_dimensions[r].height = 18
            continue
        cells = [(col, Font(bold=True, size=10, color="1F3A5F")), (definition, Font(size=10, color="222222")),
                 (computed, Font(size=9, color="555555"))]
        for j, (val, fnt) in enumerate(cells, 1):
            c = ws.cell(r, j, val); c.font = fnt; c.alignment = WRAP; c.border = THIN
        ws.row_dimensions[r].height = max(28, 12 * (1 + max(len(definition) // 48, len(computed) // 58)))
    for j, wd in enumerate([34, 50, 62], 1):
        ws.column_dimensions[get_column_letter(j)].width = wd
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


# ---------------- Sheet 7: Current Lift (live ghost-bid) ----------------
CL_COLS = [
    ("rank", "#", 5), ("final_tier", "Tier", 6), ("advertiser_id", "Advertiser ID", 11),
    ("advertiser_name", "Advertiser", 26), ("value_score", "Value score", 10),
    ("ivr", "IVR baseline", 9), ("current_rel_lift", "Real lift, relative (lead)", 13),
    ("current_abs_lift_pp", "Real lift, abs pp", 12), ("ghost_vis_clean", "Holdout visits (trust)", 12),
    ("current_z", "z (floor only)", 9), ("current_p", "p", 8), ("prior_lift_pp", "Prior lift", 9),
]


def sheet_current_lift(wb, final_rows):
    ws = wb.create_sheet("7. Current Lift (ghost-bid)")
    confirmed = [r for r in final_rows if r.get("current_lift_confirms") == "confirmed +"
                 and r["final_tier"] in ("Top", "Mid")]
    confirmed.sort(key=lambda r: -(fnum(r.get("current_rel_lift")) or -1e9))
    VS = ("confirmed +", "flat so far", "too early", "no data yet")
    counts = {}
    for t in ("Top", "Mid", "Low"):
        sub = [r for r in final_rows if r["final_tier"] == t]
        counts[t] = {k: sum(1 for r in sub if r.get("current_lift_confirms") == k) for k in VS}
    title_block(ws, len(CL_COLS),
                "INCR-75 — Current Measured Lift (live ghost-bid holdout) + the staged gate",
                "Actual treatment-vs-holdout visit ITT from silver enriched.lift__ghost_bid_visits. Method per Matt Brorby "
                "(2026-07-02): entry-anchored at first bid per advertiser×campaign×IP, 7-day-from-first-bid window, EXCLUDING "
                "the 2026-06-22 left-edge day. On the clean set the holdout fraction lands on 0.10 (design) and pooled lift = "
                "+5% (z≈26). STAGED GATE: the 17 significant-NEGATIVE advertisers are EXCLUDED; TOP now requires a 'confirmed +' "
                "lift (so Top = 21, all confirmed); 'flat so far'/'too early' stay eligible and re-gate as the window grows to "
                "30d (~late-July). NB: the gold ghost_bid_rollup is all-time / can't drop 06-22 → reads spuriously negative — "
                "don't use it. READ RELATIVE LIFT + DIRECTION; z is a floor (N-inflated); absolute pp are bid-grain ITT.")
    # summary block
    r = 4
    ws.cell(r, 1, "Measured-lift verdict by tier (after the gate)").font = Font(bold=True, size=11, color="1F3A5F"); r += 1
    header_row(ws, ["Tier", "confirmed +", "flat so far", "too early", "no data yet"], r)
    for t in ("Top", "Mid", "Low"):
        r += 1
        vals = [t] + [counts[t][k] for k in VS]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v); c.border = THIN; c.alignment = CTR; c.fill = TIER_FILL[t]
            if j == 2 and v: c.fill = GREEN
    r += 2
    ws.cell(r, 1, f"Strongest candidates — Top/Mid tier with a 'confirmed +' measured lift "
                  f"({len(confirmed)}), ranked by relative lift").font = Font(bold=True, size=11, color="1F3A5F")
    r += 1
    header_row(ws, [h for _, h, _ in CL_COLS], r)
    for i, row in enumerate(confirmed, 1):
        r += 1
        fill = TIER_FILL.get(row["final_tier"], LOW)
        for j, (key, _h, _w) in enumerate(CL_COLS, 1):
            if key == "rank":
                c = ws.cell(r, j, i); c.alignment = CTR; c.border = THIN; c.fill = fill; continue
            fmt = {"advertiser_id": "int", "advertiser_name": "text", "value_score": "num",
                   "ivr": "pct", "current_rel_lift": "relx", "current_abs_lift_pp": "pp3",
                   "ghost_vis_clean": "int", "current_z": "num", "current_p": "num",
                   "prior_lift_pp": "pp", "final_tier": "text"}[key]
            c = write_value_cell(ws, r, j, key, fmt, row); c.fill = fill
            if key == "current_rel_lift":
                c.font = Font(bold=True, color="1F6B2E")
            if key == "final_tier":
                c.fill = TIER_FILL.get(row["final_tier"], LOW)
    for j, (_k, _h, wd) in enumerate(CL_COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = wd
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def main():
    all_rows = read_csv("incr_75_all_flagged.csv")
    final_rows = read_csv("incr_75_eligible_with_current_lift.csv")  # tiered + current-lift columns
    funnel = read_csv("incr_75_funnel_counts.csv")
    tiers = {t: sum(1 for x in final_rows if x["final_tier"] == t) for t in ("Top", "Mid", "Low")}
    n_prior = sum(1 for x in final_rows if x.get("has_prior_lift", "").upper() == "TRUE")

    elig_ivr = [fnum(x["ivr"]) for x in final_rows if fnum(x["ivr"])]
    elig_cvr = [fnum(x["cvr"]) for x in final_rows if fnum(x["cvr"])]
    elig_cpm = [fnum(x["cpm"]) for x in final_rows if fnum(x["cpm"])]
    elig_ipi = [fnum(x["imps_per_ip"]) for x in final_rows if fnum(x["imps_per_ip"])]
    medians = {"ivr": statistics.median(elig_ivr), "cvr": statistics.median(elig_cvr),
               "cpm": statistics.median(elig_cpm), "ipi": statistics.median(elig_ipi)}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_funnel(wb, funnel, tiers, n_prior)
    sheet_all(wb, all_rows)
    sheet_final(wb, final_rows)
    sheet_method(wb, medians)
    sheet_curve(wb, medians)
    sheet_glossary(wb)
    sheet_current_lift(wb, final_rows)

    path = OUT / "incr_75_eligible_advertisers.xlsx"
    wb.save(path)
    print(f"wrote {path}")
    print(f"  Sheet 2 rows: {len(all_rows)}  Sheet 3 rows: {len(final_rows)}  tiers: {tiers}  prior-lift: {n_prior}")
    print(f"  cohort medians: IVR {medians['ivr']*100:.2f}% CVR {medians['cvr']*100:.3f}% "
          f"CPM ${medians['cpm']:.2f} imps/IP {medians['ipi']:.1f}")


if __name__ == "__main__":
    main()
