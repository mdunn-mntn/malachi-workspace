#!/usr/bin/env python3
"""AUDI-1194 optimizer explainer + proof deliverable.

Builds the branded multi-sheet .xlsx for the SUCCESS-SWEEP optimizer (O1-O4): how it works,
the detector catalog, and real prod findings (the 242x skew etc.). The failure DEBUGGER has
its own workbook under AUDI-1191.

Regenerate: python3 tickets/audi_1194_optimizer_efficiency_crawler/artifacts/audi_1194_how_it_works_xlsx.py
"""

import os
import sys

import pandas as pd
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))
from lib.mntn_xlsx import BRAND, FMT, MntnWorkbook

GEN = "2026-08-20"
GH = "https://github.com/mdunn-mntn/malachi-workspace/blob/main/"
ART = "https://claude.ai/code/artifact/878ac222-4ed6-4376-aea5-cd1772308cca"

wb = MntnWorkbook(
    title="Airflow/Spark Efficiency Optimizer",
    ticket="AUDI-1194",
    subtitle="Reads jobs that SUCCEEDED and finds waste: skew, spill, fetch wait, idle fleets, missing stats. Ranks a daily backlog worst-first and reports what it could not see.",
    period="Daily since 2026-08-20",
    generated=GEN,
    status="Daily, full fleet",
)

# ---------------------------------------------------------------- 1. How it works (step map)
STEPS = [
    # (step, name, what/how, code display, code repo-path#anchor, test command, proven)
    ("O1", "Get the job's Spark data",
     "Runs on SUCCEEDED jobs. Dataproc: the .zstd Spark event log from gs://mntn-data-archive-{env}/spark-events "
     "(fleet-enabled by PR #1169) or the per-batch history-server folders for ipdsc/tpa (temp bucket, needs the "
     "standing read grant). Databricks: the EXPLAIN COST plan from a SQL warehouse (jobs get-run-output carries no plan).",
     "oncall_daily_optimizer.sh", ".claude/scripts/oncall_daily_optimizer.sh",
     "bash .claude/scripts/oncall_daily_optimizer.sh --selftest",
     "2026-08-20: daily at cap 200 (~160 logs/day, so weekly cap-40 saw ~4% of the fleet). PHS half proven end-to-end under a 1h PAM grant: 22/22 batches enumerated, fetched and parsed. Standing read pends mntn-devops#4724"),
    ("O2", "Parse 7 surfaces",
     "Full Spark event-log parse into structured metrics: jobs, stages, tasks, executors, environment, SQL node "
     "metrics, storage.",
     "eventlog.py:147", "airflow_optimizer/eventlog.py#L147",
     "python3 -m airflow_optimizer.tests.test_eventlog",
     "Hardened 2026-08-07: 41 corpus-confirmed defects fixed; 48-log/611MB real corpus parses 48/48 in 58s at 49MB RSS (was 18GB); multi-frame zstd streamed, corrupt logs error instead of passing clean"),
    ("O3", "Detect waste",
     "Plan detectors (missing stats, broadcast candidate, shuffle sizing, window full-sort, repeated scan) + run "
     "detectors (skew, spill, GC pressure, spot-preemption cost, fetch instability), each with real numbers and a "
     "concrete fix. See the Detector catalog tab.",
     "optimizations.py:97", "airflow_optimizer/optimizations.py#L97",
     "python3 -m airflow_optimizer.tests.test_optimizations",
     "Live ask 2026-08-07 (intent_score_map): straggler + pinned fleet root-caused, verified against Spark source; new fetch-wait detector immediately found 6 jobs at 53-72% fetch-wait"),
    ("O4", "Rank the fleet backlog",
     "Crawls every event log, ranks jobs worst-first, groups each finding CODE / INFRA / FAILURE so the owner "
     "knows the kind of fix. This is the 'check every DAG automatically' mode.",
     "crawl.py:54", "airflow_optimizer/crawl.py#L54",
     "python3 -m airflow_optimizer.crawl <dir-or-glob>",
     "First daily run 2026-08-20: 214 jobs, 278 findings, 197 high (vs 37/59/42 on the last weekly run), including 21 PHS-only jobs the archive sweep had never seen"),
    ("O5", "Record what changed",
     "Appends every finding to a ledger keyed on job + detector + stage, and replays it to derive a state: new, "
     "chronic (3+ consecutive sweeps), owner_notified, wont_fix, resolved (stopped firing for 3 sweeps). Without it a "
     "backlog rebuilt each day cannot say whether anything improved.",
     "ledger.py", "airflow_optimizer/ledger.py",
     "python3 -m airflow_optimizer.ledger",
     "26 findings across 25 job-logs collapse to 4 distinct keys: an hourly job contributes ~24 logs a day and must count once"),
    ("O6", "Report coverage and the delta",
     "Enumerates every unpaused DAG from the Airflow API and names the ones with no Spark task, so the backlog is never "
     "mistaken for the fleet. Then renders a digest that leads with what changed, links each job to its Airflow page, and "
     "points at the full backlog.",
     "coverage.py / digest.py", "airflow_optimizer/coverage.py",
     "python3 -m airflow_optimizer.coverage <api_base> <date>",
     "2026-08-20 live: 62 active DAGs, 24 have a Spark task, 38 are structurally invisible to this tool and are listed by name"),
]
steps_df = pd.DataFrame(
    [{"Step": s, "Name": n, "What it does and how": w, "Code": d, "Test it": t, "Proven": p}
     for (s, n, w, d, _, t, p) in STEPS]
)
ws_steps = wb.table(
    "How it works",
    steps_df,
    finding="Four small steps: acquire, parse, detect, rank — all plain code, no LLM anywhere",
    method="Sweeps jobs that SUCCEEDED (the failure debugger is AUDI-1191). Code links to the exact source "
           "line on GitHub. Test it = the command that exercises just that step.",
    kind="headline",
    widths={"Step": 6, "Name": 22, "What it does and how": 64, "Code": 24, "Test it": 44, "Proven": 26},
    toc="The step map — every chunk, how it's done, the code link, and how to test it",
)
for i, (_, _, _, disp, repo_path, _, _) in enumerate(STEPS, 1):  # data rows start at 5
    c = ws_steps.cell(row=4 + i, column=list(steps_df.columns).index("Code") + 1)
    c.hyperlink = Hyperlink(ref=c.coordinate, target=GH + repo_path, display=disp)
    c.font = Font(name=c.font.name, size=10, color=BRAND["LINK"], underline="single")

# ---------------------------------------------------------------- 2. Real findings
finds = pd.DataFrame([
    {"Job": "site_network_hourly", "Engine": "Dataproc", "Fix type": "code",
     "Finding (real numbers)": "Stage 9 waits 44-73% of task time on shuffle fetch, every run, 3+ weeks. 8,663 DCU-h across 17 runs in a day.",
     "Fix": "Map-side spread: the feeding stage always starts at initialExecutors=50 and lands 90% of output on 48-105 executors.",
     "Where it stands": "Verified 4 logs; one-hour experiment drafted for Ryan"},
    {"Job": "fangorn_score_monitor", "Engine": "Dataproc", "Fix type": "code",
     "Finding (real numbers)": "Stage 17 spilled 1,138 GiB to disk (3,924 GiB in memory at spill time).",
     "Fix": "Raise shuffle partitions first; if it persists, raise executor memory.",
     "Where it stands": "In the 2026-08-20 backlog, unrouted"},
    {"Job": "intent_score_household_map", "Engine": "Dataproc", "Fix type": "code",
     "Finding (real numbers)": "Stage 10 wide shuffle, 3,758 GiB at ~3,848 MiB per partition.",
     "Fix": "Set shuffle partitions to ~256 MiB each, or enable AQE coalesce.",
     "Where it stands": "In the 2026-08-20 backlog, unrouted"},
    {"Job": "intent_score_map", "Engine": "Dataproc", "Fix type": "infra",
     "Finding (real numbers)": "A 67-min straggler on uniform data pinned 240 executors at 32% utilization; ~$175 of a ~$260 run idle.",
     "Fix": "spark.speculation=true (quantile ~0.9). shuffleTracking.timeout was REFUTED against Spark source.",
     "Where it stands": "Delivered to Ryan 2026-08-07"},
    {"Job": "aug_log_ip_vertical_id_hourly", "Engine": "Dataproc", "Fix type": "infra",
     "Finding (real numbers)": "4-11% executor utilization, 21-52 idle executor-hours per run; Stage 11 fetch wait 31-45% on 11 of 11 runs.",
     "Fix": "Cut min/initialExecutors. Ranking buries it: 11 chronic mediums sort below one job with 2 highs (IMP-046).",
     "Where it stands": "Second in line, unrouted"},
    {"Job": "materialize_mntn_select", "Engine": "Dataproc (PHS)", "Fix type": "code",
     "Finding (real numbers)": "Stage 6 waits 40-78% on shuffle fetch. One of 21 findings on jobs the archive sweep had never seen.",
     "Fix": "Same map-side check as site_network_hourly before touching partition counts.",
     "Where it stands": "Found under a 1h PAM grant; needs mntn-devops#4724"},
    {"Job": "Update Vertical Categorization", "Engine": "Dataproc", "Fix type": "code",
     "Finding (real numbers)": "Chronic Stage-0 skew up to 242x max-vs-median task, with GC pressure alongside.",
     "Fix": "Salt the skewed key or enable AQE skew join; a plain repartition will not fix a value-skewed key.",
     "Where it stands": "Owner Ryan/targeting; manual-only DAG, re-profile first"},
    {"Job": "product_categorization + product_uniques", "Engine": "Databricks", "Fix type": "code",
     "Finding (real numbers)": "EXPLAIN COST reports missing = both tables, so the CBO estimates sizeInBytes=2.27E+22 and picks the join blind.",
     "Fix": "ANALYZE TABLE COMPUTE STATISTICS FOR ALL COLUMNS, then re-check the plan.",
     "Where it stands": "First live Databricks pull, 2026-08-20"},
])
wb.table(
    "Real findings",
    finds,
    finding="Every one of these came off a job that ran GREEN, on a fleet nobody was reviewing",
    method="From the 2026-08-20 daily crawl (214 jobs, 278 findings, 197 high), the PHS batches fetched under a 1h PAM grant, and the first live Databricks EXPLAIN COST. Every number is from a real run.",
    kind="headline",
    widths={"Job": 32, "Engine": 15, "Fix type": 10, "Finding (real numbers)": 54, "Fix": 46, "Where it stands": 30},
    toc="What it has already found on real prod jobs",
)

# ---------------------------------------------------------------- 3. Detector catalog
dets = pd.DataFrame([
    {"Detector": "skew", "Reads": "event log (run)", "Flags": "One task far slower AND reading far more data than the median (60s floor; data cross-check separates true skew from stragglers).", "Typical fix": "Salt the skewed key / AQE skew join."},
    {"Detector": "straggler", "Reads": "event log (run)", "Flags": "One task far slower on the SAME data as its peers - a slow node or IO stall, not skew.", "Typical fix": "spark.speculation=true (quantile ~0.9)."},
    {"Detector": "shuffle fetch-wait", "Reads": "event log (run)", "Flags": "Tasks spending 30%+ of their time waiting on shuffle fetch instead of computing.", "Typical fix": "Check how many executors hold the map output FIRST. More partitions makes it worse when blocks are already small."},
    {"Detector": "idle reserved executors", "Reads": "event log (run)", "Flags": "A fleet held (billed) at low utilization, incl. zero-task allocations.", "Typical fix": "Fix the tail (speculation/skew); cut min executors."},
    {"Detector": "disk spill", "Reads": "event log (run)", "Flags": "Stages writing spill to disk (reported separately from in-memory-at-spill size - summing double-counts).", "Typical fix": "More partitions, more memory, or cache less."},
    {"Detector": "GC pressure", "Reads": "event log (run)", "Flags": "High share of task time spent in garbage collection (memory-starved).", "Typical fix": "Raise executor memory / fewer, larger partitions."},
    {"Detector": "spot preemption cost", "Reads": "event log (run)", "Flags": "Task re-runs on preempted spot executors (normal serverless scale-downs excluded).", "Typical fix": "first_on_demand / on-demand fallback."},
    {"Detector": "shuffle fetch instability", "Reads": "event log (run)", "Flags": "FetchFailed tasks (executors losing shuffle blocks).", "Typical fix": "Reduce shuffle block size; steadier nodes."},
    {"Detector": "missing statistics", "Reads": "plan text", "Flags": "Tables scanned without stats, so the optimizer guesses sizes and picks bad joins/sorts.", "Typical fix": "ANALYZE TABLE COMPUTE STATISTICS."},
    {"Detector": "shuffle partition sizing", "Reads": "plan text", "Flags": "Very large shuffles at a default/low partition count (huge partitions).", "Typical fix": "Set shuffle partitions (~256 MiB each) / AQE coalesce."},
    {"Detector": "broadcast candidate", "Reads": "plan text", "Flags": "A small table joined the expensive way when it could be broadcast.", "Typical fix": "Broadcast hint / raise the broadcast threshold."},
    {"Detector": "window full-sort", "Reads": "plan text", "Flags": "A window function forcing a full sort of the data.", "Typical fix": "Partition the window; pre-bucket the data."},
    {"Detector": "repeated scan", "Reads": "plan text", "Flags": "The same source read multiple times in one job.", "Typical fix": "Cache the reused frame."},
    {"Detector": "cache ineffective", "Reads": "event log (run)", "Flags": "Cached blocks evicted under memory pressure, so the data is recomputed on the next read.", "Typical fix": "More storage memory, MEMORY_AND_DISK, or cache a narrower projection."},
])
wb.table(
    "Detector catalog",
    dets,
    finding="Fourteen detectors: nine read how the job RAN, five read what the plan INTENDED",
    method="Hand-maintained list matching airflow_optimizer/optimizations.py. Every finding carries measured numbers and a concrete fix. The fetch-wait fix text was corrected 2026-08-20 (see Read me).",
    kind="data",
    widths={"Detector": 22, "Reads": 16, "Flags": 52, "Typical fix": 40},
    toc="What each detector looks for and the fix it recommends",
)

# ---------------------------------------------------------------- 4. Worked example: Dataproc
wb.notes(
    "Ex — site_network_hourly",
    intro=f"The full walkthrough, verified 2026-08-20 on four real event logs. Line-by-line version with the raw JSON and the code: {ART}",
    blocks=[
        ("The symptom", "Stage 9 spends 44-73% of task time waiting on shuffle fetch, on every hourly run, for at least three weeks. The job has never failed, so nothing ever paged. It is the fleet's largest measured consumer: 8,663 DCU-h across 17 runs in one day."),
        ("The evidence, from one task", "Task 23201: 28.3s alive, 1.63s of CPU (5.8%), 51ms GC, 16.3s blocked fetching 14,292 remote blocks averaging 1,400 bytes. Reported Success. Note Executor CPU Time is NANOseconds while Run Time is milliseconds."),
        ("Refuted #1 - the tool's own fix", "The detector said raise spark.sql.shuffle.partitions. Stages 29 and 35 in the SAME app fetch 23.4M blocks of the same size and wait 1 second, against Stage 9's 4.2M and 73%. Block count is not the cause, and raising partitions multiplies it. Fix text corrected."),
        ("Refuted #2 - the obvious source read", "site_network_hourly.py coalesces current_partitions // 33 after shuffle.partitions=5000, predicting ~151 reducers. The log shows 366, 128 and 622 across runs. Plausible mechanism, wrong stage."),
        ("What survived", "Map-side spread. Blocks are served by the executor that WROTE them. The map stage feeding Stage 9 starts with exactly 50 executors (initialExecutors) and lands 90% of its output on 48-105, hottest holding 24.6%. The later map stages start with 306-500, spread over ~480, hottest 0.3%, and wait 0%."),
        ("What does not fit, and the ask", "Stage 15 reads the same concentrated output and waits ~0% (likely cold vs warm first read; the log cannot settle it). So the ask to the owner is a one-hour initialExecutors experiment, not a config prescription, and the DCU attributable to the stall is stated as not established."),
    ],
)

# ---------------------------------------------------------------- 5. Worked example: Databricks
wb.notes(
    "Ex — Databricks",
    intro="The same detectors on Databricks, on a real job (targeted_signal in keyword_ddp_reporting). Databricks runs ~66 dbt models plus a handful of PySpark jobs.",
    blocks=[
        ("Input", "A Databricks EXPLAIN COST plan, pulled live from a SQL warehouse through the Statement Execution API. No GCS event log needed. jobs get-run-output carries no plan, so it is not the route."),
        ("Parse", "The plan gives per-node operators and optimizer statistics; the metrics give stage shuffle sizes, failed tasks, and executor removals."),
        ("CODE fixes (actual output)",
         "Missing table stats on product_categorization (13.5B rows scanned) so the optimizer defaults to full sorts, fix ANALYZE TABLE COMPUTE STATISTICS. Wide shuffles 768/72/182 GiB at the default partition count, fix set spark.sql.shuffle.partitions (~256 MiB each) or enable AQE coalesce."),
        ("INFRA / FAILURE fixes (actual output)",
         "161 task re-runs from 7 spot-reclaimed executors, fix raise first_on_demand or add on-demand fallback. 168 FetchFailed tasks (shuffle instability), route as infra and reduce shuffle block size."),
        ("Grouping", "Each finding is tagged CODE (a query/config PR), INFRA (a cluster change), or FAILURE (route it) so the owner knows the kind of fix."),
        ("Outcome", "Same detectors as Dataproc, different acquisition (the plan plus Spark metrics instead of the GCS event log). This is what proves it works on both engines."),
    ],
)

# ---------------------------------------------------------------- 5b. Ledger + digest
wb.notes(
    "Ledger + digest",
    intro="How a daily sweep says what CHANGED, and what it admits it cannot see.",
    blocks=[
        ("Why a ledger", "A backlog rebuilt from scratch each day cannot answer the three questions that decide whether anyone acts: how long has this been true, is it new, and did the fix work. Every finding is appended to optimization_ledger.jsonl keyed on job + detector + stage."),
        ("The states", "new (first sweep) - chronic (3+ consecutive sweeps) - owner_notified (a person sent the ask) - wont_fix (owner declined, with a reason) - resolved (stopped firing for 3 sweeps). The two human decisions are sticky; everything else is recomputed from what the detectors saw."),
        ("One entry per job per sweep", "An hourly job contributes ~24 logs a day. Counting each as a finding makes every job look like a crisis, so entries collapse to one per job+detector+stage, keeping the highest severity seen."),
        ("Identity is evidence-based", "Spark app names carry per-run stamps (materialize_mntn_select_16, segment-updates-to-parquet-2026-08-20-[19]). Stamps are stripped, but a trailing _<n> is only removed when the base is a DAG the coverage pass actually saw, so ipdsc_ds_67 and ipdsc_ds_13 stay distinct."),
        ("Coverage, stated not implied", "The optimizer reads Spark event logs, so BigQuery operators, sensors and Python callables produce nothing. Live 2026-08-20: 62 active DAGs, 24 with a Spark task, 38 invisible - all 38 listed by name, plus the Databricks tasks whose plans are not reachable yet."),
        ("Delivery", "The digest is written to outputs/optimizer_digest_<date>.md and links each job to its Airflow page, but only when the name matches a DAG coverage saw - a dead link costs the reader trust. MNTN retired local Slack apps in June, so posting runs through an approved server-side path, never a credential on a laptop."),
    ],
)

# ---------------------------------------------------------------- 6. Status + next
wb.notes(
    "Status + next",
    intro="Where the build stands (2026-08-20) and what remains.",
    blocks=[
        ("Running today", "Daily at 11:00 PT over a full day of logs from both sources, with a ledger, a coverage report and a digest. 40 tests, ruff clean. First daily run: 214 jobs, 278 findings, 197 high."),
        ("1. Standing read grant", "mntn-devops#4724 is out of draft with @SteelHouse/devops and csz-mntn (Cristina Szumilo) requested. It sat 13 days as a draft with zero reviewers; nothing technical was blocking it. On merge the PHS half starts producing with no code change."),
        ("2. site_network_hourly", "Verified and drafted for Ryan Kleck (JobTeamConfig.TPA_EXPORT to Team.TARGETING). The ask is a one-hour initialExecutors experiment; re-profiling that run is what turns a 44-73% stall into a DCU number."),
        ("3. Databricks live pull", "EXPLAIN COST runs live against a SQL warehouse and the detectors fire on real production plans. Not yet wired into the sweep, and there is no bridge from an enumerated run to the query it ran."),
        ("4. Plan-shuffle detectors", "Three of the five plan detectors read shuffle sizes in a Spark-UI rendering that neither Dataproc nor Photon EXPLAIN COST emits, so only the stats detector fires today (IMP-033)."),
        ("5. Off the laptop", "The sweep runs on a Mac against expiring SSO: the Databricks OAuth refresh and the astro token both die unattended. Moving it to a runner with workload identity is the next infrastructure step, not minting service-account keys."),
    ],
)

# ---------------------------------------------------------------- 7. Read me / glossary
wb.glossary(
    "Read me",
    intro="What this tool is, and the terms used across the tabs.",
    rows=[
        ("What it is", "An efficiency sweep for Airflow/Spark jobs that SUCCEEDED. It reads each finished job's Spark data, finds waste (skew, spill, GC pressure, spot churn, missing stats), and ranks a fleet backlog worst-first."),
        ("Why it exists", "Green jobs still burn money. Nobody has time to open every job's Spark UI looking for waste; this checks every job automatically and points at the worst first."),
        ("The other half", "The failure DEBUGGER (fires on failed tasks, returns a root-cause report) is a separate tool and workbook: AUDI-1191, in the same Drive folder set."),
        ("", ""),
        ("Event log", "A file Spark writes as a job runs, recording 7 layers: jobs, stages, tasks, executors, environment, SQL node metrics, storage. The optimizer reads all 7."),
        ("Skew", "One partition holds far more data than the others, so one task runs much longer than the rest. Measured as max-vs-median task time per stage."),
        ("Spill", "A stage ran out of memory and wrote shuffle data to disk, which is much slower."),
        ("AQE", "Adaptive Query Execution: Spark features that fix skew and partition sizing at runtime when enabled."),
        ("EXPLAIN COST", "A Databricks plan dump that includes the optimizer's size estimates; how the tool reads Databricks jobs without a GCS event log."),
        ("CODE / INFRA / FAILURE", "Each finding's fix type: a code or config change, a cluster change, or something to route to the owning team."),
        ("Code links", "The Code column on 'How it works' links to the exact source line on GitHub (mdunn-mntn/malachi-workspace). Repo access required; the file:line shown still works as the reference."),
    ],
)

# ---------------------------------------------------------------- Cover (LAST)
wb.cover(
    takeaways=[
        "Reads every Spark job that ran GREEN for hidden waste. Daily, key-free, 3 CPU-minutes: 214 jobs and 278 findings, against 37 and 59 weekly.",
        "The verify pass refuted the tool's OWN recommendation on the fleet's biggest job before it reached the owner. Third time it has caught one.",
        "Says what it cannot see: 62 active DAGs, 24 with a Spark task, 38 named as invisible rather than silently omitted.",
    ]
)

scratch = os.environ.get("CLAUDE_SCRATCH", "/private/tmp/claude-501/-Users-malachi-Developer-work-mntn-workspace/3c4f6695-7891-4554-8d46-623110bfd018/scratchpad")
local = wb.save_local(os.path.join(scratch, "audi_1194_how_it_works.xlsx"))
print("saved local :", local)
try:
    drive = wb.save_drive("AUDI-1194", "Optimizer How It Works")
    print("saved drive :", drive)
except Exception as e:
    print("drive save skipped:", e)

from openpyxl import load_workbook  # noqa: E402
print("tabs        :", load_workbook(local).sheetnames)
