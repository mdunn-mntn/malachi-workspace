#!/usr/bin/env python3
"""TI-1027 — build the 5x5 / data-provider evaluation workbook (.xlsx). Multi-tab:
  1. Summary           — question, bottom line, key numbers
  2. Provider Scorecard — all MM site-visit DDPs rated (value/uniqueness/quality + cost) w/ verdict
  3. 5x5 Deep-Dive      — scale, leverage, uniqueness, IP overlap
  4. Vertical Impact    — verticals most dependent on 5x5-unique domains
  5. Cost Landscape     — all data partners + billing structure
  6. Methodology        — sources, windows, caveats
Data from outputs/*.csv (reproducible)."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE.parent / "outputs"
XLSX = HERE.parent / "artifacts" / "ti_1027_5x5_data_provider_scorecard.xlsx"

NAVY = "1F3864"; RED = "C00000"; GREEN = "375623"; GREY = "808080"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

def autowidth(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def verdict_fill(val):
    v = str(val).upper()
    if "DROP" in v: return PatternFill("solid", fgColor="F4CCCC")
    if "REVIEW" in v: return PatternFill("solid", fgColor="FFF2CC")
    if "KEEP" in v: return PatternFill("solid", fgColor="D9EAD3")
    return None

def load(name):
    with (OUT / name).open() as f:
        return list(csv.DictReader(f))

def main():
    wb = Workbook()

    # ---- Tab 1: Summary ----
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "5x5 (DS 25) Data Evaluation & Provider Scorecard (TI-1027)"; ws["A1"].font = TITLE_FONT
    ws["A2"] = "Audience Intelligence (AUDI) · estimation exercise · 7-day window 2026-06-09→15 (scale: 2026-06-15)"
    ws["A2"].font = Font(italic=True, color=GREY)
    rows = [
        ["", ""],
        ["BOTTOM LINE", "Recommend KEEP (renew) 5x5. Outsized ~3.4x its data scale, #2 most-unique data partner, "
         "B2B-concentrated, flat fee (cost does not scale). Final sign-off pending the flat-fee amount from billing."],
        ["", ""],
        ["Question", "Answer"],
        ["What is 5x5?", "A flat-fee data partner sending IP->website-visit records that feed MNTN Matched "
         "(domain->vertical classification). One of ~8 external + 2 internal sources in site_visit_signal."],
        ["Scale (% of raw data)", "~3.6% of raw site-visit records (93M rows/day, 20.8M IPs/day, 93K domains/day)."],
        ["Outsized or in line?", "OUTSIZED ~3.4x. 68.5% of its domains are unique; 47,069 are MM-usable = ~12% of "
         "the classified-domain universe, from a 3.6%-of-data partner."],
        ["Value = domains, not reach", "73.8% of its IPs we already see via our own bidstream/pixel. It surfaces "
         "DIFFERENT sites for users we already know."],
        ["vs other partners", "#2 unique contributor (Predactiv #1, also flat-fee). The $0.50-CPM per-use vendors "
         "(33Across API/Sovrn/Cybba) are largely redundant -> the real cost-review targets."],
        ["Verticals impacted if dropped", "Overwhelmingly B2B (Hiring, Logistics, Data&Analytics, Sales&Marketing, "
         "IT&Engineering) + premium retail — i.e., advertisers running B2B-audience campaigns. These are our "
         "customers' targeting verticals, NOT MNTN's own mid-market-B2B acquisition target."],
        ["Domain-only complaint?", "CONFIRMED (3.8% of URLs carry a path vs 67-100% elsewhere) but MOOT for MM — the "
         "vertical classifier strips every URL to domain anyway."],
        ["Cost structure", "Flat fee (fixed; marginal cost $0). Peer MM-DDP rate = $0.50 CPM. Absolute $ pending billing."],
        ["Is it worth it?", "MM is worth tens of $M/yr; 5x5 supplies ~12% of its domain signal (B2B-weighted higher) "
         "-> clears a typical DDP flat fee with margin. Keep unless the fee is unusually large (then renegotiate)."],
    ]
    r = 4
    for a, b in rows:
        ws.cell(r, 1, a); ws.cell(r, 2, b)
        ws.cell(r, 1).alignment = WRAP; ws.cell(r, 2).alignment = WRAP
        if a in ("BOTTOM LINE",):
            ws.cell(r, 1).font = Font(bold=True, size=12, color=RED); ws.cell(r, 2).font = Font(bold=True, color=NAVY)
        elif a == "Question":
            ws.cell(r, 1).font = HDR_FONT; ws.cell(r, 2).font = HDR_FONT
            ws.cell(r, 1).fill = HDR_FILL; ws.cell(r, 2).fill = HDR_FILL
        elif a:
            ws.cell(r, 1).font = Font(bold=True, color=NAVY)
        r += 1
    autowidth(ws, [26, 96])

    # ---- Tab 2: Provider Scorecard ----
    ws2 = wb.create_sheet("Provider Scorecard")
    sc = load("ti_1027_vendor_scorecard.csv")
    hdr = ["Rank", "Provider", "DS", "Cost", "CPM", "Rows/day", "IPs/day", "% w/ path",
           "Domains (7d)", "Class. rate %", "Unique %", "Unique MM domains", "Score", "Verdict"]
    ws2.append(hdr); style_header(ws2, 1, len(hdr))
    for i, row in enumerate(sc, 1):
        cost = "internal $0" if row["internal"] == "True" else (row["billing"] or "")
        ws2.append([i, row["partner"], int(row["ds"]), cost,
                    (("$"+row["cpm"]) if row["cpm"] else ""),
                    int(row["rows_day"]), int(row["ips_day"]), float(row["pct_path"]),
                    int(row["total_domains"]), float(row["class_rate"]), float(row["pct_unique"]),
                    int(row["unique_classified"]), float(row["score"]), row["verdict"]])
        rr = ws2.max_row
        for c in (6, 7, 9, 12):
            ws2.cell(rr, c).number_format = "#,##0"
        ws2.cell(rr, 14).alignment = WRAP
        f = verdict_fill(row["verdict"])
        if f: ws2.cell(rr, 14).fill = f
        if row["ds"] == "25":
            ws2.cell(rr, 2).font = Font(bold=True, color=RED)
    autowidth(ws2, [5, 14, 5, 12, 6, 12, 11, 9, 12, 12, 9, 16, 7, 46])
    ws2.freeze_panes = "A2"

    # ---- Tab 3: 5x5 Deep-Dive ----
    ws3 = wb.create_sheet("5x5 Deep-Dive")
    ws3["A1"] = "5x5 (DS 25) — the measurable read"; ws3["A1"].font = TITLE_FONT
    ws3.append([]); ws3.append(["Metric", "Value", "Read"]); style_header(ws3, 3, 3)
    dd = [
        ["Share of raw site-visit records", "3.6%", "~93M rows/day. The leverage denominator."],
        ["Distinct IPs / day", "20.8M", "Mid-pack scale."],
        ["Distinct domains / day", "93K", "Substantial domain breadth."],
        ["% URLs with a page path", "3.8%", "Domain-only feed (vs 67-100% for other partners). Moot for MM."],
        ["Domains unique to 5x5 (7d)", "68.5%", "138,496 of 202,299 — provided by no other partner."],
        ["Unique AND MM-classifiable", "47,069", "~12% of the classified-domain universe. The net MM contribution."],
        ["Leverage ratio", "~3.4x", "12% unique MM signal from 3.6% of raw data -> OUTSIZED."],
        ["IPs already seen internally", "73.8%", "Value is incremental DOMAINS, not reach."],
        ["IPs unique to 5x5", "19.8%", "4.1M/day — modest incremental reach."],
        ["Concentration", "B2B", "20-34% of B2B verticals' fresh domain coverage is 5x5-unique."],
    ]
    for a, b, c in dd:
        ws3.append([a, b, c]); ws3.cell(ws3.max_row, 3).alignment = WRAP
        ws3.cell(ws3.max_row, 1).font = Font(bold=True, color=NAVY)
    autowidth(ws3, [32, 14, 80])

    # ---- Tab 4: Vertical Impact ----
    ws4 = wb.create_sheet("Vertical Impact")
    vt = load("ti_1027_vertical_dependence_7d.csv")
    ws4["A1"] = "Verticals most dependent on 5x5-unique domains (lost if 5x5 dropped)"; ws4["A1"].font = TITLE_FONT
    ws4.append([]); ws4.append(["Bucket", "Vertical", "Classified domains", "5x5-unique", "% dependent on 5x5"])
    style_header(ws4, 3, 5)
    for row in vt:
        ws4.append([int(row["bucket_id"]), row["vertical_name"], int(row["classified_domains"]),
                    int(row["d_5x5_unique"]), float(row["pct_dependent_on_5x5"])])
        rr = ws4.max_row
        ws4.cell(rr, 3).number_format = "#,##0"; ws4.cell(rr, 4).number_format = "#,##0"
        ws4.cell(rr, 5).number_format = "0.0"
        if row["vertical_name"].startswith("B2B"):
            ws4.cell(rr, 2).font = Font(bold=True, color=RED)
    autowidth(ws4, [8, 44, 18, 12, 18]); ws4.freeze_panes = "A4"

    # ---- Tab 5: Cost Landscape ----
    ws5 = wb.create_sheet("Cost Landscape")
    ws5["A1"] = "Data partner cost landscape (tpa.direct_data_partners, is_current=true)"; ws5["A1"].font = TITLE_FONT
    ws5.append([]); ws5.append(["Group", "Partner", "DS", "Billing", "CPM", "Feeds", "Note"])
    style_header(ws5, 3, 7)
    land = [
        ["MM site-visit", "Predactiv", 26, "flat_fee", "", "MNTN Matched", "Best value (#1 unique)"],
        ["MM site-visit", "5x5", 25, "flat_fee", "", "MNTN Matched", "KEEP (#2 unique, B2B)"],
        ["MM site-visit", "Justuno", 24, "fixed_cpm", "$0.50", "MNTN Matched", "Efficient (small, 84% unique)"],
        ["MM site-visit", "33Across", 28, "fixed_cpm", "$0.50", "MNTN Matched", "REVIEW (high CPM volume, 30% unique)"],
        ["MM site-visit", "33Across API", 40, "fixed_cpm", "$0.50", "MNTN Matched", "DROP-CANDIDATE (3% unique)"],
        ["MM site-visit", "Sovrn", 33, "fixed_cpm", "$0.50", "MNTN Matched", "DROP-CANDIDATE (2% unique)"],
        ["MM site-visit", "Cybba", 36, "fixed_cpm", "$0.50", "MNTN Matched", "REVIEW (6% unique, low volume)"],
        ["MM site-visit", "Klickly", 39, "flat_fee", "", "MNTN Matched", "REVIEW (negligible)"],
        ["MM site-visit", "LaunchLabs", 27, "fixed_cpm", "$0.50", "MNTN Matched", "DISABLED"],
        ["Internal", "augmentor_log", 30, "internal", "$0", "MNTN Matched", "Our own bidstream"],
        ["Internal", "guid_log", 23, "internal", "$0", "MNTN Matched", "Our own pixel"],
        ["Interest 3P", "LiveRamp", "11/35", "variable_cpm", "", "Interests", "Dominant 3P spend. Rated by TI-956/999."],
        ["Interest 3P", "ShareThis", 17, "fixed_cpm", "$0.95", "Interests", "Was $1.20. Rated by TI-956/999."],
        ["Interest 3P", "Dstillery", 18, "(unset)", "", "Interests", "Rated by TI-956/999."],
        ["Interest 3P", "OnAudience", 20, "(unset)", "", "Interests", "Dormant."],
        ["CRM", "Experian", 22, "flat_fee", "", "CRM", "Not scored data."],
        ["CRM", "deepsync", 29, "fixed_cpm", "$0.50", "CRM", "Not scored data."],
    ]
    for row in land:
        ws5.append(row); ws5.cell(ws5.max_row, 7).alignment = WRAP
        f = verdict_fill(row[6])
        if f: ws5.cell(ws5.max_row, 7).fill = f
    autowidth(ws5, [14, 14, 8, 14, 7, 14, 44]); ws5.freeze_panes = "A4"

    # ---- Tab: Score Tiers ----
    wst = wb.create_sheet("Score Tiers")
    st = load("ti_1027_vendor_score_tiers_7d.csv")
    st.sort(key=lambda r: int(r["hi_10000"]) / int(r["delivered_ips"]), reverse=True)
    wst["A1"] = "Score-tier mix of each vendor's IPs (scored ≠ high-value)"; wst["A1"].font = TITLE_FONT
    wst["A2"] = ("Of each vendor's IPs that MNTN served an impression to (cost_impression_log, 7d), the household-score "
                 "tier mix. The score is a household property — ~uniform across vendors. % = of delivered IPs.")
    wst["A2"].font = Font(italic=True, color=GREY); wst["A2"].alignment = WRAP
    wst.append([]); wst.append(["Vendor", "Vendor IPs", "% delivered", "HI 10000 %", "PP 8000 %",
                                "High grad %", "Mid %", "Max Reach %", "Unscored %", "% high (>=6666)"])
    style_header(wst, 4, 10)
    for row in st:
        d = int(row["delivered_ips"]); pc = lambda k: round(100 * int(row[k]) / d, 1)
        wst.append([row["partner"], int(row["vendor_ips"]), float(row["pct_delivered"]),
                    pc("hi_10000"), pc("pp_8000"), pc("high_grad"), pc("mid"), pc("maxreach"),
                    pc("unscored_delivered"), float(row["pct_delivered_high"])])
        rr = wst.max_row
        wst.cell(rr, 2).number_format = "#,##0"
        if row["data_source_id"] == "25":
            wst.cell(rr, 1).font = Font(bold=True, color=RED)
    wst.append([]); wst.append(["Read: 5x5's IPs are as high-value as any vendor's — 39.4% land in top-tier High "
                                "Intent (highest among the high-volume sources). No vendor brings low-value households; "
                                "the differentiation is unique DOMAINS, not IP quality. (Delivered scores; the full "
                                "all-IP scored universe is 19.4 TB/day and out of scope.)"])
    wst.cell(wst.max_row, 1).alignment = WRAP
    autowidth(wst, [16, 13, 12, 11, 11, 12, 8, 13, 11, 14]); wst.freeze_panes = "A5"

    # ---- Tab: Data Inventory (richness + volume) ----
    wdi = wb.create_sheet("Data Inventory")
    wdi["A1"] = "What's in the data — raw richness & volume (TI-1027 Phase 2)"; wdi["A1"].font = TITLE_FONT
    rich = load("ti_1027_vendor_richness.csv"); card = {r["data_source_id"]: r for r in load("ti_1027_cardinality_2026-06-15.csv")}
    wdi.append([]); wdi.append(["DS", "Partner", "Raw columns", "Metadata beyond ip/url/time", "Profile",
                                "Schema risk", "Events/day", "(IP×domain) pairs/day"])
    style_header(wdi, 3, 8)
    for r in rich:
        c = card.get(r["data_source_id"], {})
        wdi.append([int(r["data_source_id"]), r["partner"], r["raw_columns"], r["other_metadata"], r["profile"],
                    r["schema_risk"], int(c.get("events", 0)), int(c.get("ip_domain_pairs", 0))])
        rr = wdi.max_row; wdi.cell(rr, 3).alignment = WRAP; wdi.cell(rr, 4).alignment = WRAP
        wdi.cell(rr, 7).number_format = "#,##0"; wdi.cell(rr, 8).number_format = "#,##0"
        if r["data_source_id"] == "25": wdi.cell(rr, 2).font = Font(bold=True, color=RED)
    wdi.append([]); wdi.append(["Discard finding: pixel vendors (24/33/39/40) send event_id/mobile/query_str/referer "
                                "and Predactiv sends user_agent — all dropped at site_visit_signal. We pay for rich, keep thin."])
    wdi.cell(wdi.max_row, 1).alignment = WRAP; wdi.cell(wdi.max_row, 1).font = Font(italic=True, color=GREY)
    autowidth(wdi, [5, 14, 38, 34, 18, 24, 14, 18]); wdi.freeze_panes = "A4"

    # ---- Tab: Per-IP Depth (raw numbers) ----
    wdd = wb.create_sheet("Per-IP Depth")
    wdd["A1"] = "Raw numbers + per-IP depth — volume ≠ value"; wdd["A1"].font = TITLE_FONT
    wdd["A2"] = ("A vendor that sees one IP visit 10 sites beats one that sees it visit 1. 'unique domains/IP' = "
                 "distinct sites the vendor ALONE contributes per household. 2026-06-15.")
    wdd["A2"].font = Font(italic=True, color=GREY); wdd["A2"].alignment = WRAP
    depth = sorted(load("ti_1027_per_ip_depth.csv"), key=lambda r: -int(r["events"]))
    wdd.append([]); wdd.append(["Vendor", "Events/day", "IPs", "Domains", "IP×domain pairs", "IP×url pairs",
                                "visits/IP", "domains/IP", "unique dom/IP", "% pairs unique"])
    style_header(wdd, 4, 10)
    for r in depth:
        wdd.append([r["partner"], int(r["events"]), int(r["ips"]), int(r["domains"]), int(r["ip_domain_pairs"]),
                    int(r["ip_url_pairs"]), float(r["visits_per_ip"]), float(r["domains_per_ip"]),
                    float(r["unique_domains_per_ip"]), float(r["pct_pairs_unique"])])
        rr = wdd.max_row
        for cc in (2, 3, 4, 5, 6): wdd.cell(rr, cc).number_format = "#,##0"
        if r["data_source_id"] == "25": wdd.cell(rr, 1).font = Font(bold=True, color=RED)
    wdd.append([]); wdd.append(["Read: 33Across is the biggest feed (834M events) but shallowest in unique depth "
        "(0.65 unique dom/IP, 27% unique pairs) — repeat-visits to common domains. Two value lenses: domain→vertical "
        "breadth (MM uses this → 5x5 wins on unique domains) vs per-IP depth (augmentor/33Across-API lead)."])
    wdd.cell(wdd.max_row, 1).alignment = WRAP; wdd.cell(wdd.max_row, 1).font = Font(italic=True, color=GREY)
    autowidth(wdd, [14, 13, 12, 10, 16, 14, 9, 10, 13, 12]); wdd.freeze_panes = "A5"

    # ---- Tab: Uniqueness Layers ----
    wul = wb.create_sheet("Uniqueness Layers")
    wul["A1"] = "5x5 uniqueness by grain — value is unique DATA, not unique reach"; wul["A1"].font = TITLE_FONT
    lay = load("ti_1027_layered_uniqueness_5x5.csv")
    wul.append([]); wul.append(["Grain", "5x5 total", "Unique to 5x5", "% unique", "Also seen internally", "% internal"])
    style_header(wul, 3, 6)
    gmap = {"ip": "IP (reach)", "domain": "Domain", "ip_domain_pair": "(IP×domain) event — the data value"}
    for r in lay:
        wul.append([gmap.get(r["grain"], r["grain"]), int(r["total_5x5"]), int(r["unique_to_5x5"]),
                    float(r["pct_unique"]), int(r["also_internal"]), float(r["pct_internal"])])
        rr = wul.max_row
        for cc in (2, 3, 5): wul.cell(rr, cc).number_format = "#,##0"
        if r["grain"] == "ip_domain_pair":
            for cc in range(1, 7): wul.cell(rr, cc).font = Font(bold=True, color=RED)
    wul.append([]); wul.append(["Read: 5x5 mostly sees households we already know (20% unique IPs), but the specific "
                                "site-visits are 77% 5x5-only. The unique data value >> the unique reach. No unique metadata."])
    wul.cell(wul.max_row, 1).alignment = WRAP; wul.cell(wul.max_row, 1).font = Font(italic=True, color=GREY)
    autowidth(wul, [34, 14, 16, 11, 18, 11]); wul.freeze_panes = "A4"

    # ---- Tab: Willingness-to-Pay ----
    wp = wb.create_sheet("Willingness-to-Pay")
    wp["A1"] = "What 5x5 is worth per year — and what to pay"; wp["A1"].font = TITLE_FONT
    wp["A2"] = ("Billing base = $0.50 CPM = per 1,000 impressions served (cost in cost_impression_log). "
                "5x5 touches ~34.35M impr/day; impr to its UNIQUE IPs = 213.5K/day.")
    wp["A2"].font = Font(italic=True, color=GREY); wp["A2"].alignment = WRAP
    wp.append([]); wp.append(["Anchor", "Value", "Basis"]); style_header(wp, 4, 3)
    for a, b, c, fill in [
        ("Floor", "~$40K / yr", "Incremental reach only: 77.9M impr/yr to 5x5-unique IPs × $0.50 CPM", "D9EAD3"),
        ("Fair price", "~$150K–$600K / yr", "5x5 = ~12% of MM unique classified-domain signal, B2B-weighted; typical DDP flat-fee range", "D9EAD3"),
        ("Walk-away max", "~$6.3M / yr", "CPM-equivalent of all 12.5B/yr touched impressions (upper bound — co-occurrence, not causal)", "F4CCCC")]:
        wp.append([a, b, c]); rr = wp.max_row; wp.cell(rr, 3).alignment = WRAP
        wp.cell(rr, 1).font = Font(bold=True); wp.cell(rr, 2).fill = PatternFill("solid", fgColor=fill)
    wp.append([]); wp.append(["Per-unit rate", "Value", "Note"]); style_header(wp, wp.max_row, 3)
    for a, b, c in [
        ("Net-new IP", "~$0.01–0.50 / IP / yr", "Low — reach is NOT where 5x5's value is; don't pay much per net-new IP"),
        ("Net-new (IP×domain) event", "~$0.03 / 1,000 events", "The real asset — 9.3B unique events/yr"),
        ("Net-new classified domain", "~$3–13 / domain / yr", "47K unique MM-usable domains, B2B coverage")]:
        wp.append([a, b, c]); wp.cell(wp.max_row, 3).alignment = WRAP; wp.cell(wp.max_row, 1).font = Font(bold=True)
    wp.append([]); wp.append(["Pricing — Monthly rate (recommended)", "Value", "Note"]); style_header(wp, wp.max_row, 3)
    for a, b, c in [
        ("Floor", "~$3K / mo", "we'd happily pay"),
        ("FAIR", "$15K–50K / mo", "anchor ask ~$25–30K/mo ($300–360K/yr)"),
        ("Walk-away", "~$525K / mo", "= the CPM ceiling"),
        ("Volume minimum", "≥2.5B rows/mo AND ≥25M unique (IP×domain) pairs/day", "so they can't throttle or pad with junk")]:
        wp.append([a, b, c]); wp.cell(wp.max_row, 3).alignment = WRAP; wp.cell(wp.max_row, 1).font = Font(bold=True)
    wp.append([]); wp.append(["Pricing — CPM (if per 1,000 impr)", "Value", "Note"]); style_header(wp, wp.max_row, 3)
    for a, b, c in [
        ("On MATCHED impr", "≤ $0.50 CPM", "peer parity — fair"),
        ("On ALL touched impr", "$0.02–0.05 CPM", "~95% redundant; >$0.10 = walk away"),
        ("Reconciliation", "$25K/mo ≈ $0.024 CPM (all touched) ≈ $0.50 CPM (matched)", "same dollars, three views")]:
        wp.append([a, b, c]); wp.cell(wp.max_row, 3).alignment = WRAP; wp.cell(wp.max_row, 1).font = Font(bold=True)
    wp.append([]); wp.append(["Recency strengthens the floor", "Over the 30-day TARGETING window, 69.8% of 5x5's "
        "(IP×domain) pairs are SOLE (no other vendor in-window) and 95.4% sole-or-freshest — only ~4.6% covered "
        "fresher elsewhere. 'Overlap' from a 7-day snapshot overstates redundancy."])
    wp.cell(wp.max_row, 1).font = Font(bold=True, color=RED); wp.cell(wp.max_row, 2).alignment = WRAP
    autowidth(wp, [24, 20, 74]);

    # ---- Tab 6: Methodology ----
    ws6 = wb.create_sheet("Methodology")
    meth = [
        ["Methodology & Caveats", ""],
        ["Substrate", "gs://mntn-data-archive-prod/signals/site_visit_signal/ (parquet on GCS, keyed by "
         "data_source_id). Queried via BigQuery temporary external-table definitions (read-only)."],
        ["Windows", "Scale: 1 day (2026-06-15). Uniqueness/quality/vertical: 7 days (2026-06-09..15). DDP delivery "
         "is bursty, so multi-day windows are used for the contribution metrics."],
        ["Quality measure", "A domain is 'MM-usable' if it appears in the production domain->vertical table "
         "(website_crawl_verticals, ~1.42M classified domains). 'Unique' = provided by no other data_source_id."],
        ["Cost", "billing_type + per-unit CPM from tpa.direct_data_partners. Absolute flat-fee dollars not yet "
         "available (with billing). Score = 0.55*value + 0.25*non-redundancy + 0.20*signal-quality (value-weighted)."],
        ["Value of MM", "MM touches ~$210-385M/yr of media; drives ~10-36% visit-rate lift (Fangorn). Estimated "
         "value (via retention) is tens of $M/yr. 5x5's attributable slice ~12% of domain signal, B2B-weighted higher."],
        ["Scope", "Scorecard covers MM site-visit DDPs (directly comparable). Interest-segment 3P providers "
         "(LiveRamp/ShareThis/Dstillery) are a different modality, rated by Alex's 9-axis framework (TI-956/TI-999)."],
        ["Not a causal claim", "This is an estimation exercise. A precise causal value would require an add/remove "
         "model ablation (re-run MM with vs without DS25 -> delta-IVR). Proposed as a follow-up only if needed."],
        ["Source files", "tickets/ti_1027_5x5_data_evaluation/ — summary.md, queries/, outputs/*.csv, "
         "artifacts/ (charts, this workbook, scorecard.md)."],
    ]
    for a, b in meth:
        ws6.append([a, b]); ws6.cell(ws6.max_row, 2).alignment = WRAP
        ws6.cell(ws6.max_row, 1).font = TITLE_FONT if b == "" else Font(bold=True, color=NAVY)
    autowidth(ws6, [18, 104])

    wb.save(XLSX)
    print(f"Wrote {XLSX}")

if __name__ == "__main__":
    main()
