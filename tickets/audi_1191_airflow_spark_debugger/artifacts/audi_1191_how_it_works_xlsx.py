#!/usr/bin/env python3
"""AUDI-1191 failure-debugger explainer + proof deliverable.

Builds the branded multi-sheet .xlsx that (1) explains how the automated Airflow/Spark
failure DEBUGGER works step by step (D1-D8), and (2) proves it on real incidents
(INC-005/009/010/011/012/013 + ds67). The signature-taxonomy sheet is built live from
airflow_debugger.signatures so it never drifts. The optimizer (success-sweep) half has its
own deliverable under AUDI-1194.

Regenerate: python3 tickets/audi_1191_airflow_spark_debugger/artifacts/audi_1191_how_it_works_xlsx.py
"""

import os
import sys

import pandas as pd
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))))
from airflow_debugger import signatures as sig
from lib.mntn_xlsx import BRAND, FMT, MntnWorkbook

GEN = "2026-08-07"
GH = "https://github.com/mdunn-mntn/malachi-workspace/blob/main/"

wb = MntnWorkbook(
    title="Airflow/Spark Failure Debugger",
    ticket="AUDI-1191",
    subtitle="How the failure debugger works, step by step, and proof on 7 real prod incidents (Dataproc + Databricks).",
    period="Validated Jul-Aug 2026",
    generated=GEN,
    status="Working",
)

# ---------------------------------------------------------------- 1. How it works (step map)
# One row per testable chunk. Fires on FAILURES only; the success-sweep optimizer is AUDI-1194.
STEPS = [
    # (step, name, what/how, code display, code repo-path#anchor, test command, proven)
    ("D1", "Detect failed tasks",
     "Queries the Airflow REST API on Astronomer directly (no Slack): POST /dags/~/dagRuns/~/taskInstances/list, "
     "windowed on the UTC day, state=failed. Auth is the short-lived astro SSO login. The Slack alert is only how "
     "a human notices; the tool finds the same failure itself from the API.",
     "airflow_api.py:203", ".claude/scripts/airflow_api.py#L203",
     "bash .claude/scripts/airflow_pull.sh --date <D> --state failed",
     "Agent-tested 2026-08-06 (6 failed tasks). Mid-retry tries need watch-mode, not the day-dump"),
    ("D2", "Download the log",
     "For each failed try: GET the task-instance log (structured JSON/NDJSON), render to plain text, save as "
     "<time>__<dag>__<task>__try<N>__<state>.log plus a _manifest.jsonl row (the pass/fail grid).",
     "airflow_api.py:294", ".claude/scripts/airflow_api.py#L294",
     "same command; logs land in on-call/airflow_logs/<date>/",
     "Agent-tested 2026-08-06 (3 logs + manifest verified) · INC-010/011/012 pulls"),
    ("D3", "Parse + route",
     "Reads the log for identity (dag, task, run, logical date), routes by operator classpath to the engine "
     "(Dataproc vs Databricks), and extracts the downstream Spark job id (batch id / run id).",
     "parse.py:51", "airflow_debugger/parse.py#L51",
     "python3 -m airflow_debugger.tests.test_parse",
     "Adversarial review 2026-08-06: 9 parse defects fixed; real-log sweep 64/64 identity, 33/33 job ids"),
    ("D4", "Engine RCA",
     "Dataproc: batches describe + Cloud Logging traceback + structural TTL check (dataproc_rca.py:156); "
     "when Cloud Logging has no error text it reads the staging driveroutput named in the batch stateMessage "
     "(needs the dataproc-debug PAM grant; notes the 403 and the unblock step without it). "
     "Databricks: jobs get-run + get-run-output on the TASK run id + cluster events (databricks_rca.py:74).",
     "dataproc_rca.py:156", "airflow_debugger/dataproc_rca.py#L156",
     "python3 -m airflow_debugger.tests.test_dataproc_rca",
     "INC-005/009/012. Driveroutput fallback (IMP-028) tested on the real INC-012 driver text: full gcs_list_timeout verdict"),
    ("D5", "Signature match",
     f"{len(sig.SIGNATURES)} regex fingerprints classify the failure; a high-confidence match returns a cached verdict with NO LLM "
     "call. Each signature's programmatic-fix flag separates a fixable root cause from a downstream symptom.",
     "signatures.py:28", "airflow_debugger/signatures.py#L28",
     "python3 -m airflow_debugger.tests.test_signatures",
     "27 cases + order-integrity test; INC-012 mixed driver blob classifies correctly"),
    ("D6", "Past-incident match",
     "Lexical matcher over the local incident corpus (on-call/incident_log.jsonl) attaches the most similar past "
     "incidents to the verdict, so a repeat is recognized instantly.",
     "incident_match.py:34", "airflow_debugger/incident_match.py#L34",
     "python3 -m airflow_debugger.tests.test_incident_match",
     "Agent-tested 2026-08-06: INC-012 query -> INC-012 top match (0.605)"),
    ("D7", "Report + troubleshooting pack",
     "Two outputs. (1) BLUF report under 500 characters: root cause + confidence + fix flag + deep link. "
     "(2) Troubleshooting pack (--troubleshoot): Problem, Solution with the known fix PR when a past incident "
     "carries one (fix_pr in the corpus), and Code links built from the traceback (framework frames skipped, "
     "file resolved against the airflow-ti tree, exact #L line).",
     "report.py:188", "airflow_debugger/report.py#L188",
     "python3 -m airflow_debugger.tests.test_report",
     "Live on INC-013: repeat alert returns PR #1179 + 3 fixed files + the #L30 link; 3 adversarial-review defects fixed w/ regressions (wrong-file collision, framework-frame leak, unrelated-PR claim)"),
    ("D8", "LLM fallback",
     "ONLY when no signature matched: one bounded LLM synthesis call over the distilled evidence (synth.py:29). "
     "Everything before this is plain deterministic code, so known failures cost nothing and are instant.",
     "orchestrate.py:16", "airflow_debugger/orchestrate.py#L16",
     "python3 -m airflow_debugger.orchestrate <log>  (--no-llm to force it off)",
     "Deterministic report never replaced by an LLM error stub; unknown-sig fallback verified"),
]
steps_df = pd.DataFrame(
    [{"Step": s, "Name": n, "What it does and how": w, "Code": d, "Test it": t, "Proven": p}
     for (s, n, w, d, _, t, p) in STEPS]
)
ws_steps = wb.table(
    "How it works",
    steps_df,
    finding="Every step is a small, separately testable chunk of plain code",
    method="Fires on failures only (the success-sweep optimizer is AUDI-1194). Code links to the exact source "
           "line on GitHub. Test it = the command that exercises just that step.",
    kind="headline",
    widths={"Step": 6, "Name": 22, "What it does and how": 64, "Code": 22, "Test it": 44, "Proven": 26},
    toc="The step map — every chunk, how it's done, the code link, and how to test it",
)
for i, (_, _, _, disp, repo_path, _, _) in enumerate(STEPS, 1):  # data rows start at 5 (title block 1-3, header 4)
    c = ws_steps.cell(row=4 + i, column=list(steps_df.columns).index("Code") + 1)
    c.hyperlink = Hyperlink(ref=c.coordinate, target=GH + repo_path, display=disp)
    c.font = Font(name=c.font.name, size=10, color=BRAND["LINK"], underline="single")

# ---------------------------------------------------------------- 2. Incidents proven
cases = pd.DataFrame([
    {"Incident": "INC-005", "Failure class": "Perf / TTL", "Engine": "Dataproc",
     "DAG / task": "tpa_mntn_id_export", "What it read": "batch describe + event log",
     "Tool verdict (output)": "Cancelled at 10800s TTL (perf regression). Profile the event log for spill/skew; a TTL bump alone rarely fixes it.",
     "Confirmed by": "Owner perf fix PR #1161"},
    {"Incident": "ds67", "Failure class": "Code bug", "Engine": "Dataproc",
     "DAG / task": "tpa_ipdsc_export / ipdsc_ds_67", "What it read": "batch describe + Cloud Logging traceback",
     "Tool verdict (output)": "Invalid GCS bucket name: write_location passed as a method, not called. Root cause at ipdsc_ds_67.py:73.",
     "Confirmed by": "Owner fix a008b2e, exact file/line"},
    {"Incident": "INC-009", "Failure class": "Orchestration only", "Engine": "Databricks",
     "DAG / task": "keyword_ddp_reporting / write_targeted_signal_ds_19", "What it read": "jobs get-run + get-run-output + cluster events",
     "Tool verdict (output)": "Databricks job SUCCEEDED; K8s pod evicted mid-run (orchestration-only). Not a code fix.",
     "Confirmed by": "Owner marked success; data verified in GCS"},
    {"Incident": "INC-010", "Failure class": "Late / missing partner data", "Engine": "Dataproc",
     "DAG / task": "tpa_ipdsc_export / wait_ds17_src", "What it read": "Airflow sensor log + GCS",
     "Tool verdict (output)": "Mandatory partner feed (ShareThis / DS17) missing; the existence sensor hard-timed out. Also caught: the unblock copy was previous-day data.",
     "Confirmed by": "Owner root cause matched + day-1 fallback shipped"},
    {"Incident": "INC-011", "Failure class": "Skip treated as failure", "Engine": "n/a (sensor)",
     "DAG / task": "hashed_email_ds_26_signals / wait_fpa", "What it read": "sensor log + external task state + producer short-circuit log",
     "Tool verdict (output)": "False alarm: producer SUCCEEDED and correctly skipped (no Predactiv file that hour); the sensor counts a skip as a failure.",
     "Confirmed by": "Fix PR #1175 merged (skipped_states, 2 DAGs)"},
    {"Incident": "INC-012", "Failure class": "GCS list timeout", "Engine": "Dataproc",
     "DAG / task": "materialize_mntn_select / materialize", "What it read": "batch describe + staging driveroutput",
     "Tool verdict (output)": "Both tries died listing augmentor_log: the region glob lists every file (~17M) to find one hour. 'Lost executors' was a red herring (idle scale-downs).",
     "Confirmed by": "Fix PR #1176 merged (literal paths); corrected the thread's preemption theory"},
    {"Incident": "INC-013", "Failure class": "GCS list timeout (sibling)", "Engine": "Dataproc",
     "DAG / task": "fpa_site_visit_batch_serverless / dsid30_augmentor_log_processing", "What it read": "batch describe + Cloud Logging driver traceback",
     "Tool verdict (output)": "Same listing timeout in a sibling augmentor_log reader (glob + basePath). Repo sweep found 2 more, one failing silently (mntn_global_data shipped a day with zero augmentor rows behind a green run).",
     "Confirmed by": "Fix PR #1179 merged + prod-verified same morning (6-min run vs 19-min deaths)"},
])
wb.table(
    "Incidents proven",
    cases,
    finding="Seven real prod incidents diagnosed correctly, including three where the tool corrected the first human read",
    method="Each row is a real failure the tool was run on. Tool verdict is the actual output; Confirmed by is the owner action or merged fix that proved it right.",
    kind="headline",
    widths={"Incident": 10, "Failure class": 20, "Engine": 12, "DAG / task": 34, "What it read": 30, "Tool verdict (output)": 58, "Confirmed by": 30},
    toc="The proof — one row per real incident, tool output vs ground truth",
)

# ---------------------------------------------------------------- 6. Signature taxonomy (live)
CLASS_MAP = {  # human label for the fix flag
    "yes": "Yes", "sometimes": "Sometimes", "no": "No",
}
tax = pd.DataFrame([
    {"Signature": s.key.replace("_", " "),
     "Engine": s.engine,
     "Class": s.sig_class,
     "Likely cause": s.likely_cause,
     "Auto-fix possible?": CLASS_MAP.get(s.programmatic_fix, s.programmatic_fix)}
    for s in sig.SIGNATURES
])
wb.table(
    "Signature taxonomy",
    tax,
    finding=f"{len(tax)} failure fingerprints — a match returns a cached verdict with no LLM",
    method="Built live from airflow_debugger.signatures. Engine 'any' applies to both. Auto-fix possible separates a fixable root cause (Yes) from a symptom or infra issue (No).",
    kind="data",
    widths={"Signature": 26, "Engine": 11, "Class": 22, "Likely cause": 62, "Auto-fix possible?": 16},
    toc=f"Coverage — the {len(tax)} fingerprints the deterministic classifier knows",
)

# ---------------------------------------------------------------- 7. Read me / glossary
wb.glossary(
    "Read me",
    intro="What this tool is, and the terms used across the tabs.",
    rows=[
        ("What it is", "An automated debugger for failed Airflow/Spark tasks (Dataproc and Databricks). On a failure it returns a root-cause report with the affected file and a confidence level, in under 500 characters."),
        ("Why it exists", "On-call debugging of a failed task was slow and expert-dependent, and the framework author left. This removes the single-person dependency and cuts time-to-diagnosis."),
        ("The other half", "The efficiency sweep (reads jobs that SUCCEEDED and finds waste like skew and spill) is a separate tool and workbook: AUDI-1194, in the same Drive folder set."),
        ("", ""),
        ("Deterministic-first", "The fetch, routing, signature match, and report are plain code. The LLM is used only as a fallback when no known signature matches, so most cases cost nothing and are instant."),
        ("Key-free", "No stored API tokens and no Slack bot. Data access uses short-lived SSO / CLI tokens (astro, gcloud, Databricks OAuth). Required by MNTN security policy."),
        ("BLUF / STAR", "Bottom Line Up Front. The report leads with the answer (root cause) in one line, then the supporting detail, kept under 500 characters."),
        ("RCA", "Root-cause analysis: naming the actual underlying cause, not just that the task failed."),
        ("Signature", "A regex fingerprint of a known failure. A high-confidence match returns a cached verdict with no LLM call."),
        ("Code links", "The Code column on 'How it works' links to the exact source line on GitHub (mdunn-mntn/malachi-workspace). Repo access required; without it, the file:line shown is still the reference."),
        ("Programmatic-fix flag", "Per signature: whether an automated code fix is even possible. It separates a fixable root cause from a downstream symptom or an infra issue."),
    ],
)

# ---------------------------------------------------------------- Cover (LAST)
wb.cover(
    takeaways=[
        "A failed Airflow task goes in; a root-cause report with the affected file comes out, for both Dataproc and Databricks.",
        "Proven on six real prod incidents, including two where it corrected the first human diagnosis; both fixes merged.",
        "Deterministic-first and key-free: known failures resolve instantly with no LLM call and no stored tokens.",
    ]
)

scratch = os.environ.get("CLAUDE_SCRATCH", "/private/tmp/claude-501/-Users-malachi-Developer-work-mntn-workspace/3c4f6695-7891-4554-8d46-623110bfd018/scratchpad")
local = wb.save_local(os.path.join(scratch, "audi_1191_how_it_works.xlsx"))
print("saved local :", local)
try:
    drive = wb.save_drive("AUDI-1191", "Failure-Debugger How It Works")
    print("saved drive :", drive)
except Exception as e:
    print("drive save skipped:", e)

from openpyxl import load_workbook  # noqa: E402
print("tabs        :", load_workbook(local).sheetnames)
