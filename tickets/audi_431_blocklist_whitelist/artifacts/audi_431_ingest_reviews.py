"""AUDI-431: ingest human 'Your call' answers from the Drive workbook back into the list files.

Reads the 'Manual review' tab of the workbook in Drive, extracts Domain + 'Your call'
(Whitelist / Blocklist / Skip), writes outputs/audi_431_human_calls.csv, then re-runs
audi_431_build_lists.py so the additions files and hygiene checks include the human calls.

Run after (partially) filling the dropdown: python3 artifacts/audi_431_ingest_reviews.py
"""

import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

TICKET = Path(__file__).resolve().parents[1]
OUT = TICKET / "outputs"
XLSX = Path.home() / (
    "Library/CloudStorage/GoogleDrive-malachi@mountain.com/My Drive/Tickets/AUDI-431/"
    "AUDI-431 Blocklist Whitelist Reassessment.xlsx"
)


def main() -> None:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Manual review"]
    header = [c.value for c in ws[4]]
    dom_i, call_i = header.index("Domain"), header.index("Your call")
    calls = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        domain, call = row[dom_i], row[call_i]
        if domain and call in ("Whitelist", "Blocklist"):
            calls.append({"domain": domain, "designation": call})
    df = pd.DataFrame(calls)
    df.to_csv(OUT / "audi_431_human_calls.csv", index=False)
    print(f"human calls ingested: {len(df)} "
          f"({(df['designation'] == 'Whitelist').sum()} WL / {(df['designation'] == 'Blocklist').sum()} BL)")
    r = subprocess.run([sys.executable, str(TICKET / "artifacts" / "audi_431_build_lists.py")])
    sys.exit(r.returncode)


if __name__ == "__main__":
    main()
