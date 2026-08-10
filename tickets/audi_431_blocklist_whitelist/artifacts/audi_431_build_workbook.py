"""AUDI-431 Phase 6: build the branded decision workbook and save to Drive.

Sheets: cover (last), Decisions (all bands, ranked desc), Manual Review (blank designations),
Auto-Whitelist, Auto-Blocklist, Junk rule, Vertical Corrections, TI-200 Unsure status,
Impact, Queries, Read me.
"""

import json
import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, "/Users/malachi/Developer/work/mntn/workspace")
from lib.mntn_xlsx import FMT, MntnWorkbook  # noqa: E402

TICKET = Path(__file__).resolve().parents[1]
OUT = TICKET / "outputs"
Q = TICKET / "queries"
TOTAL_VOL = 16_046_965_205


def main() -> None:
    sheet = pd.read_csv(OUT / "audi_431_decision_sheet.csv")
    sheet["designation"] = sheet["designation"].fillna("")
    dem_path = OUT / "audi_431_qc_demotions.csv"
    if dem_path.exists():
        dem = pd.read_csv(dem_path)
        demoted = set(dem["domain"])
        sheet.loc[sheet["domain"].isin(demoted), "designation"] = ""
        sheet.loc[sheet["domain"].isin(demoted), "band"] = "manual"
        sheet.loc[sheet["domain"].isin(demoted), "designation_source"] = "qc-demoted"
    impact = json.loads((OUT / "audi_431_impact.json").read_text())
    qc = json.loads((OUT / "audi_431_qc_report.json").read_text())
    corr = pd.read_csv(OUT / "audi_431_vertical_corrections.csv")
    unsure = pd.read_csv(OUT / "audi_431_ti200_unsure_status.csv")

    def slim(df: pd.DataFrame) -> pd.DataFrame:
        out = df[["rank", "domain", "total_count", "days_seen", "n_urls",
                  "med_score", "pct_ge_04", "band", "designation", "designation_source", "band_rule"]].copy()
        out.columns = ["Rank", "Domain", "28d volume", "Days seen", "URLs scored",
                       "Median score", "Share >= 0.4", "Band", "Designation", "Source", "Rule"]
        return out

    wb = MntnWorkbook(
        title="Blocklist / Whitelist Re-assessment",
        ticket="AUDI-431",
        subtitle="Most-common uncategorized domains adjudicated from prod ecommerce-model scores; vertical corrections for top wcv domains",
        period="28d window 2026-07-13 to 2026-08-09; scores 7d to 2026-08-08",
    )

    fmt = {"28d volume": FMT.INT, "URLs scored": FMT.INT, "Median score": FMT.NUM2,
           "Share >= 0.4": FMT.PCT1, "Rank": FMT.INT, "Days seen": FMT.INT}

    dec = slim(sheet)
    wb.table("Decisions", dec,
             finding=f"{impact['bl_additions'] + impact['wl_additions']} auto-adds resolve {impact['pct_missing_volume_resolved']:.0%} of uncategorized visit volume",
             method="All 3,024 adjudicated domains, ranked by 28d missing volume. Bands from prod model score aggregates; blank designation = manual review.",
             formats=fmt, heat={"28d volume": "high"}, kind="headline",
             toc="Every adjudicated domain with its band, scores, and designation",
             query="audi_431_qa_score_aggregates.sql")

    man = slim(sheet[sheet["designation"] == ""]).drop(columns=["Designation", "Source", "Rule", "Band"])
    ai_path = OUT / "audi_431_ai_review.csv"
    man_rag = None
    if ai_path.exists():
        ai = pd.read_csv(ai_path)
        ai.columns = ["Domain", "AI verdict", "AI conf", "AI reason"]
        ai["AI verdict"] = ai["AI verdict"].map({"ecommerce": "Ecommerce", "not_ecommerce": "Not ecommerce", "unsure": "Unsure"})
        man = man.merge(ai, on="Domain", how="left")
        man_rag = {"AI verdict": lambda v: "POS" if v == "Ecommerce" else ("NEG" if v == "Not ecommerce" else ("WARN" if v == "Unsure" else None))}
    man["Your call"] = None
    ws_man = wb.table("Manual review", man,
             finding=f"{len(man)} ambiguous domains await your call in the dropdown ({impact['manual_volume_share']:.0%} of volume)",
             method="Scores inconclusive (between confident bands). AI verdict is ADVISORY - you decide. Pick Whitelist/Blocklist/Skip in 'Your call'; blank or Skip rows do not ship. Work top-down by volume.",
             formats={**fmt, "AI conf": FMT.NUM2}, heat={"28d volume": "high"}, rag=man_rag, kind="data",
             toc="Your review queue: AI advisory verdict + a dropdown for your decision",
             query="audi_431_qa_score_aggregates.sql")
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Whitelist,Blocklist,Skip"', allow_blank=True, showDropDown=False)
    ws_man.add_data_validation(dv)
    call_col = get_column_letter(list(man.columns).index("Your call") + 1)
    dv.add(f"{call_col}5:{call_col}{4 + len(man)}")

    wl = slim(sheet[(sheet["designation"] == "Whitelist")])
    wb.table("Whitelist adds", wl,
             finding=f"{len(wl)} domains are confidently ecommerce (median score >= 0.9181)",
             method="TGT-4016 P90 band + >=90% of URLs over the prod 0.4 gate; adversarial QC applied. Note: whitelist adds still lack a wcv vertical until the next crawl.",
             formats=fmt, kind="data", toc="Confident ecommerce additions",
             query="audi_431_qa_score_aggregates.sql")

    bl = slim(sheet[(sheet["designation"] == "Blocklist") & (sheet["band"] == "auto_blocklist")])
    wb.table("Blocklist adds", bl,
             finding=f"{len(bl)} domains are confidently non-ecommerce ({qc['bl_dispute_rate']:.1%} QC dispute rate)",
             method="Median score <= 0.05 and <5% of URLs clear the prod 0.4 gate - the blocklist codifies what the model already decides, saving daily re-scoring.",
             formats=fmt, kind="data", toc="Confident non-ecommerce additions",
             query="audi_431_qa_score_aggregates.sql")

    junk = sheet[sheet["band"] == "junk_rule"][["domain", "total_count", "days_seen", "junk_tier", "designation"]].copy()
    junk.columns = ["Domain", "28d volume", "Days seen", "Junk tier", "Designation"]
    wb.table("Junk strings", junk,
             finding="24 stable URL-parse artifacts carry 7% of all uncategorized volume",
             method="Trailing-dot tldextract artifacts (comhttps., android-app., ...) seen >=14 days at >=1M rows. Blocklistable stable strings - the shipped localhost. precedent.",
             formats={"28d volume": FMT.INT, "Days seen": FMT.INT}, kind="detail",
             toc="Stable parse-garbage strings proposed for the blocklist")

    n_live = int((corr["impact"] == "Live").sum())
    cw = corr[["domain", "vertical_name", "final_verdict", "impact", "suggested_vertical", "n_urls",
               "judge_confidence", "defend_confidence", "judge_reason"]].copy()
    cw.columns = ["Domain", "Current vertical", "Verdict", "Impact", "Suggested vertical", "7d URLs",
                  "Judge conf", "Defend conf", "Reason"]
    wb.table("Vertical corrections", cw,
             finding=f"{n_live} wrong verticals are reaching IP associations today (36 of 76; the other 40 are blocklisted and inert)",
             method="Two independent LLM passes (judge + defend); 'wrong' requires both. Suggested vertical is enum-constrained to the 152-name wcv roster. ip_vertical_associations anti-joins the blocklist, so blocklisted rows never reach IPs.",
             formats={"7d URLs": FMT.INT, "Judge conf": FMT.NUM2, "Defend conf": FMT.NUM2},
             rag={"Verdict": lambda v: "NEG" if v == "wrong" else ("WARN" if v == "unsure" else None),
                  "Impact": lambda v: "NEG" if v == "Live" else None},
             kind="data", toc="Top-traffic wcv domains whose vertical is wrong (Live rows are the actionable ones)",
             query="audi_431_qb_wcv_traffic.sql")

    conflict = pd.read_csv(OUT / "audi_431_whitelist_blocklist_conflict.csv")
    conflict.columns = ["Domain"]
    wb.table("List conflicts", conflict,
             finding=f"{len(conflict)} domains sit in BOTH the whitelist and the blocklist today",
             method="Pre-existing contradiction in the shipped 2025-09-23 files, not introduced here. Blocklist wins at every consumer (checked first), so these are effectively blocked. Worth reconciling at deploy.",
             kind="detail", toc="Domains in both lists (blocklist currently wins)")

    us = unsure[["domain_name", "vertical_name", "now_blocklist", "now_whitelist", "now_wcv"]].copy()
    us.columns = ["Domain", "TI-200 vertical", "In blocklist", "In whitelist", "In wcv"]
    wb.table("TI-200 unsure", us,
             finding="All 149 TI-200 'Unsure' domains were categorized by the 2025-11-07 crawl refresh",
             method="Every row is now in website_crawl_verticals; 9 also in the blocklist, 24 in the whitelist. No re-adjudication needed.",
             kind="detail", toc="The 149 leftover TI-200 rows - all self-resolved")

    imp = pd.DataFrame([
        ("Blocklist additions (auto + junk)", impact["bl_additions"], None),
        ("Whitelist additions", impact["wl_additions"], None),
        ("Merged blocklist size", impact["merged_blocklist_size"], None),
        ("28d uncategorized volume resolved", None, impact["pct_missing_volume_resolved"]),
        ("Volume left in manual band", None, impact["manual_volume_share"]),
        ("QC dispute rate (blocklist sample)", None, qc["bl_dispute_rate"]),
        ("QC dispute rate (whitelist, full)", None, qc["wl_dispute_rate"]),
    ], columns=["Measure", "Count", "Share"])
    wb.table("Impact", imp,
             finding=f"The lists move from 11 months stale to covering {impact['pct_missing_volume_resolved']:.0%} of today's gap",
             method="Volume = missing_domains rows over the 28d window (16.0B total). Both lists were last touched 2025-09-23.",
             formats={"Count": FMT.INT, "Share": FMT.PCT1}, kind="detail",
             toc="What ships and how much of the gap it closes")

    wb.sql_dir("Queries", str(Q),
               note="BigQuery external-table queries over GCS parquet (us-central1). Candidate build + list hygiene are pandas scripts in artifacts/.")

    wb.glossary("Read me", intro="How this workbook was produced and how to act on it.", rows=[
        ("Pipeline", "site_visit_signal URL -> blocklist (stop) -> whitelist (= ecommerce) -> else ecommerce model @0.4 -> if ecomm, vertical from website_crawl_verticals (wcv)."),
        ("missing_domains", "Daily prod job: svs domains not in wcv, net of both lists. Candidate source, 28d window."),
        ("Median score", "Median prod ecommerce-model score across the domain's URLs (7d). 1.0 = certainly ecommerce."),
        ("Share >= 0.4", "Fraction of the domain's URLs at/above the prod ecommerce cutoff."),
        ("Bands", "Whitelist: med >= 0.9181 (TGT-4016 P90) & share >= 90%. Blocklist: med <= 0.05 & share <= 5%. Between: manual."),
        ("QC", "Adversarial LLM pass tried to refute every auto-whitelist row and a 100-row blocklist sample; disputed rows were demoted to manual."),
        ("To deploy", "Ryan Kleck: append additions to the two files under vertical_categorizations/ecommerce_domain_whitelist/. Corrections via vertical overrides."),
    ])

    wb.cover(takeaways=[
        f"{impact['bl_additions']} blocklist + {impact['wl_additions']} whitelist adds resolve {impact['pct_missing_volume_resolved']:.0%} of uncategorized visit volume",
        f"{int((corr['impact'] == 'Live').sum())} wrong verticals reach IP associations today (facebook.com as B2B Sales & Marketing leads); 40 more are blocklisted and inert",
        f"{len(man)} ambiguous domains ship blank for hand review, volume-sorted; every auto-decision carries its rule",
    ])
    path = wb.save_drive("AUDI-431", "Blocklist Whitelist Reassessment")
    print(f"saved: {path}")


if __name__ == "__main__":
    main()
