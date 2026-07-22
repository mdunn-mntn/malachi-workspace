#!/usr/bin/env python3
"""
Rebuild the GOAL-ATTAINMENT "Customer Goal Data Map" .xlsx through the styled MNTN builder
(lib/mntn_xlsx.py), so it inherits the current standard (brand-green shading, em-dash strip, wrapping).

Reproducible: the content is captured once to goal_attainment_data.json (next to this script). If that
JSON exists we build from it; otherwise we extract it from the live Drive workbook first. Re-run any time
to re-apply the latest format:  python3 tickets/.../artifacts/goal_attainment_build_xlsx.py
"""
import json
import os
import sys

import pandas as pd

ROOT = "/Users/malachi/Developer/work/mntn/workspace"
sys.path.insert(0, ROOT)
from lib.mntn_xlsx import FMT, MntnWorkbook

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_JSON = os.path.join(HERE, "goal_attainment_data.json")
DRIVE = os.path.expanduser("~/Library/CloudStorage/GoogleDrive-malachi@mountain.com/My Drive/Tickets")
SRC_XLSX = os.path.join(DRIVE, "GOAL-ATTAINMENT", "GOAL-ATTAINMENT Customer Goal Data Map.xlsx")

TABLE_SHEETS = ["Table map", "Goal types"]
NOTES_SHEETS = ["Already built", "Join map", "How to answer it"]
ORDER = TABLE_SHEETS + NOTES_SHEETS
WIDTHS = {
    "Table map": {"Table": 48, "Role": 15, "What's in it": 48, "Grain": 18, "Join key(s)": 28, "Joins to": 32},
    "Goal types": {"ID": 6, "Goal type": 20, "What the customer is aiming for": 64},
}


def _extract():
    """Read the live workbook once into a plain dict (values only, styling-agnostic)."""
    from openpyxl import load_workbook
    wb = load_workbook(SRC_XLSX)
    ov = wb["Overview"]

    def cell(ws, coord):
        return ws[coord].value

    # cover meta + takeaways + contents map
    toc = {}
    for r in range(1, ov.max_row + 1):
        a, b = ov.cell(r, 1).value, ov.cell(r, 2).value
        if a in ORDER and b:
            toc[a] = b
    takeaways = [ov.cell(r, 2).value for r in range(1, ov.max_row + 1)
                 if str(ov.cell(r, 1).value).strip() in ("1", "2", "3")]
    cover = {
        "title": cell(ov, "A6"), "subtitle": cell(ov, "A7"),
        "ticket": cell(ov, "B9") or "GOAL-ATTAINMENT", "period": cell(ov, "B10"),
        "generated": cell(ov, "B11"), "status": cell(ov, "B13") or "Final",
        "takeaways": [t for t in takeaways if t], "toc": toc,
    }

    sheets = []
    for name in ORDER:
        ws = wb[name]
        finding, intro = ws["A1"].value, ws["A2"].value
        if name in TABLE_SHEETS:
            headers = [ws.cell(4, c).value for c in range(1, ws.max_column + 1) if ws.cell(4, c).value]
            rows = []
            r = 5
            while True:
                a = ws.cell(r, 1).value
                if a is None or str(a).startswith("Source:"):
                    break
                rows.append({h: ws.cell(r, c + 1).value for c, h in enumerate(headers)})
                r += 1
            sheets.append({"kind": "table", "name": name, "finding": finding,
                           "method": intro, "columns": headers, "rows": rows})
        else:
            body = [ws.cell(r, 1).value for r in range(4, ws.max_row + 1)
                    if ws.cell(r, 1).value and not str(ws.cell(r, 1).value).startswith("Source:")]
            blocks = [[body[i], body[i + 1] if i + 1 < len(body) else ""] for i in range(0, len(body), 2)]
            sheets.append({"kind": "notes", "name": name, "finding": finding,
                           "intro": intro, "blocks": blocks})
    return {"cover": cover, "sheets": sheets}


def load_data():
    if os.path.exists(DATA_JSON):
        with open(DATA_JSON) as f:
            return json.load(f)
    data = _extract()
    with open(DATA_JSON, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print("captured content ->", DATA_JSON)
    return data


def build(data):
    cov = data["cover"]
    wb = MntnWorkbook(title=cov["title"], ticket=cov["ticket"], subtitle=cov["subtitle"],
                      period=cov["period"], generated=cov["generated"], status=cov["status"])
    for sh in data["sheets"]:
        toc = cov["toc"].get(sh["name"], "")
        if sh["kind"] == "table":
            df = pd.DataFrame(sh["rows"], columns=sh["columns"])
            fmts = {"ID": FMT.INT} if "ID" in sh["columns"] else {}
            wb.table(sh["name"], df, finding=sh["finding"], method=sh.get("method", ""),
                     kind="headline" if sh["name"] == "Table map" else "data",
                     formats=fmts, widths=WIDTHS.get(sh["name"]), toc=toc)
        else:
            wb.notes(sh["name"], blocks=[tuple(b) for b in sh["blocks"]],
                     intro=sh.get("intro", ""), toc=toc)
    wb.cover(takeaways=cov["takeaways"])
    out = wb.save_drive("GOAL-ATTAINMENT", "Customer Goal Data Map")
    wb.save_local(os.path.join(HERE, "..", "outputs", "goal_attainment_customer_goal_map.xlsx"))
    return out


if __name__ == "__main__":
    out = build(load_data())
    from openpyxl import load_workbook
    print("rebuilt ->", out, "| tabs:", load_workbook(out).sheetnames)
