#!/usr/bin/env python3
"""AUDI-1115 — true willingness-to-pay CPM per vendor, 3 lenses.

Builds outputs/audi_1115_wtp_cpm.xlsx from measured AUDI-1089 outputs plus this
ticket's L2 flow-filtered coverage scan. Per paid vendor:

  value band   = solo-cohort media $/wk x 52 x internal margin band (10-30%) —
                 the measured $ the vendor uniquely generates (q8b machinery);
                 the max total bill at which net >= 0.
  lens units   L1 = all rows ingested (q1 median rows/day x 365)
               L2 = flow-filtered unique usable triples (audi_1115_l2 x 365/30)
               L3 = won impressions touched (deck_d2 week x 52)
  effective CPM = bill_annualized / (units/1000)     (what we pay today)
  WTP CPM band  = value band   / (units/1000)        (ceiling per lens)

Flat-fee vendors (5x5, Predactiv, Klickly) have no bill on the meter (amounts
pending Maya) -> effective CPMs PENDING; WTP ceilings still computed.
Percent-like and dollar values stay raw numbers (no string formatting) so
Excel number formats work.
"""

import csv
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE.parent / "outputs"
A1089 = HERE.parent.parent.parent / "audi_1089_ddp_vendor_evaluations" / "outputs" / "run_2026_07_10"
L2_CSV = OUT / "audi_1115_l2_flow_coverage.csv"

PAID_DS = [28, 40, 33, 24, 36, 25, 26, 39]  # ranked metered-by-bill first, then flat-fee
MARGIN_LO, MARGIN_HI = 0.10, 0.30  # internal margin parameters — not in shared queries
WEEKS_YR = 52
DAYS_YR = 365
L2_WINDOW_DAYS = 30

PENDING = "PENDING"


def read_csv(path):
    with open(path) as f:
        return list(csv.DictReader(f))


def load_inputs():
    v = {ds: {} for ds in PAID_DS}

    for r in read_csv(A1089 / "deck_d3_bills_cpm.csv"):
        ds = int(r["data_source_id"])
        if ds in v:
            v[ds]["name"] = r["data_partner_name"]
            v[ds]["billing_type"] = r["billing_type"]
            v[ds]["bill_yr"] = float(r["bill_annualized"]) if r["bill_annualized"] else None
            v[ds]["l0_units_yr"] = float(r["billed_imps_month"]) * 12 if r["billed_imps_month"] else None
            v[ds]["contract_cpm"] = float(r["contract_cpm"]) if r["contract_cpm"] else None

    rows_by_ds = {}
    for r in read_csv(A1089 / "q1_scale_by_day.csv"):
        rows_by_ds.setdefault(int(r["data_source_id"]), []).append(float(r["n_rows"]))
    for ds in PAID_DS:
        days = rows_by_ds.get(ds, [])
        v[ds]["l1_units_yr"] = statistics.median(days) * DAYS_YR if days else None
        v[ds]["l1_days_measured"] = len(days)

    for r in read_csv(A1089 / "deck_d2_touched_won_bids.csv"):
        if r["rec"] == "ds" and int(r["ds"]) in v:
            v[int(r["ds"])]["l3_units_yr"] = float(r["imps_touched"]) * WEEKS_YR

    # free co-hold share (deck_d1): the credit slice preemption removes from the meter
    for r in read_csv(A1089 / "deck_d1_universe_coverage.csv"):
        if r["rec"] == "source" and int(r["ds"]) in v and r["free_cohold_pct"]:
            ds = int(r["ds"])
            cohold = float(r["free_cohold_pct"]) / 100
            if v[ds].get("l0_units_yr"):
                v[ds]["l0p_units_yr"] = v[ds]["l0_units_yr"] * (1 - cohold)

    for r in read_csv(A1089 / "q8b_solo_perf.csv"):
        ds = int(r["ds"]) if r["ds"].isdigit() else None
        if ds in v and r["rec"] == "serve":
            if r["k"] == "media":
                v[ds]["solo_media_wk"] = float(r["v"])
            elif r["k"] == "imps":
                v[ds]["solo_imps_yr"] = float(r["v"]) * WEEKS_YR

    if L2_CSV.exists():
        for r in read_csv(L2_CSV):
            if r["rec"] == "vendor" and int(r["ds"]) in v:
                v[int(r["ds"])]["l2_units_yr"] = float(r["flow_cnt"]) * DAYS_YR / L2_WINDOW_DAYS
                v[int(r["ds"])]["l2_sameday_anchor"] = float(r["sameday_cnt"])
    return v


def per_mille(numer, units):
    if numer is None or not units:
        return None
    return numer / (units / 1000.0)


def build(v):
    wb = Workbook()
    ws = wb.active
    ws.title = "wtp_cpm"

    head_font = Font(bold=True, size=10)
    group_fill = PatternFill("solid", fgColor="DDE6F0")
    pend_fill = PatternFill("solid", fgColor="FFF3CD")
    wrap = Alignment(wrap_text=True, vertical="top")

    cols = [
        ("Vendor", None, 16),
        ("DS", "0", 5),
        ("Billing", None, 10),
        ("Bill $/yr", "#,##0", 11),
        ("Vendor value $/yr LOW (10% margin)", "#,##0", 13),
        ("Vendor value $/yr HIGH (30% margin)", "#,##0", 13),
        ("L0 units/yr (current billing meter, credited imps)", "#,##0", 15),
        ("L0 contract CPM today", "$0.00", 10),
        ("L0 WTP CPM LOW", "$0.000", 11),
        ("L0 WTP CPM HIGH", "$0.000", 11),
        ("L0p units/yr (POST-PREEMPTION meter: free-covered credit removed)", "#,##0", 15),
        ("L0p WTP CPM LOW", "$0.000", 11),
        ("L0p WTP CPM HIGH", "$0.000", 11),
        ("L1 units/yr (rows ingested)", "#,##0", 16),
        ("L1 effective CPM", "$0.00000", 12),
        ("L1 WTP CPM LOW", "$0.00000", 12),
        ("L1 WTP CPM HIGH", "$0.00000", 12),
        ("L2 units/yr (flow-filtered unique triples)", "#,##0", 16),
        ("L2 effective CPM", "$0.00000", 12),
        ("L2 WTP CPM LOW", "$0.00000", 12),
        ("L2 WTP CPM HIGH", "$0.00000", 12),
        ("L3 units/yr (won imps touched)", "#,##0", 16),
        ("L3 effective CPM", "$0.0000", 12),
        ("L3 WTP CPM LOW", "$0.0000", 12),
        ("L3 WTP CPM HIGH", "$0.0000", 12),
        ("Ref: unique won imps/yr (solo cohort)", "#,##0", 14),
    ]
    for j, (title, _, width) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=title)
        c.font = head_font
        c.alignment = wrap
        c.fill = group_fill
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = "A2"

    pending_cells = 0
    order = sorted(
        PAID_DS,
        key=lambda d: -(v[d]["bill_yr"] if v[d].get("bill_yr") is not None else -1),
    )
    for i, ds in enumerate(order, start=2):
        d = v[ds]
        val_lo = d.get("solo_media_wk", 0) * WEEKS_YR * MARGIN_LO if "solo_media_wk" in d else None
        val_hi = d.get("solo_media_wk", 0) * WEEKS_YR * MARGIN_HI if "solo_media_wk" in d else None
        bill = d.get("bill_yr")
        row = [
            d.get("name", f"ds{ds}"), ds, d.get("billing_type"),
            bill, val_lo, val_hi,
            d.get("l0_units_yr"), d.get("contract_cpm"),
            per_mille(val_lo, d.get("l0_units_yr")), per_mille(val_hi, d.get("l0_units_yr")),
            d.get("l0p_units_yr"),
            per_mille(val_lo, d.get("l0p_units_yr")), per_mille(val_hi, d.get("l0p_units_yr")),
            d.get("l1_units_yr"), per_mille(bill, d.get("l1_units_yr")),
            per_mille(val_lo, d.get("l1_units_yr")), per_mille(val_hi, d.get("l1_units_yr")),
            d.get("l2_units_yr"), per_mille(bill, d.get("l2_units_yr")),
            per_mille(val_lo, d.get("l2_units_yr")), per_mille(val_hi, d.get("l2_units_yr")),
            d.get("l3_units_yr"), per_mille(bill, d.get("l3_units_yr")),
            per_mille(val_lo, d.get("l3_units_yr")), per_mille(val_hi, d.get("l3_units_yr")),
            d.get("solo_imps_yr"),
        ]
        for j, (val, (_, fmt, _w)) in enumerate(zip(row, cols), start=1):
            c = ws.cell(row=i, column=j)
            if val is None:
                c.value = PENDING
                c.fill = pend_fill
                pending_cells += 1
            else:
                c.value = val
                if fmt:
                    c.number_format = fmt

    notes = [
        "Value band = q8b solo-cohort media $/wk x 52 x internal margin band (10-30%). Solo cohort "
        "= IPs the vendor delivered that NEITHER FREE LOG delivered (37d membership); other PAID "
        "vendors are NOT excluded, so value bands OVERLAP across paid vendors - do not sum them. "
        "This makes each break-even verdict conservative in the vendor's favor.",
        "L1 = median rows/day over the 30d q1 window x 365 (all rows ingested, pre-filter).",
        "L2 = flow-filtered unique usable triples (ip x REG_DOMAIN x date, 30d x 365/30): free-log "
        "credit requires the pair in that log during [D-30, D-1]; same-day-only presence earns no "
        "credit (2026-07-16 meeting rule, both free logs). Source: audi_1115_l2_flow_coverage.sql.",
        "L3 = won impressions (cost_impression_log, valuation week x 52) landing on IPs the vendor "
        "delivered in the 37d membership window (deck_d2). Overlapping, non-additive across vendors.",
        "Effective CPM = bill / (units/1000): what we pay today per 1,000 units at that lens. "
        "WTP CPM = value / (units/1000): the per-unit price at which the vendor nets zero at that "
        "margin. Bill above value (or effective CPM above WTP HIGH) = vendor is net-negative.",
        "L0 = the RENEGOTIATION lens: the vendor's own billing meter (deck_d3 billed credited "
        "imps x 12, current regime). L0 WTP = the contract CPM at which the vendor breaks even — "
        "compare directly against the $0.50 contract rate.",
        "L0p = L0 on the POST-PREEMPTION meter: units x (1 - free co-hold share, deck_d1) — the "
        "meter if we stop crediting vendors for signal the free logs also captured (AUDI-1113). "
        "The VALUE side already excludes free logs in every lens (q8b solo cohort); L0p removes "
        "them from the DENOMINATOR too, i.e. price per exclusively-unique credited imp. On this "
        "lens the 33Across family reaches fair at the top of the margin band (33Across HIGH "
        "~$0.54 > $0.50; 33A API HIGH ~$0.50) — preemption + renegotiation STACK to fair for the "
        "33Across pair ONLY; Sovrn/Justuno/Cybba stay far under (tiny co-hold, junk/unique credit).",
        "Flat-fee vendors (5x5, sharethis_predactiv, Klickly): bill amounts pending finance (Maya) "
        "-> Bill and effective-CPM cells PENDING; WTP ceilings computed from the value side alone.",
        "Windows: q1/q8b/deck_d2 measured on outputs/run_2026_07_10 (svs 30d 2026-06-02..07-01, "
        "37d membership, CIL week 2026-07-02..08); L2 scan adds a 30d lookback (2026-05-03+).",
        "Extrapolation caveat: value = ONE week of media x 52 (the 07-02..08 week contains the "
        "July 4th US holiday - value likely UNDERSTATED, i.e. conservative); meter = June 2026 x 12 "
        "(first full month of the integer-credit regime). Point-in-time ranges - re-measure on a "
        "non-holiday week before quoting in a renegotiation.",
        "Meter cross-check (2026-07-17): L0 meter ballpark-confirmed against the BAE billing table "
        "(dw-main-gold.reporting.ddp_mm_winners_imp_202606); no simple aggregation reproduces it "
        "exactly (+-7-42% by vendor; credit splits across matched data paths incl. 3P segments - "
        "exact BAE rule pending the 2026-07-20 billing sync). Verdicts are ROBUST to this: even the "
        "most vendor-favorable candidate leaves every metered vendor far under $0.50 break-even. "
        "Recon: queries/audi_1115_l0b_bae_winners_recon.sql + the epic workbook's bae_billing_recon sheet.",
    ]
    nrow = len(order) + 3
    for note in notes:
        c = ws.cell(row=nrow, column=1, value=note)
        c.alignment = wrap
        c.font = Font(size=9, italic=True)
        ws.merge_cells(start_row=nrow, start_column=1, end_row=nrow, end_column=len(cols))
        ws.row_dimensions[nrow].height = 26
        nrow += 1

    qs = wb.create_sheet("queries")
    qmap = [
        ("Bill $/yr, billing type", "audi_1089 runbook/queries/deck_d3_bills_cpm.sql", "LANDED"),
        ("Vendor value band (solo media)", "audi_1089 runbook/queries/q8a_solo_stock.sql + q8b (solo perf)", "LANDED"),
        ("L1 rows ingested", "audi_1089 runbook/queries/q1_scale_by_day.sql", "LANDED"),
        ("L2 flow-filtered unique triples", "audi_1115_wtp_cpm/queries/audi_1115_l2_flow_coverage.sql",
         "LANDED" if L2_CSV.exists() else "IN FLIGHT"),
        ("L3 won imps touched", "audi_1089 runbook/queries/deck_d2_touched_won_bids.sql", "LANDED"),
    ]
    qs.cell(row=1, column=1, value="Column(s)").font = head_font
    qs.cell(row=1, column=2, value="Producing query").font = head_font
    qs.cell(row=1, column=3, value="Status").font = head_font
    for i, (what, q, status) in enumerate(qmap, start=2):
        qs.cell(row=i, column=1, value=what)
        qs.cell(row=i, column=2, value=q)
        qs.cell(row=i, column=3, value=status)
    for j, w in enumerate([34, 66, 10], start=1):
        qs.column_dimensions[get_column_letter(j)].width = w

    out = OUT / "audi_1115_wtp_cpm.xlsx"
    wb.save(out)
    print(f"wrote {out}")
    print(f"pending cells: {pending_cells}")
    for ds in order:
        d = v[ds]
        anchor = d.get("l2_sameday_anchor")
        if anchor is not None:
            print(f"  ds{ds} L2 landed; sameday anchor cnt = {anchor:,.0f} (reconcile vs deck_d1 standalone)")


if __name__ == "__main__":
    build(load_inputs())
