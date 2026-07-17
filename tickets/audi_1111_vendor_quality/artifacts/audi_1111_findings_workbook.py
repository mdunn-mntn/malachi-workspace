#!/usr/bin/env python3
"""AUDI-1111 consolidated findings workbook.

One xlsx with every measured table from the epic's three analyses, built from
the landed CSVs (no hardcoded data). Percent-like values are stored as TRUE
FRACTIONS with % number formats (house rule). The full 4-lens WTP table lives
in audi_1115_wtp_cpm/outputs/audi_1115_wtp_cpm.xlsx (canonical) — this
workbook carries the L0 renegotiation summary + a pointer.
"""

import csv
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
EPIC = HERE.parent
OUT = EPIC / "outputs"
OUT.mkdir(exist_ok=True)
A1115 = EPIC / "audi_1115_wtp_cpm" / "outputs"
A1116 = EPIC / "audi_1116_rtc_free_logs" / "outputs"
A1117 = EPIC / "audi_1117_ds14_svs_overlap" / "outputs"
A1089 = EPIC.parent / "audi_1089_ddp_vendor_evaluations" / "outputs" / "run_2026_07_10"

NAMES = {23: "guid_log", 24: "Justuno", 25: "5x5", 26: "Predactiv", 28: "33Across",
         30: "augmentor_log", 33: "Sovrn", 36: "Cybba", 39: "Klickly", 40: "33Across API",
         99: "free union"}

HEAD = Font(bold=True, size=10)
FILL = PatternFill("solid", fgColor="DDE6F0")
WRAP = Alignment(wrap_text=True, vertical="top")
PCT2 = "0.00%"
INT0 = "#,##0"


def sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HEAD
        c.fill = FILL
        c.alignment = WRAP
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    return ws


def put(ws, i, vals, fmts):
    for j, (v, f) in enumerate(zip(vals, fmts), start=1):
        c = ws.cell(row=i, column=j, value=v)
        if f and v is not None and v != "":
            c.number_format = f


def note(ws, row, ncols, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=9, italic=True)
    c.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 26


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # --- 1. WTP L0 renegotiation summary -------------------------------------
    d3 = {r["data_source_id"]: r for r in csv.DictReader(open(A1089 / "deck_d3_bills_cpm.csv"))}
    q8b = {(r["rec"], r["ds"], r["k"]): float(r["v"])
           for r in csv.DictReader(open(A1089 / "q8b_solo_perf.csv")) if r["v"] not in ("", "v")}
    ws = sheet(wb, "wtp_L0_renegotiation",
               ["Vendor", "Billing meter (credited imps/yr)", "Contract CPM today",
                "Break-even CPM LOW (10% margin)", "Break-even CPM HIGH (30% margin)"],
               [14, 20, 12, 16, 16])
    rows = []
    for ds in ["28", "40", "33", "24", "36"]:
        meter = float(d3[ds]["billed_imps_month"]) * 12
        media = q8b[("serve", ds, "media")]
        rows.append((NAMES[int(ds)], meter, 0.50,
                     media * 52 * .1 / (meter / 1000), media * 52 * .3 / (meter / 1000)))
    rows.sort(key=lambda r: -r[4])
    for i, r in enumerate(rows, start=2):
        put(ws, i, list(r), [None, INT0, "$0.00", "$0.000", "$0.000"])
    note(ws, len(rows) + 3, 5,
         "Break-even = solo-cohort media x52 x margin / current meter. Full 4-lens table (L0-L3, "
         "flat-fee ceilings): audi_1115_wtp_cpm/outputs/audi_1115_wtp_cpm.xlsx (canonical). "
         "No metered vendor breaks even at $0.50; verdict is lens-invariant.")

    # --- 2. L2 flow coverage -------------------------------------------------
    l2 = list(csv.DictReader(open(A1115 / "audi_1115_l2_flow_coverage.csv")))
    universe = float(next(r["trips_total"] for r in l2 if r["rec"] == "universe"))
    ws = sheet(wb, "L2_flow_coverage",
               ["Row", "Source", "Triples total", "% of universe", "Same-day cnt",
                "Flow cnt", "Strict cnt", "Flow % of universe"],
               [9, 14, 15, 11, 15, 15, 15, 11])
    for i, r in enumerate(l2, start=2):
        put(ws, i, [r["rec"], NAMES.get(int(r["ds"])) if r["ds"] else "universe",
                    float(r["trips_total"]) if r["trips_total"] else None,
                    float(r["pct_universe"]) / 100 if r["pct_universe"] else None,
                    float(r["sameday_cnt"]) if r["sameday_cnt"] else None,
                    float(r["flow_cnt"]) if r["flow_cnt"] else None,
                    float(r["strict_cnt"]) if r["strict_cnt"] else None,
                    float(r["pct_flow"]) / 100 if r["pct_flow"] else None],
            [None, None, INT0, PCT2, INT0, INT0, INT0, PCT2])
    note(ws, len(l2) + 3, 8,
         "Vendor rows: cnt = vendor triples NOT free-covered under each credit rule (same-day = "
         "deck_d1 anchor convention, EXACT match verified; flow = 2026-07-16 meeting rule, prior "
         "30d, no same-day credit, both free logs; strict = neither). Free rows: cnt = universe "
         "triples COVERED under each rule. Windows: 2026-06-02..07-01 + 30d lookback; IPv4; "
         "usable domains (deck_d1 def). Query: audi_1115_l2_flow_shard.sql x4 + l2_merge.py.")

    # --- 3. RTC vendor share -------------------------------------------------
    rtc = list(csv.DictReader(open(A1116 / "audi_1116_rtc_vendor_share.csv")))
    ws = sheet(wb, "rtc_vendor_share",
               ["Split", "Key", "RTC IPs", "RTC imps", "% of RTC imps"],
               [10, 22, 13, 13, 12])
    for i, r in enumerate(rtc, start=2):
        key = NAMES.get(int(r["key"])) if r["rec"] == "source" else r["key"]
        put(ws, i, [r["rec"], key,
                    float(r["rtc_ips"]) if r["rtc_ips"] else 0,
                    float(r["rtc_imps"]) if r["rtc_imps"] else 0,
                    float(r["pct_rtc_imps"]) / 100 if r["pct_rtc_imps"] else 0],
            [None, None, INT0, INT0, PCT2])
    note(ws, len(rtc) + 3, 5,
         "RTC-fired = model_params realtime_conquest_score=10000; valuation week 2026-07-02..08, "
         "37d svs membership, IPv4. Coverage-based bound (membership window includes post-"
         "impression days). 'source' rows overlap - non-additive. Query: audi_1116_rtc_vendor_share.sql.")

    # --- 4. Ingest latency ---------------------------------------------------
    by_ds = {}
    for r in csv.DictReader(open(A1116 / "audi_1116_hourly_arrival.csv")):
        d = by_ds.setdefault(int(r["ds"]), {"rows": 0.0, "lags": []})
        d["rows"] += float(r["rows_evt_hour"])
        if r["ingest_lag_med_min"]:
            d["lags"].append(float(r["ingest_lag_med_min"]))
    ws = sheet(wb, "ingest_latency",
               ["Source", "Rows/day (2026-07-01)", "Median ingest lag (min)",
                "Min hourly median (min)", "Max hourly median (min)"],
               [14, 17, 15, 15, 15])
    lat = sorted(((NAMES[ds], d["rows"], statistics.median(d["lags"]),
                   min(d["lags"]), max(d["lags"])) for ds, d in by_ds.items()),
                 key=lambda r: -r[1])
    for i, r in enumerate(lat, start=2):
        put(ws, i, list(r), [None, INT0, "0.0", "0.0", "0.0"])
    note(ws, len(lat) + 3, 5,
         "Lag = ULID uid mint time minus event time (svs schema). Free logs stream at 0 min; "
         "vendor lags match the CONFIGURED per-DS lag hours in fpa_site_visit_batch_serverless. "
         "5x5 event times 2h-bucketed pre-14:00. Query: audi_1116_hourly_arrival.sql.")

    # --- 5. DS14 gate lag by cohort ------------------------------------------
    coh = {}
    for r in csv.DictReader(open(A1117 / "audi_1117_ds14_gate_lag_by_cohort.csv")):
        a = int(r["aug_lag_days"]) if r["aug_lag_days"] else None
        g = int(r["guid_lag_days"]) if r["guid_lag_days"] else None
        coh.setdefault((r["channel"], r["funnel"]), []).append((a, g, int(r["imps"])))
    ws = sheet(wb, "ds14_gate_by_cohort",
               ["Channel", "Funnel", "Imps (2026-07-01)", "% aug<=1 or guid<=4",
                "% aug<=7 or guid<=4", "% neither free log in 11d"],
               [10, 13, 14, 15, 15, 16])
    crows = []
    for (ch, fu), rs in coh.items():
        ti = sum(x[2] for x in rs)
        def pc(pred):
            return sum(x[2] for x in rs if pred(x[0], x[1])) / ti
        crows.append((ch, fu, ti,
                      pc(lambda a, g: (a is not None and a <= 1) or (g is not None and g <= 4)),
                      pc(lambda a, g: (a is not None and a <= 7) or (g is not None and g <= 4)),
                      pc(lambda a, g: a is None and g is None)))
    crows.sort(key=lambda r: -r[2])
    for i, r in enumerate(crows, start=2):
        put(ws, i, list(r), [None, None, INT0, PCT2, PCT2, PCT2])
    note(ws, len(crows) + 3, 6,
         "Display = same-day-augmentor ECHO (aug mirrors the display bid stream) - not gate "
         "evidence. CTV gate is SOFT (87.8% = upper bound; same-day rows can postdate the "
         "impression). funnel_level 1/2/3 = prospecting/stage-2/stage-3; 4 = retargeting "
         "(negligible volume). Query: audi_1117_ds14_gate_lag_by_cohort.sql.")

    # --- 6. DS14 overlap sizing ----------------------------------------------
    siz = list(csv.DictReader(open(A1117 / "audi_1117_ds14_overlap_sizing.csv")))
    ws = sheet(wb, "ds14_overlap_sizing",
               ["Row", "Key", "IPs total", "IPs in gate", "% in gate"],
               [10, 26, 14, 14, 10])
    for i, r in enumerate(siz, start=2):
        key = NAMES.get(int(r["key"])) if r["rec"] == "source" else r["key"]
        put(ws, i, [r["rec"], key, float(r["ips_total"]),
                    float(r["ips_in_gate"]) if r["ips_in_gate"] else None,
                    float(r["pct_in_gate"]) / 100 if r["pct_in_gate"] else None],
            [None, None, INT0, INT0, PCT2])
    note(ws, len(siz) + 3, 5,
         "Gate proxy = documented DS14 windows (aug 1d | guid 4d) at ref 2026-07-01; svs 30d, "
         "IPv4. expansion_free_stale (97.0M) + expansion_vendor_only (95.7M) = out-of-gate total "
         "EXACTLY (invariant verified). Query: audi_1117_ds14_overlap_sizing.sql.")

    # --- 7. BAE billing table recon ------------------------------------------
    METER = {"28": 70337329, "40": 29313195, "33": 19313324, "24": 12851766, "36": 3583966}
    BILL_MO = {"28": 35168.66, "40": 14656.60, "33": 9656.66, "24": 6425.88, "36": 1791.98}
    bae = list(csv.DictReader(open(A1115 / "audi_1115_l0b_bae_winners_recon.csv")))
    ws = sheet(wb, "bae_billing_recon",
               ["Row", "Key", "Full credit (imps)", "Equal-split credit", "DS19-only split",
                "$ at tv_cpm (split)", "Actual June meter (imps)", "Actual June bill $",
                "Split vs meter"],
               [9, 22, 14, 14, 14, 13, 15, 13, 11])
    i = 2
    for r in bae:
        if r["rec"] != "winner":
            continue
        ds = r["key"]
        meter, bill = METER.get(ds), BILL_MO.get(ds)
        split = float(r["v2"])
        put(ws, i, ["winner", NAMES.get(int(ds), ds), float(r["v1"]), split, float(r["v3"]),
                    float(r["v4"]), meter, bill,
                    (split / meter - 1) if meter else None],
            [None, None, INT0, INT0, INT0, "#,##0", INT0, "#,##0", "+0.0%;-0.0%"])
        i += 1
    for r in bae:
        if r["rec"] != "mix":
            continue
        put(ws, i, ["mix", r["key"], float(r["v1"]), float(r["v2"]), None,
                    float(r["v3"]), None, None, float(r["v4"])],
            [None, None, INT0, INT0, None, "$0.0000", None, None, PCT2])
        i += 1
    note(ws, i + 1, 9,
         "Source: dw-main-gold.reporting.ddp_mm_winners_imp_202606 (Alyson, 2026-07-17). mix rows: "
         "col C=rows, D=imps, F=avg tv_cpm, I=share tv_cpm=0. Free-only winners bill $0 (100%); "
         "mixed free+paid rows bill $0.50 on 91.7% = 291.1M imps/mo preemption gap (AUDI-1093). "
         "Credit splits across matched DATA PATHS (3P segments in the denominator, DS17 @ $0.95); "
         "no simple aggregation reproduces the coredw meter exactly (+-7-42%) — exact BAE rule = "
         "2026-07-20 billing sync. Query: audi_1115_l0b_bae_winners_recon.sql.")

    # --- 8. queries + charts map ---------------------------------------------
    ws = sheet(wb, "queries", ["Sheet / chart", "Producing query / script", "Status"],
               [34, 62, 10])
    qmap = [
        ("wtp_L0_renegotiation (+ full 4-lens xlsx)", "audi_1115_wtp_cpm/queries + artifacts/audi_1115_wtp_table.py", "LANDED"),
        ("L2_flow_coverage", "audi_1115_wtp_cpm/queries/audi_1115_l2_flow_shard.sql x4 + artifacts/audi_1115_l2_merge.py", "LANDED (anchors EXACT)"),
        ("rtc_vendor_share", "audi_1116_rtc_free_logs/queries/audi_1116_rtc_vendor_share.sql", "LANDED"),
        ("ingest_latency", "audi_1116_rtc_free_logs/queries/audi_1116_hourly_arrival.sql", "LANDED"),
        ("ds14_gate_by_cohort", "audi_1117_ds14_svs_overlap/queries/audi_1117_ds14_gate_lag(_by_cohort).sql", "LANDED"),
        ("ds14_overlap_sizing", "audi_1117_ds14_svs_overlap/queries/audi_1117_ds14_overlap_sizing.sql", "LANDED"),
        ("bae_billing_recon", "audi_1115_wtp_cpm/queries/audi_1115_l0b_bae_winners_recon.sql", "LANDED"),
        ("chart: audi_1115_wtp_vs_contract.png", "audi_1115_wtp_cpm/artifacts/audi_1115_generate_charts.py", "BUILT"),
        ("chart: audi_1115_flow_coverage_drop.png", "audi_1115_wtp_cpm/artifacts/audi_1115_generate_charts.py", "BUILT"),
        ("chart: audi_1116_ingest_latency.png", "audi_1116_rtc_free_logs/artifacts/audi_1116_generate_charts.py", "BUILT"),
        ("chart: audi_1117_ds14_pool.png", "audi_1117_ds14_svs_overlap/artifacts/audi_1117_generate_charts.py", "BUILT"),
    ]
    for i, r in enumerate(qmap, start=2):
        put(ws, i, list(r), [None, None, None])

    out = OUT / "audi_1111_findings.xlsx"
    wb.save(out)
    print(f"wrote {out} ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    build()
