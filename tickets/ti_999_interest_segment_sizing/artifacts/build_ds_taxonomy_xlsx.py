"""Build the TI-999 DS taxonomy spreadsheet (xlsx) — one row per DS, grouped
by locked taxonomy. Reads the per-DS CSV in outputs/ and the Pass 18 bucket
CSV; writes a formatted xlsx with three sheets.

Run from workspace root:
  python3 tickets/ti_999_interest_segment_sizing/artifacts/build_ds_taxonomy_xlsx.py
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

WORKSPACE = Path(__file__).resolve().parents[3]
OUTPUTS = WORKSPACE / "tickets/ti_999_interest_segment_sizing/outputs"
SRC_CSV = OUTPUTS / "ti_999_ds_with_categories_2026_05_29.csv"
PASS_CSV = OUTPUTS / "ti_999_pass21_buckets_2026_05_29.csv"
PASS_LABEL = "Pass 21"
PASS_NOTE = (
    "Four buyer-pickable axes: MM {13,19,38,46} · MNTN Select {9,42} · "
    "3P {17,18,35} (Oracle carved out) · Advertiser CRM {4,8,47}. "
    "RTC dropped as an axis — it's in 99.9% of expressions and belongs with platform plumbing "
    "(geo, DS14 freshness filter, holdout) rather than as a bucket category. "
    "16 no-RTC anomalies on the 'RTC anomalies' sheet."
)
ANOMALIES_CSV = OUTPUTS / "ti_999_pass20_anomalies_2026_05_29.csv"
POLARITY_KPI_CSV = OUTPUTS / "ti_999_pass22c_polarity_aware_buckets_2026_05_29.csv"
GEO_RESTRICTION_CSV = OUTPUTS / "ti_999_pass24_geo_restriction_2026_06_01.csv"
EXCL_AXES_CSV = OUTPUTS / "ti_999_pass25_full_polarity_2026_06_01.csv"
PASS26_CSV = OUTPUTS / "ti_999_pass26_or_vs_and_include_2026_06_01.csv"
PASS27_CSV = OUTPUTS / "ti_999_pass27_mm_and_or_split_2026_06_01.csv"
OUT_XLSX = OUTPUTS / "ti_999_ds_taxonomy_2026_05_29.xlsx"

# Locked taxonomy assignments (post-Pass 18, post-Jordan/Sean clarifications).
# Order here controls group sort order in the sheet.
GROUP_ASSIGNMENTS: dict[int, str] = {
    13: "MM", 19: "MM", 38: "MM", 46: "MM",
    9:  "MNTN Select", 42: "MNTN Select",
    2:  "MNTN Pixel", 21: "MNTN Pixel", 34: "MNTN Pixel", 43: "MNTN Pixel",
    17: "3P", 18: "3P", 35: "3P",
    4:  "Advertiser CRM", 8: "Advertiser CRM", 47: "Advertiser CRM",
    14: "Bid mechanics", 16: "Bid mechanics",
    1:  "3P (Oracle deprecated)",  # carved out per Sean Yang
    11: "3P (LiveRamp legacy)",    # deprecated per Sean Yang
}

GROUP_SORT_ORDER = [
    "MM",
    "PP",
    "MNTN Select",
    "MNTN Pixel",
    "3P",
    "3P (Oracle deprecated)",
    "3P (LiveRamp legacy)",
    "Advertiser CRM",
    "Bid mechanics",
    "Dormant / out-of-scope",
]

GROUP_FILL = {
    "MM":                        PatternFill("solid", fgColor="DCEFFC"),  # light blue
    "PP":                        PatternFill("solid", fgColor="BBDEFB"),  # slightly deeper blue
    "MNTN Select":               PatternFill("solid", fgColor="FFF3E0"),  # light orange
    "MNTN Pixel":                PatternFill("solid", fgColor="E1F5E1"),  # light green
    "3P":                        PatternFill("solid", fgColor="F3E5F5"),  # light purple
    "3P (Oracle deprecated)":    PatternFill("solid", fgColor="EDE7F6"),  # paler purple
    "3P (LiveRamp legacy)":      PatternFill("solid", fgColor="EDE7F6"),
    "Advertiser CRM":            PatternFill("solid", fgColor="FFF9C4"),  # light yellow
    "Bid mechanics":             PatternFill("solid", fgColor="ECEFF1"),  # light gray-blue
    "Dormant / out-of-scope":    PatternFill("solid", fgColor="F5F5F5"),  # neutral light gray
}

HEADER_FILL = PatternFill("solid", fgColor="263238")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)


def group_for_ds(ds_id: int) -> str:
    return GROUP_ASSIGNMENTS.get(ds_id, "Dormant / out-of-scope")


def visible_str(raw: str) -> str:
    s = raw.strip().lower()
    if s in ("true", "1", "t", "yes"):
        return "Yes"
    if s in ("false", "0", "f", "no"):
        return "No"
    return raw


def to_int(s: str, default: int = 0) -> int:
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return default


def to_float(s: str, default: float = 0.0) -> float:
    try:
        return float(s)
    except (TypeError, ValueError):
        return default


def load_per_ds_rows() -> list[dict]:
    rows: list[dict] = []
    with SRC_CSV.open() as f:
        reader = csv.DictReader(f)
        for r in reader:
            ds_id = to_int(r["ds"])
            rows.append({
                "ds": ds_id,
                "ds_name": r.get("ds_name", ""),
                "display_name": r.get("display_name", "") or "",
                "visible": visible_str(r.get("visible", "")),
                "group": group_for_ds(ds_id),
                "pos_camps": to_int(r.get("prospecting_pos_camps", "")),
                "pos_spend_M": to_float(r.get("prospecting_pos_spend_M", "")),
                "neg_camps": to_int(r.get("prospecting_neg_camps", "")),
                "category_count": to_int(r.get("category_count", "")),
                "sample_categories": r.get("sample_categories", ""),
            })
    rows.sort(key=lambda r: (GROUP_SORT_ORDER.index(r["group"]) if r["group"] in GROUP_SORT_ORDER else 99, r["ds"]))
    return rows


def write_taxonomy_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "DS taxonomy"

    title = "TI-999 DS taxonomy — locked 2026-05-29 (Pass 18)"
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    note = (
        "Source: bronze.integrationprod.data_sources (canonical type=1 DSes only, 68 total) + "
        "bronze.tpa.categories. Usage = 30d prospecting window 2026-04-29 to 2026-05-28. "
        "PP shares DSes with MM (DS13 vertical + DS19 keyword → score-tier inside MM 2.0). "
        "Oracle (DS1) and LiveRamp legacy (DS11) are deprecated 3P; called out separately."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 48

    headers = [
        "DS ID", "Group", "Name (canonical)", "Display name", "Visible (buyer pick?)",
        "+camps (30d)", "+spend (30d, $M)", "−camps (30d)", "Categories in tpa.categories",
        "Sample categories",
    ]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 28

    for i, r in enumerate(rows, start=header_row + 1):
        fill = GROUP_FILL.get(r["group"], GROUP_FILL["Dormant / out-of-scope"])
        values = [
            r["ds"],
            r["group"],
            r["ds_name"],
            r["display_name"],
            r["visible"],
            r["pos_camps"],
            r["pos_spend_M"],
            r["neg_camps"],
            r["category_count"],
            r["sample_categories"],
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=v)
            cell.fill = fill
            if col_idx == 7:  # +spend
                cell.number_format = '"$"#,##0.00'
            elif col_idx in (1, 6, 8, 9):
                cell.number_format = "#,##0"
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 10))

    widths = [8, 24, 36, 22, 12, 12, 16, 12, 16, 80]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A5"


def write_group_summary_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Group summary")

    by_group: dict[str, dict] = {}
    for r in rows:
        g = r["group"]
        if g not in by_group:
            by_group[g] = {"n_ds": 0, "n_active": 0, "n_visible": 0}
        by_group[g]["n_ds"] += 1
        is_active = (r["pos_camps"] > 0) or (r["neg_camps"] > 0)
        if is_active:
            by_group[g]["n_active"] += 1
        if r["visible"] == "Yes":
            by_group[g]["n_visible"] += 1

    ws.cell(row=1, column=1, value="Group summary").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    note = (
        "Active (30d) = at least one campaign referenced this DS positively or negatively in 30d prospecting. "
        "Visible to buyers = data_sources.visible=true (buyer-pickable in the UI)."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.row_dimensions[2].height = 28

    headers = ["Group", "# DSes in group", "# active (30d)", "# visible to buyers"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    row_i = 5
    for g in GROUP_SORT_ORDER:
        if g not in by_group:
            continue
        d = by_group[g]
        fill = GROUP_FILL.get(g, GROUP_FILL["Dormant / out-of-scope"])
        vals = [g, d["n_ds"], d["n_active"], d["n_visible"]]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_idx, value=v)
            cell.fill = fill
            if col_idx in (2, 3, 4):
                cell.number_format = "#,##0"
        row_i += 1

    for col_idx, w in enumerate([28, 18, 18, 22], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A5"


def write_pass_sheet(wb: Workbook) -> None:
    if not PASS_CSV.exists():
        return
    ws = wb.create_sheet(f"{PASS_LABEL} buckets")

    ws.cell(row=1, column=1, value=f"{PASS_LABEL} audience-bucket results (30d prospecting, 2026-04-29 to 2026-05-28)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    ws.cell(row=2, column=1, value=PASS_NOTE).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 32

    headers = ["Bucket", "n_campaigns", "% campaigns", "n_advertisers", "Spend (30d, $M)", "% spend", "Annualized ($M)"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    # Highlight 3P-touching buckets only (the focus of TI-999's curation case)
    fill_3p = PatternFill("solid", fgColor="F3E5F5")

    with PASS_CSV.open() as f:
        reader = csv.DictReader(f)
        row_i = 5
        for r in reader:
            is_3p_bucket = "3P" in r["bucket"]
            fill = fill_3p if is_3p_bucket else None
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_i, column=col_idx)
                if fill:
                    cell.fill = fill
            ws.cell(row=row_i, column=1, value=r["bucket"])
            ws.cell(row=row_i, column=2, value=to_int(r["n_campaigns"])).number_format = "#,##0"
            ws.cell(row=row_i, column=3, value=to_float(r["pct_campaigns"])).number_format = "0.0"
            ws.cell(row=row_i, column=4, value=to_int(r["n_advertisers"])).number_format = "#,##0"
            ws.cell(row=row_i, column=5, value=to_float(r["spend_30d_M"])).number_format = '"$"#,##0.000'
            ws.cell(row=row_i, column=6, value=to_float(r["pct_spend"])).number_format = "0.0"
            ws.cell(row=row_i, column=7, value=to_float(r["spend_annualized_M"])).number_format = '"$"#,##0.0'
            if fill:
                for col_idx in range(1, 8):
                    ws.cell(row=row_i, column=col_idx).fill = fill
            row_i += 1

    for col_idx, w in enumerate([36, 14, 14, 16, 18, 12, 18], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A5"


def write_anomalies_sheet(wb: Workbook) -> None:
    if not ANOMALIES_CSV.exists():
        return
    ws = wb.create_sheet("RTC anomalies")

    title = "Pass 20 anomalies — 16 prospecting campaigns with no score_type=rtc (concentrated in 3 advertisers)"
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    note = (
        "99.9% of prospecting expressions have score_type=rtc auto-attached. The 16 below are exceptions, "
        "concentrated in AID 36678 (9 campaigns), AID 37336 (6), AID 42097 (1). "
        "Both 36678 and 37336 are heavy MNTN Select household users (DS9). "
        "MM-without-RTC = buyer added DS19 batch keywords but no RTC flag. "
        "no-RTC no-MM = custom audience targeting via DS2/DS9/DS8 only. "
        "Likely deliberately bypassing RTC default via API or custom tooling — RTC isn't UI-selectable. "
        "AUD team to confirm the opt-out mechanism."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 72

    headers = ["Anomaly class", "Campaign ID", "Audience Segment ID", "Advertiser ID", "Spend 30d ($)", "Has RTC", "Has MM", "DS refs"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 28

    mm_fill = PatternFill("solid", fgColor="FFE0B2")
    nomm_fill = PatternFill("solid", fgColor="F8BBD0")

    with ANOMALIES_CSV.open() as f:
        reader = csv.DictReader(f)
        row_i = 5
        for r in reader:
            fill = mm_fill if r["anomaly_class"] == "ANOMALY_MM_without_RTC" else nomm_fill
            ws.cell(row=row_i, column=1, value=r["anomaly_class"]).fill = fill
            ws.cell(row=row_i, column=2, value=to_int(r["campaign_id"])).fill = fill
            ws.cell(row=row_i, column=2).number_format = "#,##0"
            ws.cell(row=row_i, column=3, value=to_int(r["audience_segment_id"])).fill = fill
            ws.cell(row=row_i, column=3).number_format = "#,##0"
            ws.cell(row=row_i, column=4, value=to_int(r["advertiser_id"])).fill = fill
            ws.cell(row=row_i, column=4).number_format = "#,##0"
            ws.cell(row=row_i, column=5, value=to_float(r["spend_30d_K"]) * 1000).fill = fill
            ws.cell(row=row_i, column=5).number_format = '"$"#,##0'
            ws.cell(row=row_i, column=6, value=r["has_rtc"]).fill = fill
            ws.cell(row=row_i, column=7, value=r["has_mm"]).fill = fill
            ws.cell(row=row_i, column=8, value=r["ds_refs"]).fill = fill
            ws.cell(row=row_i, column=8).alignment = Alignment(wrap_text=True)
            row_i += 1

    for col_idx, w in enumerate([28, 14, 18, 14, 14, 10, 10, 56], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A5"


def write_polarity_kpi_sheet(wb: Workbook) -> None:
    if not POLARITY_KPI_CSV.exists():
        return
    ws = wb.create_sheet("Polarity KPIs")

    ws.cell(row=1, column=1, value="Pass 22c — polarity-aware bucket KPIs (ratios)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    note = (
        "Pass 21 buckets BUT 3P and CRM are split by polarity so MM combos are properly disambiguated. "
        "MM + CRM-incl = customer-list-seeded MM scoring (adds non-MM-scored customer IPs as positive layer). "
        "MM + CRM-excl = MM scoring with customer suppression (drilling down on MM-scored audience). "
        "All KPIs are ratios: CVR/IVR/CTR as percentages, CPM and cost-per-conv in dollars. "
        "3P-only (no MM, no CRM, no Select) is the cleanest baseline for measuring 3P quality interventions."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 72

    headers = [
        "Bucket", "n_campaigns", "n_advertisers", "Spend (30d, $M)", "% spend",
        "CVR (ratio)", "IVR (ratio)", "CTR (ratio)", "CPM ($)", "Cost/conv ($)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    # Minimal coloring — only the rows that matter for the deck story
    fill_baseline = PatternFill("solid", fgColor="F3E5F5")    # 3P-incl alone = clean baseline for 3P quality work
    fill_high_cvr = PatternFill("solid", fgColor="FFF9C4")    # CRM-incl combos = high CVR cohort
    fill_default = None

    with POLARITY_KPI_CSV.open() as f:
        reader = csv.DictReader(f)
        row_i = 5
        for r in reader:
            bucket = r["bucket"]
            if bucket == "3P-incl":
                fill = fill_baseline
            elif "CRM-incl" in bucket and "CRM-excl" not in bucket:
                fill = fill_high_cvr
            else:
                fill = fill_default
            def apply_fill(col, value, fmt=None):
                cell = ws.cell(row=row_i, column=col, value=value)
                if fill:
                    cell.fill = fill
                if fmt:
                    cell.number_format = fmt
                return cell
            apply_fill(1, bucket)
            apply_fill(2, to_int(r["n_campaigns"]), "#,##0")
            apply_fill(3, to_int(r["n_advertisers"]), "#,##0")
            apply_fill(4, to_float(r["spend_30d_M"]), '"$"#,##0.000')
            apply_fill(5, to_float(r["pct_spend"]), "0.0")
            apply_fill(6, to_float(r["cvr"]), "0.000000")
            apply_fill(7, to_float(r["ivr"]), "0.000000")
            apply_fill(8, to_float(r["ctr"]), "0.000000")
            apply_fill(9, to_float(r["cpm_dollars"]), '"$"#,##0.00')
            cpc_val = r.get("cost_per_conv_dollars", "")
            if cpc_val and cpc_val.strip():
                apply_fill(10, to_float(cpc_val), '"$"#,##0.00')
            else:
                apply_fill(10, "")
            row_i += 1

    for col_idx, w in enumerate([44, 14, 14, 16, 12, 12, 12, 12, 12, 16], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "B5"


def write_geo_restriction_sheet(wb: Workbook) -> None:
    if not GEO_RESTRICTION_CSV.exists():
        return
    ws = wb.create_sheet("Pass 24 — Geo restriction")

    title = "Pass 24 — bucket KPIs split by geo restriction (Alyson + Toph ask, 2026-06-01)"
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    note = (
        "geo_restricted = expression has explicit location_ids[] in the geos clause (buyer-picked specific locations). "
        "geo_broad_or_default = expression has a geos clause but no explicit location_ids[] (likely country-level / wildcard). "
        "Toph's hypothesis: MM+3P+geo-restriction audiences perform poorly. Empirically: this IS the biggest single audience-targeted cohort ($5.62M / 17.6% of prospecting spend), but its cost-per-conv ($65.54) is actually BETTER than MM-only with geo restriction ($93.53). Geo restriction itself is the bigger drag — geo-broad versions consistently outperform geo-restricted across MM, MM+3P, and Geo-only. TI-956 curation lands at maximum leverage within this cohort."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws.row_dimensions[2].height = 90

    headers = [
        "Bucket", "Geo status", "n_campaigns", "n_advertisers", "Spend (30d, $M)", "% total spend",
        "CVR (ratio)", "IVR (ratio)", "CTR (ratio)", "CPM ($)", "Cost/conv ($)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    fill_largest = PatternFill("solid", fgColor="FFF9C4")  # the headline MM+3P+geo-restricted row
    fill_default = None

    with GEO_RESTRICTION_CSV.open() as f:
        reader = csv.DictReader(f)
        row_i = 5
        for r in reader:
            bucket = r["bucket"]
            geo_status = r["geo_status"]
            is_headline = bucket == "MM + 3P-incl" and geo_status == "geo_restricted"
            fill = fill_largest if is_headline else fill_default

            def apply(col, value, fmt=None):
                cell = ws.cell(row=row_i, column=col, value=value)
                if fill:
                    cell.fill = fill
                if fmt:
                    cell.number_format = fmt
                return cell
            apply(1, bucket)
            apply(2, geo_status)
            apply(3, to_int(r["n_campaigns"]), "#,##0")
            apply(4, to_int(r["n_advertisers"]), "#,##0")
            apply(5, to_float(r["spend_30d_M"]), '"$"#,##0.000')
            apply(6, to_float(r["pct_total_spend"]), "0.0")
            apply(7, to_float(r["cvr"]), "0.000000")
            apply(8, to_float(r["ivr"]), "0.000000")
            apply(9, to_float(r["ctr"]), "0.000000")
            apply(10, to_float(r["cpm_dollars"]), '"$"#,##0.00')
            cpc = r.get("cost_per_conv_dollars", "")
            if cpc and cpc.strip():
                apply(11, to_float(cpc), '"$"#,##0.00')
            else:
                apply(11, "")
            row_i += 1

    for col_idx, w in enumerate([34, 22, 14, 16, 18, 14, 14, 14, 14, 12, 16], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "C5"


def write_excl_axes_sheet(wb: Workbook) -> None:
    if not EXCL_AXES_CSV.exists():
        return
    ws = wb.create_sheet("Pass 25 — full polarity")

    title = "Pass 25 — full polarity split with plain-English cell labels (Alyson refinement, 2026-06-01)"
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)

    note = (
        "6 binary axes: MM × 3P-include × 3P-exclude × CRM-include × CRM-exclude × geo-restricted. "
        "Up to 64 possible cells; only cells with ≥5 campaigns shown (34 total). Each cell carries a plain-English label in column A describing exactly what the campaign is doing. "
        "Polarity check (verified 2026-06-01): excludes are wrapped in op:not in the expression JSON — confirmed by sampling actual expressions. "
        "Yellow row = largest single audience-targeted cell by spend. Red rows = cells with 3P-exclude (the destroyer pattern — removes IPs from MM and consistently hurts CVR). All KPIs are ratios (CVR/IVR/CTR) or dollars (CPM/cost-per-conv)."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=18)
    ws.row_dimensions[2].height = 92

    headers = [
        "What this cell says (plain English)",
        "MM", "3P-incl", "3P-excl", "CRM-incl", "CRM-excl", "Geo",
        "n_campaigns", "% campaigns", "n_advertisers", "% advertisers", "Spend (30d, $M)", "% spend",
        "CVR (ratio)", "IVR (ratio)", "CTR (ratio)", "CPM ($)", "Cost/conv ($)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 36

    fill_biggest = PatternFill("solid", fgColor="FFF9C4")      # largest cell by spend
    fill_destroyer = PatternFill("solid", fgColor="FFCDD2")    # 3P-exclude rows
    fill_default = None

    with EXCL_AXES_CSV.open() as f:
        reader = csv.DictReader(f)
        row_i = 5
        for idx, r in enumerate(reader):
            is_biggest = (idx == 0)  # CSV is sorted desc by spend
            has_3p_excl = (r["f_3p_excl"] == "3P-excl")
            if is_biggest:
                fill = fill_biggest
            elif has_3p_excl:
                fill = fill_destroyer
            else:
                fill = fill_default

            def apply(col, value, fmt=None):
                cell = ws.cell(row=row_i, column=col, value=value)
                if fill:
                    cell.fill = fill
                if fmt:
                    cell.number_format = fmt
                cell.alignment = Alignment(vertical="top", wrap_text=(col == 1))
                return cell
            apply(1, r["plain_english"])
            apply(2, r["f_mm"])
            apply(3, r["f_3p_incl"])
            apply(4, r["f_3p_excl"])
            apply(5, r["f_crm_incl"])
            apply(6, r["f_crm_excl"])
            apply(7, r["f_geo"])
            apply(8, to_int(r["n_campaigns"]), "#,##0")
            apply(9, to_float(r["pct_campaigns"]), "0.0")
            apply(10, to_int(r["n_advertisers"]), "#,##0")
            apply(11, to_float(r["pct_advertisers"]), "0.0")
            apply(12, to_float(r["spend_30d_M"]), '"$"#,##0.000')
            apply(13, to_float(r["pct_spend"]), "0.0")
            apply(14, to_float(r["cvr"]), "0.000000")
            apply(15, to_float(r["ivr"]), "0.000000")
            apply(16, to_float(r["ctr"]), "0.000000")
            apply(17, to_float(r["cpm_dollars"]), '"$"#,##0.00')
            cpc = r.get("cost_per_conv_dollars", "")
            if cpc and cpc.strip():
                apply(18, to_float(cpc), '"$"#,##0.00')
            else:
                apply(18, "")
            row_i += 1

    widths = [78, 8, 10, 10, 12, 12, 16, 12, 12, 14, 14, 16, 12, 14, 14, 14, 12, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "B5"


def write_or_vs_and_explainer_sheet(wb: Workbook) -> None:
    """The three-tables-stacked explainer the user asked for."""
    ws = wb.create_sheet("OR vs AND — explainer")

    # Title
    ws.cell(row=1, column=1, value="OR-include vs AND-include vs AND-exclude — layman's terms").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Setup paragraph
    note = (
        "Concrete example: buyer is running an MM (Mountain Match) campaign for 'auto-intent users' and wants to layer the Ford F-150 Intender 3P segment. "
        "Two circles: A = MM audience (everyone MNTN scores as auto-intent), B = 3P segment (everyone in Ford F-150 Intender list). "
        "There are four ways to combine them in the audience expression. Three of them are common in our data."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 56

    # ---- TABLE 1: The four patterns ----
    ws.cell(row=4, column=1, value="Table 1 — What each pattern means").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)

    t1_headers = ["Pattern", "Plain English", "Bidder math", "Audience size effect", "Example", "Common in our data?"]
    for col_idx, h in enumerate(t1_headers, start=1):
        c = ws.cell(row=5, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 30

    t1_rows = [
        ["AND-include", "Only target people who are in BOTH circles", "A ∩ B (intersection)", "Smaller — narrows to overlap only", "Target auto-intent users who ARE ALSO Ford F-150 Intenders", "5% of MM+3P-incl spend"],
        ["OR-include", "Target circle A, AND ALSO add circle B", "A ∪ B (union)", "Bigger in theory — both circles combined", "Target auto-intent users, AND ALSO add Ford F-150 Intenders to the audience", "80% of MM+3P-incl spend"],
        ["AND-exclude", "Target circle A, AND remove circle B", "A \\ B (subtraction)", "Smaller — removes overlap from A", "Target auto-intent users, EXCEPT those who are also Ford F-150 Intenders", "Yes (3P-exclude column)"],
        ["OR-exclude", "(rare / structurally weird)", "NOT(A ∪ B) ≈ neither", "Almost nobody", "Doesn't really show up in our data", "No"],
    ]
    for r_idx, row in enumerate(t1_rows, start=6):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r_idx].height = 40

    # ---- TABLE 2: Buyer thinks vs Bidder reality ----
    ws.cell(row=12, column=1, value="Table 2 — What buyer thinks vs what actually happens at the bidder (HHST > 0)").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=12, start_column=1, end_row=12, end_column=6)

    note2 = (
        "Critical context: the bidder needs a score on an IP to actually bid (when HHST > 0). Only people in MM (DS13/19/38/46) get scored. "
        "So even if the expression LISTS the 3P-only IPs as eligible, the bidder skips them because they have no score."
    )
    ws.cell(row=13, column=1, value=note2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=6)
    ws.row_dimensions[13].height = 40

    t2_headers = ["Pattern", "What buyer THINKS happens", "What actually happens at the bidder", "Match?"]
    for col_idx, h in enumerate(t2_headers, start=1):
        c = ws.cell(row=14, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[14].height = 30

    t2_rows = [
        ["AND-include", "I narrowed to auto-intent + Ford intenders", "Correct. Bidder bids on the intersection (MM ∩ 3P). Real narrowing.", "✓ Matches"],
        ["OR-include", "I expanded my audience to include 3P", "Wrong. Bidder only bids on MM-scored IPs. The 3P-only people aren't scored, fail HHST. Audience-size theater.", "✗ Mismatch — THEATER"],
        ["AND-exclude", "I removed Ford intenders from my MM", "Correct. Bidder bids on (MM minus Ford). Real subtraction.", "✓ Matches"],
        ["OR-exclude", "(n/a)", "(n/a — doesn't happen in our data)", "—"],
    ]
    for r_idx, row in enumerate(t2_rows, start=15):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Highlight the mismatch row
            if r_idx == 16:  # OR-include row
                cell.fill = PatternFill("solid", fgColor="FFE0B2")
        ws.row_dimensions[r_idx].height = 50

    # ---- TABLE 3: Portfolio totals ----
    ws.cell(row=21, column=1, value="Table 3 — What our portfolio actually does (% of MM+3P-include spend)").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=21, start_column=1, end_row=21, end_column=6)

    t3_headers = ["Pattern", "% of MM+3P-include spend", "What it really is"]
    for col_idx, h in enumerate(t3_headers, start=1):
        c = ws.cell(row=22, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[22].height = 30

    t3_rows = [
        ["OR-include", "80%", "Audience-size theater — buyer thought they were expanding, but with HHST > 0 the 3P clause changes nothing"],
        ["AND-include", "5%", "Real narrowing — bidder genuinely bids only on MM ∩ 3P"],
        ["Mixed", "8%", "Some 3P clauses are OR, some are AND in the same expression"],
        ["3P-include WITHOUT MM (separate cohort)", "11.8% of all prospecting spend", "Pure 3P prospecting — no MM scoring at all"],
    ]
    for r_idx, row in enumerate(t3_rows, start=23):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r_idx == 23:  # OR-include row — the headline finding
                cell.fill = PatternFill("solid", fgColor="FFE0B2")
        ws.row_dimensions[r_idx].height = 40

    # ---- Spend semantics clarification ----
    ws.cell(row=28, column=1, value="Important: spend is NOT driven by audience size").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=28, start_column=1, end_row=28, end_column=6)

    spend_note = (
        "Advertisers are charged only when MNTN bids on and WINS an impression. Adding a 3P clause to MM doesn't 'cost more' — "
        "same spend, same delivery, just a UI label that misrepresents what was targeted. So 'audience-size theater' is about "
        "decoupled targeting intent (buyer believes they're targeting MM+3P, but mechanically gets MM-only delivery), NOT about wasted dollars."
    )
    ws.cell(row=29, column=1, value=spend_note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=29, start_column=1, end_row=29, end_column=6)
    ws.row_dimensions[29].height = 56

    # ---- The deck-headline callout ----
    ws.cell(row=31, column=1, value="THE LAYMAN HEADLINE FOR THE DECK").font = Font(bold=True, size=13, color="C62828")
    ws.merge_cells(start_row=31, start_column=1, end_row=31, end_column=6)

    headline = (
        "When buyers layer interest segments on Mountain Match, 80% of the spend on those combos delivers as if the 3P clause "
        "wasn't there. Same impressions, same costs — but the buyer believes they're targeting MM+3P when the bidder is only "
        "bidding on MM-scored people (3P-only people don't have scores). Only 5% of the spend is doing what the buyer thought "
        "(real narrowing of MM by the 3P segment). The rest is decoupled targeting intent — a UI vs reality mismatch, not wasted budget."
    )
    ws.cell(row=32, column=1, value=headline).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=32, start_column=1, end_row=32, end_column=6)
    ws.row_dimensions[32].height = 90
    ws.cell(row=32, column=1).fill = PatternFill("solid", fgColor="FFF9C4")

    # ---- TI-956 curation argument refined ----
    ws.cell(row=34, column=1, value="TI-956 curation argument (refined for OR vs AND cohorts)").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=34, start_column=1, end_row=34, end_column=6)

    ti956_note = (
        "For AND-include (5% of MM+3P-incl spend): segment quality determines delivery quality → curation has real lift on KPIs (CVR, cost/conv). "
        "For OR-include (80%): segment quality doesn't affect delivery, but curation prevents buyers from believing they're targeting low-quality "
        "segments when they're actually getting MM-only delivery → curation has attribution / UI honesty value. Both warrant the TI-956 build; the framing is different per cohort."
    )
    ws.cell(row=35, column=1, value=ti956_note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=35, start_column=1, end_row=35, end_column=6)
    ws.row_dimensions[35].height = 70

    # Column widths
    widths = [20, 36, 28, 32, 50, 26]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def write_pass26_detail_sheet(wb: Workbook) -> None:
    """Full Pass 26 cell-level detail with OR/AND split applied."""
    if not PASS26_CSV.exists():
        return
    ws = wb.create_sheet("Pass 26 — OR vs AND detail")

    ws.cell(row=1, column=1, value="Pass 26 — full polarity split WITH OR-include vs AND-include semantic classification").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)

    note = (
        "Same axes as Pass 25 plus the OR vs AND classification for 3P-include. "
        "Classification logic: for each MM-positive clause and 3P-positive clause, walk the JSON tree to find the lowest common ancestor (LCA) operator. "
        "LCA = 'or' → OR_include (additive, theater with HHST > 0). LCA = 'and' → AND_include (intersect, real narrowing). "
        "Mixed = some pairs OR, some AND in the same expression. Yellow rows = OR_include cells. Pink rows = AND_include cells."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=18)
    ws.row_dimensions[2].height = 80

    headers = [
        "What this cell says (plain English)",
        "MM", "3P semantics", "3P-excl", "CRM-incl", "CRM-excl", "Geo",
        "n_campaigns", "% campaigns", "n_advertisers", "% advertisers", "Spend (30d, $M)", "% spend",
        "CVR (ratio)", "IVR (ratio)", "CTR (ratio)", "CPM ($)", "Cost/conv ($)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 36

    fill_or = PatternFill("solid", fgColor="FFF9C4")    # OR_include cells (theater)
    fill_and = PatternFill("solid", fgColor="F8BBD0")    # AND_include cells (real narrowing)
    fill_mixed = PatternFill("solid", fgColor="FFE0B2")  # mixed
    fill_default = None

    with PASS26_CSV.open() as f:
        reader = csv.DictReader(f)
        row_i = 5
        for r in reader:
            sem = r["f_3p_semantics"]
            if sem == "OR_include":
                fill = fill_or
            elif sem == "AND_include":
                fill = fill_and
            elif sem == "mixed":
                fill = fill_mixed
            else:
                fill = fill_default

            def apply(col, value, fmt=None):
                cell = ws.cell(row=row_i, column=col, value=value)
                if fill:
                    cell.fill = fill
                if fmt:
                    cell.number_format = fmt
                cell.alignment = Alignment(vertical="top", wrap_text=(col == 1))
                return cell
            apply(1, r["plain_english"])
            apply(2, r["f_mm"])
            apply(3, r["f_3p_semantics"])
            apply(4, r["f_3p_excl"])
            apply(5, r["f_crm_incl"])
            apply(6, r["f_crm_excl"])
            apply(7, r["f_geo"])
            apply(8, to_int(r["n_campaigns"]), "#,##0")
            apply(9, to_float(r["pct_campaigns"]), "0.0")
            apply(10, to_int(r["n_advertisers"]), "#,##0")
            apply(11, to_float(r["pct_advertisers"]), "0.0")
            apply(12, to_float(r["spend_30d_M"]), '"$"#,##0.000')
            apply(13, to_float(r["pct_spend"]), "0.0")
            apply(14, to_float(r["cvr"]), "0.000000")
            apply(15, to_float(r["ivr"]), "0.000000")
            apply(16, to_float(r["ctr"]), "0.000000")
            apply(17, to_float(r["cpm_dollars"]), '"$"#,##0.00')
            cpc = r.get("cost_per_conv_dollars", "")
            if cpc and cpc.strip():
                apply(18, to_float(cpc), '"$"#,##0.00')
            else:
                apply(18, "")
            row_i += 1

    widths = [78, 8, 22, 10, 12, 12, 16, 12, 12, 14, 14, 16, 12, 14, 14, 14, 12, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "B5"


def write_pass27_mm_and_or_sheet(wb: Workbook) -> None:
    """Pass 27 — MM-touching campaigns split by AND-include / AND-exclude / OR-include-only.

    Per user direction (2026-06-01): on an MM campaign, separate out AND-include and
    AND-exclude patterns regardless of source DS (CRM / 3P / Select / whatever).
    OR-include only gets its own bucket when there are no AND patterns at all,
    because under HHST > 0 it's bidder-inert and doesn't change delivery.
    """
    if not PASS27_CSV.exists():
        return
    ws = wb.create_sheet("Pass 27 — MM AND OR split")

    ws.cell(row=1, column=1, value="Pass 27 — MM campaigns split by AND-include / AND-exclude / OR-include").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    note = (
        "MM-touching campaigns bucketed by the structural pattern of any non-MM clauses: "
        "AND-include = positive non-MM clause AND-connected with MM (narrows MM to ∩ that segment). "
        "AND-exclude = any negative non-MM clause (always AND-wrapped via op:not — removes that segment from MM). "
        "OR-include = positive non-MM clause OR-connected with MM (additive in theory, but bidder-inert under HHST > 0 because non-MM IPs have no score). "
        "Per user direction, OR-include is treated as fluff when AND-include or AND-exclude is also present — it gets its own bucket only when no AND patterns exist. "
        "DS source (3P / CRM / Select) does not affect the bucketing here — what matters is the connector and polarity."
    )
    ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.row_dimensions[2].height = 110

    headers = [
        "Pattern", "Plain English",
        "n_campaigns", "% campaigns", "n_advertisers", "% advertisers",
        "Spend (30d, $M)", "% spend",
        "CVR (ratio)", "IVR (ratio)", "CTR (ratio)", "CPM ($)", "Cost/conv ($)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 36

    pattern_fill = {
        "MM_only":                     PatternFill("solid", fgColor="DCEFFC"),  # baseline MM blue
        "MM_OR_include_only":          PatternFill("solid", fgColor="FFF9C4"),  # theater yellow
        "MM_AND_include_only":         PatternFill("solid", fgColor="F8BBD0"),  # real narrowing pink
        "MM_AND_exclude_only":         PatternFill("solid", fgColor="FFE0B2"),  # exclusion orange
        "MM_AND_include_AND_exclude":  PatternFill("solid", fgColor="E1BEE7"),  # both purple
        "no_MM":                       PatternFill("solid", fgColor="F5F5F5"),  # context gray
    }

    with PASS27_CSV.open() as f:
        reader = csv.DictReader(f)
        row_i = 5
        for r in reader:
            pat = r["pattern"]
            fill = pattern_fill.get(pat)

            def apply(col, value, fmt=None):
                cell = ws.cell(row=row_i, column=col, value=value)
                if fill:
                    cell.fill = fill
                if fmt:
                    cell.number_format = fmt
                cell.alignment = Alignment(vertical="top", wrap_text=(col == 2))
                return cell

            apply(1, pat)
            apply(2, r["plain_english"])
            apply(3, to_int(r["n_campaigns"]), "#,##0")
            apply(4, to_float(r["pct_campaigns"]), "0.0")
            apply(5, to_int(r["n_advertisers"]), "#,##0")
            apply(6, to_float(r["pct_advertisers"]), "0.0")
            apply(7, to_float(r["spend_30d_M"]), '"$"#,##0.000')
            apply(8, to_float(r["pct_spend"]), "0.0")
            apply(9, to_float(r["cvr"]), "0.000000")
            apply(10, to_float(r["ivr"]), "0.000000")
            apply(11, to_float(r["ctr"]), "0.000000")
            apply(12, to_float(r["cpm_dollars"]), '"$"#,##0.00')
            cpc = r.get("cost_per_conv_dollars", "")
            if cpc and cpc.strip():
                apply(13, to_float(cpc), '"$"#,##0.00')
            else:
                apply(13, "")
            row_i += 1

    # Headline takeaway under the table
    ws.cell(row=row_i + 1, column=1, value="Headline").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row_i + 1, start_column=1, end_row=row_i + 1, end_column=12)
    headline = (
        "Of all MM-touching campaigns: MM-only (pure scoring) and MM + OR-include-only (theater) together are 18% of campaigns and 36% of spend "
        "— bidder behavior is identical between them. MM + AND-include (real narrowing) is a small slice (1.2% of campaigns, 3.7% of spend) and shows "
        "very low CVR — when buyers genuinely intersect MM with a non-MM segment, KPIs drop. MM + AND-exclude is the bigger real-narrowing pattern "
        "(7% of campaigns, 27% of spend) and is mainly CRM suppression. Both AND patterns combined still trail MM-only and theater by spend."
    )
    ws.cell(row=row_i + 2, column=1, value=headline).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row_i + 2, start_column=1, end_row=row_i + 2, end_column=12)
    ws.row_dimensions[row_i + 2].height = 100
    ws.cell(row=row_i + 2, column=1).fill = PatternFill("solid", fgColor="FFF9C4")

    widths = [30, 70, 14, 12, 14, 14, 16, 12, 14, 14, 14, 12, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "C5"


def main() -> None:
    rows = load_per_ds_rows()
    wb = Workbook()
    write_taxonomy_sheet(wb, rows)
    write_group_summary_sheet(wb, rows)
    write_pass_sheet(wb)
    write_polarity_kpi_sheet(wb)
    write_geo_restriction_sheet(wb)
    write_excl_axes_sheet(wb)
    write_or_vs_and_explainer_sheet(wb)
    write_pass26_detail_sheet(wb)
    write_pass27_mm_and_or_sheet(wb)
    write_anomalies_sheet(wb)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({OUT_XLSX.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
