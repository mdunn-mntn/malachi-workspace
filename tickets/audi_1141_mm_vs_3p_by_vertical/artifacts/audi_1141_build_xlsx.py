#!/usr/bin/env python3
"""AUDI-1141 shareable .xlsx. Tabs: Read me / MM vs 3P by vertical / Full scorecard / Overall /
Campaign detail / Queries. Rates stored as decimals, formatted as % and $. Content-sized columns."""
import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "tickets/audi_1141_mm_vs_3p_by_vertical/outputs/"
DEST = OUT + "audi_1141_mm_vs_3p_scorecard.xlsx"
SQLFILE = "tickets/audi_1141_mm_vs_3p_by_vertical/queries/audi_1141_cohort_scorecard.sql"
ADV_MIN_IMPS = 20000
BO = ["MM (gated)", "MM (no gate)", "MM restricted", "3P"]

HEAD="1a3c5e"; BAND="eef3f4"; GREY="666666"
HFONT=Font(bold=True,color="FFFFFF",size=11); TITLE=Font(bold=True,size=15,color=HEAD)
SUB=Font(italic=True,size=10,color=GREY); BOLD=Font(bold=True); MONO=Font(name="Consolas",size=9)
HFILL=PatternFill("solid",fgColor=HEAD); BANDF=PatternFill("solid",fgColor=BAND)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LFT=Alignment(horizontal="left",vertical="top",wrap_text=True)
THIN=Side(style="thin",color="CCCCCC"); BORD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
PCT2="0.00%"; PCT3="0.000%"; USD="\"$\"#,##0.00"; USD0="\"$\"#,##0"; NUM2="0.00"; INT="#,##0"

wb=Workbook()

def autosize(ws, df, wide=None):
    """Width fits the header's longest word (so multi-word headers wrap cleanly and single words
    like 'Advertisers' never clip). wide = {col: value_display_len} for columns whose DATA is
    longer than any header word (e.g. Spend, company names)."""
    wide = wide or {}
    for j,col in enumerate(df.columns,1):
        words=[len(w) for w in str(col).split()] or [len(str(col))]
        w=max(max(words), int(wide.get(col,0)))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w+3, 10), 46)

def write_table(ws, df, start_row, fmts, band=True, header_wrap=True):
    for j,col in enumerate(df.columns,1):
        c=ws.cell(row=start_row,column=j,value=col); c.font=HFONT; c.fill=HFILL
        c.alignment=CEN if header_wrap else Alignment(horizontal="center"); c.border=BORD
    for i,(_,r) in enumerate(df.iterrows(),1):
        rr=start_row+i
        for j,col in enumerate(df.columns,1):
            v=r[col]
            if isinstance(v,(np.integer,)): v=int(v)
            elif isinstance(v,(np.floating,)): v=float(v) if pd.notna(v) else None
            elif isinstance(v,(np.bool_,)): v=bool(v)
            c=ws.cell(row=rr,column=j,value=v); c.border=BORD
            if col in fmts: c.number_format=fmts[col]
            if j==1: c.font=BOLD
            if band and i%2==0: c.fill=BANDF
    ws.row_dimensions[start_row].height=30
    return start_row+len(df)+1

# ---------------- load + transform ----------------
df=pd.read_csv(OUT+"audi_1141_campaign_grain.csv").dropna(subset=["vertical_id"])
names=pd.read_csv(OUT+"audi_1141_advertiser_names.csv").drop_duplicates("advertiser_id")
df=df.merge(names,on="advertiser_id",how="left")
df["gated_frac"]=np.where(df.hhst_writes>0, df.hhst_writes_gated/df.hhst_writes, 0.0)
df["capped"]=df.gated_frac>=0.5
df=df[df.bucket!="Neither"].copy()
df["bucket_detail"]=np.where(df.bucket=="MM", np.where(df.capped,"MM (gated)","MM (no gate)"), df.bucket)

ov=pd.read_csv(OUT+"audi_1141_scorecard_overall.csv")
bv=pd.read_csv(OUT+"audi_1141_scorecard_by_vertical.csv")
ov2=pd.read_csv(OUT+"audi_1141_scorecard2_overall.csv")
bv2=pd.read_csv(OUT+"audi_1141_scorecard2_by_vertical.csv")
def _int0(x): return int(x) if pd.notna(x) else 0

# ================= 1) READ ME =================
ws=wb.active; ws.title="Read me"; ws.sheet_view.showGridLines=False
ws["A1"]="MNTN Matched vs 3P Segments: Prospecting Performance by Vertical"; ws["A1"].font=TITLE
ws["A2"]="Trailing 6 months. Stage-1 prospecting campaigns. Source AUDI-1141. Generated 2026-07-20."; ws["A2"].font=SUB
rows=[
 ("",""),
 ("Summary","Across all MM campaigns, MNTN Matched delivers about 4x the visit rate of 3P segments at about one third the cost per visit and roughly double the ROAS, for the typical (median) advertiser, and leads visit rate in every vertical. When MM is configured well (intent gate on, audience not over-narrowed) the gap widens to about 6x on visit rate."),
 ("",""),
 ("Two ways to read this",""),
 ("MM (all) vs 3P","The realistic average: every MM campaign vs 3P, including mis-configured ones. Use this for the general 'MM vs 3P' comparison."),
 ("MM (gated) vs 3P","The best case: MM campaigns where the intent gate is on and the audience is broad. Use this to show what MM does when configured correctly."),
 ("",""),
 ("How campaigns are grouped","Each Stage-1 prospecting campaign is classified from its live bidder audience expression."),
 ("MM (gated)","MNTN scoring (DS13/19/38/46) with the intent score threshold above 0, so the bidder only bids on IPs the model scored as high intent. 3P (if present) is joined by OR (adds reach), and the geo is broad."),
 ("MM (no gate)","MM with the intent score threshold at 0. This bypasses the model and bids broadly, similar to a 3P segment. The threshold drops to 0 for a few reasons: Max Reach, many short flights (lowered to keep delivery flowing), or an audience narrowed so far that the high-intent pool is exhausted and the score is lowered for deliverability. Note: the gate is the score threshold setting, not the scoring model itself."),
 ("MM restricted","MM whose audience is narrowed by an AND-required 3P clause, or by a sub-DMA geo (zip, city, or radius)."),
 ("3P","Bought interest segments only (ShareThis, Dstillery, LiveRamp), no MM signal."),
 ("MM (all)","Every campaign with an MM signal: MM (gated), MM (no gate), and MM restricted combined."),
 ("Why OR vs AND matters","About 85% of campaigns that carry a 3P segment join it with OR, which only adds reach and leaves MM doing the scoring. Only an AND-required 3P actually narrows the audience. Grouping all 3P-carrying campaigns together would misread additive 3P as restrictive."),
 ("",""),
 ("Metrics","All rates below are over impressions."),
 ("IVR","Visit rate. Visits divided by impressions (visits = views + clicks)."),
 ("CVR","Conversion rate. Conversions divided by impressions."),
 ("CTR","Click rate. Clicks divided by impressions."),
 ("CPV","Cost per visit. Spend divided by visits."),
 ("CPM","Cost per thousand impressions."),
 ("ROAS","Revenue divided by spend."),
 ("Weighting","Headline numbers are advertiser-weighted medians (each advertiser counts once, so one large account cannot set the result). Columns labeled 'pooled' are impression or spend weighted. Advertisers with under 20,000 impressions in a group are excluded from the median so a tiny campaign cannot swing it."),
 ("Advertisers / Campaigns","Counts of advertisers and campaigns in each group. Use them to see how much a number rests on: a small group is less reliable."),
 ("",""),
 ("Read before quoting",""),
 ("1. Pooled 3P is one account","If all impressions are pooled, 3P visit rate looks competitive, but roughly 39% of 3P impressions come from a single large, non-representative account. Use the per-advertiser (median) numbers."),
 ("2. ROAS is directional","Prospecting only, last-touch. Revenue mostly lands in retargeting, which is excluded, and some verticals have unreliable conversion pixels (one account shows over 800x). ROAS is computed only over advertisers with revenue, so where few have it the number is rough. Visit rate and cost per visit are the solid metrics."),
 ("3. Vertical mapping","The 37 MNTN internal verticals are rolled up into the 8 sales verticals using an interim crosswalk, not an official one. Provide the RevOps crosswalk and it is a one-line swap."),
 ("4. Restricted is expected for local","Auto and ProServ legitimately use local (zip, radius) targeting, so their 'MM restricted' share is high by design. In other verticals it flags narrowing."),
 ("5. MM (no gate) is small","144 advertisers. Directionally clear, thinner than the other groups."),
 ("",""),
 ("Tabs","'MM vs 3P by vertical' is the headline. 'Full scorecard' has every group by vertical. 'Overall' is all verticals combined. 'Campaign detail' is the raw campaign rows for your own pivots. 'Queries' has the SQL used to produce these numbers."),
]
r=4
for k,v in rows:
    ws.cell(row=r,column=1,value=k).font=BOLD
    c=ws.cell(row=r,column=2,value=v); c.alignment=LFT
    r+=1
ws.column_dimensions["A"].width=max((len(str(k)) for k,_ in rows), default=30)+2
ws.column_dimensions["B"].width=104

# ================= 2 & 3) MM vs 3P by vertical (blended + gated) =================
def compare_tab(wsname, tt, sub, src, gcol, gA, gB, ov_src=None):
    def c(v,g,col):
        x=src[(src.sales_vertical==v)&(src[gcol]==g)]; return x[col].iloc[0] if len(x) else np.nan
    verts=[v for v in src.sales_vertical.unique() if v!="Other / Unmapped"]
    verts=sorted(verts,key=lambda v:-(c(v,gA,"IVR_med")/max(c(v,gB,"IVR_med"),1e-9)))
    aI,bI=f"{gA} IVR",f"{gB} IVR"; aC,bC=f"{gA} CPV",f"{gB} CPV"; aR,bR=f"{gA} ROAS",f"{gB} ROAS"
    aN,bN=f"{gA} advertisers",f"{gB} advertisers"
    hd=pd.DataFrame([{
        "Sales vertical":v, aI:c(v,gA,"IVR_med"), bI:c(v,gB,"IVR_med"),
        "IVR advantage":c(v,gA,"IVR_med")/c(v,gB,"IVR_med") if c(v,gB,"IVR_med") else np.nan,
        aC:c(v,gA,"CPV_med"), bC:c(v,gB,"CPV_med"),
        aR:c(v,gA,"ROAS_med"), bR:c(v,gB,"ROAS_med"),
        aN:_int0(c(v,gA,"n_adv")), bN:_int0(c(v,gB,"n_adv")),
    } for v in verts])
    ws=wb.create_sheet(wsname); ws.sheet_view.showGridLines=False
    ws["A1"]=tt; ws["A1"].font=TITLE; ws["A2"]=sub; ws["A2"].font=SUB
    start=4
    if ov_src is not None:  # overall summary line(s) above the by-vertical table
        ows=ov_src.set_index(gcol)
        allrow=pd.DataFrame([{
            "Sales vertical":"ALL VERTICALS", aI:ows.loc[gA,"IVR_med"], bI:ows.loc[gB,"IVR_med"],
            "IVR advantage":ows.loc[gA,"IVR_med"]/ows.loc[gB,"IVR_med"],
            aC:ows.loc[gA,"CPV_med"], bC:ows.loc[gB,"CPV_med"],
            aR:ows.loc[gA,"ROAS_med"], bR:ows.loc[gB,"ROAS_med"],
            aN:_int0(ows.loc[gA,"n_adv"]), bN:_int0(ows.loc[gB,"n_adv"])}])
        start=write_table(ws,allrow,4,{aI:PCT2,bI:PCT2,"IVR advantage":"0.0\"x\"",aC:USD,bC:USD,aR:NUM2,bR:NUM2})+1
    fmts={aI:PCT2,bI:PCT2,"IVR advantage":"0.0\"x\"",aC:USD,bC:USD,aR:NUM2,bR:NUM2}
    write_table(ws,hd,start,fmts)
    ws.freeze_panes=f"A{start+1}"; ws.auto_filter.ref=f"A{start}:{get_column_letter(len(hd.columns))}{start+len(hd)}"
    autosize(ws,hd,{"Sales vertical":26})
    return ws

compare_tab("MM vs 3P by vertical",
    "MM vs 3P by vertical (all MM campaigns), median advertiser",
    "Every MM campaign vs 3P: the realistic average. Visit rate higher is better, cost per visit lower is better. ROAS is directional (see Read me).",
    bv2,"bucket2","MM (all)","3P",ov_src=ov2)
compare_tab("MM gated vs 3P by vertical",
    "MM (gated) vs 3P by vertical, median advertiser",
    "MM configured with the intent gate on (best case) vs 3P. ROAS is directional (see Read me).",
    bv,"bucket_detail","MM (gated)","3P",ov_src=ov)

# ================= 3) Full scorecard =================
fcols={"sales_vertical":"Sales vertical","bucket_detail":"Group","n_adv":"Advertisers",
 "n_camp":"Campaigns","spend":"Spend","imps":"Impressions","IVR_med":"IVR (median)","CVR_med":"CVR (median)",
 "CTR_med":"CTR (median)","CPV_med":"CPV (median)","ROAS_med":"ROAS (median)",
 "IVR_pooled":"IVR (pooled)","CPV_pooled":"CPV (pooled)","CPM_pooled":"CPM (pooled)","ROAS_pooled":"ROAS (pooled)"}
full=bv[list(fcols)].rename(columns=fcols)
ws3=wb.create_sheet("Full scorecard"); ws3.sheet_view.showGridLines=False
ws3["A1"]="Full scorecard: every group by vertical"; ws3["A1"].font=TITLE
ws3["A2"]="Median = advertiser-weighted (whale-robust). Pooled = impression or spend weighted. ROAS directional only."; ws3["A2"].font=SUB
ffmt={"Spend":USD0,"Impressions":INT,"IVR (median)":PCT2,"CVR (median)":PCT3,"CTR (median)":PCT3,
 "CPV (median)":USD,"ROAS (median)":NUM2,"IVR (pooled)":PCT2,"CPV (pooled)":USD,"CPM (pooled)":USD,"ROAS (pooled)":NUM2}
write_table(ws3,full,4,ffmt)
ws3.freeze_panes="C5"; ws3.auto_filter.ref=f"A4:{get_column_letter(len(full.columns))}{4+len(full)}"
autosize(ws3,full,{"Sales vertical":26,"Spend":11,"Impressions":13})

# ================= 4) Overall =================
oc=[c for c in fcols if c!="sales_vertical"]
overall=ov[oc].rename(columns=fcols)
ws4=wb.create_sheet("Overall"); ws4.sheet_view.showGridLines=False
ws4["A1"]="Overall: all verticals combined"; ws4["A1"].font=TITLE
ws4["A2"]="The intent gate and audience breadth both matter: un-gated and restricted MM fall well below gated MM."; ws4["A2"].font=SUB
write_table(ws4,overall,4,ffmt)
ws4.freeze_panes="A5"
autosize(ws4,overall,{"Spend":11,"Impressions":13})

# ================= 5) Campaign detail =================
det=df.copy()
det["IVR"]=det.visits/det.imps; det["CVR"]=det.conv/det.imps; det["CTR"]=det.clicks/det.imps
det["CPV"]=(det.spend/det.visits).replace([np.inf,-np.inf],np.nan)
det["ROAS"]=(det.revenue/det.spend).replace([np.inf,-np.inf],np.nan)
det=det[["company_name","advertiser_id","campaign_id","sales_vertical","vertical_name","bucket_detail",
  "semantics","capped","zip_narrow","city_narrow","radius_narrow","imps","visits","clicks","conv","revenue","spend",
  "IVR","CVR","CTR","CPV","ROAS"]].rename(columns={
  "company_name":"Advertiser","advertiser_id":"Adv ID","campaign_id":"Campaign ID","sales_vertical":"Sales vertical",
  "vertical_name":"MNTN vertical","bucket_detail":"Group","semantics":"3P semantics","capped":"HHST gated",
  "zip_narrow":"Zip","city_narrow":"City","radius_narrow":"Radius","imps":"Impressions","visits":"Visits",
  "clicks":"Clicks","conv":"Conversions","revenue":"Revenue","spend":"Spend"}).sort_values(
  ["Sales vertical","Group","Spend"],ascending=[True,True,False])
ws5=wb.create_sheet("Campaign detail"); ws5.sheet_view.showGridLines=False
write_table(ws5,det,1,{"Impressions":INT,"Visits":INT,"Clicks":INT,"Conversions":INT,"Revenue":USD0,"Spend":USD0,
  "IVR":PCT2,"CVR":PCT3,"CTR":PCT3,"CPV":USD,"ROAS":NUM2},band=False,header_wrap=True)
ws5.freeze_panes="A2"; ws5.auto_filter.ref=f"A1:{get_column_letter(len(det.columns))}{len(det)+1}"
autosize(ws5,det,{"Advertiser":30,"Sales vertical":24,"MNTN vertical":24,"3P semantics":21,"Group":13,
  "Impressions":11,"Revenue":11,"Spend":11,"Visits":8,"Conversions":9})

# ================= 6) Queries =================
ws6=wb.create_sheet("Queries"); ws6.sheet_view.showGridLines=False
ws6["A1"]="Queries used (for validation)"; ws6["A1"].font=TITLE
ws6["A2"]="Cohort SQL below (BigQuery). Aggregation to medians/pooled done in audi_1141_aggregate.py; see the ticket folder."; ws6["A2"].font=SUB
with open(SQLFILE) as fh: sql_lines=fh.read().split("\n")
r=4
for ln in sql_lines:
    c=ws6.cell(row=r,column=1,value=ln); c.font=MONO; r+=1
ws6.column_dimensions["A"].width=120

wb.save(DEST)
print("saved",DEST,"|",len(det),"campaign rows |",wb.sheetnames)
