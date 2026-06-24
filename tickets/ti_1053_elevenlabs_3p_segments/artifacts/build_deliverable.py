import json, csv
final = json.load(open('tickets/ti_1053_elevenlabs_3p_segments/outputs/final_ranked.json'))

THEME_RATIONALE = {
 'AI & Machine Learning':"Bullseye. AI/ML practitioners are ElevenLabs' core developer/API ICP — highly relevant and largely non-customers, so strong incremental headroom (not demand-harvesting).",
 'Developers & Software Engineering':"Core ICP (API/SDK users). Reaches builders who may not yet use ElevenLabs — high relevance + room to move = strong incremental fit.",
 'Audio · Podcast · Audiobook · Voiceover':"Direct use-case audience (audio creators/listeners). On-theme and broad enough to power a visit-based incrementality test while staying relevant.",
 'Content Creation & Media Production':"Film/video/multimedia producers need TTS, dubbing & voiceover. Relevant creators, broad enough to contain non-customers — good incremental reach.",
 'Gaming & Game Dev':"Game studios use AI voice for NPCs/dialogue. Niche-relevant creator audience.",
 'Marketing & Advertising (creative)':"Produce ad/social voiceover with AI audio. Adjacent creative buyers — decent relevance, moderate incremental fit.",
 'Design & Creative':"Creative professionals adjacent to AI-audio tooling — supporting reach.",
 'Education & EdTech':"Course/e-learning narration use-case — niche-relevant, smaller scale.",
 'Publishing & Media':"Publishers/media orgs adopting AI narration — relevant adjacent.",
 'IT & Software Industry (broad)':"Reach/scale filler. Broad firmographic — relevance diluted for a niche product, so weaker incrementality. Use only to top up scale.",
 'Telecom & Customer Service (CX/IVR)':"Conversational-AI / IVR use-case — relevant enterprise angle, broad.",
 'General B2B / Decision-Makers (broad)':"Broadest. Low specificity for a niche product → highest dilution / weakest incrementality. Avoid unless desperate for scale.",
}
PRIORITY = {1:'Add — core niche', 2:'Add — adjacent', 3:'Optional — reach filler'}

for i,x in enumerate(final,1):
    x['rank']=i
    x['priority']=PRIORITY[x['tier']]
    x['rationale']=THEME_RATIONALE.get(x['theme'],'Relevant to ICP.')

cols=['rank','tier','priority','theme','path','cat','seg_type','name_score','rationale']
hdr=['Rank','Tier','Priority','Theme','LiveRamp Segment (DS35 path)','Category ID','Segment Type','Name Score','Why it fits an incrementality campaign']

# ---- CSV ----
with open('tickets/ti_1053_elevenlabs_3p_segments/outputs/ti_1053_elevenlabs_3p_recommendations.csv','w',newline='') as f:
    w=csv.writer(f); w.writerow(hdr)
    for x in final: w.writerow([x[c] for c in cols])

# ---- XLSX ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Top 3P Segments"
navy=PatternFill('solid',fgColor='1F3A5F'); t1=PatternFill('solid',fgColor='E8F0E3'); t2=PatternFill('solid',fgColor='FBF3E2'); t3=PatternFill('solid',fgColor='F0F0F0')
white=Font(color='FFFFFF',bold=True,size=11); bold=Font(bold=True); wrap=Alignment(wrap_text=True,vertical='top')
thin=Border(*[Side(style='thin',color='D9D9D9')]*4)

# title block
ws.merge_cells('A1:I1'); ws['A1']="ElevenLabs (AID 51660) — 3P Interest-Segment Recommendations for an Incrementality-Focused CTV Campaign"
ws['A1'].font=Font(bold=True,size=13,color='1F3A5F')
ws.merge_cells('A2:I2'); ws['A2']="Ranked by NAME-based ICP relevance + incrementality fit (true quality/lift scores not run). Niche B2B: AI voice/audio for developers, content creators, media/production, AI/ML. Broad firmographic deliberately down-ranked (dilution → weaker incrementality). All segments are LiveRamp (DS35). Measure lift on VISITS (CVR underpowered — TI-1044). For Edgar von Trotha & Lauren Reedy."
ws['A2'].font=Font(italic=True,size=9,color='555555'); ws['A2'].alignment=Alignment(wrap_text=True,vertical='top'); ws.row_dimensions[2].height=58
r0=4
for j,h in enumerate(hdr,1):
    c=ws.cell(r0,j,h); c.fill=navy; c.font=white; c.alignment=wrap; c.border=thin
for x in final:
    r0+=1; fill={1:t1,2:t2,3:t3}[x['tier']]
    for j,col in enumerate(cols,1):
        c=ws.cell(r0,j,x[col]); c.fill=fill; c.alignment=wrap; c.border=thin
        if col in('rank','cat','name_score','tier'): c.alignment=Alignment(horizontal='center',vertical='top')
widths=[6,5,16,30,46,13,15,9,60]
for j,wd in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=wd
ws.freeze_panes='A5'

# methodology tab
ms=wb.create_sheet("Method & Caveats")
notes=[
 ("Method & Caveats",""),
 ("Request","Recommend 3P interest segments for ElevenLabs suited to an incrementality-focused CTV campaign (Edgar von Trotha, Lauren Reedy). Jira TI-1053; follow-on to TI-1044."),
 ("Scoring basis","NAME of the segment only (per request). We did NOT run true per-segment quality/lift scores or reach sizing (would need a holdout test / heavy ipdsc scans)."),
 ("Relevance","Keyword tiers mapped to ElevenLabs' niche ICP: AI/ML, developers/engineering, audio/podcast/voiceover, content/media production, gaming, marketing-creative, design, edtech. Off-target (healthcare, manufacturing, trades, SIC noise, TV-title/sports content) excluded."),
 ("Incrementality logic","For a NICHE product, broad reach = dilution = weaker incrementality (TI-1044: national broad scale diluted a working high-intent geo campaign). So broad 'general B2B / IT industry' firmographic is down-ranked to Tier 3 'reach filler'. In-market/shopper and demographic-only segments penalized (demand-harvesting / too broad)."),
 ("Tiers","Tier 1 = core niche (add first). Tier 2 = adjacent creative/edtech. Tier 3 = broad reach filler (use sparingly to top up scale)."),
 ("Not scored here","Reach/size (ipdsc) and CPM — intentionally skipped (expensive / not cleanly joinable). If needed, size only this shortlist before launch."),
 ("How to validate incrementality","The true test of a segment's incremental value is a holdout/ghost-bid lift test measured on VISITS (ElevenLabs CVR base ~0.062% is underpowered; visits are well-powered — TI-1044). Name+ICP fit is a pre-screen, not proof."),
 ("Their current 3P","ElevenLabs currently runs ~116 third-party segments (112 LiveRamp DS35 + 4 ShareThis) — they are MNTN's #2 stale-3P advertiser (TI-999). Recommendation is to replace breadth with the focused niche set below."),
]
for i,(a,b) in enumerate(notes,1):
    ca=ms.cell(i,1,a); ca.font=Font(bold=True,color='1F3A5F'); ca.alignment=wrap
    cb=ms.cell(i,2,b); cb.alignment=wrap
ms.column_dimensions['A'].width=22; ms.column_dimensions['B'].width=95
for i in range(1,len(notes)+1): ms.row_dimensions[i].height=42
ms['A1'].font=Font(bold=True,size=13,color='1F3A5F')

wb.save('tickets/ti_1053_elevenlabs_3p_segments/outputs/ti_1053_elevenlabs_3p_recommendations.xlsx')
print("WROTE csv + xlsx |", len(final),"segments")
print("themes:", {t: sum(1 for x in final if x['theme']==t) for t in dict.fromkeys(x['theme'] for x in final)})
print("\nRANKED:")
for x in final: print(f"  #{x['rank']:>2} T{x['tier']} [{x['theme'][:30]:30}] {x['path'][:62]}")
