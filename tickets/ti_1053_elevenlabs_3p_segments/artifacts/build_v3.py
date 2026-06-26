import json,csv
cur=json.load(open('tickets/ti_1053_elevenlabs_3p_segments/outputs/curated_v3.json'))
sz=json.load(open('tickets/ti_1053_elevenlabs_3p_segments/outputs/sizes_7d.json'))
EDGAR={'1012567311':15_000_000,'1012745401':12_000_000}
def seg_type(p):
    p=p.lower()
    if 'remarketing' in p or 'retargeting' in p: return 'retargeting-tech'
    if 'podcast:' in p or 'podcast enthusiast' in p or 'crime junkie' in p: return 'consumer-affinity'
    if 'intent' in p: return 'b2b-intent'
    if any(k in p for k in['title','profession','job role','job function','decision maker','blue collar','white collar','developer /','engineer /']): return 'b2b-role'
    if 'industry' in p or 'naics' in p or 'sic' in p: return 'b2b-firmographic'
    if 'interest' in p: return 'interest/affinity'
    return 'b2b-other'
INCR={'b2b-role':0.95,'b2b-firmographic':0.85,'interest/affinity':0.80,'b2b-intent':0.65,'b2b-other':0.70,
      'consumer-affinity':0.15,'retargeting-tech':0.10}
def size_score(r): return 0.10 if r<50_000 else 0.40 if r<250_000 else 0.70 if r<1_000_000 else (1.00 if r<=20_000_000 else 0.85)
def verdict(r,incr):
    if r==0: return (5,'DROP — no 7d delivery')
    if incr<0.3: return (5,'DROP — not a buyer audience')
    if r>=1_000_000 and incr>=0.7: return (1,'PRIMARY — relevant + scaled')
    if r>=1_000_000: return (3,'SCALE — but weaker fit (intent/broad)')
    if r>=250_000: return (2,'SECONDARY')
    return (4,'ADDITIVE ONLY — sub-scale')
out=[]
for x in cur:
    s=sz.get(x['cat'],{'reach':0,'days':0}); reach=EDGAR.get(x['cat']) or s['reach']
    st=seg_type(x['readable']); incr=INCR[st]; rel=min(1.0,x['name_score']/12.0); ss=size_score(reach)
    vr,vl=verdict(reach,incr); comp=round(0.30*rel+0.32*incr+0.38*ss,3)
    fl=[]
    if x['cat'] in EDGAR: fl.append('platform size (Edgar)')
    elif s['days']<=1 and reach>0: fl.append('bursty (1/7d) — load-day reach')
    if st=='b2b-intent': fl.append('intent=harvesting risk')
    if st=='consumer-affinity': fl.append('CONSUMER not buyer')
    if st=='retargeting-tech': fl.append('not relevant')
    out.append({'theme':x['theme'],'segment':x['readable'],'provider':x['provider'],'cat':x['cat'],
      'reach':reach,'days7':s['days'],'seg_type':st,'relevance':round(rel,2),'incrementality_fit':incr,
      'size_score':ss,'composite':comp,'verdict':vl,'flags':' | '.join(fl),'_vr':vr})
out.sort(key=lambda z:(z['_vr'],-z['composite']))
for i,r in enumerate(out,1): r['rank']=i
json.dump(out,open('tickets/ti_1053_elevenlabs_3p_segments/outputs/final_v3_scored.json','w'))
cols=['rank','verdict','theme','segment','provider','cat','reach','days7','seg_type','relevance','incrementality_fit','size_score','composite','flags']
hdr=['Rank','Verdict','Theme','LiveRamp Segment (DS35)','Provider','Cat ID','Reach (7d IPs)','Days /7','Type','Relevance','Incr. Fit','Size','Composite','Flags']
with open('tickets/ti_1053_elevenlabs_3p_segments/outputs/ti_1053_elevenlabs_3p_recommendations.csv','w',newline='') as f:
    w=csv.writer(f); w.writerow(hdr); [w.writerow([r[c] for c in cols]) for r in out]
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
wb=openpyxl.Workbook(); ws=wb.active; ws.title="3P Segments (recall-fixed)"
navy=PatternFill('solid',fgColor='1F3A5F'); white=Font(color='FFFFFF',bold=True)
VF={'PRIMARY — relevant + scaled':'D6E8D5','SCALE — but weaker fit (intent/broad)':'FBEFD3','SECONDARY':'EAF1F8','ADDITIVE ONLY — sub-scale':'F2F2F2','DROP — no 7d delivery':'F7D9D9','DROP — not a buyer audience':'F7D9D9'}
wrap=Alignment(wrap_text=True,vertical='top'); ctr=Alignment(horizontal='center',vertical='top'); thin=Border(*[Side('thin',color='DDDDDD')]*4)
ws.merge_cells('A1:N1'); ws['A1']="ElevenLabs (51660) — 3P Segment Recommendations for Incrementality CTV (recall-corrected)"; ws['A1'].font=Font(bold=True,size=13,color='1F3A5F')
ws.merge_cells('A2:N2')
ws['A2']=("CORRECTION: the first pass missed every 'struct-path' LiveRamp provider (ZoomInfo, Anteriad/180byTwo, Alliant, LBDigital, OnAudience...) due to a field bug — that hid most premium B2B inventory. "
 "Corrected pool = 1,759 relevant; this is the curated 44. Reach = distinct IPs in ipdsc 7d (2026-06-17→23); many providers load 1 day/week so reach = load-day size (bursty). Edgar's 2 use platform size. "
 "Authoritative sizes live in external_ddm.data_source_category_sizes (access-gated). Validate incrementality with a VISIT-based holdout (CVR underpowered — TI-1044). For Edgar von Trotha & Lauren Reedy.")
ws['A2'].font=Font(italic=True,size=9,color='555555'); ws['A2'].alignment=wrap; ws.row_dimensions[2].height=70
r0=4
for j,h in enumerate(hdr,1):
    c=ws.cell(r0,j,h); c.fill=navy; c.font=white; c.alignment=wrap; c.border=thin
for r in out:
    r0+=1; fill=PatternFill('solid',fgColor=VF.get(r['verdict'],'FFFFFF'))
    for j,col in enumerate(cols,1):
        c=ws.cell(r0,j,r[col]); c.fill=fill; c.border=thin
        c.alignment=ctr if col in('rank','cat','reach','days7','relevance','incrementality_fit','size_score','composite') else wrap
        if col=='reach': c.number_format='#,##0'
for j,wd in enumerate([5,30,26,40,16,11,12,7,16,9,9,7,10,26],1): ws.column_dimensions[get_column_letter(j)].width=wd
ws.freeze_panes='A5'
ms=wb.create_sheet("Method & Bottom Line")
notes=[("Method & Bottom Line",""),
 ("BOTTOM LINE (revised)","3P is actually a VIABLE lever for ElevenLabs once the recall bug is fixed — there are ~15+ large, on-target B2B segments (Software Developers, AI/ML professionals, Animation/SDK software, Business Software, Content-Creation intent). Lead with PRIMARY rows. (This revises the earlier 'thin/weak' read, which was an artifact of the bug.)"),
 ("The bug","Candidate filter matched keywords on COALESCE(path_from_root, names, name). For ~half of LiveRamp providers path_from_root is an unreadable struct {pathFromRoot:[ids]}, so COALESCE returned no words -> those providers (ZoomInfo, Anteriad, Alliant, LBDigital, OnAudience, NetWise...) were silently dropped. Fix: match on all readable fields concatenated. Pool 24 -> 1,759 relevant."),
 ("Scoring","Composite = 0.30*Relevance(name/ICP) + 0.32*Incrementality-fit + 0.38*Size. Incrementality-fit: b2b-role 0.95 > firmographic 0.85 > interest 0.80 > b2b-intent 0.65 (intent = closer to demand-harvesting) > consumer 0.15."),
 ("Verdicts","PRIMARY = relevant + >=1M. SCALE = >=1M but weaker fit (intent/broad). SECONDARY = 250K-1M. ADDITIVE = <250K. DROP = no delivery or not-a-buyer (consumer podcast fans, remarketing tech)."),
 ("Size caveat","ipdsc 7-day reach; bursty (Days/7). A '0' may mean 'did not load this week,' not 'small.' Authoritative platform sizes = external_ddm.data_source_category_sizes (request Storage Object Viewer on gs://mntn-data-monitoring; matches the UI numbers Edgar quoted). This was the LAST ipdsc sizing run (6.7TB)."),
 ("Incrementality caveat","Name+size is a pre-screen. True incremental value needs a visit-based holdout/ghost-bid test (ElevenLabs CVR ~0.062% underpowered — TI-1044). Prefer role/professional + interest audiences over pure 'intent' (harvesting)."),
]
for i,(a,b) in enumerate(notes,1):
    ms.cell(i,1,a).font=Font(bold=True,color='1F3A5F'); ms.cell(i,1).alignment=wrap; ms.cell(i,2,b).alignment=wrap
ms.column_dimensions['A'].width=20; ms.column_dimensions['B'].width=104
for i in range(1,len(notes)+1): ms.row_dimensions[i].height=58
ms['A1'].font=Font(bold=True,size=13,color='1F3A5F')
wb.save('tickets/ti_1053_elevenlabs_3p_segments/outputs/ti_1053_elevenlabs_3p_recommendations.xlsx')
print("WROTE final v3:",len(out),"segments")
from collections import Counter
for v,n in Counter(r['verdict'] for r in out).most_common(): print(f"  {n:>2}  {v}")
print("\nPRIMARY:")
for r in out:
    if r['verdict'].startswith('PRIMARY'): print(f"  {r['reach']:>11,}  {r['provider'][:14]:14} {r['segment'][-46:]}")
