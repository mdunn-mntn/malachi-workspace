import json, csv
gl=json.load(open('tickets/ti_1053_elevenlabs_3p_segments/outputs/genuine_shortlist.json'))
sz=json.load(open('tickets/ti_1053_elevenlabs_3p_segments/outputs/sizes_30d.json'))
def provider(path):
    for p in ['Clickagy','Datasys','AtoZ','Start.io','InfutorData','HCS']:
        if p.lower() in path.lower(): return p
    return path.split('>')[0].strip()[:14]
INCR={'AI & Machine Learning':0.85,'Developers & Software Engineering':0.95,'Audio/Voice Production':0.95,
 'Film/Video/Media Production':0.90,'Marketing & Advertising (creative)':0.75,'Design & Creative':0.80,
 'Education & EdTech':0.80,'IT & Software Industry (broad)':0.40}
def size_score(r): return 0.10 if r<50_000 else 0.40 if r<250_000 else 0.70 if r<1_000_000 else (1.00 if r<=10_000_000 else 0.85)
def verdict(reach,days,incr):
    if reach==0: return (5,'DROP — no 30d delivery')
    if reach>=1_000_000 and incr>=0.70: return (1,'PRIMARY — relevant + scaled')
    if reach>=1_000_000: return (3,'SCALE FILLER — broad, dilution risk')
    if reach>=250_000: return (2,'SECONDARY' + (' (bursty)' if days<5 else ''))
    return (4,'ADDITIVE ONLY — sub-scale')
out=[]
for x in gl:
    s=sz.get(x['cat'],{'reach':0,'days':0}); reach=s['reach']
    rel=min(1.0,x['name_score']/10.0); incr=INCR.get(x['theme'],0.6); ss=size_score(reach)
    vr,vlabel=verdict(reach,s['days'],incr)
    comp=round(0.32*rel+0.30*incr+0.38*ss,3)
    out.append({'theme':x['theme'],'segment':x['path'],'provider':provider(x['segment'] if 'segment' in x else x['path']),
      'cat':x['cat'],'reach_30d':reach,'days_present':s['days'],'relevance':round(rel,2),
      'incrementality_fit':incr,'size_score':ss,'composite':comp,'verdict':vlabel,'_vr':vr})
out.sort(key=lambda z:(z['_vr'],-z['composite']))
for i,r in enumerate(out,1): r['rank']=i
json.dump(out,open('tickets/ti_1053_elevenlabs_3p_segments/outputs/final_scored.json','w'))
cols=['rank','verdict','theme','segment','provider','cat','reach_30d','days_present','relevance','incrementality_fit','size_score','composite']
hdr=['Rank','Verdict','Theme','LiveRamp Segment (DS35)','Provider','Cat ID','Reach (30d IPs)','Days Live /30','Relevance','Incrementality Fit','Size Score','Composite']
with open('tickets/ti_1053_elevenlabs_3p_segments/outputs/ti_1053_elevenlabs_3p_recommendations.csv','w',newline='') as f:
    w=csv.writer(f); w.writerow(hdr); [w.writerow([r[c] for c in cols]) for r in out]
# XLSX
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
wb=openpyxl.Workbook(); ws=wb.active; ws.title="3P Segments (size-aware)"
navy=PatternFill('solid',fgColor='1F3A5F'); white=Font(color='FFFFFF',bold=True)
VF={'PRIMARY — relevant + scaled':'D6E8D5','SCALE FILLER — broad, dilution risk':'FBEFD3','SECONDARY':'EAF1F8','SECONDARY (bursty)':'EAF1F8','ADDITIVE ONLY — sub-scale':'F2F2F2','DROP — no 30d delivery':'F7D9D9'}
wrap=Alignment(wrap_text=True,vertical='top'); ctr=Alignment(horizontal='center',vertical='top'); thin=Border(*[Side('thin',color='DDDDDD')]*4)
ws.merge_cells('A1:L1'); ws['A1']="ElevenLabs (AID 51660) — 3P Segment Recommendations for an Incrementality-Focused CTV Campaign"; ws['A1'].font=Font(bold=True,size=13,color='1F3A5F')
ws.merge_cells('A2:L2')
ws['A2']=("Honest bottom line: LiveRamp 3P is a THIN lever for this niche product. Of ~210K DS35 segments, almost none are AI-voice-specific "
 "(0 real voice/speech, 1 AI/ML concept, no Bombora). Only ~4 segments are BOTH relevant AND big enough to feed an ~800K-imps/day campaign. "
 "The precise niche segments (motion-picture/video production, sound recording, multimedia) are real-but-tiny (<20K IPs) → additive layer only. "
 "Reach = distinct IPs in ipdsc over 30d (2026-05-25→06-23). Validate true incremental value with a VISIT-based holdout test (CVR underpowered — TI-1044). For Edgar von Trotha & Lauren Reedy.")
ws['A2'].font=Font(italic=True,size=9,color='555555'); ws['A2'].alignment=wrap; ws.row_dimensions[2].height=72
r0=4
for j,h in enumerate(hdr,1):
    c=ws.cell(r0,j,h); c.fill=navy; c.font=white; c.alignment=wrap; c.border=thin
for r in out:
    r0+=1; fill=PatternFill('solid',fgColor=VF.get(r['verdict'],'FFFFFF'))
    for j,col in enumerate(cols,1):
        c=ws.cell(r0,j,r[col]); c.fill=fill; c.border=thin
        c.alignment=ctr if col in('rank','cat','reach_30d','days_present','relevance','incrementality_fit','size_score','composite') else wrap
        if col=='reach_30d': c.number_format='#,##0'
widths=[5,30,24,44,11,12,13,11,10,12,10,10]
for j,wd in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=wd
ws.freeze_panes='A5'
# method tab
ms=wb.create_sheet("Method & Bottom Line")
notes=[("Method & Bottom Line",""),
 ("Request","3P segments for ElevenLabs suited to an incrementality-focused CTV campaign (Edgar von Trotha, Lauren Reedy). TI-1053, follow-on to TI-1044."),
 ("BOTTOM LINE","3P (LiveRamp) is a weak lever for a niche AI-voice product. Lead with the AI/ML-interest + Software-Developer + Advertising-industry segments (the only relevant ones at scale); treat everything else as additive-only or skip. Bigger win is MM keywords / contextual, not bought 3P (ElevenLabs is MNTN's #2 stale-3P advertiser — TI-999)."),
 ("Coverage check","We did NOT eyeball 210K names; we keyword-filtered to ~7.7K then scored, THEN profiled term coverage across all 210K: 0 real voice/speech-tech, 3 AI/ML, 0 Bombora, and most 'creator/gamer' matches are CONSUMERS (YouTube/Twitch viewers, gamers) not buyers. So the relevant universe is genuinely small."),
 ("Scoring","Composite = 0.32*Relevance(name/ICP) + 0.30*Incrementality-fit(buyer-firmographic & niche; broad-IT penalized) + 0.38*Size(reach band). Size weighted highest — it is the binding constraint at ~800K imps/day."),
 ("Verdicts","PRIMARY = relevant + ≥1M reach. SCALE FILLER = ≥1M but broad (dilution). SECONDARY = 250K-1M. ADDITIVE ONLY = <250K (real but too small to drive delivery). DROP = no 30d delivery."),
 ("Reach caveat","Reach = distinct IPs in ipdsc 2026-05-25→06-23. 3P delivery is bursty (Days Live /30 shows consistency). Sizing this shortlist cost ~30TB — do NOT re-run wide ipdsc DISTINCT-IP scans casually."),
 ("Incrementality caveat","Name + size is a PRE-SCREEN, not proof. True incremental value needs a holdout/ghost-bid lift test on VISITS (ElevenLabs CVR ~0.062% underpowered — TI-1044). The trap is picking high-reach demand-harvesting audiences; favor relevant-but-not-already-in-market."),
]
for i,(a,b) in enumerate(notes,1):
    ms.cell(i,1,a).font=Font(bold=True,color='1F3A5F'); ms.cell(i,1).alignment=wrap; ms.cell(i,2,b).alignment=wrap
ms.column_dimensions['A'].width=20; ms.column_dimensions['B'].width=100
for i in range(1,len(notes)+1): ms.row_dimensions[i].height=54
ms['A1'].font=Font(bold=True,size=13,color='1F3A5F')
wb.save('tickets/ti_1053_elevenlabs_3p_segments/outputs/ti_1053_elevenlabs_3p_recommendations.xlsx')
print("WROTE csv+xlsx, 24 segments")
print(f"{'#':>2} {'verdict':38} {'reach30d':>11}  segment")
for r in out: print(f"  {r['rank']:>2} {r['verdict'][:38]:38} {r['reach_30d']:>11,}  {r['segment'][:40]}")
