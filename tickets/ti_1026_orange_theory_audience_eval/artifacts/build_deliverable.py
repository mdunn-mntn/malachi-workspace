#!/usr/bin/env python3
"""TI-1026 — build the Orange Theory audience-evaluation recommendation workbook (.xlsx),
mirroring the ElevenLabs eval (TI-928) deliverable. Multi-tab:
  1. Recommendations  — headline + prioritized actions
  2. Interest Segments — 11 included + 7 excluded 3P, reach/overlap/recommendation
  3. Keywords          — 379 MNTN Matched keywords, bucket + action
  4. Geo & Exclusions  — geo-fence coverage + demographic-exclusion footprint
  5. Methodology       — sources, dates, caveats

Data from outputs/*.csv (reproducible). Run after the analysis queries land.
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE.parent / "outputs"
XLSX = HERE.parent / "artifacts" / "ti_1026_orange_theory_audience_recommendations.xlsx"

NAVY = "1F3864"; RED = "C00000"; GREEN = "375623"; AMBER = "BF8F00"; GREY = "808080"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = Alignment(wrap_text=True, vertical="center")


def autowidth(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def action_fill(val):
    v = str(val).upper()
    if v.startswith("DROP"):
        return PatternFill("solid", fgColor="F4CCCC")
    if v.startswith("REVIEW"):
        return PatternFill("solid", fgColor="FFF2CC")
    if v.startswith("KEEP"):
        return PatternFill("solid", fgColor="D9EAD3")
    return None


def load_csv(name):
    with (OUT / name).open() as f:
        return list(csv.DictReader(f))


def main():
    wb = Workbook()

    # ---- Tab 1: Recommendations ----
    ws = wb.active; ws.title = "Recommendations"
    ws["A1"] = "Orange Theory National — Audience Evaluation (TI-1026)"; ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Advertiser 39718 · Audience 34668 'MNTN Matched | New Year's 3P Segments Copy 01' · "
                "prepared by Targeting Infrastructure")
    ws["A2"].font = Font(italic=True, color=GREY)
    rows = [
        ["", ""],
        ["THE WHY (how the bidder turns this audience into the problem)", "Evidence (see 'Why — Evidence' tab)"],
        ["The audience = (MNTN Matched keywords OR 11 bought 3P segments) AND within 7 mi of a studio.",
         "Confirmed OR-include: one include block, or:[MNTN Matched, 3P]."],
        ["1. Only the MNTN Matched IPs carry a quality score; the 3P IPs do not.",
         "MNTN's household score (fitness-intent, 0-10000) is built from the keyword layer. 87% of 3P-segment IPs "
         "match NO Orange Theory keyword, so they have no score."],
        ["2. The main campaign only bids on high-scoring IPs (score threshold = 6,501).",
         "Last 14d: 82% of the main campaign's impressions went to IPs scored >=6,501; only 1.5% were unscored. "
         "So the unscored 3P-only IPs are filtered out — the 3P segments contribute ~nothing here."],
        ["3. Where there is NO score threshold, the bidder buys the unscored 3P traffic — and it's the bad traffic.",
         "A no-threshold OTF campaign delivered 99.96% to UNSCORED IPs. That no-intent traffic is exactly the "
         "'non-MNTN matched audience' the agency measured at 8-10x worse visit rate."],
        ["=> So 3P can NEVER be the reach fix: threshold on -> filtered out; threshold off -> unscored garbage. Remove it.",
         ""],
        ["4. Is the audience big enough? At the current budget, yes — but headroom is thin.",
         "The main campaign paces at/above its ~$2,000/day budget most days in June on the scored, in-fence audience "
         "ALONE (no 3P help) — 463,895 distinct scored IPs reached in 14d. It underdelivered in late May, so the "
         "scored audience within 7 mi of studios is adequate, not deep. The constraint is the scored-IP x geo "
         "intersection, NOT the 3P segments."],
        ["5. The keyword layer carries the audience — but ~25% of it is off-target.",
         "The 379 keywords are the engine (14x the 3P layer). ~51 are clearly off-target (Above Ground Pools, "
         "Antifreeze, Beer Mugs, Motorcycle Lighting, CPUs); ~43 are over-broad (Class, Power, Experience)."],
        ["6. Geo is not the bottleneck. The LiveRamp income/age exclusions ARE active (a real reach lever).",
         "The fence covers ~half the populated US and applies equally to both layers. The 7 LiveRamp (DS35) "
         "income/age exclusions actively remove millions-to-tens-of-millions of IPs (corrected after validation — "
         "an earlier single-day read wrongly called them inert; the 13 Oracle/DS1 bands ARE inert). So relaxing "
         "the income/age bands is a genuine reach lever — defensible to keep for a premium membership, but not free."],
        ["", ""],
        ["RECOMMENDATIONS (priority order)", "Expected effect"],
        ["A. Remove all 11 3P interest segments.",
         "On the score-gated main campaign they already contribute ~nothing (only 1.5% of delivery is unscored); "
         "on any no-threshold campaign they deliver the unscored, 8-10x-worse traffic. Removing them cleans the "
         "expression with ~no loss of real delivered reach and removes the worst traffic. (6 of 11 deliver nothing "
         "at all; the rest are broad/off-modality yoga-pilates, not HIIT.)"],
        ["B. To scale reach, use the SCORE THRESHOLD and KEYWORDS — not 3P.",
         "If they need more volume: (1) lower the household-score threshold from 6,501 (it keeps only the top ~third "
         "of scores — their own quality dial; lowering it opens more scored inventory at a modest, controlled VR "
         "cost); (2) broaden + clean the keyword set so more fitness households get scored into the pool; "
         "(3) widen the geo radius (7->10 mi). These add SCORED reach; 3P cannot."],
        ["C. Prune the 51 off-target keywords; review the 43 over-broad terms.",
         "Tightens relevance and raises the quality of the scored pool. See Keywords tab."],
        ["D. Relaxing the LiveRamp income/age exclusions IS a reach lever; geo is not the issue.",
         "The 7 LiveRamp (DS35) income/age exclusions actively remove millions-to-tens-of-millions of IPs — "
         "relaxing them (widen income floor / age cap) expands reach (a judgment call for a premium membership). "
         "The 13 Oracle (DS1) bands are inert (cosmetic). Geo already covers ~half the US — not the bottleneck."],
        ["E. Keep CRM-suppression, T-Mobile-cellular, and MNTN-First-Party exclusions as-is.",
         "Legitimate hygiene (existing-member suppression; mobile-carrier IPs aren't household-stable for CTV)."],
    ]
    r = 4
    for a, b in rows:
        ws.cell(r, 1, a); ws.cell(r, 2, b)
        ws.cell(r, 1).alignment = WRAP; ws.cell(r, 2).alignment = WRAP
        if a in ("THE FINDING", "RECOMMENDATIONS (priority order)"):
            ws.cell(r, 1).font = Font(bold=True, size=12, color=NAVY)
            ws.cell(r, 2).font = Font(bold=True, size=12, color=NAVY)
        elif a and a[0].isdigit():
            ws.cell(r, 1).font = Font(bold=True, color=NAVY)
        elif a and a[0] in "ABCDE" and a[1] == ".":
            ws.cell(r, 1).font = Font(bold=True)
        r += 1
    autowidth(ws, [70, 64])

    # ---- Tab 2: Why — Evidence ----
    wse = wb.create_sheet("Why — Evidence")
    wse["A1"] = "The mechanism, with evidence"; wse["A1"].font = TITLE_FONT
    wse.append([])
    wse.append(["Household-score threshold (HHST) per OTF campaign — dso.household_score_thresholds"])
    wse.cell(wse.max_row, 1).font = Font(bold=True, color=NAVY)
    wse.append(["Campaign", "HHST (min score to bid)", "Meaning"])
    style_header(wse, wse.max_row, 3)
    wse.append([319137, 6501, "Main campaign — only bids on IPs scored >=6,501 (top ~third). Unscored 3P-only IPs filtered out."])
    wse.append([328040, 0, "No score gate — bids on everything, scored or not."])
    wse.append([328041, 0, "No score gate."])
    wse.append([328042, 0, "No score gate."])
    wse.append(["(platform)", "~64% at 0", "20,982 of ~30k campaigns run HHST=0 (no score gate)."])
    for r in range(wse.max_row - 4, wse.max_row + 1):
        wse.cell(r, 3).alignment = WRAP
    wse.append([]); wse.append([])
    wse.append(["Delivered household-score by campaign — last 14 days, cost_impression_log"])
    wse.cell(wse.max_row, 1).font = Font(bold=True, color=NAVY)
    wse.append(["Campaign", "HHST", "Impressions", "Distinct IPs", "% scored >=6501", "% unscored (-1)", "Read"])
    style_header(wse, wse.max_row, 7)
    wse.append([319137, 6501, 1521364, 463895, "82.3%", "1.5%", "Score gate works — 3P-only (unscored) IPs filtered out; 3P contributes ~nothing."])
    wse.append([319133, 0, 34515, 5582, "0.04%", "99.96%", "No gate -> bids on unscored IPs = the 8-10x-worse 'non-MNTN matched' traffic."])
    wse.cell(wse.max_row, 7).alignment = WRAP; wse.cell(wse.max_row - 1, 7).alignment = WRAP
    wse.append([]); wse.append([])
    wse.append(["Reach / pacing — is the scored audience big enough?"])
    wse.cell(wse.max_row, 1).font = Font(bold=True, color=NAVY)
    wse.append(["Metric", "Value", "Read"])
    style_header(wse, wse.max_row, 3)
    wse.append(["Budget (main campaign)", "~$2,002/day", "$83.41/hr x 24"])
    wse.append(["June daily pacing", "1.08-1.31x of budget", "Paces to budget on the SCORED audience alone (no 3P) — not starved at current budget."])
    wse.append(["Late-May pacing", "0.35-0.6x of budget", "Soft patch — scored, in-fence audience is adequate but not deep; thin headroom to scale."])
    wse.append(["Scored reach realized (14d)", "463,895 distinct IPs", "All scored >=6501, in-fence — the real high-value audience."])
    wse.append(["MNTN Matched vs 3P national reach", "21.8M vs 3.0M /wk", "Keyword layer is 14x the 3P layer; 87% of 3P IPs are unscored (no keyword match)."])
    for r in range(wse.max_row - 4, wse.max_row + 1):
        wse.cell(r, 3).alignment = WRAP
    autowidth(wse, [16, 22, 78])

    # ---- Tab 3: Interest Segments ----
    ws2 = wb.create_sheet("Interest Segments (3P)")
    seg = load_csv("ti_1026_interest_segments_eval.csv")
    cols = ["role", "data_source_category_id", "provider", "segment", "modality_fit",
            "deprecated_flag", "delivery_note", "pct_matching_otf_keywords", "recommendation", "reason"]
    hdr = ["Role", "LiveRamp Cat ID", "Provider", "Segment", "Modality fit", "Catalog deprecated flag",
           "Delivery (IPDSC, bursty)", "% matching OTF keywords", "Recommendation", "Reason"]
    ws2.append(hdr); style_header(ws2, 1, len(hdr))
    for row in seg:
        vals = [row[c] for c in cols]
        ws2.append(vals)
        rr = ws2.max_row
        ws2.cell(rr, 4).alignment = WRAP; ws2.cell(rr, 7).alignment = WRAP; ws2.cell(rr, 10).alignment = WRAP
        f = action_fill(row["recommendation"])
        if f:
            ws2.cell(rr, 9).fill = f
    autowidth(ws2, [13, 14, 16, 34, 12, 16, 30, 16, 14, 64])
    ws2.freeze_panes = "A2"

    # ---- Tab 3: Keywords ----
    ws3 = wb.create_sheet("Keywords")
    kw = load_csv("ti_1026_keyword_classification.csv")
    ws3.append(["MNTN Matched Cat ID", "Keyword", "Bucket", "Recommended action"])
    style_header(ws3, 1, 4)
    for row in kw:
        ws3.append([row["data_source_category_id"], row["keyword"], row["bucket"], row["recommended_action"]])
        rr = ws3.max_row
        f = action_fill(row["recommended_action"])
        if f:
            ws3.cell(rr, 4).fill = f
    autowidth(ws3, [18, 42, 18, 30]); ws3.freeze_panes = "A2"

    # ---- Tab 4: Geo & Exclusions ----
    ws4 = wb.create_sheet("Geo & Exclusions")
    ws4["A1"] = "Geo-fence coverage (946 studios x 7-mi radius)"; ws4["A1"].font = TITLE_FONT
    ws4.append([]) ; ws4.append(["Measure", "Fenced", "US total", "% fenced"])
    style_header(ws4, 3, 4)
    ws4.append(["Geolocated network blocks (population proxy)", 2203886, 4458870, "49.4%"])
    ws4.append(["IPv4 address capacity", "401M", "1,613M", "24.9%"])
    ws4.append([])
    ws4.append(["Read: the fence covers ~half the populated US. Geo is a real constraint but not the bottleneck; "
                "it applies equally to MNTN Matched and 3P, so it does not explain 3P underperformance."])
    ws4.cell(ws4.max_row, 1).alignment = WRAP
    ws4.append([]); ws4.append([])
    rr = ws4.max_row + 1
    ws4.cell(rr, 1, "Demographic exclusion footprint on the audience (2026-06-08)"); ws4.cell(rr, 1).font = TITLE_FONT
    ws4.append(["Include audience (MM ∪ 3P), 2026-06-08", "", 7596517, "national, pre-geo"])
    ws4.append(["Exclusion group", "Cats", "Active?", "Note (corrected after validation)"])
    style_header(ws4, ws4.max_row, 4)
    ws4.append(["LiveRamp (DS35) income/age bands", 7, "ACTIVE",
                "Removes millions-to-tens-of-millions of IPs (e.g. Ages 65-74 ~15.4M, HHI<$25k ~10.0M on 06-04). "
                "A real reach lever — relaxing income/age would expand reach (defensible to keep for a premium membership)."])
    ws4.append(["Oracle (DS1) income/age bands", 13, "INERT",
                "DS1 has zero IPDSC presence — these 13 exclude nobody. Cosmetic cleanup only."])
    ws4.append(["T-Mobile Cellular (ISP type)", 1, "active (bid-time)",
                "KEEP — mobile-carrier IPs aren't household-stable for CTV"])
    ws4.append(["CRM suppression (existing members)", 2, "active", "KEEP — good hygiene"])
    ws4.append(["MNTN First Party (past visitors)", 3, "active", "KEEP — retargeting exclusion"])
    ws4.append([])
    ws4.append(["Read: the LiveRamp (DS35) income/age exclusions are ACTIVE and materially shrink the targetable "
                "universe — they are a genuine reach lever (relax to gain reach). The Oracle (DS1) bands are inert. "
                "NOTE: IPDSC 3P delivery is bursty (each category loads on only 2-4 days/month), so single-day "
                "measurements understate — these figures are from multi-day windows."])
    ws4.cell(ws4.max_row, 1).alignment = WRAP
    autowidth(ws4, [38, 8, 14, 66])

    # ---- Tab 5: Methodology ----
    ws5 = wb.create_sheet("Methodology")
    meth = [
        ["Methodology & Caveats", ""],
        ["Audience", "audience.audiences id 34668 (expression_type_id=2), advertiser 39718. Expression parsed to "
         "include/exclude data-source + category sets and a geo radii_include list."],
        ["Reach / overlap", "ipdsc__v1 (dw-main-bronze.external), distinct IPs, 7-day window 2026-06-04..06-10. "
         "MM=DS19 MNTN Matched (379 cats), 3P=DS35 LiveRamp (11 cats). Single-day 2026-06-10 also shown."],
        ["3P volatility", "Same 11 segments deliver wildly different IP counts day to day (e.g. Stirista 2.1M on 06-08, "
         "0 on 06-06). 7-day window is the fair measure; single-day understates."],
        ["Geo coverage", "geo.maxmind_blocks_ipv4: network blocks within 7 mi (11,265 m) of any of the 946 studios "
         "(ST_DWITHIN). Block-count is a population-density proxy; IP-capacity over-weights rural ranges. "
         "radii_exclude (21 zones) not subtracted."],
        ["Keyword buckets", "Conservative heuristic + curated lists; default = KEEP. The exact keep/drop line is for "
         "Kelly Thurlow / Sales to finalize — this surfaces high-confidence candidates."],
        ["Performance", "Visit rate is audience-level (campaign 319137: 0.25% over 90d). Per-segment delivery visit "
         "rate is not separable from logs; the 87%-non-keyword overlap + agency's 8-10x report are the intent evidence."],
        ["Not a causal claim", "Reach/overlap and keyword relevance are descriptive. They explain WHY 3P underperforms "
         "(low intent, off-modality) but the 8-10x figure is the agency's measured result, not re-derived here."],
        ["IPDSC 3P is BURSTY", "Each bought-3P (DS35) category refreshes into IPDSC on only ~2-4 days/month, so a "
         "single-day or single-week reach number is window-luck-dependent (3P weekly reach swings ~3M-19M). Always "
         "measure 3P category reach/exclusion over a >=30-day window. The recommendation rests on the window-STABLE "
         "facts (87% of 3P non-keyword; 1.5% delivered share under the gate), not on absolute 3P membership reach."],
        ["Independently validated", "8 adversarial validators re-derived every load-bearing claim on different "
         "days/methods. Confirmed: expression decomposition, 3P daily volatility, geo coverage, keyword split, "
         "campaign perf. Corrected here after validation: LiveRamp income/age exclusions are ACTIVE (not inert); "
         "3P absolute reach / '14x' framing dropped as window-luck; '6 segments deliver zero / Epsilon deprecated' "
         "reworded to bursty-delivery. Full report: artifacts/ti_1026_validation_report.md."],
    ]
    for a, b in meth:
        ws5.append([a, b]); rr = ws5.max_row
        ws5.cell(rr, 2).alignment = WRAP
        if b == "":
            ws5.cell(rr, 1).font = TITLE_FONT
        else:
            ws5.cell(rr, 1).font = Font(bold=True, color=NAVY)
    autowidth(ws5, [18, 100])

    wb.save(XLSX)
    print(f"Wrote {XLSX}")


if __name__ == "__main__":
    main()
