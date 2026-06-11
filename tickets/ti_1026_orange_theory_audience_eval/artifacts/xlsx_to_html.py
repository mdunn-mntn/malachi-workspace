#!/usr/bin/env python3
"""Render the recommendation workbook as a single self-contained HTML page (all tabs),
so it can be opened in any browser without Excel. Preserves cell fills for the
KEEP/DROP/REVIEW color coding."""
from pathlib import Path
from openpyxl import load_workbook
from html import escape

HERE = Path(__file__).resolve().parent
XLSX = HERE / "ti_1026_orange_theory_audience_recommendations.xlsx"
HTML = HERE / "ti_1026_orange_theory_audience_recommendations.html"

wb = load_workbook(XLSX)
parts = ["""<!doctype html><meta charset=utf-8><title>OTF Audience Recommendations (TI-1026)</title>
<style>
 body{font:14px/1.5 -apple-system,Helvetica,Arial,sans-serif;margin:24px;color:#1a1a1a;background:#fafafa}
 h1{color:#1F3864;font-size:22px} h2{color:#1F3864;margin-top:34px;border-bottom:2px solid #1F3864;padding-bottom:4px}
 nav a{margin-right:14px;color:#1F3864} table{border-collapse:collapse;margin:10px 0;width:100%}
 th{background:#1F3864;color:#fff;text-align:left;padding:6px 9px;font-size:12px;position:sticky;top:0}
 td{border:1px solid #e0e0e0;padding:6px 9px;vertical-align:top;font-size:13px}
 tr:nth-child(even) td{background:#f4f4f4}
</style>
<h1>Orange Theory National — Audience Evaluation (TI-1026)</h1>
<p><em>Workbook preview. Color coding: <span style="background:#F4CCCC">DROP</span> /
<span style="background:#FFF2CC">REVIEW</span> / <span style="background:#D9EAD3">KEEP</span>.</em></p>
<nav>"""]
parts += [f'<a href="#{escape(ws.title)}">{escape(ws.title)}</a>' for ws in wb.worksheets]
parts.append("</nav>")

for ws in wb.worksheets:
    parts.append(f'<h2 id="{escape(ws.title)}">{escape(ws.title)}</h2><table>')
    for ri, row in enumerate(ws.iter_rows()):
        parts.append("<tr>")
        tag = "th" if ri == 0 and ws.title not in ("Recommendations", "Geo & Exclusions", "Methodology") else "td"
        for cell in row:
            v = "" if cell.value is None else escape(str(cell.value))
            fill = ""
            try:
                rgb = cell.fill.fgColor.rgb
                if rgb and rgb not in ("00000000", None) and isinstance(rgb, str):
                    hexc = rgb[-6:]
                    if hexc.upper() not in ("FFFFFF", "000000"):
                        fill = f' style="background:#{hexc}"'
            except Exception:
                pass
            parts.append(f"<{tag}{fill}>{v}</{tag}>")
        parts.append("</tr>")
    parts.append("</table>")

HTML.write_text("".join(parts))
print(f"Wrote {HTML}")
