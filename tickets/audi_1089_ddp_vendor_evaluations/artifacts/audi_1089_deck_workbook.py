#!/usr/bin/env python3
"""AUDI-1089 deck workbook: the user's 6-block deck sheet as an .xlsx, every cell
filled from the canonical run CSVs (outputs/run_2026_07_10/).

Blocks and their supporting queries (deck_d1..d6, see runbook/queries/MANIFEST.md):
  1 coverage + touched won imps + CPM   <- q3c masks (== deck_d1), q6/q15/q7d
                                           (== deck_d2), deck_d3
  2 CPM / bill per year                 <- deck_d3 (Profit column left for the
                                           user's sheet formula — margins internal)
  3 bill if free_logs preempted         <- deck_d3 bill x (1 - free-cohold share
                                           from q3c masks == deck_d1)
  4 scenario ladder                     <- q3c masks (triples) + deck_d4 (HI/PP)
  5 tier coverage, ALL member IPs       <- deck_d5
  6 tier coverage, IPs that got bid on  <- deck_d6

Cells whose scan hasn't landed yet print PENDING; rerun after the deck_d4/d5/d6
CSVs land and they fill in. Deterministic: rerun = identical file.
"""
import csv
import os

HERE = os.path.dirname(os.path.abspath(__file__))
RUN = os.path.join(HERE, "..", "outputs", "run_2026_07_10")
OUT = os.path.join(HERE, "..", "outputs", "audi_1089_deck_sheet.xlsx")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BITSQ = [23, 24, 25, 26, 28, 30, 33, 36, 39, 40]
BIT = {ds: i for i, ds in enumerate(BITSQ)}
FREE_MASK = 33
PAID_MASK = 990

NAME = {23: "guid_log (free)", 30: "augmentor (free)", 24: "Justuno", 25: "5x5",
        26: "Predactiv", 28: "33Across", 33: "Sovrn", 36: "Cybba", 39: "Klickly",
        40: "33Across API", 99: "free_logs UNION (guid+aug)"}
METERED = {24, 28, 33, 36, 40}
FLAT = {25, 26, 39}


def read_csv(name):
    p = os.path.join(RUN, name)
    if not os.path.exists(p) or os.path.getsize(p) == 0:
        return None
    with open(p) as fh:
        rows = list(csv.DictReader(fh))
    return rows or None


# ---------------------------------------------------------------- inputs
mh = {}
for r in read_csv("q3c_visit_grain_uniqueness.csv"):
    if r["rec"] == "mask":
        mh[int(r["k1"])] = int(float(r["n"]))
UNIVERSE = sum(mh.values())

q6 = {int(r["data_source_id"]): r for r in read_csv("q6_value_tiers.csv")}
q7d = read_csv("q7d_platform_week.csv")[0]
PLAT_IMPS = float(q7d["imps_week"])
q15s = {(r["k1"], r["k2"]): float(r["v"]) for r in read_csv("q15_free_union_perf.csv")
        if r["rec"] == "serve"}
q15_touched_imps = q15s[("touched", "imps")]
q8b_serve = {}
for r in read_csv("q8b_solo_perf.csv"):
    if r["rec"] == "serve":
        q8b_serve.setdefault(int(r["ds"]), {})[r["k"]] = float(r["v"])

MARGIN_LO, MARGIN_HI = 0.10, 0.30  # blended margin band (internal)


def touched_media_imps(ds):
    if ds == 99:
        return q15s[("touched", "media")], q15s[("touched", "imps")]
    return float(q6[ds]["media_touched"]), float(q6[ds]["imps_touched"])


def standalone_media(ds):
    """weekly media on the vendor-as-only-paid-source cohort (q8b solo; union:
    q15 sole-vs-paid)"""
    return q15s[("sole", "media")] if ds == 99 else q8b_serve[ds]["media"]


def unique_cpm(ds):
    """media CPM on the UNIQUE imps only (IPs outside guid+aug; q8b solo)"""
    if ds == 99:
        return q15s[("sole", "media")] / q15s[("sole", "imps")] * 1000
    return q8b_serve[ds]["media"] / q8b_serve[ds]["imps"] * 1000


def profit_band(media_week):
    lo, hi = media_week * 52 * MARGIN_LO, media_week * 52 * MARGIN_HI
    return f"${lo:,.0f} - ${hi:,.0f}"

d3 = {int(r["data_source_id"]): r for r in read_csv("deck_d3_bills_cpm.csv")
      if r["data_source_id"] and int(r["data_source_id"]) != 27}  # DS27 = context row

d4 = read_csv("deck_d4_scenario_ladder.csv")
d4 = {r["scenario"]: r for r in d4} if d4 else None
def load_tier_csv(name):
    rows = read_csv(name)
    if not rows:
        return None
    # superaggregate row: '' in the first-run CSVs (alias-shadow), '1_all_ips' after
    key = "tier_row" if "tier_row" in rows[0] else "tier"
    return {(r[key] or "1_all_ips"): r for r in rows}


d5 = load_tier_csv("deck_d5_tier_free_coverage_all_ips.csv")
d6 = load_tier_csv("deck_d6_tier_free_coverage_bid_ips.csv")
d8 = read_csv("deck_d8_signal_volume_served.csv")
d8 = {r["tier_row"]: r for r in d8} if d8 else None

# q3d: score-tier holder masks (hi/pp/hg) at scored-IP grain, 37d window — the
# ALREADY-MEASURED source for HI/PP coverage; used until the deck_d4/d6 scans
# land at the sheet-exact grain (values differ only by window, <0.1pp)
q3d = {"hi": {}, "pp": {}, "hg": {}}
for r in read_csv("q3d_score_vertical_coverage.csv"):
    if r["rec"] in q3d:
        q3d[r["rec"]][int(r["k1"])] = int(float(r["n"]))


def q3d_cov(tier, keepmask):
    tot = sum(q3d[tier].values())
    kept = sum(n for m, n in q3d[tier].items() if m & keepmask)
    return kept / tot


def q3d_split(tier):
    free = sum(n for m, n in q3d[tier].items() if m & FREE_MASK)
    vonly = sum(n for m, n in q3d[tier].items() if not (m & FREE_MASK))
    return free, vonly, free / (free + vonly)


PENDING = "PENDING (scan running)"


# ---------------------------------------------------------------- mask algebra
def msum(pred):
    return sum(n for m, n in mh.items() if pred(m))


def src_stats(ds):
    b = 1 << BIT[ds]
    fb = {23: 1 << BIT[30], 30: 1 << BIT[23]}.get(ds, FREE_MASK)
    total = msum(lambda m: m & b)
    cohold = msum(lambda m: (m & b) and (m & fb))
    return total, cohold


stats = {ds: src_stats(ds) for ds in BITSQ}
order = sorted(BITSQ, key=lambda d: -stats[d][0])  # rank by total desc (== deck_d1)
FU_TOTAL = msum(lambda m: m & FREE_MASK)

# post-preemption bills (block 3): bill x (1 - free-cohold share), metered only
bill_after = {}
for ds in METERED:
    total, cohold = stats[ds]
    bill_after[ds] = float(d3[ds]["bill_annualized"]) * (1 - cohold / total)
POST_TOTAL = sum(bill_after.values())
assert abs(POST_TOTAL - 538726) < 100, f"preemption tripwire: {POST_TOTAL:,.0f} vs known ~538,726"

SCEN = [  # (sheet label, kept-vendors text, keepmask, deck_d4 scenario key)
    ("Today (all 8)", "Justuno + 5x5 + Predactiv + 33Across + Sovrn + Cybba + Klickly + 33A API", 1023, "today_all_8_paid"),
    ("Drop Sovrn + Cybba", "Justuno + 5x5 + Predactiv + 33Across + Klickly + 33A API", 831, "drop_sovrn_cybba"),
    ("+ drop Klickly", "Justuno + 5x5 + Predactiv + 33Across + 33A API", 575, "plus_drop_klickly"),
    ("+ drop Justuno (knee k=4)", "5x5 + Predactiv + 33Across + 33A API", 573, "plus_drop_justuno_k4"),
    ("33Across combined only", "33Across + 33A API", 561, "33across_combined_only"),
    ("Flat-fee only (5x5 + Predactiv)", "5x5 + Predactiv", 45, "flat_fee_only_5x5_predactiv"),
    ("Free logs only (guid + augmentor)", "(none)", 33, "free_logs_only"),
    ("Free: augmentor DS30 only", "(none)", 32, "augmentor_ds30_only"),
    ("Free: guid_log DS23 only", "(none)", 1, "guid_log_ds23_only"),
]

TIER_ROWS = [("1_all_ips", "Free Logs ONLY (all IPs)"), ("2_hi_10000", "HI10000"),
             ("3_pp_8000", "PP8000"), ("4_high_graduated", "High-Graduated"),
             ("5_mid", "Mid"), ("6_max_reach", "MAX REACH"), ("7_unscored", "Unscored")]


# ---------------------------------------------------------------- xlsx helpers
wb = Workbook()
ws = wb.active
ws.title = "deck"

HDR_FILL = PatternFill("solid", fgColor="203864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
N0 = "#,##0"
MONEY = "$#,##0"
# percent columns store TRUE FRACTIONS (0.488) with a percent number format —
# they render 48.8% and survive the user re-applying Excel's percent data type
PCT1 = "0.0%"
PCT2 = "0.00%"
PCT4 = "0.0000%"

row = 1
col_max = {}  # data-driven column widths, collected while emitting


def _track(c, v, cap=46):
    if v is None or v == "":
        return
    if isinstance(v, float):
        s = f"{v:,.2f}"
    elif isinstance(v, int):
        s = f"{v:,}"
    else:
        s = str(v)
    col_max[c] = max(col_max.get(c, 0), min(len(s), cap))


def header(cells):
    global row
    for c, text in enumerate(cells, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP
        # headers wrap over 2-3 lines — contribute half their length to width
        _track(c, text[: max(len(text) // 2, 14)])
    ws.row_dimensions[row].height = 40
    row += 1


def emit(cells, fmts=None):
    global row
    for c, v in enumerate(cells, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fmts and fmts.get(c) and isinstance(v, (int, float)):
            cell.number_format = fmts[c]
        _track(c, v)
    row += 1


def note(text, span):
    """footnote merged across the block's width so it wraps inside the table
    instead of blowing out column A"""
    global row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(italic=True, color="666666", size=9)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    est_chars_per_line = span * 14
    lines = max(1, -(-len(text) // est_chars_per_line))
    ws.row_dimensions[row].height = lines * 12 + 4
    row += 1


def gap(n=2):
    global row
    row += n


def cpm_label(ds):
    if ds in METERED:
        return 0.50
    if ds in FLAT:
        return "flat (pending)"
    return "$0 (internal)"


def bill_label(ds):
    if ds in METERED:
        return float(d3[ds]["bill_annualized"])
    if ds in FLAT:
        return "flat (pending)"
    return 0


# ---------------------------------------------------------------- block 1
header(["Vendor", "Total IP x Domain x Date triples (usable, 30d)", "% of total universe",
        "Cumulative % of universe (union, top-N by total)",
        "Touched won imps (valuation wk)",
        "Share of ALL our won imps that landed on this source's IPs"])
fmts1 = {2: N0, 3: PCT2, 4: PCT2, 5: N0, 6: PCT1}
km = 0
for ds in order:
    total, _ = stats[ds]
    km |= 1 << BIT[ds]
    cum = msum(lambda m: m & km)
    media, imps = touched_media_imps(ds)
    emit([NAME[ds], total, total / UNIVERSE, cum / UNIVERSE,
          imps, imps / PLAT_IMPS], fmts1)
media99, imps99 = touched_media_imps(99)
emit([NAME[99], FU_TOTAL, FU_TOTAL / UNIVERSE, "—",
      imps99, imps99 / PLAT_IMPS], fmts1)
gap()

# ---------------------------------------------------------------- block 2
header(["Vendor", "Media CPM (touched, $)", "Media CPM (UNIQUE imps only, $)",
        "Bill / yr cost",
        "Profit on UNIQUE media — won imps on IPs outside guid_log+augmentor — "
        "x 10-30% margin ($/yr)",
        "Vendor contract CPM ($)"])
fmts2 = {2: "0.00", 3: "0.00", 4: MONEY}
b2_order = sorted(METERED, key=lambda d: -float(d3[d]["bill_annualized"])) + sorted(FLAT)
for ds in b2_order:
    media, imps = touched_media_imps(ds)
    emit([NAME[ds], media / imps * 1000, unique_cpm(ds), bill_label(ds),
          profit_band(standalone_media(ds)), cpm_label(ds)], fmts2)
note("* UNIQUE = the vendor's value beyond the free logs: won imps whose IP neither"
     " guid_log nor augmentor delivered (measured, q8b solo cohort). UNIQUE CPM = media"
     " / imps x 1000 on that cohort alone; profit = its media x52 x 10-30% blended"
     " margin. The free logs' own value is a DIFFERENT cohort — see the FREE LOGS table"
     " below.", 6)
gap()

# ---------------------------------------------------------------- block 3
header(["Vendor", "Bill/yr if free_logs preempted from billing",
        "Profit on UNIQUE media (outside free logs) x 10-30% ($/yr) — unchanged by preemption",
        "free co-hold % (share of vendor's visit-days a free log also holds)"])
fmts3 = {2: MONEY, 4: PCT1}
for ds in b2_order:
    if ds in METERED:
        total, cohold = stats[ds]
        emit([NAME[ds], bill_after[ds], profit_band(standalone_media(ds)),
              cohold / total], fmts3)
    else:
        emit([NAME[ds], "flat (pending) — preemption does not change flat fees",
              profit_band(standalone_media(ds)), ""], fmts3)
emit(["TOTAL metered (today $812,397)", POST_TOTAL, "",
      f"savings ${812397 - POST_TOTAL:,.0f}/yr = the AUDI-1093 preemption"], fmts3)
gap()

# ------------------------------------------------ free-logs value block (own query: d7)
header(["FREE LOGS — value beyond ALL 8 paid vendors combined",
        "Media $/yr on IPs no paid vendor covers",
        "Profit x 10-30% margin ($/yr)", "Media CPM on this cohort ($)",
        "Cohort measured"])
fmtsF = {2: MONEY, 4: "0.00"}
guid_w = float(q6[23]["media_sole"])
aug_w = float(q6[30]["media_sole"])
union_w = q15s[("sole", "media")]
emit([NAME[23], guid_w * 52, profit_band(guid_w),
      guid_w / float(q6[23]["imps_sole"]) * 1000,
      "IPs ONLY guid_log holds (no paid vendor, no augmentor)"], fmtsF)
emit([NAME[30], aug_w * 52, profit_band(aug_w),
      aug_w / float(q6[30]["imps_sole"]) * 1000,
      "IPs ONLY augmentor holds (no paid vendor, no guid_log)"], fmtsF)
emit([NAME[99], union_w * 52, profit_band(union_w),
      union_w / q15s[("sole", "imps")] * 1000,
      "IPs NO paid vendor holds (either free log has them)"], fmtsF)
note(f"* The union row (${union_w * 52:,.0f}) EXCEEDS guid+aug summed"
     f" (${(guid_w + aug_w) * 52:,.0f}) — IPs BOTH free logs hold with no paid co-holder"
     " count only in the union row (measured cohort algebra, q6/q15; single query:"
     " deck_d7). This is what the free logs alone protect — no paid roster money"
     " currently buys this slice.", 5)
gap()

# ---------------------------------------------------------------- block 4
TODAY_TRIPS = float(d4["today_all_8_paid"]["trips_kept"]) if d4 else UNIVERSE
header(["Scenario", "Paid vendors kept",
        f"Total triples kept (of {TODAY_TRIPS:,.0f} possible today)",
        "Coverage (% of today)", "Triples LOST vs today",
        "HI triples kept", "HI-IP coverage %", "PP triples kept", "PP-IP coverage %"])
fmts4b = {3: N0, 4: PCT2, 5: N0, 6: N0, 7: PCT4, 8: N0, 9: PCT4}
for label, kept_txt, keepmask, key in SCEN:
    if d4 and key in d4:
        # every block-4 cell from the ONE supporting query once it lands
        r4 = d4[key]
        kept, pct = float(r4["trips_kept"]), float(r4["pct_of_today"]) / 100
        hi_t, hi_p = float(r4["hi_trips_kept"]), float(r4["hi_ip_coverage_pct"]) / 100
        pp_t, pp_p = float(r4["pp_trips_kept"]), float(r4["pp_ip_coverage_pct"]) / 100
    else:
        # pre-landing: triples from the q3c masks; HI/PP coverage from q3d (37d)
        kept = msum(lambda m: m & keepmask)
        pct = kept / UNIVERSE
        hi_t = pp_t = PENDING
        hi_p, pp_p = q3d_cov("hi", keepmask), q3d_cov("pp", keepmask)
    emit([label, kept_txt, kept, pct, TODAY_TRIPS - kept,
          hi_t, hi_p, pp_t, pp_p], fmts4b)
if not d4:
    note("* HI/PP coverage % = the already-measured q3d masks (37d window, IP grain) —"
         " deck_d4 (running) replaces them at the sheet's 30d grain; expect <0.1pp"
         " movement. HI/PP triple counts have no prior measurement (pending d4).", 8)
gap()

# ---------------------------------------------------------------- blocks 5 & 6
Q3D_TIER = {"2_hi_10000": "hi", "3_pp_8000": "pp", "4_high_graduated": "hg"}
for title, data, allow_q3d in [
    ("Of ALL member IPs, served or not (audience-size coverage)", d5, False),
    ("Of member IPs that actually got bid on (won impressions)", d6, True),
]:
    header([title, "free-covered IPs", "vendor-only IPs", "% free covered"])
    fmts56 = {2: N0, 3: N0, 4: PCT2}
    used_q3d = False
    for key, label in TIER_ROWS:
        if data and key in data:
            r56 = data[key]
            emit([label, float(r56["free_covered_ips"]), float(r56["vendor_only_ips"]),
                  float(r56["pct_free_covered"]) / 100], fmts56)
        elif allow_q3d and key in Q3D_TIER:
            # q3d's scored-IP population IS this block's population for scored
            # tiers (scored in CIL week => served)
            free, vonly, pct = q3d_split(Q3D_TIER[key])
            emit([label, free, vonly, pct], fmts56)
            used_q3d = True
        else:
            emit([label, PENDING, PENDING, PENDING], fmts56)
    if used_q3d:
        note("* HI/PP/High-Graduated rows = the already-measured q3d masks (37d, scored-IP"
             " grain); deck_d6 (running) replaces at sheet-exact grain. Mid/MAX REACH/"
             "Unscored have no prior measurement at this population.", 4)
    gap()

# data-driven widths: wide enough that nothing clips, capped so no column hogs
# ------------------------------------------- signal-volume lens (from deck_d4)
if d4:
    t_row, f_row = d4["today_all_8_paid"], d4["free_logs_only"]
    hi_tot, hi_free = float(t_row["hi_trips_kept"]), float(f_row["hi_trips_kept"])
    pp_tot, pp_free = float(t_row["pp_trips_kept"]), float(f_row["pp_trips_kept"])
    all_tot, all_free = float(t_row["trips_kept"]), float(f_row["trips_kept"])
    ot_tot = all_tot - hi_tot - pp_tot
    ot_free = all_free - hi_free - pp_free
    header(["SIGNAL VOLUME by tier — IP x Domain x Date triples (usable, 30d), "
            "free logs vs vendors", "Triples today (all sources)",
            "Free-covered triples (kept under free-only)",
            "Vendor-only triples (LOST under free-only)", "% free covered"])
    fmtsS = {2: N0, 3: N0, 4: N0, 5: PCT2}
    for label, tot, free in [("HI10000 IPs' triples", hi_tot, hi_free),
                             ("PP8000 IPs' triples", pp_tot, pp_free),
                             ("All other tiers' triples", ot_tot, ot_free),
                             ("ALL triples", all_tot, all_free)]:
        emit([label, tot, free, tot - free, free / tot], fmtsS)
    note("* The THIRD lens: blocks 5/6 count IPs (membership — would we still know the"
         " household?); this table counts triples (signal — would we still have the same"
         " domain-and-date evidence about it?). Both are true at once: free-only keeps"
         " 99.76% of HI member IPs but only 65.7% of the signal volume on them — the"
         " scores survive because membership needs SOME qualifying signal, not ALL of it."
         " Tier assignment = per-IP MAX(household_score), CIL valuation week; rows split"
         " hi / pp / everything-else because deck_d4's histogram carries exactly those"
         " three tiers at triple grain. Derived from deck_d4's today + free_logs_only"
         " rows — no separate query.", 5)
    gap()

# ---------------------------------- signal volume on SERVED IPs only (deck_d8)
D8_ROWS = [("2_hi_10000", "HI10000 IPs' triples"),
           ("3_pp_8000", "PP8000 IPs' triples"),
           ("4_served_other_tiers", "Other served tiers' triples"),
           ("5_not_served_this_week", "Not served this week (context)")]
header(["SIGNAL VOLUME on SERVED IPs only — triples on IPs that actually got won"
        " impressions (valuation wk)", "Triples (all sources)",
        "Free-covered triples", "Vendor-only triples (LOST under free-only)",
        "% free covered"])
fmtsS8 = {2: N0, 3: N0, 4: N0, 5: PCT2}
if d8:
    served_tot = served_free = 0
    for key, label in D8_ROWS:
        r8 = d8[key]
        tot, free = float(r8["triples_total"]), float(r8["triples_free_covered"])
        if key != "5_not_served_this_week":
            served_tot += tot
            served_free += free
        emit([label, tot, free, tot - free, free / tot], fmtsS8)
    emit(["ALL SERVED IPs' triples", served_tot, served_free,
          served_tot - served_free, served_free / served_tot], fmtsS8)
else:
    for key, label in D8_ROWS + [("all_served", "ALL SERVED IPs' triples")]:
        emit([label, PENDING, PENDING, PENDING, PENDING], fmtsS8)
note("* Same lens as the table above, restricted to IPs we actually spent money on"
     " (the served population of block 6). Splits deck_d4's mixed 'other' bucket into"
     " served-other vs never-served. HI/PP rows should match the table above within"
     " snapshot drift (scored IPs are served by definition). ALL SERVED = the three"
     " served tiers summed (sheet arithmetic). Query: deck_d8.", 5)
gap()

for c, m in col_max.items():
    cap = 38 if c == 1 else 30
    ws.column_dimensions[get_column_letter(c)].width = max(min(m + 3, cap), 11)

# ---------------------------------------------------------------- queries sheet
qs = wb.create_sheet("queries")
qrow = [["Block", "Supporting query (runbook/queries/)", "What it computes", "Status"],
        ["1 cols B-D", "deck_d1_universe_coverage.sql", "triple holder-mask histogram -> totals, % of universe, cumulative union % (numbers here derive from the identical measured q3c histogram)", "measured (q3c run_2026_07_10)"],
        ["1 cols E-F", "deck_d2_touched_won_bids.sql", "won imps on touched IPs + % of platform won imps (RETITLED col F: share of platform, NOT a win rate)", "measured (q6/q15/q7d)"],
        ["2", "deck_d3_bills_cpm.sql", "registry roster, contract/implied CPM, June 2026 meter bill x 12; media CPMs from q6/q15 media+imps", "run 2026-07-16"],
        ["3", "deck_d3 x deck_d1", "bill_after = bill x (1 - free-cohold share); sheet formula, inputs in this workbook", "computed"],
        ["FREE LOGS table", "deck_d7_free_logs_value.sql", "the reverse cohort: free-side media on IPs OUTSIDE the paid roster (guid strictly-sole / augmentor strictly-sole / union-no-paid); union > sum of the two by cohort algebra", "measured (q6 sole + q15; d7 = optional independent re-run)"],
        ["2/3 Profit cols", "q8b_solo_perf.sql (serve media per ds); union: q15_free_union_perf.sql (sole media)", "profit band = weekly UNIQUE media x 52 x 10-30% margin — computed offline; margin band internal, never in shared queries", "measured"],
        ["4", "deck_d4_scenario_ladder.sql", "9 keep-set scenarios: triples kept, % of today, triples LOST (= today row minus scenario, sheet arithmetic), HI/PP triples + IP-grain coverage",
         "landed — all cells measured" if d4 else "triples measured; HI/PP coverage shown from q3d (37d) until the d4 scan lands"],
        ["5", "deck_d5_tier_free_coverage_all_ips.sql", "ALL member IPs by score tier: free-covered vs vendor-only",
         "landed — all cells measured" if d5 else "scan running"],
        ["6", "deck_d6_tier_free_coverage_bid_ips.sql", "same split, only IPs with won impressions",
         "landed — all cells measured" if d6 else "scan running"],
        ["SIGNAL VOLUME table", "deck_d4_scenario_ladder.sql (today + free_logs_only rows)", "triple-grain (ip x domain x date) free coverage per tier — the signal lens next to blocks 5/6's membership lens; sheet arithmetic over d4's output", "derived (d4 landed)"],
        ["SIGNAL VOLUME, SERVED IPs", "deck_d8_signal_volume_served.sql", "same triple-grain lens restricted to IPs with won impressions; splits d4's 'other' into served-other vs never-served",
         "landed — all cells measured" if d8 else "scan running"]]
for i, rr in enumerate(qrow, 1):
    for c, v in enumerate(rr, 1):
        cell = qs.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        if i == 1:
            cell.fill, cell.font = HDR_FILL, HDR_FONT
for c, w in enumerate([12, 38, 60, 28], 1):
    qs.column_dimensions[get_column_letter(c)].width = w

# ---------------------------------------------------------------- notes sheet
ns = wb.create_sheet("notes")
notes = [
    "Windows: triples/uniqueness 30d svs (dt 2026-06-02..07-01); serving membership 37d (..07-08); CIL valuation week 2026-07-02..08; bills June 2026 x 12.",
    "Universe = 13,286,670,656 distinct usable (ip x domain x date) triples across all 10 sources. 'Usable' = domain consumable by DS13 or DS19.",
    "Block 1 rows are ranked by total triples desc (matches deck_d1's rank order); cumulative % = deduplicated UNION coverage of the top-N rows, NOT a running sum of column C.",
    "Free-log rows' numbers are their own totals; the free_logs UNION row is measured on the union (never a sum of the two rows). A component can exceed the union on 'standalone'-type cuts — see artifacts/audi_1089_deck_coverage.md for that decomposition.",
    "Column 'Touched won imps' is NOT additive across rows (an IP delivered by several sources counts for each). Its % column = share of ALL 398,301,655 won imps (val wk) that landed on the source's IPs — a reach share, not a win rate.",
    "Block 2 has THREE CPMs: touched media CPM (revenue rate on everything the source touched), UNIQUE media CPM (revenue rate on just the imps landing on IPs outside guid+aug — the slice the profit column values), and the vendor's contract CPM ($0.50 = what we PAY per billed impression).",
    "Blocks 2/3 Profit = the vendor's UNIQUE contribution: media on won imps whose IP neither free log delivered (q8b solo cohort = 'the vendor as our only paid source'), annualized x52, x 10-30% blended margin (INTERNAL — do not quote the margin band outside the team). Touched-media profit is deliberately NOT shown: every large vendor touches ~90-98% of platform imps, so touched bands overlap almost entirely and misread as huge per-vendor value.",
    "FREE LOGS table = the REVERSE cohort of blocks 2/3: free-side media on IPs no PAID vendor covers (guid strictly-sole $277.5K/yr, augmentor strictly-sole $167.5K/yr, union-no-paid $602.9K/yr, each x margin). The union exceeds the two summed because IPs both free logs hold (no paid) count only in the union. Supporting query: deck_d7 (values already measured by q6/q15).",
    f"Block 3: preemption applies to METERED vendors only (flat fees don't meter). bill_after = bill x (1 - free co-hold share of the vendor's visit-days), the AUDI-1093 fix. Metered total falls $812,397 -> ${POST_TOTAL:,.0f} (-${812397 - POST_TOTAL:,.0f}/yr).",
    "Block 4: 'HI/PP triples kept' counts signal volume on HI/PP IPs (drops roughly with column C); 'HI/PP-IP coverage %' counts audience MEMBERS still covered (stays 99%+ in every paid-drop scenario). Different grains on purpose — audience size is the IP-grain number.",
    "Blocks 5 vs 6 are the same split over two populations: ALL member IPs (audience-size lens) vs only IPs that got won impressions (delivery-reality lens). Coverage % can differ sharply between them — that gap is the point.",
    "Block 5/6 '% free covered' denominator = member IPs of that row (free-covered + vendor-only). Unscored in block 5 includes IPs never served in the valuation week; in block 6 every row was served.",
    "'High-Graduated' (6666-9999 excl. 8000) is a score-value bucket, not a product tier: legacy scoring pins HI/PP at exactly 10000/8000; Fangorn (newer scoring) emits graduated scores filling those bands (8001-9999 = HI band, 6666-7999 = PP band). HI10000 + PP8000 + High-Graduated together ~= the HI/PP audiences across both scoring generations. Mid (3333-6665) ~= MI; Max Reach = 1-3332.",
    "DS27 LaunchLabs appears in deck_d3's CSV as a registry context row (disabled, never metered) — deliberately excluded from this workbook.",
    "Rebuild: python3 artifacts/audi_1089_deck_workbook.py (rerun after the deck_d4/d5/d6 scans land to replace PENDING cells).",
]
ns.cell(row=1, column=1, value="Reading notes").font = Font(bold=True, size=12)
for i, t in enumerate(notes, 3):
    cell = ns.cell(row=i, column=1, value="• " + t)
    cell.alignment = WRAP
    ns.row_dimensions[i].height = max(1, -(-len(t) // 110)) * 13 + 4
ns.column_dimensions["A"].width = 120

wb.save(OUT)
pend = 0
for sheet in wb.worksheets:
    for r in sheet.iter_rows():
        for c in r:
            if c.value == PENDING:
                pend += 1
print(f"wrote {OUT}")
print(f"pending cells: {pend} (deck_d4 {'LANDED' if d4 else 'running'}, "
      f"d5 {'LANDED' if d5 else 'running'}, d6 {'LANDED' if d6 else 'running'})")
print(f"post-preemption metered total ${POST_TOTAL:,.0f} (savings ${812397 - POST_TOTAL:,.0f})")
