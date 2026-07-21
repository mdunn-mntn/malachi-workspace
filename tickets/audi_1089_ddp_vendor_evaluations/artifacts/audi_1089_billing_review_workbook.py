#!/usr/bin/env python3
"""Build the AUDI-1089 DDP billing-review workbook (.xlsx) for the billing team.
Reproducible: numbers from ../outputs/*.csv, SQL from the .sql files. Rerun -> identical file.
Sheets: Overview, 1 Meter Proof, 2 Preemption (fair), 3 Augmentor Fix, 4 Worth vs Bill,
5 Recommendations, 6 Audit Map, then one sheet per runnable query."""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
TICKET = os.path.dirname(HERE)
OUT = os.path.join(TICKET, "outputs")

NAVY, RED, GRAY, LGRAY = "1B2A4A", "C0392B", "666666", "EDEFF2"
HDR = PatternFill("solid", fgColor=NAVY)
ALT = PatternFill("solid", fgColor="F5F7FA")
KICK = PatternFill("solid", fgColor=LGRAY)
WHITE_B = Font(color="FFFFFF", bold=True, size=11)
TITLE = Font(color=NAVY, bold=True, size=15)
SUB = Font(color=GRAY, size=10, italic=True)
BOLD = Font(bold=True)
REDB = Font(color=RED, bold=True)
NAVYB = Font(color=NAVY, bold=True)
MONO = Font(name="Consolas", size=9)
THIN = Side(style="thin", color="D6DBE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right")

VNAME = {24: "Justuno", 25: "5x5", 26: "Predactiv", 28: "33Across", 33: "Sovrn",
         36: "Cybba", 39: "Klickly", 40: "33Across API"}
BILL = {"33Across": 422024, "33Across API": 175879, "Sovrn": 115880, "Justuno": 77111, "Cybba": 21504}


def load(name):
    with open(os.path.join(OUT, name)) as f:
        return list(csv.DictReader(f))


def hrow(ws, r, headers, start=1):
    for j, h in enumerate(headers):
        c = ws.cell(r, start + j, h)
        c.fill = HDR; c.font = WHITE_B; c.alignment = CTR; c.border = BORDER
    ws.row_dimensions[r].height = 26


def drow(ws, r, vals, start=1, fmts=None, alt=False, bolds=None, colors=None):
    for j, v in enumerate(vals):
        c = ws.cell(r, start + j, v)
        c.border = BORDER
        if alt:
            c.fill = ALT
        if fmts and fmts[j]:
            c.number_format = fmts[j]
        if isinstance(v, (int, float)) and (not fmts or "@" not in (fmts[j] or "")):
            c.alignment = RIGHT
        if bolds and bolds[j]:
            c.font = BOLD
        if colors and colors[j]:
            c.font = Font(color=colors[j], bold=(bolds[j] if bolds else False))


def title_block(ws, title, sub, kicker=None, query_ref=None):
    if kicker:
        ws["A1"] = kicker; ws["A1"].font = Font(color=RED, bold=True, size=9)
    ws["A2"] = title; ws["A2"].font = TITLE
    ws["A3"] = sub; ws["A3"].font = SUB
    if query_ref:
        ws["A4"] = "Query ▸ " + query_ref
        ws["A4"].font = Font(color="1E7A34", bold=True, size=9)
    ws.row_dimensions[2].height = 22


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


wb = Workbook()

# ---------------- Overview ----------------
ws = wb.active; ws.title = "Overview"
widths(ws, {"A": 30, "B": 60, "C": 18})
ws["A1"] = "AUDI-1089 — DDP Vendor Billing Review"; ws["A1"].font = Font(color=NAVY, bold=True, size=16)
ws["A2"] = "Audit workbook for the DDP billing team — every number traces to a read-only query."
ws["A2"].font = SUB
r = 4
ws.cell(r, 1, "The four findings").font = NAVYB; r += 1
for a, b in [
    ("1. The meter does NOT preempt",
     "tv_cpm = $0.50 whenever ANY paid vendor wins, $0 only when none does. Free co-presence is ignored — 268.9M June impressions billed $0.50 despite a free log co-winning the same impression."),
    ("2. Fair preemption = $200.4K/yr (-24.7%)",
     "Roster $812.4K -> $612.0K, keeping ALL vendor data. Conservative (free-dominant); upper bound $243.5K. A naive same-day test says $273.7K but over-credits augmentor (the bid stream) — see sheet 3."),
    ("3. No metered vendor paid for itself",
     "On money-made (profit its unique data produced), every metered vendor is < 1.0x after preemption: API 0.94x, 33Across 0.83x, Sovrn 0.29x, Cybba 0.17x, Justuno 0.15x. (Justuno/Cybba have separate domain-licensing value if kept for coverage — see sheet 4.)"),
    ("4. Recommendation",
     "Preempt (-$200K) -> renegotiate the two 33Across feeds toward the fair cap -> drop Sovrn + Cybba. Lock flat-fee (5x5/Predactiv) prices first."),
]:
    ws.cell(r, 1, a).font = BOLD; ws.cell(r, 1).alignment = Alignment(vertical="top")
    c = ws.cell(r, 2, b); c.alignment = LEFT
    ws.row_dimensions[r].height = 58; r += 1
r += 1
ws.cell(r, 1, "Windows").font = NAVYB
ws.cell(r, 2, "Delivery/uniqueness = 30d svs (2026-06-02..07-01); fair-preemption scan = 37d (05-26..07-01) measured over the last 7d; serving = valuation week 07-02..08; bills = June 2026 x 12. All read-only."); ws.cell(r, 2).alignment = LEFT
ws.row_dimensions[r].height = 42; r += 2
ws.cell(r, 1, "Caveats").font = NAVYB
ws.cell(r, 2, "GRAIN (conservative): coverage is matched on the exact DOMAIN, but targeting keys off the CATEGORY the domain "
              "falls into (DS13 vertical / DS19 keyword) — 'did this IP have a prior visit in this vertical/keyword'. A free log seeing a "
              "DIFFERENT same-category domain prior already makes the IP targetable, but the domain grain doesn't count it — so $200K is a "
              "FLOOR; the category grain recovers more. Also: N=1 valuation week (July trough) -> dependency = envelope not a CI; meter regime "
              "changed May 2026 (never mix months); flat-fee amounts pending finance; winners-table imp counts over-count the meter (prove the RATE, not the $)."); ws.cell(r, 2).alignment = LEFT
ws.row_dimensions[r].height = 92; r += 2
ws.cell(r, 1, "Sheets").font = NAVYB
ws.cell(r, 2, "Data: 1 Meter Proof · 2 Preemption (fair) · 3 Augmentor Fix · 4 Worth vs Bill · 5 Recommendations · 6 Audit Map. "
              "Each data sheet names its query on the green 'Query ▸' line. SQL: all 8 backing queries are embedded as "
              "'Qn ...' sheets (Q1 Meter Proof, Q2 Targeted Signal, Q3 Fair Prior-Day, Q4 Bills, Q5 Dependency, Q6 Domain band, Q7 Free coverage, Q8 Drop savings) — paste-and-run.")
ws.cell(r, 2).alignment = LEFT; ws.row_dimensions[r].height = 56

# ---------------- 1 Meter Proof ----------------
ws = wb.create_sheet("1 Meter Proof")
widths(ws, {"A": 42, "B": 20, "C": 16, "D": 40})
title_block(ws, "The meter does not preempt", "gold.reporting.ddp_mm_winners_imp_202606 (June, verified live). tv_cpm tracks paid-vendor presence only.", "PROOF 1", query_ref="sheet 'Q1 Meter Proof'")
r = 5; hrow(ws, r, ["Winners on the impression", "June impressions", "tv_cpm charged", "Reading"])
mp = [
    ("Free log (23/30) AND paid vendor both win", 268886220, "$0.50 on 100%", "PAID STILL BILLS on a free-covered impression"),
    ("Free log wins alone", 165726546, "$0 on 100%", "free never bills (correct)"),
    ("Paid vendor wins alone", 38160144, "$0.50 on 100%", "legitimately paid"),
    ("Neither (3P / other path)", 20662953, "$0 on 100%", "no MM winner"),
]
for i, (a, imp, cpm, rd) in enumerate(mp):
    r += 1; drow(ws, r, [a, imp, cpm, rd], fmts=[None, "#,##0", None, None], alt=(i % 2))
    ws.cell(r, 4).alignment = LEFT
    if i == 0:
        for cc in range(1, 5): ws.cell(r, cc).font = REDB
r += 2
ws.cell(r, 1, "If the meter preempted, the top row's tv_cpm would be $0. It is $0.50 — a free log co-winning the exact impression has ZERO effect on the charge. Query: sheet 'Q1 Meter Proof'.").font = SUB
ws.cell(r, 1).alignment = LEFT; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); ws.row_dimensions[r].height = 44

# ---------------- 2 Preemption (fair) ----------------
q3e = {int(x["ds"]): x for x in load("q3e_v2_free_prior_lookback.csv")}
ws = wb.create_sheet("2 Preemption (fair)")
widths(ws, {"A": 18, "B": 15, "C": 16, "D": 15, "E": 15})
title_block(ws, "Fair preemption — prior-day, recency-credited", "Recoverable = bill x free_prior_dominant share (q3e_v2). Ranked by recoverable $.", "FAIR = $200.4K/yr", query_ref="'Q3 Fair Prior-Day' (shares) × 'Q4 Bills (q0)' (bills)")
r = 5; hrow(ws, r, ["Vendor", "Bill / yr", "Fair prior-day %", "Recoverable", "Bill after"])
order = ["33Across", "33Across API", "Cybba", "Justuno", "Sovrn"]
dsmap = {"33Across": 28, "33Across API": 40, "Cybba": 36, "Justuno": 24, "Sovrn": 33}
tot_bill = tot_rec = 0
for i, v in enumerate(order):
    b = BILL[v]; share = float(q3e[dsmap[v]]["pct_prior_dominant"]) / 100
    rec = round(b * share); after = b - rec; tot_bill += b; tot_rec += rec
    r += 1; drow(ws, r, [v, b, share, rec, after],
                 fmts=[None, "$#,##0", "0.0%", "$#,##0", "$#,##0"], alt=(i % 2))
    ws.cell(r, 4).font = REDB
r += 1
drow(ws, r, ["ROSTER", tot_bill, tot_rec / tot_bill, tot_rec, tot_bill - tot_rec],
     fmts=[None, "$#,##0", "0.0%", "$#,##0", "$#,##0"], bolds=[1, 1, 1, 1, 1])
for cc in range(1, 6): ws.cell(r, cc).fill = KICK; ws.cell(r, cc).font = NAVYB
r += 2
ws.cell(r, 1, "Sovrn & Justuno are barely overlap-driven — preemption does not fix them (see sheet 4). Same-day (naive) = $273.7K; upper bound (incl. vendor-fresher slice) = $243.5K.").font = SUB
ws.cell(r, 1).alignment = LEFT; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); ws.row_dimensions[r].height = 40

# ---------------- 3 Augmentor Fix ----------------
ws = wb.create_sheet("3 Augmentor Fix")
widths(ws, {"A": 16, "B": 13, "C": 13, "D": 13, "E": 15, "F": 15, "G": 12})
title_block(ws, "Why $200K and not $274K — the augmentor fix",
            "augmentor (DS30) is the SSP bid stream, so same-day cohold is circular. Fair test = prior-day within 30d AND free still as fresh. q3e_v2: 37d scan / 7d measure / full lookback / ALL IPs.", "METHODOLOGY", query_ref="sheet 'Q3 Fair Prior-Day' (q3e_v2)")
r = 5; hrow(ws, r, ["Vendor", "same-day %", "prior-30d %", "FAIR %", "triples", "prior_dominant", "no_free"])
rows3 = sorted(q3e.values(), key=lambda x: -float(x["pct_sameday_old"]))
for i, x in enumerate(rows3):
    v = VNAME[int(x["ds"])]
    r += 1
    drow(ws, r, [v, float(x["pct_sameday_old"]) / 100, float(x["pct_prior30_fair"]) / 100,
                 float(x["pct_prior_dominant"]) / 100, int(x["triples"]),
                 int(x["free_prior_dominant"]), int(x["no_free"])],
         fmts=[None, "0.0%", "0.0%", "0.0%", "#,##0", "#,##0", "#,##0"], alt=(i % 2))
    ws.cell(r, 4).font = NAVYB
r += 2
for line in [
    "same-day % = a free log has the SAME (ip,domain,DATE) — reproduces q3c; inflated for augmentor (bid stream).",
    "prior-30d % = a free log had the pair on a PRIOR day within the 30-day window (household already targetable).",
    "FAIR % (prior_dominant) = prior-30d AND the free log is still >= as fresh as the vendor -> fully preemptable.",
    "The gap prior-30d - FAIR = the slice where the VENDOR is the freshest source (its recency value; credited to the vendor).",
]:
    ws.cell(r, 1, line).font = SUB; ws.cell(r, 1).alignment = LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); r += 1

# ---------------- 4 Worth vs Bill ----------------
ws = wb.create_sheet("4 Worth vs Bill")
widths(ws, {"A": 15, "B": 16, "C": 22, "D": 11, "E": 21, "F": 26})
title_block(ws, "Did the vendor pay for itself? — money-made vs bill",
            "WORTH = money-made = (media revenue on the vendor's UNIQUE serves) × margin, most generous. Data-licensing value (unique domains) is a SEPARATE coverage/keep comp, not money-made. Ranked by worth/bill.", "RESULT",
            query_ref="'Q5 Dependency (q6)' = money-made · 'Q6 Domain band (q4)' = licensing")
r = 5; hrow(ws, r, ["Vendor", "Bill after preempt", "Value produced (money-made)", "Worth / bill", "Data-licensing (domains)", "Read"])
# (vendor, bill_after, money_made_ceiling, domain_value, read)
worth = [("33Across API", 142814, 134000, 36000, "just under — pays only at the ceiling"),
         ("33Across", 259967, 217000, 89000, "~1.2x over"),
         ("Sovrn", 115764, 34000, 2400, "~3x over"),
         ("Cybba", 17698, 3000, 4700, "~6x over"),
         ("Justuno", 75800, 11000, 60000, "~7x over on profit; domain value $60K if kept for coverage")]
for i, (v, after, money, dom, rd) in enumerate(worth):
    ratio = money / after
    r += 1; drow(ws, r, [v, after, money, ratio, dom, rd],
                 fmts=[None, "$#,##0", "$#,##0", '0.00"x"', "$#,##0", None], alt=(i % 2))
    ws.cell(r, 4).font = REDB if ratio < 0.5 else NAVYB
    ws.cell(r, 6).alignment = LEFT
r += 2
ws.cell(r, 1, "Money-made = the profit the vendor's UNIQUE (sole) data actually produced: unique won impressions × MNTN's ~$11.5 media eCPM × margin, x52, most-generous (solo counterfactual). Every metered vendor < 1.0x -> none paid for itself even at the ceiling. The domain column is a DATA-LICENSING comp (unique classified domains x per-domain rate) — it does NOT reflect money made; it's the coverage value that can justify KEEPING a vendor (the reason the flat-fee 5x5/Predactiv are kept). Caveats: N=1 week, sole cut only, margin is an internal assumption -> generous.").font = SUB
ws.cell(r, 1).alignment = LEFT; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); ws.row_dimensions[r].height = 88

# ---------------- 5 Recommendations ----------------
ws = wb.create_sheet("5 Recommendations")
widths(ws, {"A": 16, "B": 6, "C": 12, "D": 14, "E": 14, "F": 13, "G": 34})
title_block(ws, "What we should pay", "Move 1: preempt (-$200K). Move 2: reprice residual toward the fair cap. Sequence: lock flats -> preempt -> renegotiate 33Across -> drop Sovrn/Cybba.", "RECOMMENDATION", query_ref="composite: 'Q4 Bills (q0)' × 'Q3 Fair Prior-Day' × 'Q5 Dependency (q6)' / 'Q6 Domain band (q4)'")
r = 5; hrow(ws, r, ["Vendor", "DS", "Billing", "Current / yr", "After preempt", "Cap at fair", "Action"])
rec = [("33Across", 28, "$0.50 CPM", 422024, 259967, "<=$217K", "Renegotiate — biggest lever"),
       ("33Across API", 40, "$0.50 CPM", 175879, 142814, "<=$134K", "Renegotiate / drop (same vendor as DS28)"),
       ("Sovrn", 33, "$0.50 CPM", 115880, 115764, "<=$34K", "DROP — not overlap-driven"),
       ("Justuno", 24, "$0.50 CPM", 77111, 75800, "<=$60K", "Trim the meter"),
       ("Cybba", 36, "$0.50 CPM", 21504, 17698, "<=$4.7K", "DROP"),
       ("Klickly", 39, "flat", None, None, "<=$1.5K", "DROP unless renewal ~free"),
       ("Predactiv", 26, "flat", None, None, "high (domain)", "KEEP / lock price (hard non-MM HEM->CRM dep.)"),
       ("5x5", 25, "flat", None, None, "high (domain)", "KEEP (TI-1027)")]
for i, (v, ds, bt, cur, aft, cap, act) in enumerate(rec):
    r += 1
    drow(ws, r, [v, ds, bt, cur if cur else "pending", aft if aft else "-", cap, act],
         fmts=[None, "0", None, "$#,##0" if cur else None, "$#,##0" if aft else None, None, None], alt=(i % 2))
    ws.cell(r, 7).alignment = LEFT
    if act.startswith("DROP"):
        ws.cell(r, 7).font = REDB
    if act.startswith("KEEP"):
        ws.cell(r, 7).font = Font(color="1E7A34", bold=True)

# ---------------- 6 Audit Map ----------------
ws = wb.create_sheet("6 Audit Map")
widths(ws, {"A": 46, "B": 34, "C": 34, "D": 10})
title_block(ws, "Every claim -> its query -> expected number", "All read-only. Fast path = deck_d1..d8 (plain bq query). Anchors in runbook/queries/VALIDATION_GUIDE.md.", "AUDIT")
r = 5; hrow(ws, r, ["Claim", "Query file", "Expected", "Cost"])
amap = [
    ("Total metered bill", "runbook/queries/q0_roster_cost.sql", "$812,398/yr (imps x $0.50 = usage)", "cheap"),
    ("Meter charges $0.50 on free-covered imps", "queries/audi_1089_preemption_proof_winners_table.sql", "268.9M imps, 100% @ $0.50", "18 GB"),
    ("Fair preemption (prior-day, recency)", "runbook/queries/q3e_v2_free_prior_lookback.sql", "-$200.4K/yr (-24.7%)", "BIG 37d"),
    ("  · 33Across / API / Cybba / Justuno / Sovrn", "  same (free_prior_dominant x q0 bill)", "162.1 / 33.1 / 3.8 / 1.3 / 0.1 ($K)", ""),
    ("Same-day (naive) / upper bound", "  q3e_v2 free_sameday / free_prior30", "$273.7K / $243.5K", ""),
    ("Free logs cover the universe", "runbook/queries/deck_d1_universe_coverage.sql", "59.4% visit-day / 60.4% pair", "BIG"),
    ("Row-level source, self-serve", "queries/audi_1089_targeted_signal_bq_per_vendor_split.sql", "per-vendor used-row split ($0)", "$0"),
    ("Dependency ceiling (worth)", "runbook/queries/q6_value_tiers.sql (+q8b solo)", "per-vendor T1/T2 x52", "BIG"),
    ("Unique-domain fee-band (worth)", "runbook/queries/q4_domain_value.sql", "per-vendor $ band", "BIG"),
    ("Drop-savings (reassignment classes)", "runbook/queries/q3b_credit_reassignment.sql", "33A $385.7K, Sovrn $109.0K ...", "BIG"),
]
for i, (cl, qf, ex, co) in enumerate(amap):
    r += 1; drow(ws, r, [cl, qf, ex, co], alt=(i % 2))
    for cc in (1, 2, 3): ws.cell(r, cc).alignment = LEFT
    ws.cell(r, 2).font = MONO

# ---------------- Query sheets ----------------
def sql_sheet(name, path, header):
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 110
    ws["A1"] = header; ws["A1"].font = NAVYB; ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 30
    with open(os.path.join(TICKET, path)) as f:
        for i, line in enumerate(f.read().rstrip("\n").split("\n")):
            c = ws.cell(3 + i, 1, line if line else " ")
            c.font = MONO; c.alignment = Alignment(horizontal="left", vertical="top")

# every backing query, embedded as a paste-and-run SQL sheet (name matches the "Query ▸" refs above)
QUERIES = [
    ("Q1 Meter Proof", "queries/audi_1089_preemption_proof_winners_table.sql",
     "Q1 (sheet 1) — does the meter skip paid credit when a free log already covers the impression? ~18 GB."),
    ("Q2 Targeted Signal", "queries/audi_1089_targeted_signal_bq_per_vendor_split.sql",
     "Q2 (audit map) — row-level used-signal split by originating vendor (BQ external, $0)."),
    ("Q3 Fair Prior-Day", "runbook/queries/q3e_v2_free_prior_lookback.sql",
     "Q3 (sheets 2 & 3) — the fair preemption scan (prior-day + recency, full 30d lookback, all IPs). BIG."),
    ("Q4 Bills (q0)", "runbook/queries/q0_roster_cost.sql",
     "Q4 (sheets 2 & 5) — roster + actual meter bills; meter check imps x $0.50 = usage. Console-cheap."),
    ("Q5 Dependency (q6)", "runbook/queries/q6_value_tiers.sql",
     "Q5 (sheet 4) — media on each vendor's sole serves -> T1/T2 dependency value (x52). BIG."),
    ("Q6 Domain band (q4)", "runbook/queries/q4_domain_value.sql",
     "Q6 (sheet 4) — sole classified domains x fee band -> the unique-domain value lens. BIG."),
    ("Q7 Free coverage (d1)", "runbook/queries/deck_d1_universe_coverage.sql",
     "Q7 (audit map) — free-log coverage of the (ip x domain x date) universe (59.4% / 60.4%). BIG."),
    ("Q8 Drop savings (q3b)", "runbook/queries/q3b_credit_reassignment.sql",
     "Q8 (audit map) — first-reporter reassignment classes -> exact drop savings per vendor. BIG."),
]
for nm, pth, hd in QUERIES:
    sql_sheet(nm, pth, hd)

# freeze header rows on the data sheets
for nm in ["1 Meter Proof", "2 Preemption (fair)", "3 Augmentor Fix", "4 Worth vs Bill", "5 Recommendations", "6 Audit Map"]:
    wb[nm].freeze_panes = "A6"

path = os.path.join(HERE, "audi_1089_billing_review.xlsx")
wb.save(path)
print("wrote", path, "-", len(wb.sheetnames), "sheets:", ", ".join(wb.sheetnames))
