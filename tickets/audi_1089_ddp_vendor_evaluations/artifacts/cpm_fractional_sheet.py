#!/usr/bin/env python3
"""Shared builder: add the AUDI-1115 fractional-credit CPM sheet to a workbook.

Imported by BOTH audi_1089_deck_workbook.py (deck sheet) and
runbook/charts/fill_template.py (quality sheet) so the "how billing works +
preempt free logs + price the residual" table is identical and reproducible in
each. Reads the measured l0f/q2b/deck_d3 CSVs — nothing hardcoded.

Verified 2026-07-17 (3-agent adversarial pass). Carries the mandatory caveats:
the residual pricing is a PRICING lens, NOT a keep/drop test (it over-credits);
keep ~$0.50 (below residual break-even), savings come from preemption volume;
residual volume depends on the preemption grain (07-20 billing sync).
"""

import csv
import os

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
TICKET_1089 = os.path.dirname(HERE)
TICKETS = os.path.dirname(TICKET_1089)
L0F_CSV = os.path.join(TICKETS, "audi_1111_vendor_quality", "audi_1115_wtp_cpm",
                       "outputs", "audi_1115_l0f_fractional_credit_cpm.csv")

NAMES = {"28": "33Across", "40": "33Across API", "33": "Sovrn", "24": "Justuno",
         "36": "Cybba", "25": "5x5", "26": "Predactiv", "39": "Klickly"}
METERED = {"28", "40", "33", "24", "36"}
MARGIN_LO, MARGIN_HI = 0.10, 0.30

HEAD = Font(bold=True, size=10)
TITLE = Font(bold=True, size=12)
FILL = PatternFill("solid", fgColor="DDE6F0")
FLATFILL = PatternFill("solid", fgColor="F2F2F2")
WRAP = Alignment(wrap_text=True, vertical="top")
PCT1 = "0.0%"
INT0 = "#,##0"


def _read(path):
    with open(path) as f:
        return list(csv.DictReader(f))


def add_cpm_fractional_sheet(wb, run_dir):
    """run_dir = absolute path to the outputs/run_YYYY_MM_DD folder (for q2b/d3)."""
    l0f = {r["vendor_ds"]: r for r in _read(L0F_CSV)}
    d3 = {r["data_source_id"]: r for r in _read(os.path.join(run_dir, "deck_d3_bills_cpm.csv"))}
    q2b = {r["ds"]: r for r in _read(os.path.join(run_dir, "q2b_daily_drops.csv"))}

    ws = wb.create_sheet("cpm_fractional")
    ws.cell(row=1, column=1, value=(
        "FAIR CPM — billing is at the WON impression; preempt free logs, price the residual "
        "(AUDI-1115, verified 2026-07-17)")).font = TITLE

    intro = (
        "Billing structure (confirmed via BAE dw-main-gold.reporting.ddp_mm_winners_imp, keyed on "
        "ad_served_id): vendors are billed ONCE, per WON+credited impression, at the contract CPM "
        "($0.50 for metered). NOT billed for ingestion or DS13/DS19 usage (those are the gate). "
        "Only ~0.2% of ingested rows ever bill. ~88-97% of every vendor's won impressions ALSO have "
        "a free-log winner -> free-log preemption (AUDI-1093) removes that overlap. On the residual, "
        "each impression earns ~$10.7 media CPM, so break-even = media CPM x internal margin band "
        "(10-30%) = ~$1-3 for EVERY vendor -> $0.50 is already below break-even.")
    ws.cell(row=2, column=1, value=intro).alignment = WRAP
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 58

    cols = [
        ("Vendor", None, 14),
        ("Billing", None, 9),
        ("Billed imps/mo (meter)", INT0, 13),
        ("Bill $/yr", INT0, 11),
        ("Won imps/mo (BAE)", INT0, 13),
        ("% won imps a free log also won", PCT1, 12),
        ("Residual imps/mo (post-preemption)", INT0, 13),
        ("Residual media CPM", "$0.00", 11),
        ("Break-even CPM @10% margin", "$0.000", 12),
        ("Break-even CPM @30% margin", "$0.000", 12),
        ("Current CPM", "$0.00", 10),
    ]
    hr = 4
    for j, (title, _f, w) in enumerate(cols, start=1):
        c = ws.cell(row=hr, column=j, value=title)
        c.font = HEAD
        c.fill = FILL
        c.alignment = WRAP
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"

    order = sorted(l0f.keys(), key=lambda ds: -float(l0f[ds]["imps_paid_eligible"]))
    r = hr + 1
    for ds in order:
        row = l0f[ds]
        any_ = float(row["imps_any_winner"])
        pre = float(row["imps_free_preempted"])
        elig = float(row["imps_paid_eligible"])
        cpm = float(row["media_cpm_frac"])
        is_metered = ds in METERED
        billed = float(d3[ds]["billed_imps_month"]) if d3.get(ds) and d3[ds]["billed_imps_month"] else None
        bill_yr = float(d3[ds]["bill_annualized"]) if d3.get(ds) and d3[ds]["bill_annualized"] else None
        vals = [
            NAMES.get(ds, ds),
            "metered" if is_metered else "flat fee",
            billed, bill_yr, any_,
            (pre / any_) if any_ else None,
            elig, cpm, cpm * MARGIN_LO, cpm * MARGIN_HI,
            0.50 if is_metered else None,
        ]
        for j, (val, (_t, fmt, _w)) in enumerate(zip(vals, cols), start=1):
            c = ws.cell(row=r, column=j, value=val)
            if val is not None and fmt:
                c.number_format = fmt
            if not is_metered:
                c.fill = FLATFILL
        r += 1

    notes = [
        "READ: the rate isn't the lever - the overlap is. $0.50 sits below the $1-3 residual "
        "break-even for every vendor. Keep ~$0.50; the savings come from PREEMPTION (removing the "
        "~90% free-log overlap volume), not a rate cut.",
        "CAVEAT (mandatory, verified): the residual break-even is a PRICING lens, NOT a keep/drop "
        "test. It values each credited impression at full media, incl. impressions we'd win anyway "
        "(other paid vendors, or free-log coverage of the same IP on another day) - so it OVER-"
        "credits. The keep/drop (marginal) value is the SOLO cohort on the numbers/solo sheets. Do "
        "NOT quote this as 'the vendor is worth $0.50 on the current full meter' - that greenlights "
        "a deal we're losing on today.",
        "CAVEAT: media_spend is the ADVERTISER's payment; the $1-3 assumes the vendor's signal is "
        "WHY we won. If not fully incremental, the residual is worth less and $0.50 approaches "
        "break-even - so do NOT cut below $0.50 without an incrementality read.",
        "GRAIN: 'won imps' free-overlap here is impression-winner grain (a free log won THAT "
        "impression) = ~90%; the visit-day grain (AUDI-1093) is 52.5% for 33Across. Same rate, "
        "different residual VOLUME - the eng team's fractional-credit system's grain sets the total "
        "(33Across residual 27.5M vs ~5.6M/mo). frac totals provisional until the 2026-07-20 sync.",
        "Vendors differ by RESIDUAL VOLUME, not the per-impression rate (all ~$1-3): 33Across 27.5M "
        "> 33A API 14.8M > Sovrn 5.3M ~ Justuno 4.7M > Cybba 0.6M imps/mo (impression grain).",
        "Windows: BAE + CIL media June 2026; meter/bill June x12. Margin band is INTERNAL - do not "
        "quote outside the team. Source query: audi_1111_vendor_quality/audi_1115_wtp_cpm/queries/"
        "audi_1115_l0f_fractional_credit_cpm.sql (+ l0b for the billing-structure evidence).",
    ]
    r += 1
    for t in notes:
        c = ws.cell(row=r, column=1, value="- " + t)
        c.alignment = WRAP
        c.font = Font(size=9, italic=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        ws.row_dimensions[r].height = max(14, -(-len(t) // 115) * 13 + 4)
        r += 1
    return ws
