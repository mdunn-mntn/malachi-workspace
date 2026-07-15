#!/usr/bin/env python3
"""Build the filled AUDI-1089 vendor-quality workbook from the canonical run CSVs.

Output workbook (outputs/audi_1089_quality_template_filled.xlsx) has two sheets:
  - numbers: ONE question per row, ONE value per cell (real numbers with Excel
    number formats — counts, %, $, decimals — so columns sort/compute). Vendors
    are columns in the template's order. "—" = not applicable for that column.
  - notes: one row per vendor with every text answer (scope, billing detail,
    renewal status, ingestion + off-switch, blast radius, full verdict, asks)
    plus a CONVENTIONS block (windows, annualization, formulas, caveats).

The question set is the user's template (outputs/audi_1089_quality_template.xlsx)
with compound rows split so each row carries exactly one metric. Sources per
section: runbook/README.md (Template map). Economics formulas:
runbook/dependency_valuation.md; margin ladder 15/20/30% (blended estimate).

Usage: python3 fill_template.py [run_dir] [bill_month]
  run_dir    outputs/ subfolder holding the q*.csv files (default run_2026_07_10)
  bill_month YYYY-MM whose meter bills fill the COST rows (default 2026-06;
             must exist as a month-end reporting_month snapshot in q0/q1d)
"""
import csv
import math
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
TICKET = os.path.dirname(os.path.dirname(HERE))
RUN = sys.argv[1] if len(sys.argv) > 1 else "run_2026_07_10"
BILL_MONTH = sys.argv[2] if len(sys.argv) > 2 else "2026-06"
RDIR = os.path.join(TICKET, "outputs", RUN)
OUT = os.path.join(TICKET, "outputs", "audi_1089_quality_template_filled.xlsx")

FREEC = 99  # pseudo-vendor: guid_log + augmentor treated as ONE source ("free_logs")
DS_COLS = [28, 40, 33, 24, 36, 25, 26, 39, 27, 30, 23, FREEC, 35, 17, 29]
ACTIVE = [28, 40, 33, 24, 36, 25, 26, 39, 30, 23]
EXT = [24, 25, 26, 28, 33, 36, 39, 40]
FREE = [23, 30, FREEC]
METERED = [24, 28, 33, 36, 40]
FLAT = [25, 26, 39]
OOS = [27, 35, 17, 29]
SHORT = {23: "guid_log", 24: "Justuno", 25: "5x5", 26: "Predactiv", 27: "LaunchLabs",
         28: "33Across", 30: "augmentor", 33: "Sovrn", 36: "Cybba", 39: "Klickly",
         40: "33A API", 35: "LiveRamp IP", 17: "ShareThis", 29: "deepsync",
         99: "free_logs"}
HDR_NAMES = {28: "33Across", 40: "33Across API", 33: "Sovrn", 24: "Justuno", 36: "Cybba",
             25: "5x5", 26: "sharethis_predactiv", 39: "Klickly", 27: "LaunchLabs",
             30: "augmentor_log", 23: "guid_log", 99: "free_logs (guid+aug)",
             35: "LiveRamp IP", 17: "ShareThis", 29: "deepsync"}
BITSQ = {23: 0, 24: 1, 25: 2, 26: 3, 28: 4, 30: 5, 33: 6, 36: 7, 39: 8, 40: 9}
FREE_MASK = (1 << 0) | (1 << 5)
ANN30 = 365.0 / 30.0
BASELINE_VR = 0.0223
NA = "—"


def _poisson_ci(k):
    z = 1.959964

    def chi2_q(pz, df):
        if df == 0:
            return 0.0
        return df * (1 - 2.0 / (9 * df) + pz * (2.0 / (9 * df)) ** 0.5) ** 3

    lo = 0.0 if k == 0 else 0.5 * chi2_q(-z, 2 * k)
    hi = 0.5 * chi2_q(z, 2 * k + 2)
    return lo, hi


# ---------------- loaders ----------------
def rows_of(fname):
    p = os.path.join(RDIR, fname)
    if not os.path.exists(p) or os.path.getsize(p) == 0:
        return []
    with open(p) as f:
        return list(csv.DictReader(f))


def by_ds(fname, key=None):
    out = {}
    for r in rows_of(fname):
        k = key or ("ds" if "ds" in r else "data_source_id")
        out[int(r[k])] = r
    return out


q0 = {}
for r in rows_of("q0_roster_cost.csv"):
    d = int(r["data_source_id"])
    ent = q0.setdefault(d, {"reg": r, "june_usd": None})
    if r.get("reporting_month", "").startswith(BILL_MONTH):
        ent["june_usd"] = float(r["usage_dollars"])
        ent["reg"] = r

q1 = {}
for r in rows_of("q1_scale_by_day.csv"):
    d = int(r["data_source_id"])
    a = q1.setdefault(d, {"days": [], "rows": 0, "ipv6": 0, "path": 0})
    n = int(r["n_rows"])
    a["days"].append(n)
    a["rows"] += n
    a["ipv6"] += int(r["ipv6_rows"])
    a["path"] += int(r["rows_with_path"])

q1b = {}
for r in rows_of("q1b_column_richness.csv"):
    q1b.setdefault(int(r["data_source_id"]), {})[r["field"]] = float(r["pct_populated"])

q1c = by_ds("q1c_content_quality.csv")
q1d = by_ds("q1d_billed_usage.csv")
q2 = by_ds("q2_window_reach.csv")
q2b = by_ds("q2b_daily_drops.csv")
q2c = by_ds("q2c_funnel.csv")
q3 = by_ds("q3_usable_uniqueness.csv")
q3r = by_ds("q3_pair_recency.csv")
q4 = by_ds("q4_domain_value.csv")
q6 = by_ds("q6_value_tiers.csv")
q7 = by_ds("q7_sole_vr.csv")

q5 = {}
for r in rows_of("q5_score_tiers.csv"):
    q5.setdefault(int(r["data_source_id"]), {})[r["cohort"]] = r

q6b = {}
for r in rows_of("q6b_sole_by_funnel.csv"):
    d = int(r["ds"])
    a = q6b.setdefault(d, {"prosp": 0, "tot": 0, "prosp_media": 0.0})
    a["tot"] += int(r["imps"])
    if r["obj_bucket"] == "prospecting_family":
        a["prosp"] += int(r["imps"])
        a["prosp_media"] += float(r["media"])

q7bd = {}
for r in rows_of("q7b_perf_by_cohort.csv"):
    q7bd.setdefault(int(r["ds"]), {})[r["cohort"]] = r

q7cd = {}
for r in rows_of("q7c_conversions.csv"):
    q7cd.setdefault(int(r["ds"]), {})[r["cohort"]] = r

q7d = rows_of("q7d_platform_week.csv")
PLAT_IPS = float(q7d[0]["ips_served_week"]) if q7d else None

masks, reassign = {}, {}
for r in rows_of("q3b_credit_reassignment.csv"):
    if r["rec"] == "mask":
        masks[int(r["k1"])] = int(r["n_pairs"])
    elif r["rec"] == "reassign":
        reassign.setdefault(int(r["k1"]), {})[r["k2"]] = int(r["n_pairs"])

hi3d, pp3d, vert3d = {}, {}, {}
for r in rows_of("q3d_score_vertical_coverage.csv"):
    if r["rec"] == "hi":
        hi3d[int(r["k1"])] = int(r["n"])
    elif r["rec"] == "pp":
        pp3d[int(r["k1"])] = int(r["n"])
    elif r["rec"] == "vert":
        vert3d.setdefault(r["k1"], {})[r["k2"]] = int(r["n"])

masks3c, vend3c = {}, {}
for r in rows_of("q3c_visit_grain_uniqueness.csv"):
    if r["rec"] == "mask":
        masks3c[int(r["k1"])] = int(r["n"])
    else:
        vend3c.setdefault(int(r["k1"]), {})[r["k2"]] = int(r["n"])
TRIP_TOTAL = sum(masks3c.values())

# ---- SOLO counterfactual (each vendor as the ONLY paid source; free logs kept) ----
q8a, q8b = {}, {}
for r in rows_of("q8a_solo_stock.csv"):
    q8a.setdefault(int(r["ds"]), {}).setdefault(r["rec"], {})[r["k"]] = float(r["v"])
for r in rows_of("q8b_solo_perf.csv"):
    q8b.setdefault(int(r["ds"]), {}).setdefault(r["rec"], {})[r["k"]] = float(r["v"])
Q8A, Q8B = bool(q8a), bool(q8b)
PENDING = "pending scan (q8)"

# ---- DS19-only universe (q13a) + membership split (q13b) ----
q13a_pair, q13a_trip, q13a_cat, q13a_ds = {}, {}, {}, {}
Q13A_PATH = None
for r in rows_of("q13a_ds19_universe.csv"):
    if r["rec"] == "pair":
        q13a_pair[int(r["k1"])] = int(float(r["v1"]))
    elif r["rec"] == "trip":
        q13a_trip[int(r["k1"])] = int(float(r["v1"]))
    elif r["rec"] == "path":
        Q13A_PATH = (int(float(r["v1"])), int(float(r["v2"])), int(float(r["v3"])))
    elif r["rec"] == "cat":
        q13a_cat[int(r["k1"])] = {"name": r["k2"], "all": int(float(r["v1"])),
                                  "free": int(float(r["v2"])), "k4": int(float(r["v3"]))}
    elif r["rec"] == "ds":
        q13a_ds[int(r["k1"])] = int(float(r["v1"]))
Q13A = bool(q13a_pair)
PENDING13 = "pending scan (q13a)"

q13b = {}
for r in rows_of("q13b_ds19_perf.csv"):
    q13b.setdefault(r["k1"], {})[r["k2"]] = float(r["v"])
Q13B = bool(q13b)


def ds19_cov(keep, universe, fm=None):
    km = (fm if fm is not None else FREE_MASK) | sum(1 << BITSQ[d] for d in keep)
    tot = sum(universe.values())
    return sum(n for m, n in universe.items() if m & km) / tot if tot else 0


# ---- WASTE sheet inputs: measured GCS bytes per vendor (svs is partitioned by
# data_source_id). Method: gsutil du "gs://mntn-data-archive-prod/signals/
# site_visit_signal/dt=<D>/hh=*/data_source_id=*/*" summed per ds; GB_DAY = avg of
# 2026-06-15 and 2026-07-01. svs has NO TTL — first partition dt=2025-08-31; GB_ACCUM
# = accumulated footprint integrated from monthly 1st-of-month samples (Sep'25-Jul'26).
GB_DAY, GB_ACCUM = {}, {}
for r in rows_of("q14_gcs_ingest_bytes.csv"):
    GB_DAY[int(r["ds"])] = float(r["gb_day"])
    GB_ACCUM[int(r["ds"])] = float(r["gb_accum"])
STORAGE_RATE = 0.02 * 12  # $/GB-YEAR, GCS standard list price ($0.020/GB-month)
WASTE_COLS = [28, 40, 33, 24, 36, 25, 26, 39, 30, 23]


def used_frac(d):
    return int(q2c[d]["rows_used"]) / int(q2c[d]["rows_raw"])


# ---- q15: free_logs COMBINED pseudo-vendor (guid+aug as ONE source) injection ----
q15 = {}
for r in rows_of("q15_free_union_perf.csv"):
    q15.setdefault(r["k1"], {}).setdefault(r["rec"], {})[r["k2"]] = float(r["v"])
Q15 = bool(q15)

q15b = {}
for r in rows_of("q15b_free_union_stock.csv"):
    q15b.setdefault(r["rec"], {})[r["k1"]] = float(r["v"])
Q15B = bool(q15b)


def _synth_free_combo():
    """Populate pseudo-ds 99 wherever the combination rule is exact.
    Additive: rows/GB (events are disjoint per source). Row-weighted: quality shares.
    Mask-exact: usable pair/visit-day union, sole-vs-paid. Union-unique counts (raw
    reach, funnel IPs/domains) are NOT derivable -> left absent (render as em-dash)
    until/unless measured. q15 injects measured serving/perf cohorts when it lands."""
    a, b = 23, 30
    q0[FREEC] = {"reg": {"billing_type": "free", "fixed_cpm": 0, "valid_from": ""},
                 "june_usd": 0.0}
    if a in q1 and b in q1 and len(q1[a]["days"]) == len(q1[b]["days"]):
        q1[FREEC] = {"days": [x + y for x, y in zip(q1[a]["days"], q1[b]["days"])],
                     "rows": q1[a]["rows"] + q1[b]["rows"],
                     "ipv6": q1[a]["ipv6"] + q1[b]["ipv6"],
                     "path": q1[a]["path"] + q1[b]["path"]}
        wa, wb = q1[a]["rows"], q1[b]["rows"]

        def _wavg(fa, fb):
            return (fa * wa + fb * wb) / (wa + wb)

        if a in q1b and b in q1b:
            q1b[FREEC] = {k: _wavg(q1b[a].get(k, 0.0), q1b[b].get(k, 0.0)) for k in q1b[a]}
        if a in q1c and b in q1c:
            # row-share metrics combine exactly as row-weighted averages; concentration
            # metrics (top_domain*) do NOT (union top-1 unknown) -> omitted
            q1c[FREEC] = {k: str(_wavg(float(q1c[a][k] or 0), float(q1c[b][k] or 0)))
                          for k in ("url_parse_fail_pct", "url_malformed_pct",
                                    "pct_googlebot_ip", "pct_private_ip", "uid_dup_pct",
                                    "time_top1_share")
                          if k in q1c[a] and k in q1c[b]}
        q15c = rows_of("q15c_free_union_hour_quality.csv")
        if q15c:
            q1c.setdefault(FREEC, {}).update({
                "top_domain": q15c[0]["top_domain"],
                "top_domain_share": q15c[0]["top_domain_share"],
                "top5_domain_share": q15c[0]["top5_domain_share"],
                "ua_bot_pct": q15c[0]["ua_bot_pct"],
            })
        if a in q2b and b in q2b:
            ra, rb = int(q2b[a]["rows_day"]), int(q2b[b]["rows_day"])
            q2b[FREEC] = {
                "rows_day": str(ra + rb),
                "rows_bot_ua": str(int(q2b[a]["rows_bot_ua"]) + int(q2b[b]["rows_bot_ua"])),
                "pct_hard_dropped": str((float(q2b[a]["pct_hard_dropped"]) * ra
                                         + float(q2b[b]["pct_hard_dropped"]) * rb) / (ra + rb)),
                "pct_blocked_ds13": str((float(q2b[a]["pct_blocked_ds13"]) * ra
                                         + float(q2b[b]["pct_blocked_ds13"]) * rb) / (ra + rb)),
            }
    if a in q2c and b in q2c:
        q2c[FREEC] = {k: str(int(q2c[a][k]) + int(q2c[b][k]))
                      for k in ("rows_raw", "rows_kept", "rows_ds13_input",
                                "rows_ds13_class", "rows_ds19_cat", "rows_used")}
    if masks:
        q3[FREEC] = {
            "usable_pairs": str(sum(n for m, n in masks.items() if m & FREE_MASK)),
            "sole_pairs": str(sum(n for m, n in masks.items()
                                  if (m & FREE_MASK) and not (m & ~FREE_MASK & 1023))),
            "netnew_vs_free_pairs": "0",
        }
    if masks3c:
        vend3c[FREEC] = {"all_triples": sum(n for m, n in masks3c.items() if m & FREE_MASK)}
    if a in GB_DAY and b in GB_DAY:
        GB_DAY[FREEC] = GB_DAY[a] + GB_DAY[b]
    if a in GB_ACCUM and b in GB_ACCUM:
        GB_ACCUM[FREEC] = GB_ACCUM[a] + GB_ACCUM[b]
    if Q15B:
        # union stock/reach/freshness vs the PAID roster (q15b) -> per-ds dicts
        rch, stk, dms = q15b["reach"], q15b["stock"], q15b["doms"]
        fp, fd = q15b["fresh_pair"], q15b["fresh_day"]
        q2[FREEC] = {k: str(int(rch[k])) for k in
                     ("ips_30d", "domains_30d", "ip_domain_pairs_30d")}
        q3[FREEC].update({
            "usable_ips": str(int(stk["usable_ips"])),
            "sole_ips": str(int(stk["sole_ips"])),
            "pairs_per_ip": str(round(stk["usable_pairs"] / stk["usable_ips"], 1)),
        })
        # anchor: measured union usable pairs must match the mask-derived value
        mask_up = int(q3[FREEC]["usable_pairs"])
        if mask_up and abs(stk["usable_pairs"] - mask_up) / mask_up > 0.001:
            raise AssertionError(
                f"q15b usable_pairs {stk['usable_pairs']:.0f} vs masks {mask_up}")
        q4[FREEC] = {"sole_domains": str(int(dms["sole_domains"])),
                     "sole_classified": str(int(dms["sole_classified"]))}
        tot_raw = rch["ip_domain_pairs_30d"]
        q3r[FREEC] = {"pct_freshest": str(round(100 * fp["fresher_than_paid"] / tot_raw, 1)),
                      "pct_tied": str(round(100 * fp["tied_with_paid"] / tot_raw, 1)),
                      "pct_stale": str(round(100 * fp["stale_vs_paid"] / tot_raw, 1))}
        vend3c[FREEC] = {"sole_new_pair": int(fd["sole_new_pair"]),
                         "sole_refresh": int(fd["refresh_of_paid_pair"]),
                         "shared_same_day": int(fd["same_day_dup_with_paid"])}
    if Q15:
        # measured union cohorts -> the existing per-ds dicts, so every row fn just works.
        # touched = guid OR aug delivered; sole = touched AND no paid vendor.
        def _pcts(c):
            served = q15[c]["serve"]
            tier = q15[c]["tier"]
            mem = q15[c]["mem"]["member_ips"]
            ips = served["ips_served"]
            out = {"vendor_ips": mem, "delivered_ips": ips,
                   "pct_delivered": 100 * ips / mem if mem else 0,
                   "hi_10000": tier["hi"], "pp_8000": tier["pp"], "high_grad": tier["hg"],
                   "mid": tier["mid"], "maxreach": tier["maxreach"],
                   "unscored_delivered": tier["unscored"]}
            for t, key in (("hi", "pct_hi"), ("pp", "pct_pp"), ("hg", "pct_high_grad"),
                           ("mid", "pct_mid"), ("maxreach", "pct_maxreach"),
                           ("unscored", "pct_unscored_delivered")):
                out[key] = 100 * tier[t] / ips if ips else 0
            return {k: str(v) for k, v in out.items()}
        q5[FREEC] = {"touched": _pcts("touched"), "sole": _pcts("sole")}
        st, ss = q15["touched"]["serve"], q15["sole"]["serve"]
        q6[FREEC] = {"ips_touched": st["ips_served"], "ips_sole": ss["ips_served"],
                     "imps_touched": st["imps"], "imps_sole": ss["imps"],
                     "media_touched": st["media"], "media_sole": ss["media"],
                     "data_touched": st["data"], "data_sole": ss["data"],
                     "imps_sole_scored_nonrtc": ss["imps_scored_nonrtc"],
                     "media_sole_scored": ss["media_scored"],
                     "data_sole_scored": ss["data_scored"]}
        for c in ("touched", "sole"):
            sv, pf = q15[c]["serve"], q15[c]["perf"]
            q7bd.setdefault(FREEC, {})[c] = {
                "imps": sv["imps"], "ips_served": sv["ips_served"],
                "visits": pf["visits"],
                "vr_pct": 100 * pf["visits"] / sv["imps"] if sv["imps"] else 0,
                "avg_hs_scored": sv["avg_hs"],
                "pct_scored": 100 * sv["imps_hs_pos"] / sv["imps"] if sv["imps"] else 0,
                "media": sv["media"]}
            q7cd.setdefault(FREEC, {})[c] = {
                "imps": sv["imps"], "conversions": pf["conversions"],
                "revenue": pf["revenue"]}
        q7[FREEC] = {"sole_ips_delivered": ss["ips_served"], "sole_imps": ss["imps"],
                     "sole_visits": q15["sole"]["perf"]["visits"],
                     "vr_overall_pct": 100 * q15["sole"]["perf"]["visits"] / ss["imps"]
                     if ss["imps"] else 0}
        # solo sheet: for the combined free column, "solo (vs free logs only)" = its
        # FULL cohort (there is no other free log to exclude) -> touched cohort.
        q8b[FREEC] = {"serve": dict(q15["touched"]["serve"]),
                      "perf": dict(q15["touched"]["perf"]),
                      "tier": dict(q15["touched"]["tier"])}


_synth_free_combo()

STOP_SENDING = {
    28: "STOP webmail (yahoo/aol ~29% of rows) + Googlebot IPs (6.4%) -> ~35% ingest-volume cut. NOTE: most of this is NOT in the thrown-away 22% - it passes the DS19 gate and BILLS; stopping it at source cuts ingestion AND junk billing",
    40: "STOP cookie-sync / ad-infra pixel URLs (nextmillmedia, programmaticx - its top BILLED domains are sync junk)",
    33: "FIX the URL-doubling bug (77% of rows malformed 'msn.comhttps://...') or stop sending until fixed - most of the feed is unclassifiable garbage we ingest and store",
    24: "None - cleanest feed on the roster. Ask to ADD user_agent instead (enables bot filtering before we pay)",
    36: "Dedupe server-side (one IP = 3.5% of rows); feed is tiny so waste cost is negligible",
    25: "STOP outbrain.com widget-iframe URLs (52.7% of rows) - send the publisher page URL instead",
    26: "Filter adult-content domains at source; otherwise strongest breadth on the roster",
    39: "94% is *.myshopify.com checkout - volume is tiny (0.26 GB/day) so waste is immaterial",
    30: "internal free log - n/a", 23: "internal free log - n/a",
}

# other_free(d): the free-log bits that count as overlap for d — both free logs for
# a paid vendor, the OTHER free log for a free column (guid vs augmentor).
OTHERFREE = {d: FREE_MASK & ~(1 << b) for d, b in BITSQ.items()}
OTHERFREE[99] = 0  # the combined free column has no OTHER free log


def bits_of(d):
    return FREE_MASK if d == FREEC else (1 << BITSQ[d])


def solo_sum(d, universe):
    b, of = bits_of(d), OTHERFREE[d]
    return sum(n for m, n in universe.items() if (m & b) and not (m & of))


def keep_cov(d, universe):
    km = FREE_MASK | bits_of(d)
    return sum(n for m, n in universe.items() if m & km)


def g8a(d, rec, k):
    return q8a[d].get(rec, {}).get(k, 0.0)


def g8b(d, rec, k):
    return q8b[d][rec][k]


def t2_solo(d):
    return g8b(d, "serve", "media") * 52


def t1_solo(d):
    return g8b(d, "serve", "media_scored") * 52


def trip_holder(d):
    return sum(n for m, n in masks3c.items() if m & bits_of(d))


def trip_free_cohold(d):
    return sum(n for m, n in masks3c.items() if (m & bits_of(d)) and (m & FREE_MASK))


# ---- net-of-free universe (free-log-touched pairs removed entirely) ----
NOF_P = {m: n for m, n in masks.items() if not (m & FREE_MASK)}
NOF_T = {m: n for m, n in masks3c.items() if not (m & FREE_MASK)}
NOF_U = sum(NOF_P.values())

FREE_COV_P = sum(n for m, n in masks.items() if m & FREE_MASK)
FREE_COV_T = sum(n for m, n in masks3c.items() if m & FREE_MASK)
PAID_BITS = sum(1 << BITSQ[d] for d in EXT)
PAID_TRIPLES = sum(n for m, n in masks3c.items() if m & PAID_BITS)
TOT_METERED_BILL = sum(q0[d]["june_usd"] * 12 for d in METERED
                       if q0.get(d, {}).get("june_usd") is not None)


# AUDI-1093 post-preemption: if free logs preempt credit on co-held (ip,domain,DATE)
# visit-days, the vendor's bill drops by its free-cohold share; its unique value is
# untouched by construction (sole/solo cohorts exclude free-coheld signal).
def preempt_cut(d):
    return q0[d]["june_usd"] * 12 * trip_free_cohold(d) / trip_holder(d)


def bill_after_preempt(d):
    return q0[d]["june_usd"] * 12 - preempt_cut(d)


def nof_cov(S, universe):
    sm = sum(1 << BITSQ[d] for d in S)
    return sum(n for m, n in universe.items() if m & sm)


# ---------------- derived ----------------
def cov(keep, fm=FREE_MASK):
    km = fm
    for d in keep:
        km |= (1 << BITSQ[d])
    return sum(p for m, p in masks.items() if m & km)


FULL_COV = cov(EXT)
add_order, add_gain = [], {}
_rem, _cur = list(EXT), cov([])
while _rem:
    _best = max(_rem, key=lambda d: cov(add_order + [d]))
    _g = cov(add_order + [_best]) - _cur
    _cur = cov(add_order + [_best])
    add_order.append(_best)
    add_gain[_best] = 100.0 * _g / FULL_COV
    _rem.remove(_best)

coverage_lost = {d: 100.0 * (FULL_COV - cov([x for x in EXT if x != d])) / FULL_COV for d in EXT}


def t2_ann(d):
    return float(q6[d]["media_sole"]) * 52


def t1_ann(d):
    return float(q6[d]["media_sole_scored"]) * 52


# measured T2-per-sole-pair density (the ladder's standalone-$ multiplier)
DENS = {d: t2_ann(d) / float(q3[d]["sole_pairs"]) for d in EXT if d in q3 and d in q6}


def curved_scores():
    max_sc = max(float(q4[d]["sole_classified"]) for d in EXT)
    max_t1 = max(float(q6[d]["imps_sole_scored_nonrtc"]) for d in EXT)
    comp = {}
    for d in EXT:
        V = math.log10(float(q4[d]["sole_classified"]) + 1) / math.log10(max_sc + 1)
        R = (float(q3r[d]["pct_sole"]) + float(q3r[d]["pct_freshest"])) / 100
        Q = 0.5 * float(q4[d]["pct_classified"]) / 100 \
            + 0.5 * (1 - float(q5[d]["sole"]["pct_unscored_delivered"]) / 100)
        D = math.log10(float(q6[d]["imps_sole_scored_nonrtc"]) + 1) / math.log10(max_t1 + 1)
        P = min(float(q7[d]["vr_overall_pct"]) / BASELINE_VR, 2) / 2 \
            if float(q7[d]["sole_imps"]) >= 5000 else 0.5
        comp[d] = 100 * (0.40 * V + 0.15 * R + 0.15 * Q + 0.10 * D + 0.20 * P)
    best = max(comp.values())
    return {d: (comp[d], 100 * comp[d] / best) for d in comp}


CURVE = curved_scores()


def _med(d):
    days = sorted(q1[d]["days"])
    n = len(days)
    return days[n // 2] if n % 2 else (days[n // 2 - 1] + days[n // 2]) / 2


def g7b(d, c, f):
    return float(q7bd[d][c][f])


def g7c(d, c, f):
    return float(q7cd[d][c][f])


def share(d, table, field):
    tot = sum(float(r[field]) for k, r in table.items() if k != FREEC and field in r)
    return 100 * float(table[d][field]) / tot


def tier_tot(f):
    return sum(float(q5[x]["touched"][f]) for x in ACTIVE if x in q5 and "touched" in q5[x])


def solo_share(d, table, field):
    if d == FREEC:
        return 100.0  # the combined free column IS the {vendor+free} world
    wset = {d, 23, 30}
    tot = sum(float(table[x][field]) for x in wset if x in table and field in table[x])
    return 100 * float(table[d][field]) / tot


def solo_tier_share(d, f):
    if d == FREEC:
        return 100.0
    wset = {d, 23, 30}
    tot = sum(float(q5[x]["touched"][f]) for x in wset if x in q5 and "touched" in q5[x])
    return 100 * float(q5[d]["touched"][f]) / tot


# ---------------- text banks (notes sheet) ----------------
VERDICT_SHORT = {28: "NEGOTIATE (w/ DS40)", 40: "NEGOTIATE (w/ DS28)", 33: "DROP",
                 24: "KEEP-trim", 36: "DROP", 25: "KEEP", 26: "KEEP (HEM)",
                 39: "DROP unless ~free", 23: "KEEP (free)", 30: "KEEP (free)",
                 99: "KEEP (free, the baseline)"}
VERDICT_FULL = {
    28: "NEGOTIATE — one deal with DS40 (~$598K/yr combined); target rate cut, cite free-log overlap",
    40: "NEGOTIATE — same vendor as DS28; fold into one combined renegotiation",
    33: "DROP — 77% malformed, near-zero sole value; exact recovery $109K/yr (sequencing-safe)",
    24: "KEEP-trim — cleanest feed, 91.6% sole pairs; trim rate if possible",
    36: "DROP — subscale (~700x below siblings); $21K/yr recovery",
    25: "KEEP — anchor vendor, 69.3% sole; lock flat price at renewal",
    26: "KEEP — #1 unique classified domains + HEM prod dependency (drop blocker); lock flat price",
    39: "DROP unless ~free — T2 ceiling $2.2K/yr revenue vs flat fee; or convert to $0.50 meter",
    23: "KEEP — free internal log", 30: "KEEP — free internal log",
}
ASKS = {
    28: "Stop sending webmail junk (yahoo.com alone 25% of rows; 29% of rows DS13-blocklisted); scrub the 6.4% Googlebot IPs",
    40: "Billed domains are cookie-sync infra junk (nextmillmedia 9%, programmaticx 8%) — send real page URLs, not sync pixels",
    33: "Fix the URL-doubling bug: 77% of URLs malformed (host concatenated twice, e.g. 'msn.comhttps') — even billed domains are corrupted",
    24: "Cleanest feed of the roster — asks: populate user_agent and advertiser_id (both 0%)",
    36: "Scale is the problem (1.5M rows/day, ~700x below siblings) — but per-unit economics are TOP-2 genuine on the roster: offer to KEEP if they 50-100x the feed at a flat/capped price; also dedupe server-side (one IP = 3.5% of rows)",
    25: "outbrain.com widget URLs = 52.7% of rows — send the publisher page URL, not the widget iframe URL",
    26: "Filter adult content at source (explicit domains in sample); otherwise strongest domain breadth on the roster",
    39: "94% of rows are *.myshopify.com checkout — diversify beyond Shopify or price accordingly; tiny scale (169K rows/day)",
    23: "internal log — n/a", 30: "internal log — n/a",
    99: "internal logs combined — n/a",
}
SCOPE = {
    28: "Active MM site-visit DDP (batch). Same vendor as DS40 (batch vs real-time feeds, per Ryan) — one combined negotiation (~$598K/yr).",
    40: "Active MM site-visit DDP (real-time). Sibling of DS28 — same vendor.",
    33: "Active MM site-visit DDP (real-time Kafka).",
    24: "Active MM site-visit DDP (real-time Kafka).",
    36: "Active MM site-visit DDP (batch).",
    25: "Active MM site-visit DDP (batch, flat fee).",
    26: "Active MM site-visit DDP (batch, flat fee).",
    39: "Active MM site-visit DDP (real-time Kafka, flat fee).",
    99: "COMBINED pseudo-vendor: guid_log + augmentor treated as ONE source. Unique counts are the UNION (not the sum of the two columns — the logs overlap heavily). Percentage rows are ROW-WEIGHTED by volume: augmentor is 71.5% of combined rows, so e.g. % advertiser_id populated reads 28.5% (guid 100% x 28.5% share + aug 0% x 71.5% share) — the column describes the COMBINED feed, not its best half; use the per-log columns for per-log capabilities. Raw-reach and funnel-unique cells are '—' (union not derivable from per-source scans); serving/performance cells are MEASURED union cohorts (q15).",
    23: "Internal free log (MNTN guid pixel) — always kept; $0.",
    30: "Internal free log (bid-time augmentor) — always kept; $0. In svs since 2026-05-12.",
    27: "Registered metered vendor, DISABLED (enabled=false) — no feed, no bill. Feed rows are '—'.",
    35: "NOT an MM site-visit feed: 3P interests source (variable CPM). Contract + bill rows only; feed rows are '—'.",
    17: "NOT an MM site-visit feed: 3P interests source. Contract + bill rows only; feed rows are '—'.",
    29: "NOT an MM site-visit feed: CRM ingestion source. Contract + bill rows only; feed rows are '—'.",
}
RATE_NOTE = {24: "$0.50 CPM metered (pay per used signal imp)", 28: "$0.50 CPM metered",
             33: "$0.50 CPM metered", 36: "$0.50 CPM metered", 40: "$0.50 CPM metered",
             25: "Flat fee — amount pending (Maya / renewal schedule)",
             26: "Flat fee — amount pending (Maya / renewal schedule)",
             39: "Flat fee — amount pending (Maya / renewal schedule)",
             23: "Free — internal", 30: "Free — internal", 99: "Free — internal (combined)",
             27: "$0.50 CPM registered, disabled",
             35: "Variable CPM, ~$1.19-1.32 realized", 17: "$0.95 CPM (was $1.20 until May)",
             29: "$0.50 CPM metered"}
RENEWAL_NOTE = {39: "Renewal LIVE NOW — pass/play answer due Mon 2026-07-13 (Paulo)",
                25: "Contract valid_from 2025-10-17; renewal date pending (Maya / renewal schedule)",
                26: "Contract valid_from 2025-10-17; renewal date pending (Maya / renewal schedule)",
                35: "Contract valid_from 2025-05-01; renewal date pending",
                27: "Disabled since delivery never scaled",
                23: "n/a — internal", 30: "n/a — internal", 99: "n/a — internal"}
INGEST_NOTE = {}
for _d in (23, 25, 26, 28, 30, 36):
    INGEST_NOTE[_d] = "Batch drop -> daily ingest DAG (ENABLED_DSIDS); off-switch: Data Eng DAG config"
for _d in (24, 33, 39, 40):
    INGEST_NOTE[_d] = f"Real-time pixel -> Kafka (fpa_dsid{_d}_kafka_log); off-switch: Kafka topic (Data Eng)"
INGEST_NOTE[27] = "Was metered batch; disabled"
INGEST_NOTE[35] = INGEST_NOTE[17] = "3P interests pipeline (not svs)"
INGEST_NOTE[29] = "CRM ingestion pipeline (not svs)"
INGEST_NOTE[99] = "guid: batch pixel log; augmentor: batch bid-time log — combined view"
BLAST_NOTE = {26: "HEM (hashed-email) feed powers CRM/identity matching in PROD — hard drop blocker beyond MM",
              23: "Internal platform log — never drop", 30: "Internal platform log — never drop",
              99: "Internal platform logs — never drop",
              35: "3P interests targeting layer (outside MM scope here)",
              17: "3P interests targeting layer (outside MM scope here)",
              29: "CRM audience ingestion (outside MM scope here)",
              27: "None (disabled)"}
for _d in (28, 40, 33, 24, 36, 25, 39):
    BLAST_NOTE[_d] = "None known — MM site-visit only"

CONVENTIONS = [
    "One question per row; one value per cell. '—' = not applicable for that column (free log, flat fee with no meter, disabled, or out-of-scope source — see Scope).",
    "Numbers are real Excel values (sortable); % cells store the percent number (21.8 = 21.8%).",
    "'share of column total' rows: sources overlap, so these sum >100% down a row... across vendors.",
    "'% of pairs usable' can read slightly >100%: q3 usable pairs marginally exceed q2 delivered pairs (<1% window mismatch between scans).",
    "'% of platform served IPs': denominator = 28.03M distinct served IPs in the valuation week (q7d).",
    "Windows: delivery metrics = 30d (2026-06-02..07-01); serving/performance = 37d svs union x valuation week 2026-07-02..08.",
    "Annualization: weekly flows x52; 30d stocks x365/30; June bills x12. Never annualize unique-IP stocks.",
    "Economics: T2 = sole-won media x52 (dependency ceiling); T1 = high-score-GATED sole media x52 (narrowest floor). Three dependency layers, don't conflate: USED = signal row consumed by DS13/DS19 (bills the meter, independent of scoring); MEMBER = IP servable only because the vendor loaded it (97-99% of sole serves, q6b) -> T2 is the decision number; SCORED = any score on a sole IP is vendor-derived (T1 uses only the >=6666 gate subset). Margins 15/20/30% on observed eCPM.",
    "Max justified CPM = (T2 x 30% margin) spread over 100% of delivered rows, or over billed/used imps (compare to the $0.50 we pay). Flat vendors on the used-imps row = hypothetical meter.",
    "Flat-contract equivalent rows: floor = T1 x 15%, fair = T2 x 20%, ceiling = T2 x 30%.",
    "T2 envelope = T2 x0.4..x1.8 (volume x0.5-1.5, CPM x0.8-1.2) — scenario range, not a confidence interval.",
    "Exact drop savings = annual bill x share of credits NOT reassigning to another metered vendor (q3b). RESOLVED 2026-07-13 (AUDI-1092): the meter switched regimes at reporting month 2026-05 - Jan-Apr rows are ~100% FRACTIONAL imps (1/N split credit), May-Jun rows are 100% INTEGER (single-vendor credit). June bills = current single-credit regime = what q3b modeled; savings are EXACT if the winner rule is first-reporter (AP-3779) and CONSERVATIVE FLOORS if it is cheapest/free-priority (free logs then displace even more). Do not use Jan-Apr bills for LOO. Flat vendors: savings = the flat fee itself either way.",
    "Sole-cohort conversion/visit counts are Poisson-tiny — read 0 as '<~1/wk', not exactly zero. CVR-sole is '—' when sole visits = 0.",
    "Sole-IP visit rates are REAL, not an attribution artifact (q7f): of 33Across's 99K served sole IPs, only 25 showed ANY clickpass event for ANY advertiser that week (0.025%), vs 1.43% for guid_log's sole IPs measured identically. The households are genuinely dark; the ad_served_id visit join even credits cross-device visits, so 116/wk is if anything generous.",
    "Touched-cohort performance mirrors the platform (pools cover 12-97% of served IPs) — vendor differentiation lives in the SOLE rows.",
    "Conversions (q7c): ui_conversions deduped to one row per conversion event preferring last-touch; assists + disputed excluded; revenue = order_amt.",
    "AUDI-1093 (Sean Yang 2026-07-13): free logs do NOT preempt paid credit today - vendors earn day-grain credit on signals guid/augmentor also capture. 'Recoverable if free logs preempt' rows = bill x visit-day free-cohold share, EXACT at (ip,domain,date) grain from q3c: roster total ~$274K/yr (33Across $222K, 33A API $42K). The fix KEEPS vendors' unique data - stacks with renegotiation, substitutes for drops on the overlap slice.",
    "Visit-grain rows (q3c, 13.29B visit-days/30d): the true value unit is (IP x domain x DATE) - new date on a known pair = recency refresh (real value: 30d scoring window + the meter pays per day); same date from two sources = duplication. Free coverage at visit grain: guid 10.7% / augmentor 48.8% / both 59.4% (pair grain: 60.4%) - augmentor DS30 is the dominant free source and is INCLUDED in every free-log number in this workbook.",
    "NET-OF-FREE LADDER (decisions sheet): universe first drops every pair guid_log/augmentor touched AT ANY DATE in the window (date-blind pair cut; the visit-day $ column is the date-aware lens - they agree within 6%). Standalone = vendor as the ONLY PAID source (overlap with other paid vendors still counts for it); the strictly-nobody-else number is its sole/T2 row. Ladder $ figures are DEPENDENT REVENUE, not kept profit: pay-up-to = x15/20/30% margin. 33Across standalone: $397K revenue -> pay-up-to $60-119K/yr vs $422K bill - converges with its independently-derived WTP band ($30-100K).",
    "Grains: pair = IP x domain (visited ever in window); visit = IP x domain x DATE (each day = distinct event; the meter's grain); IP-alone only used for stock counts. RECENCY IS CREDITED AT VISIT GRAIN: a vendor delivering a FRESHER date for a pair free logs saw earlier counts to the VENDOR (sole_refresh), not to free - free visit-grain coverage is 59.4% AFTER that credit (barely below the 60.4% pair figure, because augmentor re-observes active households daily). Within-day frequency collapses: same pair, same day, N events = one billable/valuable unit.",
    "q3d (HI/PP + verticals): scored-audience coverage is vendor-independent (k=4 keeps 99.9991% of HI - exactly 53 of 5,959,159 HI IPs lost, the dropped vendors' sole-HI singletons + 5 combination-only IPs; free-only keeps 99.94% HI / 99.25% PP - pinned households are active, our logs see them). Vendor reach value concentrates in ad-invisible verticals: health/personal-care 0-26%, cosmetics/pharmacy ~25%, airlines 23% retained under free-only. k=4 keep-set: >=94% everywhere. Vertical size = svs-reachable IPs on the vertical's domains (proxy; audience builder layers recency on top).",
    "Row sources: runbook/README.md 'Template map' (q0..q7d, one SQL + one CSV each).",
    "SOLO sheet: the whole numbers grid recomputed under the counterfactual that the column's vendor is the ONLY paid source - overlap counts only vs our free logs (guid_log DS23 + augmentor DS30; the free-log columns measure vs the OTHER free log). 'Solo' >= 'sole' everywhere (sole excludes ALL 9 other sources); solo == the decisions-ladder 'standalone' at pair grain. Rows that don't change under the counterfactual (feed scale, quality, funnel, touched-cohort) are copied so the sheet reads standalone; LOO/portfolio rows are replaced by {free+vendor} coverage rows.",
    "SOLO bill is a BOUNDED ESTIMATE, not a quote: LOW = today's run-rate (removing competitors can only ADD credit to the survivor - hard floor); HIGH = max(LOW, total metered bills x the vendor's share of paid-held visit-days) (q3c masks; proportional-consumption assumption; all metered CPMs $0.50 so imp share = $ share). The proportional term under-runs junk-billing vendors (Sovrn/Cybba/Justuno credit sits on domains outside the usable universe, uncontested by other vendors) - their bounds collapse to today's bill; only 33Across (+8%) and 33A API (+75%) have material solo-bill upside. Flat vendors: fee pending either way.",
    "free_logs (guid+aug) COLUMN: the two internal logs treated as ONE pseudo-vendor (union semantics). Unique counts are the UNION, NOT guid+aug summed (the logs overlap heavily); '\u2014' cells = union-uniques not derivable from per-source scans (raw 30d reach, funnel IPs/domains, freshness). Rows/GB/bytes are exact sums; quality shares are row-weighted; pair/visit-day/tier rows are mask-exact; SERVING/SCORE/PERFORMANCE are MEASURED union cohorts (q15: touched = either log delivered; sole = no paid vendor delivered). On the solo sheet this column reads as the free-only counterfactual (nothing to exclude).",
    "'pending scan (q8)' cells = measured solo serving/performance awaiting the q8a/q8b background scans; mask-derived, rebased and copied cells are final. Rerun fill_template.py after the q8 CSVs land (the pending count must print 0; anchors: q8a solo pairs == q3b mask solo pairs == q3 net-new-vs-free, q8b >= q6 sole everywhere; q8b HI/PP vs q3d-mask gap is DIAGNOSTIC, not an error - raw vs usable membership lenses: clean vendors read 3-10% low in q8b, Sovrn reads +55-68% HIGH because its malformed-URL rows carry IPs that never reach a usable domain).",
    "POST-PREEMPTION (AUDI-1093 applied): if free logs preempt co-held credit, the meter stops paying for (ip,domain,DATE) visit-days guid_log/augmentor also captured - bills drop by each vendor's free-cohold share, roster $812K -> $539K (-$274K, -33.7%), while vendors KEEP their unique data (pay ranges unchanged by construction: sole/solo value never included free-coheld signal). Visit grain = the fair version (vendor still credited for FRESHER dates on known pairs); strict pair-grain version is barely larger (~$284K proxy). VERDICT: nobody flips on the portfolio lens; 33A API lands exactly AT its measured-solo ceiling top ($134K == $134K), 33Across inside its ceiling range near the top - preemption and renegotiation STACK; Sovrn/Justuno bills are nearly untouched (their credit is junk/unique, not free-coheld).",
    "SOLO NON-ADDITIVITY + measured-vs-estimate: NEVER sum T2_solo (or any solo column) across vendors - solo cohorts overlap heavily (the same multi-paid-vendor IP is 'solo' for every vendor vs free logs); the ladder's MARGINAL column is the only additive lens. Measured T2_solo runs 3-5x ABOVE the density estimate everywhere (33Across $724K vs $397K est) because the estimate inherits sole-cohort adverse selection (dark households) while the solo cohort includes livelier multi-vendor IPs. Even so, no metered vendor's 10-30% pay range reaches its bill on the measured basis. T2_solo is a generous ceiling: the solo cohort's prospecting-attribution share is unmeasured (sole cohort's was 97-99%; the livelier solo IPs likely include more retargeting-driven serves that never needed the vendor).",
]


# ---------------- SPEC: (label, fmt, fn, oos_ok) ----------------
def S(label):
    return (label, None, None, False)


def R(label, fmt, fn, oos_ok=False):
    return (label, fmt, fn, oos_ok)


SPEC = [
    S("CONTRACT & IDENTITY"),
    R("Data source ID", "int", lambda d: d, True),
    R("Billing type", "txt", lambda d: "free" if d in FREE else
      {"fixed_cpm": "metered CPM", "flat_fee": "flat fee", "variable_cpm": "variable CPM"}
      .get(q0[d]["reg"]["billing_type"]), True),
    R("Contract rate ($ CPM; flat amounts pending)", "usd2", lambda d:
      0.0 if d in FREE else ("pending" if d in FLAT else
      (1.28 if d == 35 else float(q0[d]["reg"]["fixed_cpm"]))), True),
    R("Renewal status", "txt", lambda d: NA if d in FREE else
      ("LIVE NOW" if d == 39 else ("disabled" if d == 27 else "pending")), True),
    R("Ingestion path", "txt", lambda d: "batch" if d in (23, 25, 26, 28, 30, 36) else
      ("Kafka RT" if d in (24, 33, 39, 40) else ("disabled" if d == 27 else NA)), True),
    R("Non-MM blast radius", "txt", lambda d: {26: "HEM -> CRM prod", 23: "internal",
      30: "internal", 99: "internal", 35: "3P interests", 17: "3P interests", 29: "CRM",
      27: "none"}.get(d, "none known"), True),

    S("FEED SCALE (30d)"),
    R("Total rows delivered", "int", lambda d: q1[d]["rows"]),
    R("Median rows/day", "int", lambda d: _med(d)),
    R("Weakest day (% of median)", "pct0", lambda d: 100 * min(q1[d]["days"]) / _med(d)),
    R("Days <50% of median (count)", "int", lambda d: sum(1 for x in q1[d]["days"] if x < 0.5 * _med(d))),
    R("Days delivered (of 30)", "int", lambda d: len(q1[d]["days"])),
    R("% IPv6 rows", "pct", lambda d: 100 * q1[d]["ipv6"] / q1[d]["rows"]),
    R("Unique IPs delivered", "int", lambda d: int(q2[d]["ips_30d"])),
    R("Unique domains delivered", "int", lambda d: int(q2[d]["domains_30d"])),
    R("Unique IP x domain pairs delivered", "int", lambda d: int(q2[d]["ip_domain_pairs_30d"])),

    S("DATA QUALITY (junk flags)"),
    R("% URLs unparseable", "pct2", lambda d: float(q1c[d]["url_parse_fail_pct"])),
    R("% URLs malformed", "pct2", lambda d: float(q1c[d]["url_malformed_pct"])),
    R("% Googlebot IPs", "pct2", lambda d: float(q1c[d]["pct_googlebot_ip"])),
    R("% bot user-agents", "pct", lambda d: float(q1c[d]["ua_bot_pct"]) if q1c[d].get("ua_bot_pct") else None),
    R("Top-1 domain", "txt", lambda d: q1c[d]["top_domain"] or "(empty url)"),
    R("Top-1 domain share", "pct", lambda d: float(q1c[d]["top_domain_share"])),
    R("Top-5 domain share", "pct", lambda d: float(q1c[d]["top5_domain_share"])),
    R("% private IPs", "pct3", lambda d: float(q1c[d]["pct_private_ip"])),
    R("% uid duplicates (clamped >=0)", "pct2", lambda d: max(0.0, float(q1c[d]["uid_dup_pct"]))),
    R("Top-1 timestamp share (stamping check)", "pct2", lambda d: float(q1c[d]["time_top1_share"])),
    R("% user_agent populated", "pct0", lambda d: q1b[d].get("user_agent", 0.0)),
    R("% url populated", "pct0", lambda d: q1b[d].get("url", 0.0)),
    R("% URLs with path", "pct0", lambda d: 100 * q1[d]["path"] / q1[d]["rows"]),
    R("% query_parameters populated", "pct0", lambda d: q1b[d].get("query_parameters", 0.0)),
    R("% advertiser_id populated", "pct0", lambda d: q1b[d].get("advertiser_id", 0.0)),

    S("USABLE FUNNEL (survives to DS13/DS19) — counts are PER DAY (1-day sample 2026-07-01); pairs row is 30d"),
    R("Rows used (per day; 1-day sample)", "int", lambda d: int(q2c[d]["rows_used"])),
    R("% of rows used (within vendor)", "pct", lambda d: 100 * int(q2c[d]["rows_used"]) / int(q2c[d]["rows_raw"])),
    R("% rows hard-dropped", "pct", lambda d: float(q2b[d]["pct_hard_dropped"])),
    R("% rows DS13-blocklisted", "pct", lambda d: float(q2b[d]["pct_blocked_ds13"])),
    R("% rows bot-UA", "pct", lambda d: 100 * int(q2b[d]["rows_bot_ua"]) / int(q2b[d]["rows_day"])),
    R("Unique IPs used (per day)", "int", lambda d: int(q2c[d]["ips_used"])),
    R("% of unique IPs used (within vendor)", "pct", lambda d: 100 * int(q2c[d]["ips_used"]) / int(q2c[d]["ips_raw"])),
    R("Unique domains used (per day, classified)", "int", lambda d: int(q2c[d]["domains_classified"])),
    R("% of domains classified (within vendor)", "pct", lambda d: 100 * int(q2c[d]["domains_classified"]) / int(q2c[d]["domains_raw"])),
    R("Usable IP x domain pairs", "int", lambda d: int(q3[d]["usable_pairs"])),
    R("% of pairs usable", "pct", lambda d: min(100.0, 100 * int(q3[d]["usable_pairs"]) / int(q2[d]["ip_domain_pairs_30d"]))),
    R("% of rows used — share of column total", "pct", lambda d: share(d, q2c, "rows_used")),
    R("% of IPs used — share of column total", "pct", lambda d: share(d, q2c, "ips_used")),
    R("% of domains used — share of column total", "pct", lambda d: share(d, q2c, "domains_classified")),

    S("UNIQUENESS & FRESHNESS (vs all other sources incl. free logs)"),
    R("Sole usable IPs", "int", lambda d: int(q3[d]["sole_ips"])),
    R("% of usable IPs sole", "pct", lambda d: 100 * int(q3[d]["sole_ips"]) / int(q3[d]["usable_ips"])),
    R("Sole usable domains", "int", lambda d: int(q4[d]["sole_domains"])),
    R("Sole CLASSIFIED domains (fee-band axis)", "int", lambda d: int(q4[d]["sole_classified"])),
    R("Pairs per IP (visit density)", "dec1", lambda d: float(q3[d]["pairs_per_ip"])),
    R("% pairs sole", "pct", lambda d: 100 * int(q3[d]["sole_pairs"]) / int(q3[d]["usable_pairs"])
      if d == FREEC else float(q3r[d]["pct_sole"])),
    R("% pairs freshest", "pct", lambda d: float(q3r[d]["pct_freshest"])),
    R("% pairs tied", "pct", lambda d: float(q3r[d]["pct_tied"])),
    R("% pairs stale", "pct", lambda d: float(q3r[d]["pct_stale"])),
    R("% net-new vs free logs", "pct", lambda d: None if d in FREE else float(q3r[d]["pct_netnew_vs_free"])),
    R("Marginal coverage when added (pp)", "dec2", lambda d:
      100.0 * FREE_COV_P / FULL_COV if d == FREEC else add_gain.get(d)),
    R("Frontier add-order rank", "int", lambda d: 0 if d == FREEC else
      (add_order.index(d) + 1 if d in add_order else None)),
    R("Unique visit-days delivered (ip x domain x date, 30d)", "int", lambda d: sum(vend3c[d].values())),
    R("% visit-days sole — new pair", "pct", lambda d: 100 * vend3c[d]["sole_new_pair"] / sum(vend3c[d].values())),
    R("% visit-days sole — recency refresh", "pct", lambda d: 100 * vend3c[d]["sole_refresh"] / sum(vend3c[d].values())),
    R("% visit-days duplicated same-day", "pct", lambda d:
      100 * (trip_holder(d) - solo_sum(d, masks3c)) / trip_holder(d)
      if d == FREEC else 100 * vend3c[d]["shared_same_day"] / sum(vend3c[d].values())),

    S("SERVING & WON BIDS (valuation week; served = won impression)"),
    R("Touched IPs (37d union)", "int", lambda d: int(q5[d]["touched"]["vendor_ips"])),
    R("Served-won IPs", "int", lambda d: int(q5[d]["touched"]["delivered_ips"])),
    R("% of touched IPs served-won", "pct", lambda d: float(q5[d]["touched"]["pct_delivered"])),
    R("Sole IPs served-won", "int", lambda d: int(q6[d]["ips_sole"])),
    R("% of sole stock served-won", "pct2", lambda d: 100 * float(q6[d]["ips_sole"]) / float(q3[d]["sole_ips"])),
    R("Shared IPs served-won", "int", lambda d: float(q6[d]["ips_touched"]) - float(q6[d]["ips_sole"])),
    R("% of served IPs shared", "pct", lambda d: 100 * (float(q6[d]["ips_touched"]) - float(q6[d]["ips_sole"])) / float(q6[d]["ips_touched"])),
    R("Sole won bids / wk", "int", lambda d: float(q6[d]["imps_sole"])),
    R("Sole won bids per served sole IP", "dec1", lambda d: float(q6[d]["imps_sole"]) / float(q6[d]["ips_sole"])),
    R("Sole won bids annualized (x52)", "int", lambda d: float(q6[d]["imps_sole"]) * 52),
    R(f"Billed domains ({BILL_MONTH})", "int", lambda d: int(q1d[d]["billed_domains"]) if d in q1d and q1d[d].get("billed_domains") else None),
    R("% of platform served IPs touched (week)", "pct", lambda d: 100 * float(q6[d]["ips_touched"]) / PLAT_IPS),

    S("SCORE QUALITY (of served/delivered IPs, touched cohort)"),
    R("HI 10000 count", "int", lambda d: int(q5[d]["touched"]["hi_10000"])),
    R("% HI (within vendor)", "pct", lambda d: float(q5[d]["touched"]["pct_hi"])),
    R("% HI — share of column total", "pct", lambda d: 100 * float(q5[d]["touched"]["hi_10000"]) / tier_tot("hi_10000")),
    R("PP 8000 count", "int", lambda d: int(q5[d]["touched"]["pp_8000"])),
    R("% PP (within vendor)", "pct", lambda d: float(q5[d]["touched"]["pct_pp"])),
    R("% PP — share of column total", "pct", lambda d: 100 * float(q5[d]["touched"]["pp_8000"]) / tier_tot("pp_8000")),
    R("High-graduated count (Fangorn band)", "int", lambda d: int(q5[d]["touched"]["high_grad"])),
    R("% high-graduated", "pct", lambda d: float(q5[d]["touched"]["pct_high_grad"])),
    R("% mid", "pct", lambda d: float(q5[d]["touched"]["pct_mid"])),
    R("% max-reach", "pct", lambda d: float(q5[d]["touched"]["pct_maxreach"])),
    R("% unscored", "pct", lambda d: float(q5[d]["touched"]["pct_unscored_delivered"])),
    R("HI 10000 count — SOLE IPs (vendor-unique, served)", "int", lambda d: int(q5[d]["sole"]["hi_10000"])),
    R("% HI of sole served IPs", "pct", lambda d: float(q5[d]["sole"]["pct_hi"])),
    R("PP 8000 count — SOLE IPs (vendor-unique, served)", "int", lambda d: int(q5[d]["sole"]["pp_8000"])),
    R("% PP of sole served IPs", "pct", lambda d: float(q5[d]["sole"]["pct_pp"])),
    R("Avg household score — touched (scored imps)", "int", lambda d: g7b(d, "touched", "avg_hs_scored")),
    R("% imps scored — touched", "pct", lambda d: g7b(d, "touched", "pct_scored")),
    R("Avg household score — sole (scored imps)", "int", lambda d: g7b(d, "sole", "avg_hs_scored")),
    R("% imps scored — sole", "pct", lambda d: g7b(d, "sole", "pct_scored")),

    S("PERFORMANCE — TOUCHED-WON COHORT (per week; mirrors platform — see notes)"),
    R("Spend (media $) — touched", "usd", lambda d: float(q6[d]["media_touched"])),
    R("Impressions (won bids) — touched", "int", lambda d: float(q6[d]["imps_touched"])),
    R("Visits — touched", "int", lambda d: g7b(d, "touched", "visits")),
    R("Conversions — touched", "int", lambda d: g7c(d, "touched", "conversions")),
    R("Revenue — touched", "usd", lambda d: g7c(d, "touched", "revenue")),
    R("CPM — touched", "usd2", lambda d: 1000 * float(q6[d]["media_touched"]) / float(q6[d]["imps_touched"])),
    R("IVR — touched", "pct3", lambda d: g7b(d, "touched", "vr_pct")),
    R("CVR (conv/visits) — touched", "pct2", lambda d: 100 * g7c(d, "touched", "conversions") / g7b(d, "touched", "visits")),
    R("AOV — touched", "usd", lambda d: g7c(d, "touched", "revenue") / g7c(d, "touched", "conversions")),
    R("ROAS — touched", "x2", lambda d: g7c(d, "touched", "revenue") / float(q6[d]["media_touched"])),

    S("PERFORMANCE — UNIQUE (SOLE) COHORT (per week; the vendor discriminator)"),
    R("Spend (media $) — sole", "usd", lambda d: float(q6[d]["media_sole"])),
    R("Impressions (won bids) — sole", "int", lambda d: float(q6[d]["imps_sole"])),
    R("Visits — sole", "int", lambda d: int(q7[d]["sole_visits"])),
    R("Visits — sole, 95% CI", "txt", lambda d: "{:.0f}-{:.0f}".format(*_poisson_ci(int(q7[d]["sole_visits"])))),
    R("Conversions — sole", "int", lambda d: g7c(d, "sole", "conversions")),
    R("Revenue — sole", "usd", lambda d: g7c(d, "sole", "revenue")),
    R("CPM — sole", "usd2", lambda d: 1000 * float(q6[d]["media_sole"]) / float(q6[d]["imps_sole"])),
    R("IVR — sole", "pct4", lambda d: float(q7[d]["vr_overall_pct"])),
    R("IVR — sole, x of 0.0223% baseline", "dec1", lambda d: float(q7[d]["vr_overall_pct"]) / BASELINE_VR),
    R("CVR (conv/visits) — sole", "pct2", lambda d: 100 * g7c(d, "sole", "conversions") / int(q7[d]["sole_visits"]) if int(q7[d]["sole_visits"]) else None),
    R("AOV — sole", "usd", lambda d: g7c(d, "sole", "revenue") / g7c(d, "sole", "conversions") if g7c(d, "sole", "conversions") else None),
    R("ROAS — sole", "x2", lambda d: g7c(d, "sole", "revenue") / float(q6[d]["media_sole"])),

    S("ECONOMICS — COST"),
    R(f"Meter bill ({BILL_MONTH})", "usd", lambda d: 0.0 if d in FREE or d == 27 else
      (q0[d]["june_usd"] if q0.get(d, {}).get("june_usd") is not None else ("pending" if d in FLAT else None)), True),
    R(f"Run-rate $/yr ({BILL_MONTH} x12)", "usd", lambda d: 0.0 if d in FREE or d == 27 else
      (q0[d]["june_usd"] * 12 if q0.get(d, {}).get("june_usd") is not None else ("pending" if d in FLAT else None)), True),
    R("% of delivered rows billed", "pct2", lambda d: None if d in FREE or d in FLAT else
      100 * float(q1d[d]["billed_imps"]) / (q1[d]["rows"] * (365.0 / 12) / 30.0)),
    R("% of sole serves via prospecting (vendor-dependent)", "pct", lambda d: 100 * q6b[d]["prosp"] / q6b[d]["tot"]),

    S("ECONOMICS — WHAT THE DATA IS WORTH (formulas in notes)"),
    R("Max justified CPM — on 100% of delivered rows", "usd4", lambda d: None if d in FREE else
      1000 * t2_ann(d) * 0.30 / (q1[d]["rows"] * ANN30)),
    R("Max justified CPM — on used/billed imps (vs $0.50)", "usd4", lambda d: None if d in FREE else
      (1000 * t2_ann(d) * 0.30 / (float(q1d[d]["billed_imps"]) * 12) if d in q1d and q1d[d].get("billed_imps")
       else 1000 * t2_ann(d) * 0.30 / (float(q2c[d]["rows_used"]) * ANN30))),
    R("Flat equivalent — floor (T1 x 15%)", "usd", lambda d: None if d in FREE else t1_ann(d) * 0.15),
    R("Flat equivalent — fair (T2 x 20%)", "usd", lambda d: None if d in FREE else t2_ann(d) * 0.20),
    R("Flat equivalent — ceiling (T2 x 30%)", "usd", lambda d: None if d in FREE else t2_ann(d) * 0.30),
    R("T2 dependent revenue $/yr (sole-won media x52)", "usd", lambda d: t2_ann(d)),
    R("T2 envelope low (x0.4)", "usd", lambda d: t2_ann(d) * 0.4),
    R("T2 envelope high (x1.8)", "usd", lambda d: t2_ann(d) * 1.8),
    R("T1 provable floor $/yr (score-gated)", "usd", lambda d: t1_ann(d)),
    R("Fee band, domain axis — low (sole classified x $3)", "usd", lambda d: None if d in FREE else float(q4[d]["sole_classified"]) * 3),
    R("Fee band, domain axis — high (sole classified x $13)", "usd", lambda d: None if d in FREE else float(q4[d]["sole_classified"]) * 13),
    R("Scale-normalized: standalone revenue per 1M delivered usable pairs", "usd", lambda d: None if d in FREE else
      (float(q3r[d]["pct_netnew_vs_free"]) / 100) * (t2_ann(d) / float(q3[d]["sole_pairs"])) * 1e6),
    R("Scale-normalized: standalone revenue per 1M RAW delivered pairs", "usd", lambda d: None if d in FREE else
      (float(q3r[d]["pct_netnew_vs_free"]) / 100) * (t2_ann(d) / float(q3[d]["sole_pairs"])) * 1e6
      * min(1.0, int(q3[d]["usable_pairs"]) / int(q2[d]["ip_domain_pairs_30d"]))),

    S("PORTFOLIO (leave-one-out, exact from q3b)"),
    R("Exact drop savings $/yr", "usd", lambda d: None if d in FREE else
      ("= flat fee" if d in FLAT else float(q1d[d]["billed_usd"]) * 12 * (1 - reassign[d].get("metered", 0) / sum(reassign[d].values())))),
    R("Drop savings as % of bill", "pct0", lambda d: None if d in FREE or d in FLAT else
      100 * (1 - reassign[d].get("metered", 0) / sum(reassign[d].values()))),
    R("% credits vanish (were sole)", "pct0", lambda d: 100 * reassign[d]["none"] / sum(reassign[d].values())),
    R("% credits -> flat-fee vendors", "pct0", lambda d: 100 * reassign[d]["flat_fee"] / sum(reassign[d].values())),
    R("% credits -> free logs", "pct0", lambda d: 100 * (reassign[d].get("free_first", 0) + reassign[d].get("free_later", 0)) / sum(reassign[d].values())),
    R("% credits -> other metered (still paid)", "pct", lambda d: 100 * reassign[d].get("metered", 0) / sum(reassign[d].values())),
    R("Coverage lost if dropped (pp of pair coverage)", "dec2", lambda d: -coverage_lost[d] if d in coverage_lost else None),
    R("% of visit-days co-held by our free logs", "pct0", lambda d: None if d in FREE or d not in vend3c else
      100 * trip_free_cohold(d) / trip_holder(d)),
    R("Recoverable $/yr if free logs preempt credit (AUDI-1093, exact visit grain)", "usd",
      lambda d: None if d in FREE or d not in vend3c or d not in q1d or not q1d[d].get("billed_usd") else
      float(q1d[d]["billed_usd"]) * 12 * trip_free_cohold(d) / trip_holder(d)),
    R("Post-preemption bill $/yr (AUDI-1093 applied)", "usd", lambda d: 0.0 if d in FREE else
      ("= flat fee (no meter)" if d in FLAT else
       (bill_after_preempt(d) if d in METERED and q0.get(d, {}).get("june_usd") is not None else None))),

    S("VERDICT"),
    R("Composite quality score (curved, best=100)", "int", lambda d: CURVE[d][1] if d in CURVE else None),
    R("Composite quality score (raw)", "dec1", lambda d: CURVE[d][0] if d in CURVE else None),
    R("Verdict", "txt", lambda d: VERDICT_SHORT.get(d)),
    R("Asks / weird things (full text in notes)", "txt", lambda d: "see notes" if d in ASKS else None),
]


# ---------------- SOLO sheet: SPEC mirrored under the solo counterfactual ----------------
# kind: copy (no entry) | derive (rebased to {vendor+free} world, exact, phase 1)
#       | mask (exact from q3b/q3c/q3d holder masks, phase 1)
#       | est (labeled bound/estimate, phase 1) | scan (q8a/q8b-fed; PENDING till CSVs land)
def _scan(need, fn):
    def wrapped(d):
        if ("a" in need and not Q8A) or ("b" in need and not Q8B):
            return PENDING
        return fn(d)
    return wrapped


def _fresh_share(d, k):
    tot = sum(g8a(d, "fresh_pair", c) for c in ("fresher_than_free", "tied_with_free", "stale_vs_free"))
    return 100 * g8a(d, "fresh_pair", k) / tot if tot else None


def _fday_share(d, k):
    tot = sum(g8a(d, "fresh_day", c)
              for c in ("solo_new_pair", "refresh_of_free_pair", "same_day_dup_with_free"))
    return 100 * g8a(d, "fresh_day", k) / tot if tot else None


def _solo_bill_low(d):
    if d in FREE or d == 27:
        return 0.0
    if d in METERED and q0.get(d, {}).get("june_usd") is not None:
        return q0[d]["june_usd"] * 12
    return "pending" if d in FLAT else None


def _solo_bill_bounds(d):
    lo = q0[d]["june_usd"] * 12
    # clamp: LOW is a hard floor (survivor keeps all current credit); the proportional
    # estimate under-runs junk-billing vendors (credit on domains outside the usable
    # universe is uncontested, so their solo bill ~ today's bill)
    hi = max(lo, TOT_METERED_BILL * trip_holder(d) / PAID_TRIPLES)
    return lo, hi


def _solo_bill_high(d):
    if d in FREE or d == 27:
        return 0.0
    if d in METERED and q0.get(d, {}).get("june_usd") is not None:
        return _solo_bill_bounds(d)[1]
    return "pending" if d in FLAT else None


def _solo_pay_range(d):
    if d in FREE:
        return None
    if Q8B:
        base, tag = t2_solo(d), ""
    elif d in DENS:
        base, tag = solo_sum(d, masks) * DENS[d], " (est)"
    else:
        return None
    return f"{money(base * 0.10)} - {money(base * 0.30)}{tag}"


def _solo_post_preempt(d):
    if d in FREE:
        return "$0"
    if d in FLAT:
        return "= flat fee (no meter)"
    if d not in METERED or q0.get(d, {}).get("june_usd") is None:
        return None
    lo, hi = _solo_bill_bounds(d)
    f = 1 - trip_free_cohold(d) / trip_holder(d)
    if hi == lo:
        return money(lo * f)
    return f"{money(lo * f)} - {money(hi * f)}"


def _solo_worth_bill(d):
    if d in FREE:
        return None
    if d in FLAT:
        return "flat (pending)"
    if d not in METERED or q0.get(d, {}).get("june_usd") is None:
        return None
    lo, hi = _solo_bill_bounds(d)
    t = t2_solo(d)
    if hi == lo:
        return f"{t / lo:.2f}x"
    return f"{t / hi:.2f}x - {t / lo:.2f}x"


SOLO_OVERRIDE = {
    # ---- USABLE FUNNEL: shares rebased to the {vendor + free logs} world ----
    "% of rows used — share of column total": ("derive", "% of rows used — share of {vendor+free} world", None,
        lambda d: solo_share(d, q2c, "rows_used")),
    "% of IPs used — share of column total": ("derive", "% of IPs used — share of {vendor+free} world", None,
        lambda d: solo_share(d, q2c, "ips_used")),
    "% of domains used — share of column total": ("derive", "% of domains used — share of {vendor+free} world", None,
        lambda d: solo_share(d, q2c, "domains_classified")),

    # ---- UNIQUENESS & FRESHNESS ----
    "Sole usable IPs": ("scan", "Solo usable IPs (vs free logs only)", None,
        _scan("a", lambda d: g8a(d, "stock", "solo_ips"))),
    "% of usable IPs sole": ("scan", "% of usable IPs solo", None,
        _scan("a", lambda d: 100 * g8a(d, "stock", "solo_ips") / int(q3[d]["usable_ips"]))),
    "Sole usable domains": ("scan", "Solo domains (vs free logs only)", None,
        _scan("a", lambda d: g8a(d, "stock", "solo_domains"))),
    "Sole CLASSIFIED domains (fee-band axis)": ("scan", "Solo CLASSIFIED domains (fee-band axis, vs free)", None,
        _scan("a", lambda d: g8a(d, "stock", "solo_classified"))),
    "% pairs sole": ("mask", "% usable pairs solo (vs free logs only)", None,
        lambda d: 100 * solo_sum(d, masks) / int(q3[d]["usable_pairs"])),
    "% pairs freshest": ("scan", "% co-held-with-free pairs where vendor fresher", None,
        _scan("a", lambda d: _fresh_share(d, "fresher_than_free"))),
    "% pairs tied": ("scan", "% co-held-with-free pairs tied same-day", None,
        _scan("a", lambda d: _fresh_share(d, "tied_with_free"))),
    "% pairs stale": ("scan", "% co-held-with-free pairs where free fresher", None,
        _scan("a", lambda d: _fresh_share(d, "stale_vs_free"))),
    "Marginal coverage when added (pp)": ("mask", "Pair-coverage gain over free-only (pp)", None,
        lambda d: 100 * (keep_cov(d, masks) - FREE_COV_P) / FULL_COV),
    "Frontier add-order rank": ("mask", "Solo share of net-of-free universe", "pct",
        lambda d: None if d in FREE else 100 * solo_sum(d, masks) / NOF_U),
    "% visit-days sole — new pair": ("scan", "% visit-days solo — new pair (vs free)", None,
        _scan("a", lambda d: _fday_share(d, "solo_new_pair"))),
    "% visit-days sole — recency refresh": ("scan", "% visit-days solo — refresh of free-held pair", None,
        _scan("a", lambda d: _fday_share(d, "refresh_of_free_pair"))),
    "% visit-days duplicated same-day": ("scan", "% visit-days duplicated same-day with free logs", None,
        _scan("a", lambda d: _fday_share(d, "same_day_dup_with_free"))),

    # ---- SERVING & WON BIDS ----
    "Sole IPs served-won": ("scan", "Solo IPs served-won (vs free logs only)", None,
        _scan("b", lambda d: g8b(d, "serve", "ips_solo"))),
    "% of sole stock served-won": ("scan", "% of solo stock served-won", None,
        _scan("ab", lambda d: 100 * g8b(d, "serve", "ips_solo") / g8a(d, "stock", "solo_ips"))),
    "Shared IPs served-won": ("scan", "Served IPs shared with free logs", None,
        _scan("b", lambda d: float(q6[d]["ips_touched"]) - g8b(d, "serve", "ips_solo"))),
    "% of served IPs shared": ("scan", "% of served IPs shared with free logs", None,
        _scan("b", lambda d: 100 * (float(q6[d]["ips_touched"]) - g8b(d, "serve", "ips_solo"))
              / float(q6[d]["ips_touched"]))),
    "Sole won bids / wk": ("scan", "Solo won bids / wk", None,
        _scan("b", lambda d: g8b(d, "serve", "imps"))),
    "Sole won bids per served sole IP": ("scan", "Solo won bids per served solo IP", None,
        _scan("b", lambda d: g8b(d, "serve", "imps") / g8b(d, "serve", "ips_solo"))),
    "Sole won bids annualized (x52)": ("scan", "Solo won bids annualized (x52)", None,
        _scan("b", lambda d: g8b(d, "serve", "imps") * 52)),

    # ---- SCORE QUALITY ----
    "% HI — share of column total": ("derive", "% HI — share of {vendor+free} world", None,
        lambda d: solo_tier_share(d, "hi_10000")),
    "% PP — share of column total": ("derive", "% PP — share of {vendor+free} world", None,
        lambda d: solo_tier_share(d, "pp_8000")),
    "HI 10000 count — SOLE IPs (vendor-unique, served)": ("mask", "HI 10000 count — SOLO IPs (vs free logs only)", None,
        lambda d: solo_sum(d, hi3d)),
    "% HI of sole served IPs": ("scan", "% HI of solo served IPs", None,
        _scan("b", lambda d: 100 * g8b(d, "tier", "hi") / g8b(d, "serve", "ips_solo"))),
    "PP 8000 count — SOLE IPs (vendor-unique, served)": ("mask", "PP 8000 count — SOLO IPs (vs free logs only)", None,
        lambda d: solo_sum(d, pp3d)),
    "% PP of sole served IPs": ("scan", "% PP of solo served IPs", None,
        _scan("b", lambda d: 100 * g8b(d, "tier", "pp") / g8b(d, "serve", "ips_solo"))),
    "Avg household score — sole (scored imps)": ("scan", "Avg household score — solo (scored imps)", None,
        _scan("b", lambda d: g8b(d, "serve", "avg_hs"))),
    "% imps scored — sole": ("scan", "% imps scored — solo", None,
        _scan("b", lambda d: 100 * g8b(d, "serve", "imps_hs_pos") / g8b(d, "serve", "imps"))),

    # ---- PERFORMANCE — SOLO COHORT ----
    "Spend (media $) — sole": ("scan", "Spend (media $) — solo (vs free logs only)", None,
        _scan("b", lambda d: g8b(d, "serve", "media"))),
    "Impressions (won bids) — sole": ("scan", "Impressions (won bids) — solo", None,
        _scan("b", lambda d: g8b(d, "serve", "imps"))),
    "Visits — sole": ("scan", "Visits — solo", None,
        _scan("b", lambda d: g8b(d, "perf", "visits"))),
    "Visits — sole, 95% CI": ("scan", "Visits — solo, 95% CI", None,
        _scan("b", lambda d: "{:.0f}-{:.0f}".format(*_poisson_ci(int(g8b(d, "perf", "visits")))))),
    "Conversions — sole": ("scan", "Conversions — solo", None,
        _scan("b", lambda d: g8b(d, "perf", "conversions"))),
    "Revenue — sole": ("scan", "Revenue — solo", None,
        _scan("b", lambda d: g8b(d, "perf", "revenue"))),
    "CPM — sole": ("scan", "CPM — solo", None,
        _scan("b", lambda d: 1000 * g8b(d, "serve", "media") / g8b(d, "serve", "imps"))),
    "IVR — sole": ("scan", "IVR — solo", None,
        _scan("b", lambda d: 100 * g8b(d, "perf", "visits") / g8b(d, "serve", "imps"))),
    "IVR — sole, x of 0.0223% baseline": ("scan", "IVR — solo, x of 0.0223% baseline", None,
        _scan("b", lambda d: (100 * g8b(d, "perf", "visits") / g8b(d, "serve", "imps")) / BASELINE_VR)),
    "CVR (conv/visits) — sole": ("scan", "CVR (conv/visits) — solo", None,
        _scan("b", lambda d: 100 * g8b(d, "perf", "conversions") / g8b(d, "perf", "visits")
              if g8b(d, "perf", "visits") else None)),
    "AOV — sole": ("scan", "AOV — solo", None,
        _scan("b", lambda d: g8b(d, "perf", "revenue") / g8b(d, "perf", "conversions")
              if g8b(d, "perf", "conversions") else None)),
    "ROAS — sole": ("scan", "ROAS — solo", None,
        _scan("b", lambda d: g8b(d, "perf", "revenue") / g8b(d, "serve", "media"))),

    # ---- ECONOMICS — COST (solo bill = bounded estimate, see notes) ----
    f"Meter bill ({BILL_MONTH})": ("est", "Solo bill $/yr — LOW bound (= today's run-rate)", "usd",
        _solo_bill_low),
    f"Run-rate $/yr ({BILL_MONTH} x12)": ("est", "Solo bill $/yr — HIGH bound (visit-day share of consumed credit)", "usd",
        _solo_bill_high),
    "% of delivered rows billed": ("est", "% of delivered rows billed — solo HIGH bound", None,
        lambda d: None if d not in METERED else
        100 * (_solo_bill_bounds(d)[1] * 2000) / (q1[d]["rows"] * ANN30)),

    # ---- ECONOMICS — WORTH (T2_solo = measured solo-won media x52) ----
    "Max justified CPM — on 100% of delivered rows": ("scan", "Max justified CPM — on 100% of delivered rows (solo)", None,
        _scan("b", lambda d: None if d in FREE else 1000 * t2_solo(d) * 0.30 / (q1[d]["rows"] * ANN30))),
    "Max justified CPM — on used/billed imps (vs $0.50)": ("scan", "Max justified CPM — on used/billed imps (solo, vs $0.50)", None,
        _scan("b", lambda d: None if d in FREE else
              (1000 * t2_solo(d) * 0.30 / (float(q1d[d]["billed_imps"]) * 12)
               if d in q1d and q1d[d].get("billed_imps")
               else 1000 * t2_solo(d) * 0.30 / (float(q2c[d]["rows_used"]) * ANN30)))),
    "Flat equivalent — floor (T1 x 15%)": ("scan", "Flat equivalent — floor (T1_solo x 15%)", None,
        _scan("b", lambda d: None if d in FREE else t1_solo(d) * 0.15)),
    "Flat equivalent — fair (T2 x 20%)": ("scan", "Flat equivalent — fair (T2_solo x 20%)", None,
        _scan("b", lambda d: None if d in FREE else t2_solo(d) * 0.20)),
    "Flat equivalent — ceiling (T2 x 30%)": ("scan", "Flat equivalent — ceiling (T2_solo x 30%)", None,
        _scan("b", lambda d: None if d in FREE else t2_solo(d) * 0.30)),
    "T2 dependent revenue $/yr (sole-won media x52)": ("scan", "T2_solo dependent revenue $/yr (solo-won media x52, MEASURED)", None,
        _scan("b", lambda d: t2_solo(d))),
    "T2 envelope low (x0.4)": ("scan", "T2_solo envelope low (x0.4)", None,
        _scan("b", lambda d: t2_solo(d) * 0.4)),
    "T2 envelope high (x1.8)": ("scan", "T2_solo envelope high (x1.8)", None,
        _scan("b", lambda d: t2_solo(d) * 1.8)),
    "T1 provable floor $/yr (score-gated)": ("scan", "T1_solo provable floor $/yr (score-gated)", None,
        _scan("b", lambda d: t1_solo(d))),
    "Fee band, domain axis — low (sole classified x $3)": ("scan", "Fee band, domain axis — low (solo classified x $3)", None,
        _scan("a", lambda d: None if d in FREE else g8a(d, "stock", "solo_classified") * 3)),
    "Fee band, domain axis — high (sole classified x $13)": ("scan", "Fee band, domain axis — high (solo classified x $13)", None,
        _scan("a", lambda d: None if d in FREE else g8a(d, "stock", "solo_classified") * 13)),
    "Scale-normalized: standalone revenue per 1M delivered usable pairs": ("est",
        "T2_solo $/yr — DENSITY ESTIMATE (= ladder standalone)", "usd",
        lambda d: None if d in FREE or d not in DENS else solo_sum(d, masks) * DENS[d]),
    "Scale-normalized: standalone revenue per 1M RAW delivered pairs": ("est",
        "SOLO PAY RANGE $/yr (10-30% margin on T2_solo; density est until scan lands)", "txt",
        _solo_pay_range),

    # ---- PORTFOLIO -> SOLO coverage block (LOO concepts don't exist solo) ----
    "Exact drop savings $/yr": ("mask", "Pair coverage under {free + vendor} (% of full roster)", "pct2",
        lambda d: 100 * keep_cov(d, masks) / FULL_COV),
    "Drop savings as % of bill": ("mask", "Visit-day coverage under {free + vendor}", "pct2",
        lambda d: 100 * keep_cov(d, masks3c) / TRIP_TOTAL),
    "% credits vanish (were sole)": ("mask", "Visit-day coverage gain over free-only (pp)", "dec2",
        lambda d: 100 * (keep_cov(d, masks3c) - FREE_COV_T) / TRIP_TOTAL),
    "% credits -> flat-fee vendors": ("mask", "HI-IP coverage under {free + vendor}", "pct3",
        lambda d: 100 * keep_cov(d, hi3d) / sum(hi3d.values())),
    "% credits -> free logs": ("mask", "PP-IP coverage under {free + vendor}", "pct3",
        lambda d: 100 * keep_cov(d, pp3d) / sum(pp3d.values())),
    "% credits -> other metered (still paid)": ("mask", "Solo visit-days (vs free logs only)", "int",
        lambda d: solo_sum(d, masks3c)),
    "Coverage lost if dropped (pp of pair coverage)": ("mask", "Share of paid-held visit-days (solo-bill HIGH-bound driver)", "pct",
        lambda d: None if d in FREE else 100 * trip_holder(d) / PAID_TRIPLES),
    "Post-preemption bill $/yr (AUDI-1093 applied)": ("est", "Post-preemption SOLO bill (bounds x (1 - free-cohold))", "txt",
        _solo_post_preempt),

    # ---- VERDICT ----
    "Composite quality score (curved, best=100)": ("scan", "Solo worth / bill — MEASURED (T2_solo vs bill LOW-HIGH)", "txt",
        _scan("b", _solo_worth_bill)),
    "Composite quality score (raw)": ("est", "Solo worth / bill — DENSITY EST (vs today's bill)", "x2",
        lambda d: None if d not in METERED or d not in DENS or q0.get(d, {}).get("june_usd") is None
        else (solo_sum(d, masks) * DENS[d]) / (q0[d]["june_usd"] * 12)),
}


def build_solo_spec():
    out, kinds = [], {}
    unmatched = set(SOLO_OVERRIDE)
    for label, fmt, fn, oos_ok in SPEC:
        if fn is None:
            out.append(S(label))
            continue
        ov = SOLO_OVERRIDE.get(label)
        if ov is None:
            out.append((label, fmt, fn, oos_ok))
            kinds[label] = "copy"
            continue
        unmatched.discard(label)
        kind, nl, nf, nfn = ov
        lbl2, fmt2 = (nl or label), (nf or fmt)
        if lbl2 != label and label in DIR and lbl2 not in DIR:
            DIR[lbl2] = DIR[label]
        out.append((lbl2, fmt2, nfn, oos_ok))
        kinds[lbl2] = kind
    assert not unmatched, f"SOLO_OVERRIDE keys not in SPEC: {unmatched}"
    assert len(out) == len(SPEC), f"SOLO_SPEC length {len(out)} != SPEC {len(SPEC)}"
    return out, kinds


WASTE_SPEC = [
    S("VOLUME DELIVERED (what we ingest)"),
    R("Rows/day (median)", "int", lambda d: _med(d)),
    R("Rows / 30d", "int", lambda d: q1[d]["rows"]),
    R("GB/day on disk (measured, GCS)", "dec1", lambda d: GB_DAY.get(d)),
    R("TB/yr ingest rate", "dec1", lambda d: GB_DAY[d] * 365 / 1000),
    S("USED vs THROWN AWAY"),
    R("% of rows USED (reaches DS13 or DS19)", "pct", lambda d: 100 * used_frac(d)),
    R("% of rows THROWN AWAY", "pct", lambda d: 100 * (1 - used_frac(d))),
    R("Wasted rows/day", "int", lambda d: _med(d) * (1 - used_frac(d))),
    R("Wasted GB/day", "dec1", lambda d: GB_DAY[d] * (1 - used_frac(d))),
    R("Wasted TB/yr", "dec2", lambda d: GB_DAY[d] * (1 - used_frac(d)) * 365 / 1000),
    S("WHY IT'S THROWN AWAY (single-day shares; categories overlap - see index)"),
    R("% hard-dropped (empty/unparseable/infra URL)", "pct", lambda d: float(q2b[d]["pct_hard_dropped"])),
    R("% bot user-agents", "pct", lambda d: 100 * int(q2b[d]["rows_bot_ua"]) / int(q2b[d]["rows_day"])),
    S("USED BUT SHOULDN'T BE (not in 'thrown away' - these rows PASS the permissive DS19 gate and BILL)"),
    R("% webmail rows (DS13-blocked, USED + BILLED via DS19)", "pct", lambda d: float(q2b[d]["pct_blocked_ds13"])),
    R("% Googlebot IPs (bot traffic that bills via DS19)", "pct2", lambda d: float(q1c[d]["pct_googlebot_ip"])),
    R("% URLs malformed (junk; mostly DS19-categorized anyway)", "pct2", lambda d: float(q1c[d]["url_malformed_pct"])),
    R("Top-1 domain share (concentration)", "pct", lambda d: float(q1c[d]["top_domain_share"])),
    S("COST OF THE WASTE (measurable floor; processing compute needs Data Eng)"),
    R("Data fee $/yr (meter bill, context)", "usd", lambda d: 0.0 if d in FREE else
      (q0[d]["june_usd"] * 12 if d in METERED and q0.get(d, {}).get("june_usd") is not None else "flat (pending)")),
    R("Accumulated on disk GB (no TTL since 2025-08-31)", "int", lambda d: GB_ACCUM.get(d)),
    R("Storage floor $/yr (current footprint x $0.02/GB-mo)", "usd", lambda d:
      GB_ACCUM[d] * STORAGE_RATE if d in GB_ACCUM else None),
    R("Wasted-share storage floor $/yr", "usd", lambda d:
      GB_ACCUM[d] * STORAGE_RATE * (1 - used_frac(d)) if d in GB_ACCUM else None),
    S("STOP-SENDING REQUEST (cut needless ingestion at source)"),
    R("The ask", "txt", lambda d: STOP_SENDING.get(d)),
]

WASTE_DEF = {
    "GB/day on disk (measured, GCS)": ("Actual parquet bytes landing per day in site_visit_signal (gsutil du on the data_source_id partitions; avg of 2026-06-15 and 2026-07-01).", "gsutil"),
    "% of rows USED (reaches DS13 or DS19)": ("A row is 'used' if it survives to either consumer (DS13 vertical classification OR DS19 product categorization). Everything else is ingested, stored, and never used. CAUTION: low thrown-away does NOT mean clean data - DS19's permissive gate USES junk (Sovrn throws away only 7.2% because its malformed rows pass DS19 and BILL; the fix-it ask matters more than the waste number there).", "q2c"),
    "% webmail rows (DS13-blocked, USED + BILLED via DS19)": ("NOT part of 'thrown away': webmail is blocked from DS13 only - DS19 has no blocklist, so yahoo/aol rows are USED and BILL (the documented route junk reaches the meter). This is why 33Across can throw away 22% while the stop-sending ask targets ~35%: the ask also kills junk we currently USE and PAY FOR.", "q2b"),
    "WHY IT'S THROWN AWAY (single-day shares; categories overlap - see index)": ("", ""),
    "Accumulated on disk GB (no TTL since 2025-08-31)": ("svs has NO TTL - every day since 2025-08-31 is still on disk. Integrated from 1st-of-month partition-size samples.", "gsutil"),
    "Storage floor $/yr (current footprint x $0.02/GB-mo)": ("GCS standard list price on today's accumulated footprint. A FLOOR on ingestion cost: excludes Kafka cluster share (real-time vendors), ingest DAG compute, DS13/DS19 classifier compute, and BQ processing - those need Data Eng numbers.", "computed"),
    "The ask": ("What to request the vendor stop sending so we do not ingest/store/process data we never use. Volume shares from the DATA QUALITY evidence.", "q1c/q2b"),
}

# waste reason-shares overlap note: hard-drop, blocklist, bot and malformed are not
# mutually exclusive and blocklisted rows may still be USED via DS19 - the decomposition
# is indicative, only '% thrown away' is exact.

def solo_anchor_checks():
    for d in EXT:
        a, b = solo_sum(d, masks), int(q3[d]["netnew_vs_free_pairs"])
        assert a == b, f"anchor1 FAIL ds{d}: mask solo pairs {a} != q3 netnew_vs_free {b}"
    tot = sum(masks.values())
    parts = (solo_sum(23, masks) + solo_sum(30, masks)
             + sum(n for m, n in masks.items() if (m & FREE_MASK) == FREE_MASK)
             + sum(n for m, n in masks.items() if not (m & FREE_MASK)))
    assert parts == tot, f"anchor6 FAIL: free-partition {parts} != mask total {tot}"
    # AUDI-1093 regression tripwire: total preemption cut must stay at the published ~$273.7K
    tot_cut = sum(preempt_cut(d) for d in METERED if q0.get(d, {}).get("june_usd") is not None)
    assert 265_000 < tot_cut < 285_000, f"preemption total drifted: ${tot_cut:,.0f}"
    print(f"preemption total: ${tot_cut:,.0f}/yr of ${TOT_METERED_BILL:,.0f} "
          f"({100 * tot_cut / TOT_METERED_BILL:.1f}%)")
    if Q8A:
        # cross-scan convergence: q8a same-day-dup share must reproduce the q3c cohold share
        for d in METERED:
            fd_tot = sum(g8a(d, "fresh_day", c) for c in
                         ("solo_new_pair", "refresh_of_free_pair", "same_day_dup_with_free"))
            if not fd_tot:
                continue
            dup = g8a(d, "fresh_day", "same_day_dup_with_free") / fd_tot
            ch = trip_free_cohold(d) / trip_holder(d)
            if abs(dup - ch) > 0.005:
                print(f"WARN preemption-share divergence ds{d}: q8a dup {100 * dup:.2f}% "
                      f"vs q3c cohold {100 * ch:.2f}%")
    if Q13A:
        # truncation guard: all 5 rec types must have survived --max_rows
        assert q13a_ds and Q13A_PATH and q13a_pair and q13a_trip and q13a_cat, \
            "q13a CSV missing rec types (silent bq truncation?)"
        for d in ACTIVE:
            if d not in q13a_ds or d not in q2c:
                continue
            # q2c is a ONE-DAY sample (2026-07-01); q13a spans 30d -> compare per-day
            # average with day-to-day variance tolerance (warn 25%, fail 60%)
            qa, qc = q13a_ds[d] / 30.0, int(q2c[d]["rows_ds19_cat"])
            if qc and abs(qa - qc) / qc > 0.60:
                raise AssertionError(f"q13a ds-anchor FAIL ds{d}: {qa:,.0f}/day vs q2c sample-day {qc:,}")
            if qc and abs(qa - qc) / qc > 0.25:
                print(f"q13a ds-anchor day-variance ds{d}: {qa:,.0f}/day avg vs q2c sample-day {qc:,} "
                      f"({100 * (qa - qc) / qc:+.0f}%)")
        pcov = ds19_cov([], q13a_pair)
        tcov = ds19_cov([], q13a_trip)
        pathcov = Q13A_PATH[1] / Q13A_PATH[0]
        print(f"DS19-only free coverage: pair {100 * pcov:.1f}% / visit-day {100 * tcov:.1f}% / "
              f"true-path {100 * pathcov:.1f}% (union anchors: 60.4 / 59.4)")
    if Q13B:
        mf = q13b.get("free_covered", {}).get("member_ips", 0)
        mv = q13b.get("vendor_only", {}).get("member_ips", 0)
        if mf + mv:
            print(f"DS19 members: free-covered {100 * mf / (mf + mv):.1f}% / "
                  f"vendor-only {100 * mv / (mf + mv):.1f}% of {mf + mv:,.0f} IPs")
        th = q13b.get("free_covered", {}).get("hi", 0)
        tv = q13b.get("vendor_only", {}).get("hi", 0)
        if th + tv:
            print(f"DS19 served-HI free-covered share: {100 * th / (th + tv):.2f}% (expect >=99)")
    if Q8A:
        for d in ACTIVE:
            qa, ms = g8a(d, "stock", "solo_pairs"), solo_sum(d, masks)
            if ms and abs(qa - ms) / ms > 0.001:
                raise AssertionError(f"anchor3 FAIL ds{d}: q8a solo_pairs {qa:.0f} vs mask {ms}")
            if qa != ms:
                print(f"anchor3 drift ds{d}: q8a {qa:.0f} vs mask {ms} ({100 * (qa - ms) / ms:+.4f}%)")
            fd = g8a(d, "fresh_day", "solo_new_pair") + g8a(d, "fresh_day", "refresh_of_free_pair")
            mt = solo_sum(d, masks3c)
            if mt and abs(fd - mt) / mt > 0.001:
                raise AssertionError(f"anchor7 FAIL ds{d}: fresh_day solo {fd:.0f} vs mask {mt}")
    if Q8B:
        for d in ACTIVE:
            assert g8b(d, "serve", "media") >= float(q6[d]["media_sole"]) - 0.01, f"anchor4 media ds{d}"
            assert g8b(d, "serve", "ips_solo") >= float(q6[d]["ips_sole"]), f"anchor4 ips ds{d}"
            assert g8b(d, "serve", "imps") >= float(q6[d]["imps_sole"]), f"anchor4 imps ds{d}"
            # q8b tiers use RAW 37d membership (q5/q6 cohort convention); q3d masks are
            # usable-gated. The gap is DEFINITIONAL and itself diagnostic: clean vendors
            # run a few % LOW in q8b (free-log webmail sightings only count raw), junk
            # vendors run far HIGH (Sovrn +68%: malformed-URL rows carry IPs that never
            # reach a usable domain). Report, never abort - the two lenses differ by design.
            for tk, uni in (("hi", hi3d), ("pp", pp3d)):
                qb, ms = g8b(d, "tier", tk), solo_sum(d, uni)
                if ms >= 100 and qb != ms:
                    print(f"anchor5 raw-vs-usable gap ds{d} {tk}: q8b {qb:.0f} vs q3d mask {ms} "
                          f"({100 * (qb - ms) / ms:+.2f}%)")


# pct* store the FRACTION (0.218) with a true percent format so Excel/Sheets
# recognize the cell as a percentage; the writer divides percent numbers by 100.
FMT = {"int": "#,##0", "pct0": "0%", "pct": "0.0%", "pct2": "0.00%",
       "pct3": "0.000%", "pct4": "0.0000%", "usd": "$#,##0", "usd2": "$#,##0.00",
       "usd4": "$#,##0.0000", "dec1": "0.0", "dec2": "0.00", "x2": '0.00"x"', "txt": None}

# Per-row heat direction for the color scale across vendor columns:
# +1 green = highest, -1 green = lowest, absent = no shading (identity/neutral rows).
DIR = {
    "Total rows delivered": 1, "Median rows/day": 1, "Weakest day (% of median)": 1,
    "Days <50% of median (count)": -1, "Days delivered (of 30)": 1, "% IPv6 rows": -1,
    "Unique IPs delivered": 1, "Unique domains delivered": 1, "Unique IP x domain pairs delivered": 1,
    "% URLs unparseable": -1, "% URLs malformed": -1, "% Googlebot IPs": -1,
    "% bot user-agents": -1, "Top-1 domain share": -1, "Top-5 domain share": -1,
    "% private IPs": -1, "% uid duplicates (clamped >=0)": -1,
    "Top-1 timestamp share (stamping check)": -1,
    "% user_agent populated": 1, "% url populated": 1, "% URLs with path": 1,
    "% query_parameters populated": 1, "% advertiser_id populated": 1,
    "Rows used (per day; 1-day sample)": 1, "% of rows used (within vendor)": 1, "% rows hard-dropped": -1,
    "% rows DS13-blocklisted": -1, "% rows bot-UA": -1, "Unique IPs used (per day)": 1,
    "% of unique IPs used (within vendor)": 1, "Unique domains used (per day, classified)": 1,
    "% of domains classified (within vendor)": 1, "Usable IP x domain pairs": 1,
    "% of pairs usable": 1, "% of rows used — share of column total": 1,
    "% of IPs used — share of column total": 1, "% of domains used — share of column total": 1,
    "Sole usable IPs": 1, "% of usable IPs sole": 1, "Sole usable domains": 1,
    "Sole CLASSIFIED domains (fee-band axis)": 1, "Pairs per IP (visit density)": 1,
    "% pairs sole": 1, "% pairs freshest": 1, "% pairs tied": -1, "% pairs stale": -1,
    "% net-new vs free logs": 1, "Marginal coverage when added (pp)": 1,
    "Frontier add-order rank": -1,
    "Touched IPs (37d union)": 1, "Served-won IPs": 1, "% of touched IPs served-won": 1,
    "Sole IPs served-won": 1, "% of sole stock served-won": 1,
    "% of served IPs shared": -1, "Sole won bids / wk": 1,
    "Sole won bids per served sole IP": 1, "Sole won bids annualized (x52)": 1,
    "% of platform served IPs touched (week)": 1,
    "HI 10000 count": 1, "% HI (within vendor)": 1, "% HI — share of column total": 1,
    "PP 8000 count": 1, "% PP (within vendor)": 1, "% PP — share of column total": 1,
    "High-graduated count (Fangorn band)": 1, "% high-graduated": 1,
    "% max-reach": -1, "% unscored": -1,
    "HI 10000 count — SOLE IPs (vendor-unique, served)": 1,
    "% HI of sole served IPs": 1,
    "PP 8000 count — SOLE IPs (vendor-unique, served)": 1,
    "% PP of sole served IPs": 1,
    "Avg household score — touched (scored imps)": 1, "% imps scored — touched": 1,
    "Avg household score — sole (scored imps)": 1, "% imps scored — sole": 1,
    "Spend (media $) — touched": 1, "Impressions (won bids) — touched": 1,
    "Visits — touched": 1, "Conversions — touched": 1, "Revenue — touched": 1,
    "IVR — touched": 1, "CVR (conv/visits) — touched": 1, "AOV — touched": 1,
    "ROAS — touched": 1,
    "Spend (media $) — sole": 1, "Impressions (won bids) — sole": 1, "Visits — sole": 1,
    "Conversions — sole": 1, "Revenue — sole": 1, "IVR — sole": 1,
    "IVR — sole, x of 0.0223% baseline": 1, "CVR (conv/visits) — sole": 1,
    "AOV — sole": 1, "ROAS — sole": 1,
    "% of delivered rows billed": -1,
    "Max justified CPM — on 100% of delivered rows": 1,
    "Max justified CPM — on used/billed imps (vs $0.50)": 1,
    "Flat equivalent — floor (T1 x 15%)": 1, "Flat equivalent — fair (T2 x 20%)": 1,
    "Flat equivalent — ceiling (T2 x 30%)": 1,
    "T2 dependent revenue $/yr (sole-won media x52)": 1,
    "T2 envelope low (x0.4)": 1, "T2 envelope high (x1.8)": 1,
    "T1 provable floor $/yr (score-gated)": 1,
    "Fee band, domain axis — low (sole classified x $3)": 1,
    "Fee band, domain axis — high (sole classified x $13)": 1,
    "Scale-normalized: standalone revenue per 1M delivered usable pairs": 1,
    "Scale-normalized: standalone revenue per 1M RAW delivered pairs": 1,
    "Exact drop savings $/yr": 1, "Drop savings as % of bill": 1,
    "Coverage lost if dropped (pp of pair coverage)": 1,
    "% of visit-days co-held by our free logs": 1,
    "Recoverable $/yr if free logs preempt credit (AUDI-1093, exact visit grain)": 1,
    "Unique visit-days delivered (ip x domain x date, 30d)": 1,
    "% visit-days sole — new pair": 1,
    "% visit-days sole — recency refresh": 1,
    "% visit-days duplicated same-day": -1,
    "Composite quality score (curved, best=100)": 1, "Composite quality score (raw)": 1,
}
DIR.update({
    "% mid": 1,
    "% of sole serves via prospecting (vendor-dependent)": 1,
    "% credits vanish (were sole)": 1,          # vanished credits = saved on drop
    "% credits -> flat-fee vendors": 1,          # absorbed free = saved on drop
    "% credits -> free logs": 1,                 # absorbed free = saved on drop
    "% credits -> other metered (still paid)": -1,
})
# Bill rows (labels are f-strings on BILL_MONTH): lower bill = green.
DIR[f"Meter bill ({BILL_MONTH})"] = -1
DIR[f"Run-rate $/yr ({BILL_MONTH} x12)"] = -1

# Scale-visibility-only rows (no good/bad direction): white -> steel-blue scale.
NEUTRAL = {"CPM — touched", "CPM — sole",
           "Contract rate ($ CPM; flat amounts pending)",
           f"Billed domains ({BILL_MONTH})"}

# Build the SOLO spec AFTER DIR/NEUTRAL exist (renamed rows inherit direction here),
# then override directions for replacement rows whose meaning flipped.
SOLO_SPEC, SOLO_KIND = build_solo_spec()
DIR.update({
    "Solo bill $/yr — LOW bound (= today's run-rate)": -1,
    "Solo bill $/yr — HIGH bound (visit-day share of consumed credit)": -1,
    "% of delivered rows billed — solo HIGH bound": -1,
    "% co-held-with-free pairs where vendor fresher": 1,
    "% co-held-with-free pairs tied same-day": -1,
    "% co-held-with-free pairs where free fresher": -1,
    "Solo share of net-of-free universe": 1,
    "Pair-coverage gain over free-only (pp)": 1,
    "Pair coverage under {free + vendor} (% of full roster)": 1,
    "Visit-day coverage under {free + vendor}": 1,
    "Visit-day coverage gain over free-only (pp)": 1,
    "HI-IP coverage under {free + vendor}": 1,
    "PP-IP coverage under {free + vendor}": 1,
    "Solo visit-days (vs free logs only)": 1,
    "Share of paid-held visit-days (solo-bill HIGH-bound driver)": -1,
    "T2_solo $/yr — DENSITY ESTIMATE (= ladder standalone)": 1,
    "Solo worth / bill — DENSITY EST (vs today's bill)": 1,
})
DIR["Post-preemption bill $/yr (AUDI-1093 applied)"] = -1
NEUTRAL.add("CPM — solo")

# WASTE sheet heat directions (defined here, after DIR/NEUTRAL exist)
DIR.update({
    "Rows/day (median)": 1, "Rows / 30d": 1,
    "% of rows USED (reaches DS13 or DS19)": 1,
    "% of rows THROWN AWAY": -1,
    "Wasted rows/day": -1, "Wasted GB/day": -1, "Wasted TB/yr": -1,
    "% hard-dropped (empty/unparseable/infra URL)": -1,
    "% webmail rows (DS13-blocked, USED + BILLED via DS19)": -1,
    "% Googlebot IPs (bot traffic that bills via DS19)": -1,
    "% URLs malformed (junk; mostly DS19-categorized anyway)": -1,
    "Top-1 domain share (concentration)": -1,
    "Data fee $/yr (meter bill, context)": -1,
    "Accumulated on disk GB (no TTL since 2025-08-31)": -1,
    "Storage floor $/yr (current footprint x $0.02/GB-mo)": -1,
    "Wasted-share storage floor $/yr": -1,
})
NEUTRAL.add("GB/day on disk (measured, GCS)")
NEUTRAL.add("TB/yr ingest rate")


def money(v):
    v = float(v)
    if abs(v) >= 1e6:
        return f"${v / 1e6:.2f}M"
    if abs(v) >= 1e3:
        return f"${v / 1e3:.1f}K"
    return f"${v:,.0f}"



# ---------------- index-sheet definitions: label -> (what it means / how computed, source) ----------------
DEF = {
    "Data source ID": ("MNTN's internal id for the feed in site_visit_signal and the billing registry. DS28+DS40 are the SAME vendor (batch vs real-time).", "q0"),
    "Billing type": ("metered CPM = pay $0.50 per USED signal impression; flat fee = fixed price regardless of use; free = internal MNTN log.", "q0"),
    "Contract rate ($ CPM; flat amounts pending)": ("$ per 1,000 used signal impressions (metered). Flat amounts pending Maya / renewal schedule.", "q0"),
    "Renewal status": ("Contract state. Only lever on flat-fee vendors is the renewal date.", "q0"),
    "Ingestion path": ("batch = daily file drop into the ingest DAG; Kafka RT = real-time pixel stream. Determines the off-switch owner.", "code audit"),
    "Non-MM blast radius": ("Production systems OUTSIDE MM that depend on this feed - check before any drop (Predactiv HEM feeds CRM/identity).", "code audit"),
    "Total rows delivered": ("Raw events landed in site_visit_signal over the 30d window (2026-06-02..07-01).", "q1"),
    "Median rows/day": ("Typical daily delivery volume (median of 30 daily counts).", "q1"),
    "Weakest day (% of median)": ("Worst delivery day as % of the median day - feed reliability.", "q1"),
    "Days <50% of median (count)": ("Partial-day incidents: days delivering under half the median.", "q1"),
    "Days delivered (of 30)": ("Liveness: days with any delivery. Gate = 29+/30.", "q1"),
    "% IPv6 rows": ("Rows with IPv6 addresses - EXCLUDED from all IP analyses (footprint undercount flag for that vendor).", "q1"),
    "Unique IPs delivered": ("Distinct IPv4 households seen in 30d.", "q2"),
    "Unique domains delivered": ("Distinct registered domains (eTLD+1) in 30d.", "q2"),
    "Unique IP x domain pairs delivered": ("Distinct (IP, domain) combinations = site-visit facts. THE unit of billing credit and uniqueness analysis.", "q2"),
    "% URLs unparseable": ("URL fails host extraction entirely.", "q1c"),
    "% URLs malformed": ("URL structurally broken (e.g. Sovrn's host-doubled concat bug).", "q1c"),
    "% Googlebot IPs": ("Rows from Google crawler IPs - bot traffic that can bill through DS19.", "q1c"),
    "% bot user-agents": ("Rows whose user_agent identifies a bot (only measurable where UA is sent).", "q1c"),
    "Top-1 domain": ("Most frequent domain - is the feed one site in a trenchcoat? '(empty url)' = the modal bucket is rows with no/unparseable URL (guid_log: ~25% of its rows).", "q1c"),
    "Top-1 domain share": ("% of rows on that single domain.", "q1c"),
    "Top-5 domain share": ("% of rows on the top five domains (concentration).", "q1c"),
    "% private IPs": ("RFC1918 unroutable addresses (10.x etc.) - junk.", "q1c"),
    "% uid duplicates (clamped >=0)": ("Repeated event uids. ~0 within sketch error; negative HLL estimates clamped to 0.", "q1c"),
    "Top-1 timestamp share (stamping check)": ("% of rows sharing one exact timestamp - batch re-stamping red flag.", "q1c"),
    "% user_agent populated": ("Field richness: does the vendor send user_agent (enables bot filtering BEFORE we get billed).", "q1b"),
    "% url populated": ("Rows with a URL at all.", "q1b"),
    "% URLs with path": ("URLs deeper than the homepage - page-level signal (BUK/DS38 input).", "q1"),
    "% query_parameters populated": ("Query-string capture (checkout tokens etc.) - nobody sends it today.", "q1b"),
    "% advertiser_id populated": ("Vendor tags which advertiser the visit belongs to (only guid_log).", "q1b"),
    "Rows used (per day; 1-day sample)": ("Rows surviving to a consumer on the SAMPLE DAY (2026-07-01): DS13 vertical classification OR DS19 product categorization. The OR defines usable = billable. PER-DAY number - compare to Median rows/day, NOT to the 30d total (839M/day used of 1.08B/day delivered = 77.6%, for 33Across).", "q2c"),
    "% of rows used (within vendor)": ("Survival rate of the vendor's own feed.", "q2c"),
    "% rows hard-dropped": ("Dropped before any consumer: unparseable / empty / infra URLs.", "q2b"),
    "% rows DS13-blocklisted": ("On DS13's domain blocklist (webmail: yahoo/aol/easybrain) - still billable via DS19 (no blocklist there).", "q2b"),
    "% rows bot-UA": ("Bot user-agent rows entering the pipeline.", "q2b"),
    "Unique IPs used (per day)": ("Distinct IPs on used rows, sample day.", "q2c"),
    "% of unique IPs used (within vendor)": ("IP-grain survival rate.", "q2c"),
    "Unique domains used (per day, classified)": ("Domains the classifiers can actually consume, sample day.", "q2c"),
    "% of domains classified (within vendor)": ("Domain-grain survival rate.", "q2c"),
    "Usable IP x domain pairs": ("(IP, domain) facts surviving to consumers - the credit-eligible pool.", "q3"),
    "% of pairs usable": ("Usable pairs / ALL delivered pairs. The filter is on DOMAINS, not IPs: a pair is usable only if its domain is in the consumable universe (wcv verticals UNION pc product categories). <100% = the vendor ships pairs on domains NEITHER classifier knows (ad-infra, sync endpoints, never-categorized sites) - ingested but can never classify or bill. Capped at 100 (q3-vs-q2 scans differ <1%).", "q3/q2"),
    "% of rows used — share of column total": ("This vendor's slice of ALL sources' used rows. Sources overlap, so the row sums >100% across vendors.", "q2c"),
    "% of IPs used — share of column total": ("Same, IP grain (overlapping).", "q2c"),
    "% of domains used — share of column total": ("Same, domain grain (overlapping).", "q2c"),
    "Sole usable IPs": ("Usable IPs seen by NO other source (incl. our free logs) in 30d - the vendor's unique household contribution.", "q3"),
    "% of usable IPs sole": ("Unique share of its usable footprint.", "q3"),
    "Sole usable domains": ("Domains only this vendor delivers.", "q4"),
    "Sole CLASSIFIED domains (fee-band axis)": ("Unique domains MM can consume - the durable value unit; drives the $3-13 fee band.", "q4"),
    "Pairs per IP (visit density)": ("Site-visit depth per household (usable pairs / usable IPs).", "q3"),
    "% pairs sole": ("Pairs where this vendor is the ONLY holder.", "q3 pair recency"),
    "% pairs freshest": ("Shared pairs where this vendor reported most recently.", "q3 pair recency"),
    "% pairs tied": ("Shared pairs reported same-day by another source.", "q3 pair recency"),
    "% pairs stale": ("Shared pairs where another source is fresher.", "q3 pair recency"),
    "% net-new vs free logs": ("Pairs our own guid/augmentor logs did NOT already have - what we could not collect ourselves.", "q3 pair recency"),
    "Marginal coverage when added (pp)": ("Usable-pair coverage the vendor adds at its position in the greedy best-first add order (free logs always in).", "q3b masks"),
    "Frontier add-order rank": ("Position in that greedy order; 1 = most additive vendor.", "q3b masks"),
    "Touched IPs (37d union)": ("Every IP the vendor delivered in the 37d union window - all it could have influenced.", "q5"),
    "Served-won IPs": ("Touched IPs that WON at least one impression in the valuation week (2026-07-02..08). Won imps = cost_impression_log.", "q5"),
    "% of touched IPs served-won": ("How much of the footprint the bidder actually reaches.", "q5"),
    "Sole IPs served-won": ("Unique-to-vendor IPs that served - the dependency stock that mattered this week.", "q6"),
    "% of sole stock served-won": ("Serve rate of the unique stock (0.2-0.3% - sole IPs barely appear in auctions).", "q6/q3"),
    "Shared IPs served-won": ("Served IPs other sources also delivered - replaceable coverage.", "q6"),
    "% of served IPs shared": ("Share of its served footprint that is replaceable (99.6%+ everywhere).", "q6"),
    "Sole won bids / wk": ("Won impressions on sole IPs per week - the dependent media flow.", "q6"),
    "Sole won bids per served sole IP": ("Frequency on those households (~4.5-5.5 everywhere).", "q6"),
    "Sole won bids annualized (x52)": ("Yearly expected won bids IF the weekly flow persists. Flow x52 - never annualize the IP stock.", "q6"),
    "% of platform served IPs touched (week)": ("Vendor footprint vs ALL 28.03M distinct served IPs that week (honest, non-overlapping denominator).", "q6/q7d"),
    "HI 10000 count": ("Served IPs pinned HI (highest-intent tier).", "q5"),
    "% HI (within vendor)": ("HI share of ITS served IPs.", "q5"),
    "% HI — share of column total": ("Slice of all sources' HI pools (overlapping - sums >100%).", "q5"),
    "PP 8000 count": ("Served IPs pinned PP (positive-prospect tier).", "q5"),
    "% PP (within vendor)": ("PP share of its served IPs.", "q5"),
    "% PP — share of column total": ("Overlapping share of the PP pool.", "q5"),
    "High-graduated count (Fangorn band)": ("Scores 6666-9999 excl. 8000 - the Fangorn continuous high band.", "q5"),
    "% high-graduated": ("Its share in that band.", "q5"),
    "% mid": ("Scores 3333-6665.", "q5"),
    "% max-reach": ("Scores 1-3332 - lowest scored tier.", "q5"),
    "% unscored": ("Served IPs with no score (<=0) - untargetable beyond max-reach.", "q5"),
    "HI 10000 count — SOLE IPs (vendor-unique, served)": ("Served HI IPs that ONLY this vendor delivered - what the roster loses in HI if the vendor drops (the q3d scenario losses are these singletons + combination-only IPs).", "q5"),
    "% HI of sole served IPs": ("HI share of the vendor's unique served IPs (vs 22-37% on touched - adverse selection).", "q5"),
    "PP 8000 count — SOLE IPs (vendor-unique, served)": ("Served PP IPs only this vendor delivered.", "q5"),
    "% PP of sole served IPs": ("PP share of the vendor's unique served IPs.", "q5"),
    "Avg household score — touched (scored imps)": ("Mean household_score on the cohort's SCORED impressions only (RT rows carry HS=-1 and are excluded).", "q7b"),
    "% imps scored — touched": ("How much of the cohort's inventory carries any score.", "q7b"),
    "Avg household score — sole (scored imps)": ("Same for the unique-IP cohort - reads much lower (adverse selection).", "q7b"),
    "% imps scored — sole": ("Scored share of sole imps (1-6% vs 28-31% touched).", "q7b"),
    "Spend (media $) — touched": ("Weekly media on ALL its served IPs. Touched rows MIRROR THE PLATFORM (pools overlap) - context only, not a discriminator.", "q6"),
    "Impressions (won bids) — touched": ("Weekly won imps, touched cohort.", "q6"),
    "Visits — touched": ("Clickpass visits joined per ad_served_id (trail to 07-10).", "q7b"),
    "Conversions — touched": ("Attributed conversions (last-touch dedup, no assists/disputed).", "q7c"),
    "Revenue — touched": ("order_amt on those conversions.", "q7c"),
    "CPM — touched": ("media / imps x 1000.", "q6"),
    "IVR — touched": ("visits / imps.", "q7b"),
    "CVR (conv/visits) — touched": ("conversions / visits.", "q7c/q7b"),
    "AOV — touched": ("revenue / conversions.", "q7c"),
    "ROAS — touched": ("revenue / media spend.", "q7c/q6"),
    "Spend (media $) — sole": ("Weekly media on its UNIQUE IPs only - the true vendor discriminator; the counterfactual spend at risk.", "q6"),
    "Impressions (won bids) — sole": ("Weekly won imps on sole IPs.", "q6"),
    "Visits — sole": ("Clickpass visits on sole-IP serves (q7 canonical measurement).", "q7"),
    "Visits — sole, 95% CI": ("Poisson (Garwood) 95% interval on that count - these are small numbers.", "q7"),
    "Conversions — sole": ("Attributed conversions on sole-IP serves. Poisson-tiny: read 0 as under ~1/wk.", "q7c"),
    "Revenue — sole": ("order_amt on those conversions.", "q7c"),
    "CPM — sole": ("Sole media / sole imps x 1000 (~$11.5-12 everywhere).", "q6"),
    "IVR — sole": ("Sole visits / sole imps.", "q7"),
    "IVR — sole, x of 0.0223% baseline": ("Multiple of the no-svs-data sole-serve baseline. ~1x = behaves like inventory nobody had data on.", "q7"),
    "CVR (conv/visits) — sole": ("Conversions per sole visit ('-' when 0 visits). Roughly platform-normal where measurable - the collapse is at the VISIT step.", "q7c/q7"),
    "AOV — sole": ("Revenue / conversions on sole serves.", "q7c"),
    "ROAS — sole": ("Sole revenue / sole media.", "q7c/q6"),
    "% of delivered rows billed": ("Billed imps / delivered rows - how little of the feed we pay for (first-reporter credit + used gate).", "q1d/q1"),
    "% of sole serves via prospecting (vendor-dependent)": ("Sole-IP serves through MM-audience-gated prospecting (incl. max-reach) - the share that genuinely required the vendor's data (97-99%).", "q6b"),
    "Max justified CPM — on 100% of delivered rows": ("Break-even CPM if we paid for EVERY delivered row: (T2 x 30% margin) / annual rows x 1000. Shows why per-row pricing must be fractions of a cent.", "computed"),
    "Max justified CPM — on used/billed imps (vs $0.50)": ("Same value spread over only billed/used imps - compare directly to the $0.50 we pay (flat vendors: hypothetical meter on used rows).", "computed"),
    "Flat equivalent — floor (T1 x 15%)": ("Flat fee justified by PROVABLE dependency at conservative margin.", "computed"),
    "Flat equivalent — fair (T2 x 20%)": ("Flat fee justified by the full dependency ceiling at mid margin.", "computed"),
    "Flat equivalent — ceiling (T2 x 30%)": ("Never-pay-more line: full ceiling at top margin.", "computed"),
    "T2 dependent revenue $/yr (sole-won media x52)": ("THE dependency ceiling: media revenue on sole-won imps, annualized. What could vanish if the vendor left (97-99% prospecting-attributed).", "q6"),
    "T2 envelope low (x0.4)": ("Stress band bottom (volume x0.5, CPM x0.8). Scenario range, NOT a confidence interval.", "computed"),
    "T2 envelope high (x1.8)": ("Stress band top (volume x1.5, CPM x1.2).", "computed"),
    "T1 provable floor $/yr (score-gated)": ("Media on sole imps where a HIGH score (>=6666, non-RTC) GATED the serve - the narrowest dependency claim. Note: ANY score on a sole IP is vendor-derived (~1-6% of sole imps), and membership itself (97-99% of sole serves, q6b) already required the vendor - which is why T2, not T1, is the decision number.", "q6"),
    "Fee band, domain axis — low (sole classified x $3)": ("Domain-axis worth: unique classified domains x $3/domain-yr (roster-calibrated).", "q4"),
    "Fee band, domain axis — high (sole classified x $13)": ("Same at the generous $13/domain-yr ceiling.", "q4"),
    "Scale-normalized: standalone revenue per 1M delivered usable pairs": ("The same-scale hypothetical: net-new-vs-free rate x measured revenue density - what 1M of this vendor's USABLE pairs (survived DS13/DS19 = the billable pool) would be worth standalone at equal volume. Numerator is measured serve revenue (T2 density). CAVEATS: Sovrn junk-uniqueness-inflated; small-vendor densities Poisson-noisy; densities revert with scale.", "q3/q3r/q6"),
    "Scale-normalized: standalone revenue per 1M RAW delivered pairs": ("Same metric per RAW shipped pair (x usable rate) - what 1M pairs of the vendor's actual feed are worth before filtering. Dirty feeds deflate here (Sovrn $255 -> $158).", "q3/q3r/q6/q2"),
    "Exact drop savings $/yr": ("Bill x share of its credits that do NOT re-race to another metered vendor (first-reporter reassignment, measured over 30d). Flat vendors: savings = the fee itself.", "q3b"),
    "Drop savings as % of bill": ("Recovery rate of the bill if dropped.", "q3b"),
    "% credits vanish (were sole)": ("Credits nobody else re-reports - saved.", "q3b"),
    "% credits -> flat-fee vendors": ("Credits absorbed by flat vendors at no marginal cost - saved.", "q3b"),
    "% credits -> free logs": ("Credits our own logs absorb - saved.", "q3b"),
    "% credits -> other metered (still paid)": ("Credits that just move to another $0.50 meter - NOT saved.", "q3b"),
    "Coverage lost if dropped (pp of pair coverage)": ("Usable-pair coverage the roster loses if this vendor alone is dropped.", "q3b masks"),
    "% of visit-days co-held by our free logs": ("Share of the vendor's (ip,domain,date) visit-days that guid_log or augmentor ALSO delivered - data we pay for but already have (free logs do not preempt paid credit; Sean Yang 2026-07-13).", "q3c"),
    "Recoverable $/yr if free logs preempt credit (AUDI-1093, exact visit grain)": ("Bill x visit-day free-cohold share: what a free-source-preemption rule recovers WITHOUT dropping the vendor (keeps their unique contribution). Roster total ~$274K/yr.", "q3c/q0"),
    "Post-preemption bill $/yr (AUDI-1093 applied)": ("The bill IF free logs preempted co-held credit: today's run-rate x (1 - free-cohold share). Cost-only change - the vendor's unique value is untouched by construction. Roster: $812K -> $539K (-33.7%). Nobody flips on the portfolio lens; see the decisions POST-PREEMPTION block.", "q3c/q0"),
    "Unique visit-days delivered (ip x domain x date, 30d)": ("Distinct visit events at the TRUE value grain - each (IP, domain, DATE) triple; a new date on a known pair is a distinct visit.", "q3c"),
    "% visit-days sole — new pair": ("Visit-days on pairs ONLY this vendor holds - brand-new signal (strongest contribution).", "q3c"),
    "% visit-days sole — recency refresh": ("Dates unique to this vendor on pairs other sources also hold - refreshes the 30d recency window MM scores on; the meter pays per (ip,url,day) so these ARE billed.", "q3c"),
    "% visit-days duplicated same-day": ("Same (ip,domain,date) delivered by another source too - pure duplication, worth ~zero.", "q3c"),
    "Composite quality score (curved, best=100)": ("100 x (0.40 unique value + 0.15 non-redundancy + 0.15 signal quality + 0.10 dependency + 0.20 performance), curved to best-in-roster = 100.", "q9b formula"),
    "Composite quality score (raw)": ("Same before curving.", "q9b formula"),
    "Verdict": ("The call. Full reasoning on the notes sheet and decisions sheet.", "eval"),
    "Asks / weird things (full text in notes)": ("What to demand from the vendor - full text on notes/decisions sheets.", "eval"),
}
DEF["Renewal status"] = DEF["Renewal status"]
DEF[f"Billed domains ({BILL_MONTH})"] = ("Distinct domains credited on the meter in the bill month (closest measure of 'domains we pay for').", "q1d")
DEF[f"Meter bill ({BILL_MONTH})"] = ("What the meter actually charged that month (usage_reporting_data month-end snapshot).", "q0")
DEF[f"Run-rate $/yr ({BILL_MONTH} x12)"] = ("That bill annualized.", "q0")

# ---- SOLO sheet definitions (non-copy rows only; copied rows share the numbers-sheet DEF) ----
SOLO_DEF = {
    "% of rows used — share of {vendor+free} world": ("Vendor's slice of used rows if the world were just {vendor + both free logs}. Sums can exceed 100% down the column only via overlap with free.", "q2c"),
    "% of IPs used — share of {vendor+free} world": ("Same, IP grain.", "q2c"),
    "% of domains used — share of {vendor+free} world": ("Same, domain grain.", "q2c"),
    "Solo usable IPs (vs free logs only)": ("Usable IPs the vendor delivers that NEITHER free log delivers (other paid vendors ignored). The IP-grain solo stock - measured, not mask-derivable.", "q8a"),
    "% of usable IPs solo": ("Solo IPs / its usable IPs.", "q8a/q3"),
    "Solo domains (vs free logs only)": ("Delivered domains neither free log delivers (all parsed domains, q4B grain).", "q8a"),
    "Solo CLASSIFIED domains (fee-band axis, vs free)": ("Solo domains present in wcv - the durable classifier-coverage unit under the solo counterfactual.", "q8a"),
    "% usable pairs solo (vs free logs only)": ("Usable pairs held by the vendor and no free log - EXACT from q3b masks; equals q3 net-new-vs-free (anchor).", "q3b masks"),
    "% co-held-with-free pairs where vendor fresher": ("Of raw pairs BOTH the vendor and a free log hold: vendor's MAX(dt) strictly newer - its recency contribution on shared ground.", "q8a"),
    "% co-held-with-free pairs tied same-day": ("Same MAX(dt) as the free log.", "q8a"),
    "% co-held-with-free pairs where free fresher": ("Free log newer - the vendor's copy is stale.", "q8a"),
    "Pair-coverage gain over free-only (pp)": ("Usable-pair coverage {free + vendor} adds over free logs alone (free-only = 60.4%).", "q3b masks"),
    "Solo share of net-of-free universe": ("Share of the 2.37B net-of-free pairs the vendor reaches alone = the ladder's standalone coverage.", "q3b masks"),
    "% visit-days solo — new pair (vs free)": ("Its (ip,domain,date) triples on pairs no free log holds at all - brand-new signal vs our own logs.", "q8a"),
    "% visit-days solo — refresh of free-held pair": ("Dates the vendor alone delivered on pairs a free log also holds - recency refreshes we could not have collected that day.", "q8a"),
    "% visit-days duplicated same-day with free logs": ("Same (ip,domain,date) a free log also captured - duplication vs our own collection.", "q8a"),
    "Solo IPs served-won (vs free logs only)": ("Served IPs in the SOLO cohort (vendor delivered, neither free log did; other paid vendors ignored). Superset of the sole cohort by construction.", "q8b"),
    "% of solo stock served-won": ("Served solo IPs / solo usable stock.", "q8b/q8a"),
    "Served IPs shared with free logs": ("Its served IPs a free log also delivered = replaceable by our own collection.", "q6/q8b"),
    "% of served IPs shared with free logs": ("Share of its served footprint our free logs already cover.", "q6/q8b"),
    "Solo won bids / wk": ("Won impressions on solo-cohort IPs per week - the media flow only this vendor (vs our own logs) enabled.", "q8b"),
    "Solo won bids per served solo IP": ("Frequency on those households.", "q8b"),
    "Solo won bids annualized (x52)": ("Yearly expected if the weekly flow persists.", "q8b"),
    "% HI — share of {vendor+free} world": ("HI-pool slice within {vendor + free logs}.", "q5"),
    "% PP — share of {vendor+free} world": ("PP-pool slice within {vendor + free logs}.", "q5"),
    "HI 10000 count — SOLO IPs (vs free logs only)": ("Served HI IPs the vendor delivered that no free log did - EXACT from q3d masks (usable-domain lens). The q8b-measured % rows use raw membership (q5/q6 convention) and run ~3-5% lower - free-log webmail sightings count only there.", "q3d masks"),
    "% HI of solo served IPs": ("HI share of the solo served cohort.", "q8b"),
    "PP 8000 count — SOLO IPs (vs free logs only)": ("Served PP IPs no free log delivered.", "q3d masks"),
    "% PP of solo served IPs": ("PP share of the solo served cohort.", "q8b"),
    "Avg household score — solo (scored imps)": ("Mean household_score on solo-cohort scored impressions (hs>0 only).", "q8b"),
    "% imps scored — solo": ("Scored share of solo-cohort impressions.", "q8b"),
    "Spend (media $) — solo (vs free logs only)": ("Weekly media on solo-cohort IPs - the solo counterfactual's dependent-revenue flow.", "q8b"),
    "Impressions (won bids) — solo": ("Weekly won imps, solo cohort.", "q8b"),
    "Visits — solo": ("Clickpass visits on solo-cohort serves (q7b join pattern).", "q8b"),
    "Visits — solo, 95% CI": ("Poisson (Garwood) 95% interval.", "q8b"),
    "Conversions — solo": ("Attributed conversions (q7c dedup) on solo-cohort serves.", "q8b"),
    "Revenue — solo": ("order_amt on those conversions.", "q8b"),
    "CPM — solo": ("Solo media / solo imps x 1000.", "q8b"),
    "IVR — solo": ("Solo visits / solo imps.", "q8b"),
    "IVR — solo, x of 0.0223% baseline": ("Multiple of the no-svs-data baseline.", "q8b"),
    "CVR (conv/visits) — solo": ("Conversions per solo visit.", "q8b"),
    "AOV — solo": ("Revenue / conversions.", "q8b"),
    "ROAS — solo": ("Solo revenue / solo media.", "q8b"),
    "Solo bill $/yr — LOW bound (= today's run-rate)": ("If the vendor were the only paid source it keeps at least its current credit (removing competitors only adds credit) - so today's run-rate is a hard FLOOR on the solo bill.", "q0"),
    "Solo bill $/yr — HIGH bound (visit-day share of consumed credit)": ("CEILING estimate: max(today's bill, total metered bills x the vendor's share of paid-held visit-days). The proportional term assumes consumption distributes with visit-day holdings (all metered CPMs $0.50); it under-runs junk-billing vendors (their credit sits on domains outside the usable universe, uncontested) so the bound clamps to today's bill - only the two 33Across feeds have real solo-bill upside. Flat vendors: fee pending.", "q3c masks/q0"),
    "% of delivered rows billed — solo HIGH bound": ("High-bound billed imps (HIGH bill / $0.50 x 1000) over annual delivered rows.", "computed"),
    "Max justified CPM — on 100% of delivered rows (solo)": ("(T2_solo x 30%) spread over every delivered row.", "q8b/q1"),
    "Max justified CPM — on used/billed imps (solo, vs $0.50)": ("Same over billed/used imps - compare to the $0.50 meter.", "q8b/q1d"),
    "Flat equivalent — floor (T1_solo x 15%)": ("Flat fee justified by provable solo dependency at conservative margin.", "q8b"),
    "Flat equivalent — fair (T2_solo x 20%)": ("Fair flat fee on the solo dependency ceiling.", "q8b"),
    "Flat equivalent — ceiling (T2_solo x 30%)": ("Never-pay-more line under the solo counterfactual.", "q8b"),
    "T2_solo dependent revenue $/yr (solo-won media x52, MEASURED)": ("THE solo decision number: measured media on solo-cohort won imps, annualized. Sits between T2-sole (measured, vs all sources) and the ladder's density-estimated standalone.", "q8b"),
    "T2_solo envelope low (x0.4)": ("Stress band bottom.", "computed"),
    "T2_solo envelope high (x1.8)": ("Stress band top.", "computed"),
    "T1_solo provable floor $/yr (score-gated)": ("Media on solo imps where a high score gated the serve.", "q8b"),
    "Fee band, domain axis — low (solo classified x $3)": ("Domain-axis worth on SOLO classified domains.", "q8a"),
    "Fee band, domain axis — high (solo classified x $13)": ("Same at the $13 ceiling.", "q8a"),
    "T2_solo $/yr — DENSITY ESTIMATE (= ladder standalone)": ("The decisions-ladder standalone number reproduced on this sheet: solo pairs x measured T2-per-sole-pair density. Compare with the MEASURED row above - the gap is the density-extrapolation error (33Across precedent: $397K est vs $270K measured-sole).", "q3b masks/q6"),
    "SOLO PAY RANGE $/yr (10-30% margin on T2_solo; density est until scan lands)": ("What to actually pay: 10-30% blended margin on T2_solo revenue. THE negotiating range under the solo counterfactual.", "q8b"),
    "Pair coverage under {free + vendor} (% of full roster)": ("Usable-pair coverage if the roster were just this vendor + free logs (the scenario-table lens, per vendor).", "q3b masks"),
    "Visit-day coverage under {free + vendor}": ("Same at (ip,domain,date) grain.", "q3c masks"),
    "Visit-day coverage gain over free-only (pp)": ("Visit-day coverage added over free logs alone (free-only = 59.4%).", "q3c masks"),
    "HI-IP coverage under {free + vendor}": ("Share of all HI IPs reachable with just this vendor + free logs (free-only = 99.94%).", "q3d masks"),
    "PP-IP coverage under {free + vendor}": ("Share of all PP IPs (free-only = 99.25%).", "q3d masks"),
    "Solo visit-days (vs free logs only)": ("Visit-day triples held by the vendor and no free log (count).", "q3c masks"),
    "Share of paid-held visit-days (solo-bill HIGH-bound driver)": ("The vendor's share of all visit-days any paid vendor holds - the multiplier behind the solo-bill HIGH bound.", "q3c masks"),
    "Solo worth / bill — MEASURED (T2_solo vs bill LOW-HIGH)": ("Measured T2_solo divided by the solo-bill range (worst case first). >1x = would cover even the high-bound bill on a revenue basis.", "q8b/q0/q3c"),
    "Post-preemption SOLO bill (bounds x (1 - free-cohold))": ("The solo-bill LOW-HIGH bounds with the AUDI-1093 free-cohold share removed - what the solo counterfactual would bill if free logs preempted co-held credit.", "q3c/q0"),
    "Solo worth / bill — DENSITY EST (vs today's bill)": ("Density-estimated standalone revenue / today's run-rate (the ladder's standalone worth/bill, reproduced per vendor).", "q3b masks/q0"),
}

# WTP bands from the eval index (q9b chart) - the official pay-up-to ranges.
WTP = {25: "$150K-600K", 26: "$0.7M-3M", 28: "$30K-100K", 24: "$14K-60K",
       39: "$0.1K-1.5K", 40: "$10K-40K", 36: "$1.1K-4.7K", 33: "$0.5K-2.4K"}

NEGOTIATION = {
    28: "ONE combined deal with DS40 (~$598K/yr today): cap the pair at <=$130-140K/yr, i.e. CPM $0.50 -> ~$0.10-0.15 or a billing cap. Leverage: 53% of its credits re-race to our FREE logs; 29% of feed is DS13-blocklisted webmail; 6.4% Googlebot IPs we currently pay for; augmentor displacement is already eroding the bill month-over-month.",
    40: "Fold into the DS28 negotiation - same vendor. Standalone worth $10-40K/yr; its billed domains are cookie-sync infra junk (strong discount evidence).",
    33: "Do NOT renegotiate - drop. Fair value $0.5-2.4K/yr vs $116K bill; a ~98% discount is not a negotiation. Exact recovery $109K/yr (sequencing-safe: only 5.9% re-races to other meters). Bug report available if they want to re-pitch after fixing URLs.",
    24: "Keep, trim toward the $14-60K band (bill $77K/yr): ask ~20-25% rate cut or a monthly cap. Cleanest feed on the roster, 91.6% sole pairs - worth keeping at the right price.",
    36: "Drop ($21.6K/yr bill vs $1.1-4.7K band). 700x below siblings in scale. Needs the ENABLED_DSIDS config change (Data Eng).",
    25: "Keep; lock the flat price at renewal. TI-1027 fair value $150-600K/yr - accept anything inside the band. Ask for URL paths + user_agent (domain-only feed today).",
    26: "Keep; lock the flat price. Band $0.7-3M/yr - #1 unique classified domain contributor. HEM feeds CRM/identity in PROD: run the blast-radius check before ANY contract change.",
    39: "Renewal is LIVE: pay <=$1.6K/yr (fee band top) or walk; alternatively convert to the $0.50 meter (~$2.1K/yr at current usable volume). T2 ceiling says its unique IPs made us at most ~$2.2K revenue last year-equivalent.",
}


def disp_len(v, fmt):
    if v is None:
        return 1
    if isinstance(v, str):
        return len(v)
    v = float(v)
    if fmt == "int":
        t = f"{v:,.0f}"
    elif fmt and fmt.startswith("pct"):
        d = {"pct0": 0, "pct": 1, "pct2": 2, "pct3": 3, "pct4": 4}[fmt]
        t = f"{v * 100:,.{d}f}%"
    elif fmt == "usd":
        t = f"${v:,.0f}"
    elif fmt == "usd2":
        t = f"${v:,.2f}"
    elif fmt == "usd4":
        t = f"${v:,.4f}"
    elif fmt == "x2":
        t = f"{v:.2f}x"
    elif fmt in ("dec1", "dec2"):
        t = f"{v:.1f}" if fmt == "dec1" else f"{v:.2f}"
    else:
        t = str(v)
    return len(t)


def main():
    import openpyxl
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    solo_anchor_checks()

    wb = openpyxl.Workbook()
    idx = wb.active
    idx.title = "index"
    dec = wb.create_sheet("decisions")
    wt = wb.create_sheet("waste")
    ws = wb.create_sheet("numbers")
    ss = wb.create_sheet("solo")
    ns = wb.create_sheet("notes")

    bold = Font(bold=True)
    section_fill = PatternFill("solid", fgColor="1F3864")
    section_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    band_fill = PatternFill("solid", fgColor="F2F2F2")
    verdict_color = {"KEEP": "1E7A1E", "NEGO": "B26B00", "DROP": "B00020"}

    # ================= numbers + solo + waste (same grid renderer) =================
    def render_grid(sheet, spec, cols=DS_COLS, max_w=36):
        sheet.cell(row=1, column=1, value="Question").font = bold
        for i, d in enumerate(cols):
            c = sheet.cell(row=1, column=2 + i, value=HDR_NAMES[d])
            c.font = bold
            c.alignment = Alignment(horizontal="right")

        r = 1
        for label, fmt, fn, oos_ok in spec:
            r += 1
            a = sheet.cell(row=r, column=1, value=label)
            if fn is None:
                a.font = section_font
                a.fill = section_fill
                for i in range(len(cols)):
                    sheet.cell(row=r, column=2 + i).fill = section_fill
                continue
            for i, d in enumerate(cols):
                cell = sheet.cell(row=r, column=2 + i)
                cell.alignment = Alignment(horizontal="right")
                if d in OOS and not oos_ok:
                    cell.value = NA
                    continue
                try:
                    v = fn(d)
                except Exception:
                    v = None
                if v is None:
                    cell.value = NA
                elif isinstance(v, str):
                    cell.value = v
                    if len(v) > max_w:
                        cell.alignment = Alignment(horizontal="left", vertical="top",
                                                   wrap_text=True)
                else:
                    v = float(v)
                    if fmt and fmt.startswith("pct"):
                        v /= 100.0
                    cell.value = round(v, 8)
                    if FMT.get(fmt):
                        cell.number_format = FMT[fmt]

        RED, YEL, GRN = "F8696B", "FFEB84", "63BE7B"
        last_col = get_column_letter(1 + len(cols))
        r = 1
        for label, fmt, fn, oos_ok in spec:
            r += 1
            if fn is None:
                continue
            rng = f"B{r}:{last_col}{r}"
            if label in NEUTRAL:
                sheet.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="min", start_color="EFF6FC",
                    end_type="max", end_color="7FB2E5"))
                continue
            direction = DIR.get(label)
            if not direction:
                continue
            lo, hi = (RED, GRN) if direction > 0 else (GRN, RED)
            sheet.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color=lo,
                mid_type="percentile", mid_value=50, mid_color=YEL,
                end_type="max", end_color=hi))

        sheet.column_dimensions["A"].width = max(len(x[0]) for x in spec) + 3
        for i, d in enumerate(cols):
            m = len(HDR_NAMES[d])
            r2 = 1
            for label, fmt, fn, oos_ok in spec:
                r2 += 1
                if fn is None:
                    continue
                m = max(m, disp_len(sheet.cell(row=r2, column=2 + i).value, fmt))
            # cap: long text wraps inside the cell instead of blowing the column wide
            sheet.column_dimensions[get_column_letter(2 + i)].width = min(round(m * 1.1) + 2, max_w)
        r2 = 1
        for label, fmt, fn, oos_ok in spec:
            r2 += 1
            if fn is None:
                continue
            lines = 1
            for i in range(len(cols)):
                v = sheet.cell(row=r2, column=2 + i).value
                if isinstance(v, str) and len(v) > max_w:
                    lines = max(lines, -(-len(v) // (max_w - 2)))
            if lines > 1:
                sheet.row_dimensions[r2].height = min(lines, 10) * 13 + 4
        sheet.freeze_panes = "B2"

    render_grid(ws, SPEC)
    render_grid(ss, SOLO_SPEC)
    render_grid(wt, WASTE_SPEC, WASTE_COLS)

    # ================= index =================
    idx_hdr = ["Section", "Question", "What it means / how computed", "Source"]
    idx.append(idx_hdr)
    for c in range(1, 5):
        cell = idx.cell(row=1, column=c)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = border
    missing_defs = []
    section = ""
    ri = 1
    band = False
    for label, fmt, fn, oos_ok in SPEC:
        if fn is None:
            section = label
            band = not band
            continue
        meaning, src = DEF.get(label, ("", ""))
        if not meaning:
            missing_defs.append(label)
        ri += 1
        vals = [section, label, meaning, src]
        for c, v in enumerate(vals, start=1):
            cell = idx.cell(row=ri, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if band:
                cell.fill = band_fill
        idx.cell(row=ri, column=2).font = bold

    # ---- SOLO sheet: banner + definitions for its recomputed (non-copy) rows ----
    ri += 1
    ban = idx.cell(row=ri, column=1, value=(
        "SOLO SHEET — counterfactual: each column assumes that vendor is the ONLY paid source; "
        "free logs (guid_log + augmentor) always kept. 'Solo' = touched by the vendor and by "
        "NEITHER free log (free-log columns: vs the OTHER free log). Differs from 'sole' "
        "(vs ALL 9 other sources) and equals the decisions-ladder 'standalone' at pair grain. "
        "Rows below = only the solo sheet's recomputed rows; all others copy the numbers sheet."))
    idx.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4)
    ban.font = section_font
    ban.fill = section_fill
    ban.alignment = Alignment(vertical="top", wrap_text=True)
    idx.row_dimensions[ri].height = 42
    section = ""
    band = False
    for label, fmt, fn, oos_ok in SOLO_SPEC:
        if fn is None:
            section = label
            band = not band
            continue
        if SOLO_KIND.get(label) == "copy":
            continue
        meaning, src = SOLO_DEF.get(label, ("", ""))
        if not meaning:
            missing_defs.append("SOLO: " + label)
        ri += 1
        vals = ["SOLO: " + section, label, meaning, src]
        for c, v in enumerate(vals, start=1):
            cell = idx.cell(row=ri, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if band:
                cell.fill = band_fill
        idx.cell(row=ri, column=2).font = bold

    # ---- WASTE sheet: banner + definitions ----
    ri += 1
    ban2 = idx.cell(row=ri, column=1, value=(
        "WASTE SHEET — how much delivered vendor data we throw away, why, what it costs to "
        "ingest/store, and what to ask each vendor to STOP sending. '% thrown away' is exact "
        "(1 - used share, q2c); the reason rows are indicative single-day shares and OVERLAP "
        "(a blocklisted webmail row can still be USED - and billed - via DS19). Storage floor "
        "= GCS list price on the measured no-TTL footprint; Kafka/DAG/classifier compute needs "
        "Data Eng numbers and is NOT included."))
    idx.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4)
    ban2.font = section_font
    ban2.fill = section_fill
    ban2.alignment = Alignment(vertical="top", wrap_text=True)
    idx.row_dimensions[ri].height = 42
    section = ""
    band = False
    for label, fmt, fn, oos_ok in WASTE_SPEC:
        if fn is None:
            section = label
            band = not band
            continue
        meaning, src = WASTE_DEF.get(label, ("", ""))
        if not meaning:
            continue
        ri += 1
        vals = ["WASTE: " + section, label, meaning, src]
        for c, v in enumerate(vals, start=1):
            cell = idx.cell(row=ri, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if band:
                cell.fill = band_fill
        idx.cell(row=ri, column=2).font = bold

    sec_w = max(len(x[0]) for x in SPEC if x[2] is None) + 3 + len("SOLO: ")
    q_w = max(len(x[0]) for x in list(SPEC) + list(SOLO_SPEC) if x[2] is not None) + 3
    src_w = max(len(v[1]) for v in list(DEF.values()) + list(SOLO_DEF.values())) + 3
    for i, w in enumerate([sec_w, q_w, 95, max(src_w, len("Source") + 2)], start=1):
        idx.column_dimensions[get_column_letter(i)].width = w
    idx.freeze_panes = "A2"

    # ================= decisions =================
    def put_row(sheet, ri, vals, fmts=None, header=False, band_row=False):
        for c, v in enumerate(vals, start=1):
            cell = sheet.cell(row=ri, column=c, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header:
                cell.font = section_font
                cell.fill = section_fill
            elif band_row:
                cell.fill = band_fill
            if fmts and not header and isinstance(v, (int, float)):
                f = fmts.get(c)
                if f:
                    cell.number_format = f
        return ri + 1

    dep_risk = {d: q6b[d]["prosp_media"] * 52 for d in q6b}

    dec.merge_cells("A1:M1")
    t = dec.cell(row=1, column=1, value=(
        f"VENDOR DECISIONS — ranked by composite score. Bills = {BILL_MONTH} x12. "
        "Free logs (guid, augmentor) always kept. DS28+DS40 = ONE vendor: negotiate combined. "
        "Savings VALIDATED 2026-07-13 (AUDI-1092): meter is single-credit since 2026-05; "
        "figures exact under first-reporter, conservative floors under free-priority."))
    t.font = section_font
    t.fill = section_fill
    dec.row_dimensions[1].height = 20

    hdr = ["Vendor", "DS", "Score (best=100)", "Bill $/yr", "Pay up to (WTP band)",
           "T1 provable floor $/yr", "T2 fair-ceiling $/yr", "Drop savings $/yr (validated; exact-to-floor)",
           "Coverage lost (pp)", "Dep. revenue at risk $/yr", "DECISION",
           "Negotiation plan / target", "Top asks to improve their data"]
    ri = put_row(dec, 2, hdr, header=True)
    dfmt = {3: "0", 4: "$#,##0", 6: "$#,##0", 8: "$#,##0", 9: "0.00", 10: "$#,##0"}
    for n, d in enumerate(sorted(EXT, key=lambda x: -CURVE[x][1])):
        jb = q0.get(d, {}).get("june_usd")
        bill = jb * 12 if jb is not None else "pending (flat)"
        t2 = t2_ann(d)
        if d in FLAT:
            savings = "= flat fee"
        else:
            rr = reassign[d]
            savings = float(q1d[d]["billed_usd"]) * 12 * (1 - rr.get("metered", 0) / sum(rr.values()))
        ri = put_row(dec, ri, [
            HDR_NAMES[d], d, round(CURVE[d][1]), bill, WTP[d], round(t1_ann(d)),
            f"{money(t2 * 0.20)} - {money(t2 * 0.30)}", savings,
            round(-coverage_lost[d], 2), round(dep_risk.get(d, 0)),
            VERDICT_SHORT[d], NEGOTIATION[d], ASKS[d],
        ], fmts=dfmt, band_row=(n % 2 == 1))
        vcell = dec.cell(row=ri - 1, column=11)
        vword = str(vcell.value or "")[:4].upper()
        if vword in verdict_color:
            vcell.font = Font(bold=True, color=verdict_color[vword])

    def saved(dropped):
        # metered->metered deduction only applies while a metered DESTINATION survives
        metered_left = [x for x in METERED if x not in dropped]
        out = 0.0
        for d in dropped:
            if d not in reassign or d not in q1d:
                continue
            rr = reassign[d]
            sh = 1 - rr.get("metered", 0) / sum(rr.values())
            if d in (28, 40) and 28 in dropped and 40 in dropped:
                sh = 1.0
            if not metered_left:
                sh = 1.0
            out += float(q1d[d]["billed_usd"]) * 12 * sh
        return out

    ri += 1
    dec.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=13)
    t = dec.cell(row=ri, column=1, value=(
        "ROSTER SCENARIOS — every keep-set evaluated EXACTLY from the q3b holder masks "
        "(coverage = usable (IP,domain) pairs reachable; free logs always kept)"))
    t.font = section_font
    t.fill = section_fill
    ri += 1
    sc_hdr = ["Scenario", "Paid vendors kept", "Coverage (% of today)",
              "HI-IP coverage %", "PP-IP coverage %",
              "DS19-only pair coverage % (MM Core / Max Reach universe)",
              "DS19-only visit-day coverage %",
              "Metered bills kept $/yr", "Metered recovery from drops $/yr (validated exact-to-floor; flat-fee savings ADDITIONAL, amounts pending)",
              "Dep. revenue at risk $/yr (T2, prospecting-attributed)", "", "", ""]
    ri = put_row(dec, ri, sc_hdr, header=True)
    sfmt = {3: "0.00%", 4: "0.000%", 5: "0.000%", 6: "0.0%", 7: "0.0%",
            8: "$#,##0", 9: "$#,##0", 10: "$#,##0"}

    def scov(tier_masks, keep, fm=FREE_MASK):
        km = fm | sum(1 << BITSQ[d] for d in keep)
        tot = sum(tier_masks.values())
        return sum(n for m, n in tier_masks.items() if m & km) / tot if tot else 0
    ALL8 = [24, 25, 26, 28, 33, 36, 39, 40]
    SC = [
        ("Today (all 8)", ALL8),
        ("Drop Sovrn + Cybba", [24, 25, 26, 28, 39, 40]),
        ("+ drop Klickly", [24, 25, 26, 28, 40]),
        ("+ drop Justuno (knee k=4)", [25, 26, 28, 40]),
        ("33Across combined only", [28, 40]),
        ("Flat-fee only (5x5 + Predactiv)", [25, 26]),
        ("Free logs only (guid + augmentor)", []),
        ("Free: augmentor DS30 only", [], 1 << 5),
        ("Free: guid_log DS23 only", [], 1 << 0),
    ]
    for n, sc_row in enumerate(SC):
        label, keep = sc_row[0], sc_row[1]
        fm = sc_row[2] if len(sc_row) > 2 else FREE_MASK
        dropped = [d for d in ALL8 if d not in keep]
        met_kept = sum(float(q1d[d]["billed_usd"]) * 12 for d in keep if d in q1d and q1d[d].get("billed_usd"))
        ri = put_row(dec, ri, [
            label, " + ".join(SHORT[d] for d in keep) if keep else "(none)",
            cov(keep, fm) / FULL_COV, scov(hi3d, keep, fm), scov(pp3d, keep, fm),
            ds19_cov(keep, q13a_pair, fm) if Q13A else PENDING13,
            ds19_cov(keep, q13a_trip, fm) if Q13A else PENDING13,
            round(met_kept) if met_kept else 0,
            round(saved(dropped)), round(sum(dep_risk.get(d, 0) for d in dropped)),
            "", "", "",
        ], fmts=sfmt, band_row=(n % 2 == 1))

    # ---- net-of-free value ladder ----
    dens = DENS
    dens_t = {d: t2_ann(d) / (vend3c[d]["sole_new_pair"] + vend3c[d]["sole_refresh"])
              for d in EXT if d in vend3c and d in q6}
    ri += 1
    dec.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=13)
    t = dec.cell(row=ri, column=1, value=(
        f"NET-OF-FREE VALUE LADDER — guid_log + augmentor ALWAYS excluded from the universe first "
        f"({NOF_U / 1e9:.2f}B pairs = {100 * NOF_U / FULL_COV:.0f}% of usable that free logs do NOT cover). "
        "Standalone = vendor as the ONLY paid source; marginal = what it adds at its ladder position. "
        "$ = DEPENDENT REVENUE, the FULL media CPM on those serves - not our cut. PAY RANGE columns apply the 10-30% blended margin (avg ~20%): that is what the data is worth TO US and the number to negotiate with. Worth/bill ratios stay revenue-basis (generous to the vendor)."))
    t.font = section_font
    t.fill = section_fill
    ri += 1
    lad_hdr = ["Step", "Vendor", "Standalone net-of-free pairs", "Standalone % of universe",
               "Marginal pairs at this step", "Marginal visit-days (recency-crediting grain)",
               "Cumulative % of universe", "Bill $/yr",
               "STANDALONE dependent REVENUE $/yr (vs free logs only)",
               "STANDALONE PAY RANGE $/yr (10-30% margin)", "STANDALONE worth / bill (revenue basis)",
               "MARGINAL dependent REVENUE $/yr (pair density)",
               "MARGINAL dependent REVENUE $/yr (visit-day density — credits refreshes)",
               "MARGINAL PAY RANGE $/yr (10-30% margin)", "MARGINAL worth / bill (revenue basis)"]
    ri = put_row(dec, ri, lad_hdr, header=True)
    lfmt = {3: "#,##0", 4: "0.0%", 5: "#,##0", 6: "#,##0", 7: "0.0%", 8: "$#,##0",
            9: "$#,##0", 11: '0.00"x"', 12: "$#,##0", 13: "$#,##0", 15: '0.00"x"'}
    order, cur_p, cur_t = [], 0, 0
    marg_val = {}
    rem = [d for d in EXT if d in dens]
    for step in range(1, len(rem) + 1 + len(order)):
        if not rem:
            break
        nxt = max(rem, key=lambda d: nof_cov(order + [d], NOF_P))
        newp = nof_cov(order + [nxt], NOF_P)
        newt = nof_cov(order + [nxt], NOF_T)
        mp, mt = newp - cur_p, newt - cur_t
        mval = mp * dens[nxt]
        jb = q0.get(nxt, {}).get("june_usd")
        bill = jb * 12 if jb is not None else "flat (pending)"
        ratio = (mval / bill) if isinstance(bill, float) else NA
        mval_t = mt * dens_t[nxt]
        aval = nof_cov([nxt], NOF_P) * dens[nxt]
        aratio = (aval / bill) if isinstance(bill, float) else NA
        ri = put_row(dec, ri, [
            step, HDR_NAMES[nxt], nof_cov([nxt], NOF_P), nof_cov([nxt], NOF_P) / NOF_U,
            mp, mt, newp / NOF_U, bill if isinstance(bill, float) else bill,
            round(aval), f"{money(aval * 0.10)} - {money(aval * 0.30)}", aratio,
            round(mval), round(mval_t), f"{money(mval * 0.10)} - {money(mval * 0.30)}", ratio,
        ], fmts=lfmt, band_row=(step % 2 == 0))
        order.append(nxt)
        rem.remove(nxt)
        marg_val[nxt] = mval
        cur_p, cur_t = newp, newt

    # ---- post-preemption economics (AUDI-1093 applied) ----
    ri += 1
    dec.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=13)
    t = dec.cell(row=ri, column=1, value=(
        "POST-PREEMPTION ECONOMICS — bills IF free logs preempt co-held credit (AUDI-1093 fix): "
        "the meter stops paying for (ip,domain,DATE) visit-days guid_log/augmentor also captured; "
        "vendors KEEP their unique data, so the PAY RANGES are unchanged by construction. Visit "
        "grain = fair version (vendor still credited for fresher dates; strict pair-grain barely "
        "larger, ~$284K proxy). Combined-33Across ceiling range is <= the sum (solo cohorts "
        "overlap). VERDICT: a cost fix, not a valuation flip — stack with renegotiation."))
    t.font = section_font
    t.fill = section_fill
    ri += 1
    pp_hdr = ["Vendor", "Bill today $/yr", "Preemption cut $/yr", "Cut % of bill",
              "Bill AFTER $/yr", "PAY RANGE sole-T2 (portfolio lens)",
              "PAY RANGE marginal (ladder position)", "PAY RANGE solo MEASURED (ceiling)",
              "Worth/bill AFTER (portfolio, pay-top)", "Worth/bill AFTER (ceiling, pay-top)",
              "VERDICT FLIP?", "", ""]
    ri = put_row(dec, ri, pp_hdr, header=True)
    pfmt = {2: "$#,##0", 3: "$#,##0", 4: "0.0%", 5: "$#,##0",
            9: '0.00"x"', 10: '0.00"x"'}
    FLIP_NOTE = {
        28: "NO on portfolio — but lands INSIDE its ceiling range near the top; preemption + renegotiation together reach fair",
        40: "NO on portfolio — lands exactly AT its ceiling top; same combined negotiation as DS28",
        33: "NO — cut is negligible (0.2%: its credit is junk/unique, not free-coheld); still DROP",
        24: "NO — cut 4.9% barely moves it; still DROP/trim",
        36: "NO — still ~5x the ceiling top; still DROP",
    }
    pp_rows = [28, 40, 33, 24, 36]
    for n, d in enumerate(pp_rows):
        bill = q0[d]["june_usd"] * 12
        cut = preempt_cut(d)
        after = bill - cut
        t2 = t2_ann(d)
        mv = marg_val.get(d, 0)
        t2s = t2_solo(d) if Q8B else None
        ceil_ratio = (t2s * 0.30 / after) if t2s else NA
        ri = put_row(dec, ri, [
            HDR_NAMES[d], round(bill), round(cut), cut / bill, round(after),
            f"{money(t2 * 0.10)} - {money(t2 * 0.30)}",
            f"{money(mv * 0.10)} - {money(mv * 0.30)}",
            f"{money(t2s * 0.10)} - {money(t2s * 0.30)}" if t2s else "pending scan (q8)",
            t2 * 0.30 / after, ceil_ratio,
            FLIP_NOTE[d], "", "",
        ], fmts=pfmt, band_row=(n % 2 == 1))
    b_pair = q0[28]["june_usd"] * 12 + q0[40]["june_usd"] * 12
    c_pair = preempt_cut(28) + preempt_cut(40)
    a_pair = b_pair - c_pair
    t2_pair, mv_pair = t2_ann(28) + t2_ann(40), marg_val.get(28, 0) + marg_val.get(40, 0)
    t2s_pair = (t2_solo(28) + t2_solo(40)) if Q8B else None
    ri = put_row(dec, ri, [
        "33Across COMBINED (one vendor)", round(b_pair), round(c_pair), c_pair / b_pair,
        round(a_pair), f"{money(t2_pair * 0.10)} - {money(t2_pair * 0.30)}",
        f"{money(mv_pair * 0.10)} - {money(mv_pair * 0.30)}",
        (f"<= {money(t2s_pair * 0.10)} - {money(t2s_pair * 0.30)} (cohorts overlap)"
         if t2s_pair else "pending scan (q8)"),
        t2_pair * 0.30 / a_pair, (t2s_pair * 0.30 / a_pair) if t2s_pair else NA,
        "the negotiation unit: bill lands at the TOP of the ceiling pay range after preemption",
        "", "",
    ], fmts=pfmt)
    tot_bill = sum(q0[d]["june_usd"] * 12 for d in pp_rows)
    tot_cut = sum(preempt_cut(d) for d in pp_rows)
    ri = put_row(dec, ri, [
        "ROSTER TOTAL (metered)", round(tot_bill), round(tot_cut), tot_cut / tot_bill,
        round(tot_bill - tot_cut), NA, NA, "non-additive (see notes)", NA, NA,
        "cost fix, not a valuation flip — flats unaffected (no meter)", "", "",
    ], fmts=pfmt)

    widths = [30, 32, 13, 14, 17, 14, 18, 16, 13, 14, 21, 64, 60]
    for i, w in enumerate(widths, start=1):
        dec.column_dimensions[get_column_letter(i)].width = w
    dec.freeze_panes = "A3"

    # ================= notes =================
    hdr = ["Vendor", "DS", "Scope", "Billing / rate", "Renewal / contract status",
           "Ingestion + off-switch", "Blast radius (non-MM prod deps)",
           "Verdict (full)", "Asks / weird things to raise with the vendor"]
    ns.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ns.cell(row=1, column=c)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    ns.row_dimensions[1].height = 22

    for i, d in enumerate(DS_COLS):
        reg = q0.get(d, {}).get("reg", {})
        vf = (reg.get("valid_from") or "")[:10]
        renewal = RENEWAL_NOTE.get(
            d, f"Contract valid_from {vf}; renewal date pending (Maya / renewal schedule)" if vf else "pending")
        vals = [HDR_NAMES[d], d, SCOPE.get(d, ""), RATE_NOTE.get(d, ""), renewal,
                INGEST_NOTE.get(d, ""), BLAST_NOTE.get(d, ""),
                VERDICT_FULL.get(d, NA), ASKS.get(d, NA)]
        r_i = 2 + i
        for c, v in enumerate(vals, start=1):
            cell = ns.cell(row=r_i, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if i % 2 == 1:
                cell.fill = band_fill
        ns.cell(row=r_i, column=1).font = bold
        vcell = ns.cell(row=r_i, column=8)
        vword = str(vcell.value or "")[:4].upper()
        if vword in verdict_color:
            vcell.font = Font(bold=True, color=verdict_color[vword])

    conv_hdr_row = len(DS_COLS) + 3
    ns.cell(row=conv_hdr_row, column=1, value="CONVENTIONS — how to read this workbook")
    ns.merge_cells(start_row=conv_hdr_row, start_column=1, end_row=conv_hdr_row, end_column=len(hdr))
    hc = ns.cell(row=conv_hdr_row, column=1)
    hc.font = section_font
    hc.fill = section_fill
    ns.row_dimensions[conv_hdr_row].height = 20
    for j, ctext in enumerate(CONVENTIONS):
        r_j = conv_hdr_row + 1 + j
        ns.cell(row=r_j, column=1, value=f"{j + 1}. {ctext}")
        ns.merge_cells(start_row=r_j, start_column=1, end_row=r_j, end_column=len(hdr))
        cc = ns.cell(row=r_j, column=1)
        cc.alignment = Alignment(vertical="top", wrap_text=True)
        cc.border = border
        if j % 2 == 1:
            cc.fill = band_fill
        ns.row_dimensions[r_j].height = max(16, (len(f"{j + 1}. {ctext}") // 300 + 1) * 15 + 3)

    widths = [16, 5, 46, 34, 42, 46, 44, 62, 70]
    for i, w in enumerate(widths, start=1):
        ns.column_dimensions[get_column_letter(i)].width = w
    ns.freeze_panes = "A2"

    wb.save(OUT)
    nrows = sum(1 for x in SPEC if x[2] is not None)
    print(f"wrote {OUT}")
    print(f"sheets: index ({nrows} definitions), decisions (8 vendors + {len(SC)} scenarios), "
          f"numbers ({nrows} rows x {len(DS_COLS)}), solo ({nrows} rows, "
          f"q8a {'loaded' if Q8A else 'PENDING'} / q8b {'loaded' if Q8B else 'PENDING'}), "
          f"waste ({sum(1 for x in WASTE_SPEC if x[2] is not None)} rows x {len(WASTE_COLS)}), notes")
    print("missing index definitions:", missing_defs if missing_defs else "none")
    for name, sheet, spec, ccols in (("numbers", ws, SPEC, DS_COLS), ("solo", ss, SOLO_SPEC, DS_COLS),
                                     ("waste", wt, WASTE_SPEC, WASTE_COLS)):
        empty, pending = [], 0
        r = 1
        for label, fmt, fn, oos_ok in spec:
            r += 1
            if fn is None:
                continue
            for i, d in enumerate(ccols):
                v = sheet.cell(row=r, column=2 + i).value
                if v in (None, ""):
                    empty.append((label, d))
                elif v == PENDING:
                    pending += 1
        print(f"empty {name} cells:", empty if empty else "none")
        print(f"{name} pending-scan cells: {pending}")


if __name__ == "__main__":
    main()
